"""Regression tests for the local API and database write/query safety boundaries."""

from __future__ import annotations

import asyncio
import os
import importlib.util
import json
import shutil
import sqlite3
import sys
import tempfile
import threading
import time
import types
import unittest
import zipfile
from concurrent.futures import ThreadPoolExecutor
from contextlib import closing
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from unittest import mock

from aiohttp import FormData, web
from aiohttp.test_utils import AioHTTPTestCase

import requests

import db_scheduler
import db_access_control
import db_audit_store
import db_chart_cache
import db_excel_transfer
import db_identity_store
import db_sessions_store
import db_semantic_store
import dbquill_core as dc
import desktop_bridge
import model_baseline_contract as model_baselines
import model_profiles
import nl2db_evaluation
import timezone_release_contract as tzcontract


def _make_db(path: Path, rows: int = 1) -> None:
    with closing(sqlite3.connect(path)) as conn:
        conn.execute("CREATE TABLE items (id INTEGER PRIMARY KEY, value TEXT)")
        conn.executemany(
            "INSERT INTO items(value) VALUES (?)",
            [(f"v{i}",) for i in range(rows)],
        )
        conn.commit()


def _add_table(path: Path, name: str = "orders") -> None:
    with closing(sqlite3.connect(path)) as conn:
        conn.execute(f'CREATE TABLE "{name}" (id INTEGER PRIMARY KEY, amount REAL)')
        conn.execute(f'INSERT INTO "{name}"(amount) VALUES (12.5)')
        conn.commit()


class RuntimeDependencyTests(unittest.TestCase):
    def test_project_fingerprint_is_stable_across_text_checkout_line_endings(self):
        root = Path(__file__).resolve().parents[3]
        gate_path = root / "scripts" / "project_gate.py"
        spec = importlib.util.spec_from_file_location(
            "project_gate_for_line_ending_test",
            gate_path,
        )
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        gate = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gate)
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "source.py"
            path.write_bytes(b"first\nsecond\n")
            lf = gate._fingerprint_content(path)
            path.write_bytes(b"first\r\nsecond\r\n")
            crlf = gate._fingerprint_content(path)
            self.assertEqual(lf, crlf)

    def test_source_release_contract_is_reproducible_and_secret_free(self):
        root = Path(__file__).resolve().parents[3]
        hygiene_path = root / "scripts" / "check_repository_hygiene.py"
        spec = importlib.util.spec_from_file_location(
            "repository_hygiene_for_test",
            hygiene_path,
        )
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        hygiene = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(hygiene)
        report = hygiene.validate_repository()
        self.assertEqual(report["direct_dependencies"], 7)
        self.assertEqual(report["locked_dependencies"], 26)
        self.assertEqual(report["credential_findings"], 0)
        self.assertEqual(report["redistributed_assets_verified"], 2)

    def test_model_profile_example_is_json_and_contains_no_live_secret(self):
        example_path = Path(__file__).resolve().parents[1] / "model_profiles.example.json"
        document = json.loads(example_path.read_text(encoding="utf-8"))
        self.assertEqual(document["version"], 1)
        self.assertEqual(document["profiles"][0]["api_mode"], "chat_completions")
        self.assertEqual(document["profiles"][0]["api_key"], "replace-with-local-api-key")
        self.assertNotRegex(example_path.read_text(encoding="utf-8"), r"sk-[A-Za-z0-9_-]{20,}")

    def test_demo_database_generator_creates_consistent_resettable_fixture(self):
        spec = importlib.util.spec_from_file_location(
            "demo_database_generator_for_test",
            Path(__file__).resolve().parents[3] / "scripts/create_demo_database.py",
        )
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        generator = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(generator)
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "demo.sqlite"
            result = generator.create_database(output)
            self.assertEqual(result["dataset_version"], "2026.08.18-v1")
            self.assertEqual(result["orders"], 24)
            with closing(sqlite3.connect(output)) as conn:
                self.assertEqual(conn.execute("PRAGMA quick_check").fetchone()[0], "ok")
                self.assertEqual(conn.execute("PRAGMA foreign_key_check").fetchall(), [])
                self.assertEqual(conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0], 12)
                self.assertEqual(conn.execute("SELECT COUNT(*) FROM products").fetchone()[0], 10)
                self.assertEqual(conn.execute("SELECT COUNT(*) FROM order_items").fetchone()[0], 43)
                paid = conn.execute(
                    "SELECT COUNT(*), ROUND(SUM(total_amount), 2) "
                    "FROM orders WHERE status = 'paid'"
                ).fetchone()
                self.assertEqual(paid, (16, 52926.85))
            with self.assertRaises(FileExistsError):
                generator.create_database(output)
            reset = generator.create_database(output, force=True)
            self.assertEqual(reset["path"], str(output.resolve()))

    def test_desktop_upload_picker_only_offers_supported_excel_format(self):
        html_path = Path(__file__).parent / "desktop" / "static" / "db.html"
        html = html_path.read_text(encoding="utf-8")
        picker = next(
            line for line in html.splitlines() if 'id="fileInput"' in line
        )
        accept_value = picker.split('accept="', 1)[1].split('"', 1)[0]
        self.assertEqual(
            accept_value.split(","),
            [".db", ".sqlite", ".sqlite3", ".csv", ".xlsx"],
        )
        self.assertIn(
            "旧版 Excel .xls 暂不支持，请先转换为 .xlsx 后再上传。",
            html,
        )

    def test_desktop_language_switch_is_local_persistent_and_dynamic(self):
        static_root = Path(__file__).parent / "desktop" / "static"
        html = (static_root / "db.html").read_text(encoding="utf-8")
        css = (static_root / "calm-theme.css").read_text(encoding="utf-8")
        i18n = (static_root / "i18n.js").read_text(encoding="utf-8")

        self.assertIn('class="locale-switch"', html)
        self.assertIn('data-locale="zh-CN"', html)
        self.assertIn('data-locale="en"', html)
        self.assertIn('<html lang="en">', html)
        self.assertIn('data-locale="en" class="active"', html)
        self.assertIn('src="i18n.js?v=20260827-5"', html)
        self.assertIn("window.DBQuillI18n.start()", html)
        self.assertIn("dbquill_locale", i18n)
        self.assertIn("return 'en';", i18n)
        self.assertIn("MutationObserver", i18n)
        self.assertIn("Open-Source AI Database Agent", i18n)
        self.assertIn("'会话': 'Conversations'", i18n)
        self.assertIn("window.confirm(uiT(message))", html)
        self.assertNotIn("fetch(", i18n)
        self.assertIn(".locale-switch button.active", css)
        self.assertIn('html[lang="en"] .nav-tab', css)
        project_gate = (static_root.parents[4] / "scripts" / "project_gate.py").read_text(
            encoding="utf-8",
        )
        self.assertIn('static_root.glob("*.js")', project_gate)

    def test_desktop_upload_streams_without_renderer_base64_copies(self):
        html_path = Path(__file__).parent / "desktop" / "static" / "db.html"
        html = html_path.read_text(encoding="utf-8")
        self.assertIn("new FormData()", html)
        self.assertIn("new XMLHttpRequest()", html)
        self.assertIn("xhr.upload.onprogress", html)
        self.assertIn("正在检查数据库", html)
        self.assertIn("bridgeStatus.uploadProtocol !== 'multipart-v1'", html)
        self.assertIn("本地服务版本过旧", html)
        self.assertNotIn("new FileReader()", html)
        self.assertNotIn("readAsDataURL", html)
        self.assertNotIn("dataUrl: reader.result", html)

    def test_desktop_exposes_expandable_audit_trace_memory_and_excel_transfer(self):
        static_root = Path(__file__).parent / "desktop" / "static"
        html = (static_root / "db.html").read_text(encoding="utf-8")
        css = (static_root / "calm-theme.css").read_text(encoding="utf-8")
        self.assertIn('<details class="audit-item">', html)
        self.assertIn('class="audit-item-expanded"', html)
        self.assertIn("auditExpandedRows(event)", html)
        self.assertIn('class="run-trace"', html)
        self.assertIn('class="answer-trace"', html)
        self.assertIn('class="memory-card"', html)
        self.assertIn('id="excelImportBtn"', html)
        self.assertIn('id="excelExportBtn"', html)
        self.assertIn('/db/excel/import/prepare?', html)
        self.assertIn('/db/excel/import/confirm', html)
        self.assertIn('/db/excel/export?', html)
        self.assertIn("不含模型隐藏思维链", html)
        self.assertIn(".audit-item-expanded", css)
        self.assertIn(".memory-layer", css)
        self.assertIn(".excel-import-modal", css)

    def test_desktop_uses_single_owner_mode_without_role_management_ui(self):
        html_path = Path(__file__).parent / "desktop" / "static" / "db.html"
        html = html_path.read_text(encoding="utf-8")
        # 2026-08-20 起“本机模式”徽标按用户要求移除：单机所有者体验不再展示角色字样，
        # 鉴权与角色边界仍由服务端执行（ADR-050 行为不变，只是不再有可见徽标）
        self.assertNotIn('id="roleBadge"', html)
        self.assertNotIn("本机模式</span>", html)
        credential_panel = next(
            line for line in html.splitlines() if 'id="credentialPanel"' in line
        )
        self.assertIn(" hidden", credential_panel)
        self.assertNotIn("if (name === 'audit') { loadAudit(); loadCredentials(); }", html)
        self.assertIn("执行前仍会重新校验", html)

    def test_write_confirmation_choices_are_clickable_and_server_authoritative(self):
        html_path = Path(__file__).parent / "desktop" / "static" / "db.html"
        html = html_path.read_text(encoding="utf-8")
        self.assertIn('data-write-decision="approve"', html)
        self.assertIn('data-write-decision="cancel"', html)
        self.assertIn("button.addEventListener('click', () => confirmWrite(", html)
        self.assertNotIn(
            "(canApprove ? '' : 'disabled title=\"当前会话不能批准此操作\" ')",
            html,
        )
        self.assertIn("buttons.forEach(item => { item.disabled = true; });", html)
        self.assertIn("buttons.forEach(item => { item.disabled = false; });", html)
        self.assertEqual(
            db_access_control.required_role("POST", "/db/write/confirm"),
            "operator",
        )

    def test_structured_insert_ui_uses_example_row_preview_and_confirmation_chain(self):
        html_path = Path(__file__).parent / "desktop" / "static" / "db.html"
        css_path = Path(__file__).parent / "desktop" / "static" / "calm-theme.css"
        html = html_path.read_text(encoding="utf-8")
        css = css_path.read_text(encoding="utf-8")
        self.assertIn('id="writeFormModal"', html)
        self.assertIn('id="writeFormTable"', html)
        self.assertIn('id="writeFormGrid"', html)
        self.assertIn('calm-theme.css?v=20260827-6', html)
        self.assertIn("原表示例", html)
        self.assertIn("生成变更预览", html)
        self.assertIn("/db/write/form?", html)
        self.assertIn("/db/write/prepare-insert", html)
        self.assertIn("mode: mode ? mode.value : 'omit'", html)
        self.assertIn("data-open-write-form", html)
        self.assertIn("awaitingWriteTarget", html)
        self.assertIn("<b class=\"op-tag warn\">待选择</b>", html)
        self.assertIn(".write-form-table", css)
        self.assertIn("回滚预览", css)
        wrapper_rule = css.split(".write-form-table-wrap {", 1)[1].split("}", 1)[0]
        table_rule = css.split(".write-form-table {", 1)[1].split("}", 1)[0]
        self.assertIn("width: auto", wrapper_rule)
        self.assertNotIn("width: max-content", wrapper_rule)
        self.assertIn("min-width: 100%", table_rule)
        self.assertIn("table-layout: auto", table_rule)
        self.assertEqual(
            db_access_control.required_role("GET", "/db/write/form"),
            "viewer",
        )
        self.assertEqual(
            db_access_control.required_role("POST", "/db/write/prepare-insert"),
            "operator",
        )
        self.assertIn('id="createTableModal"', html)
        self.assertIn('id="createTableBtn"', html)
        self.assertIn("/db/write/prepare-create-table", html)
        self.assertIn("STRUCTURED DDL", html)
        self.assertIn("CREATE_TABLE_TYPES", html)
        self.assertIn(".create-column-row", css)
        self.assertEqual(
            db_access_control.required_role("POST", "/db/write/prepare-create-table"),
            "admin",
        )

    def test_chart_ui_uses_multiple_colors_value_context_and_lazy_rendering(self):
        html_path = Path(__file__).parent / "desktop" / "static" / "db.html"
        css_path = Path(__file__).parent / "desktop" / "static" / "calm-theme.css"
        html = html_path.read_text(encoding="utf-8")
        css = css_path.read_text(encoding="utf-8")
        self.assertIn("const CHART_PALETTE = [", html)
        self.assertGreaterEqual(html.count("'#"), 12)
        self.assertIn("IntersectionObserver", html)
        self.assertIn("chart-card-context", html)
        self.assertIn("chart-insight", html)
        self.assertIn("m.overallValue", html)
        self.assertIn(".chart-stats", css)
        self.assertIn("--chart-accent", css)

    def test_desktop_renders_basic_conversation_as_a_normal_answer(self):
        html_path = Path(__file__).parent / "desktop" / "static" / "db.html"
        css_path = Path(__file__).parent / "desktop" / "static" / "calm-theme.css"
        html = html_path.read_text(encoding="utf-8")
        css = css_path.read_text(encoding="utf-8")
        self.assertIn("conversation: ['conversation', '交流'", html)
        self.assertIn("['conversation', 'clarification', 'schema'", html)
        self.assertIn("CONVERSE / QUERY / RETRIEVE", html)
        self.assertIn("基础沟通和结构查看仍可使用", html)
        self.assertIn("'你能做什么？'", html)
        self.assertIn(".kind-conversation", css)

    def test_session_switch_restores_bounded_rich_readonly_results(self):
        html_path = Path(__file__).parent / "desktop" / "static" / "db.html"
        css_path = Path(__file__).parent / "desktop" / "static" / "calm-theme.css"
        html = html_path.read_text(encoding="utf-8")
        css = css_path.read_text(encoding="utf-8")
        self.assertIn("m.display && typeof m.display === 'object'", html)
        self.assertIn("addAnswerCard(m.display, s.dbId || state.dbId, true)", html)
        self.assertIn("function addAnswerCard(res, sourceDbId = state.dbId, historyMode = false)", html)
        self.assertIn("renderTable(res.columns, res.rows, false, res.row_count)", html)
        self.assertIn("dataset.row_count", html)
        self.assertIn("历史快照", html)
        self.assertIn(".table-snapshot-note", css)
        self.assertIn("const TABLE_DEFAULT_ROWS = 10", html)
        self.assertIn("data-table-expand", html)
        self.assertIn("rowIndex >= TABLE_DEFAULT_ROWS", html)
        self.assertIn("row.hidden = !expanded", html)
        self.assertIn("'展开全部 ' + total + ' 行'", html)
        self.assertIn("'收起到 10 行'", html)
        self.assertIn("aria-expanded", html)
        self.assertIn(".table-expand-toggle", css)

    def test_chart_snapshots_persist_and_refresh_only_after_source_change_or_request(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database_path = root / "charts.sqlite"
            _make_db(database_path, rows=3)
            _add_table(database_path, "orders")
            cache_path = root / "chart-cache.sqlite"
            db_id = "chart-cache-test"
            desktop_bridge._DB_AGENT_DBS[db_id] = {
                "id": db_id,
                "name": "charts.sqlite",
                "path": str(database_path),
                "tables": ["items", "orders"],
                "kind": "sqlite",
                "attachedAt": 0,
            }
            try:
                with mock.patch.object(db_chart_cache, "_DATA_DIR", root), \
                        mock.patch.object(db_chart_cache, "_DB_PATH", cache_path), \
                        mock.patch.object(desktop_bridge, "_chart_cache", db_chart_cache), \
                        mock.patch.object(
                            desktop_bridge,
                            "_db_charts_auto",
                            wraps=desktop_bridge._db_charts_auto,
                        ) as generate:
                    first = desktop_bridge._db_charts_cached(db_id)
                    self.assertEqual(first["cache"]["status"], "generated")
                    self.assertEqual(
                        {chart["meta"]["table"] for chart in first["charts"]},
                        {"items", "orders"},
                    )
                    self.assertEqual(generate.call_count, 1)

                    cached = desktop_bridge._db_charts_cached(db_id)
                    self.assertEqual(cached["cache"]["status"], "hit")
                    self.assertEqual(cached["charts"], first["charts"])
                    self.assertEqual(generate.call_count, 1)
                    self.assertFalse(
                        desktop_bridge._db_charts_cache_status(db_id)["changed"]
                    )

                    with closing(sqlite3.connect(database_path)) as conn:
                        conn.execute("INSERT INTO items(value) VALUES ('changed')")
                        conn.commit()
                    current = database_path.stat()
                    os.utime(
                        database_path,
                        ns=(current.st_atime_ns, current.st_mtime_ns + 2_000_000_000),
                    )
                    self.assertTrue(
                        desktop_bridge._db_charts_cache_status(db_id)["changed"]
                    )
                    changed = desktop_bridge._db_charts_cached(db_id)
                    self.assertEqual(changed["cache"]["status"], "database_changed")
                    self.assertEqual(generate.call_count, 2)
                    self.assertFalse(
                        desktop_bridge._db_charts_cache_status(db_id)["changed"]
                    )

                    manual = desktop_bridge._db_charts_cached(db_id, force_refresh=True)
                    self.assertEqual(manual["cache"]["status"], "manual_refresh")
                    self.assertEqual(generate.call_count, 3)
                    stored = db_chart_cache.load_snapshot(
                        desktop_bridge._database_scope_ref(
                            desktop_bridge._DB_AGENT_DBS[db_id]
                        ),
                        "all",
                    )
                    self.assertEqual(len(stored["charts"]), 2)
                    self.assertIsNone(db_chart_cache.load_snapshot(
                        desktop_bridge._database_scope_ref(
                            desktop_bridge._DB_AGENT_DBS[db_id]
                        ),
                        "a" * 64,
                    ))
            finally:
                desktop_bridge._DB_AGENT_DBS.pop(db_id, None)

    def test_auto_charts_exclude_fts_storage_and_emit_bounded_business_metadata(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            database_path = Path(temp_dir) / "fts-charts.sqlite"
            with closing(sqlite3.connect(database_path)) as conn:
                conn.execute("CREATE TABLE sales(category TEXT, amount REAL, sold_at TEXT)")
                conn.executemany(
                    "INSERT INTO sales VALUES (?, ?, ?)",
                    [("office", 12.5, "2026-08-24"), ("software", 99, "2026-08-25")],
                )
                conn.execute("CREATE VIRTUAL TABLE search_docs USING fts5(title, body)")
                conn.executemany(
                    "INSERT INTO search_docs(title, body) VALUES (?, ?)",
                    [(f"doc-{index}", "x" * 12000) for index in range(40)],
                )
                conn.commit()
            db_id = "fts-chart-test"
            desktop_bridge._DB_AGENT_DBS[db_id] = {
                "id": db_id, "name": database_path.name, "path": str(database_path),
                "tables": desktop_bridge._db_validate_sqlite(str(database_path)),
                "kind": "sqlite", "attachedAt": 0,
            }
            try:
                charts = desktop_bridge._db_charts_auto(db_id)
                self.assertEqual([chart["meta"]["table"] for chart in charts], ["sales"])
                chart = charts[0]
                self.assertEqual(chart["meta"]["profile"], "trend")
                self.assertEqual(chart["meta"]["rowCount"], 2)
                self.assertEqual(chart["meta"]["pointCount"], 2)
                self.assertIn("summary", chart["meta"])
                self.assertLess(len(json.dumps(charts, ensure_ascii=False).encode("utf-8")), 100_000)
                visible = desktop_bridge._db_validate_sqlite(str(database_path))
                self.assertIn("search_docs", visible)
                self.assertFalse(any(name.startswith("search_docs_") for name in visible))
            finally:
                desktop_bridge._DB_AGENT_DBS.pop(db_id, None)

    def test_pinned_timezone_runtime_and_dst_boundaries(self):
        status = dc.TimezoneRuntime.status()
        self.assertTrue(status["available"], status["error"])
        self.assertEqual(status["tzdata_version"], "2026.3")
        self.assertEqual(status["iana_version"], "2026c")
        self.assertEqual(status["release_id"], "tzdata-2026.3-iana-2026c")
        self.assertEqual(status["zones_count"], 598)
        self.assertEqual(status["release_count"], 1)
        self.assertRegex(status["archive_sha256"], r"^[0-9a-f]{64}$")
        contract = dc.TimezoneRuntime.validate_contract()
        self.assertEqual(contract["active"]["probes_passed"], 8)
        self.assertEqual(contract["active"]["probes_total"], 8)
        convert = dc.TimezoneRuntime.utc_to_local_date
        token = dc.TimezoneRuntime.VERSION_TOKEN
        self.assertEqual(convert("2024-01-01T04:30:00Z", "America/New_York", token), "2023-12-31")
        self.assertEqual(convert("2024-01-01T05:30:00Z", "America/New_York", token), "2024-01-01")
        self.assertEqual(convert("2024-07-01T03:30:00Z", "America/New_York", token), "2024-06-30")
        self.assertEqual(convert("2024-07-01T04:30:00Z", "America/New_York", token), "2024-07-01")
        self.assertIsNone(convert("2024-07-01T04:30:00+08:00", "America/New_York", token))
        self.assertIsNone(convert("2024-07-01T04:30:00Z", "America/New_York", "wrong"))


class ModelProfileStoreTests(unittest.TestCase):
    def test_crud_is_atomic_and_public_views_never_return_the_secret(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = model_profiles.ModelProfileStore(Path(temp_dir))
            created = store.add_model_profile({
                "name": "Local provider",
                "model": "example-model",
                "baseUrl": "https://models.example/v1/",
                "apiKey": "private-test-value",
            })
            key = created["profileKey"]
            self.assertTrue((Path(temp_dir) / "model_profiles.json").is_file())
            self.assertNotIn("private-test-value", json.dumps(created))
            self.assertNotIn("private-test-value", json.dumps(store.list_model_profiles()))
            public = store.get_model_profile(key)
            self.assertEqual(public["baseUrl"], "https://models.example/v1")
            self.assertEqual(public["keyTail"], "alue")
            self.assertNotIn("apiKey", public)

            store.update_model_profile(key, {
                "name": "Renamed",
                "model": "example-model-v2",
                "baseUrl": "https://models.example/v1",
                "apiKey": "",
            })
            runtime = store.get_runtime_profile(key)
            self.assertEqual(runtime["api_key"], "private-test-value")
            self.assertEqual(runtime["model"], "example-model-v2")
            store.delete_model_profile(key)
            self.assertEqual(store.list_model_profiles(), [])

    def test_invalid_or_embedded_credential_url_is_rejected_before_write(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = model_profiles.ModelProfileStore(Path(temp_dir))
            with self.assertRaisesRegex(model_profiles.ModelProfileError, "embedded credentials"):
                store.add_model_profile({
                    "model": "example-model",
                    "baseUrl": "https://user:pass@models.example/v1",
                    "apiKey": "private-test-value",
                })
            self.assertFalse((Path(temp_dir) / "model_profiles.json").exists())


class TimezoneReleaseContractTests(unittest.TestCase):
    def _portable_manifest(self) -> dict:
        manifest = tzcontract.load_manifest()
        return {key: value for key, value in manifest.items() if key != "manifest_path"}

    def test_manifest_rejects_archive_path_escape(self):
        raw = self._portable_manifest()
        active = raw["active_release_id"]
        raw["releases"][active]["archive"] = "../escaped.zip"
        with self.assertRaisesRegex(ValueError, "archive"):
            tzcontract.validate_manifest(raw)

    def test_release_hash_tampering_is_rejected(self):
        source_manifest = tzcontract.load_manifest()
        source_archive = tzcontract.archive_path(source_manifest, source_manifest["active_release_id"])
        with tempfile.TemporaryDirectory() as temp_dir:
            target_dir = Path(temp_dir)
            target_manifest = target_dir / "manifest.json"
            archive = target_dir / source_archive.name
            shutil.copy2(source_archive, archive)
            target_manifest.write_text(
                json.dumps(self._portable_manifest(), ensure_ascii=False), encoding="utf-8",
            )
            with archive.open("ab") as handle:
                handle.write(b"tampered")
            with self.assertRaisesRegex(ValueError, "哈希"):
                tzcontract.validate_contract(target_manifest)

    def test_release_build_is_deterministic(self):
        manifest = tzcontract.load_manifest()
        archive = tzcontract.archive_path(manifest, manifest["active_release_id"])
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source"
            zone_path = source / "America" / "New_York"
            zone_path.parent.mkdir(parents=True)
            with zipfile.ZipFile(archive, "r") as package:
                zone_path.write_bytes(package.read("zoneinfo/America/New_York"))
            zones = root / "zones.txt"
            zones.write_text("America/New_York\n", encoding="utf-8")
            first = tzcontract.build_release_archive(
                source, zones, root / "first.zip", "test.1", "testa",
            )
            second = tzcontract.build_release_archive(
                source, zones, root / "second.zip", "test.1", "testa",
            )
            self.assertEqual(first["sha256"], second["sha256"])
            self.assertEqual((root / "first.zip").read_bytes(), (root / "second.zip").read_bytes())

    def test_activation_and_rollback_keep_both_releases(self):
        raw = self._portable_manifest()
        active = raw["active_release_id"]
        candidate = "tzdata-2026.4-iana-2026d"
        raw["releases"][candidate] = {
            **raw["releases"][active],
            "tzdata_version": "2026.4",
            "iana_version": "2026d",
            "archive": f"{candidate}.zip",
            "sha256": "a" * 64,
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "manifest.json"
            tzcontract.write_manifest_atomic(path, raw)
            with mock.patch.object(tzcontract, "validate_release_archive", return_value={}), \
                    mock.patch.object(tzcontract, "validate_contract", return_value={}):
                switched = tzcontract.switch_active_release(path, candidate)
                self.assertEqual(switched["active_release_id"], candidate)
                self.assertEqual(switched["rollback_release_id"], active)
                self.assertEqual(len(switched["releases"]), 2)
                rolled_back = tzcontract.switch_active_release(path, active)
            self.assertEqual(rolled_back["active_release_id"], active)
            self.assertEqual(rolled_back["rollback_release_id"], candidate)
            self.assertEqual(len(rolled_back["releases"]), 2)

    def test_pinned_calendar_release_survives_a_new_active_release(self):
        original = tzcontract.load_manifest()
        original_id = original["active_release_id"]
        original_archive = tzcontract.archive_path(original, original_id)
        probes = [
            {"id": "ny-winter-before", "zone": "America/New_York", "utc": "2024-01-01T04:30:00Z", "expected_date": "2023-12-31"},
            {"id": "ny-winter-after", "zone": "America/New_York", "utc": "2024-01-01T05:30:00Z", "expected_date": "2024-01-01"},
            {"id": "ny-summer-before", "zone": "America/New_York", "utc": "2024-07-01T03:30:00Z", "expected_date": "2024-06-30"},
            {"id": "ny-summer-after", "zone": "America/New_York", "utc": "2024-07-01T04:30:00Z", "expected_date": "2024-07-01"},
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source" / "America"
            source.mkdir(parents=True)
            with zipfile.ZipFile(original_archive, "r") as package:
                (source / "New_York").write_bytes(package.read("zoneinfo/America/New_York"))
            zones = root / "zones.txt"
            zones.write_text("America/New_York\n", encoding="utf-8")
            built = tzcontract.build_release_archive(
                root / "source", zones, root / "tzdata-2026.4-iana-2026d.zip", "2026.4", "2026d",
            )
            shutil.copy2(original_archive, root / original_archive.name)
            candidate_id = built["release_id"]
            portable = self._portable_manifest()
            portable["active_release_id"] = candidate_id
            portable["rollback_release_id"] = original_id
            portable["releases"][candidate_id] = {
                "tzdata_version": "2026.4",
                "iana_version": "2026d",
                "archive": built["archive"],
                "sha256": built["sha256"],
                "zones_count": 1,
                "probes": probes,
            }
            manifest_path = root / "manifest.json"
            tzcontract.write_manifest_atomic(manifest_path, portable)
            manifest = tzcontract.load_manifest(manifest_path)
            tzcontract.validate_contract(manifest_path)

            runtime = dc.TimezoneRuntime
            saved = (
                runtime._MANIFEST, runtime._ACTIVE_RELEASE_ID, runtime._ACTIVE_RELEASE,
                runtime.TZDATA_VERSION, runtime.IANA_VERSION, runtime.VERSION_TOKEN,
            )
            try:
                runtime._MANIFEST = manifest
                runtime._ACTIVE_RELEASE_ID = candidate_id
                runtime._ACTIVE_RELEASE = dict(manifest["releases"][candidate_id])
                runtime.TZDATA_VERSION = "2026.4"
                runtime.IANA_VERSION = "2026d"
                runtime.VERSION_TOKEN = "tzdata-2026.4/iana-2026d"
                runtime._zone.cache_clear()
                old_token = "tzdata-2026.3/iana-2026c"
                self.assertEqual(
                    runtime.utc_to_local_date("2024-01-01T04:30:00Z", "America/New_York", old_token),
                    "2023-12-31",
                )
                resolved_id, _ = runtime.resolve_release("2026.3", "2026c")
                self.assertEqual(resolved_id, original_id)
            finally:
                (
                    runtime._MANIFEST, runtime._ACTIVE_RELEASE_ID, runtime._ACTIVE_RELEASE,
                    runtime.TZDATA_VERSION, runtime.IANA_VERSION, runtime.VERSION_TOKEN,
                ) = saved
                runtime._zone.cache_clear()

    def test_prepare_rejects_same_version_overwrite_without_mutation(self):
        spec = importlib.util.spec_from_file_location(
            "timezone_release_manager_for_test",
            Path(__file__).resolve().parents[3] / "scripts/manage_timezone_release.py",
        )
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        manager = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(manager)
        source_manifest = tzcontract.load_manifest()
        source_archive = tzcontract.archive_path(source_manifest, source_manifest["active_release_id"])
        initial_probes = [
            {"id": "initial-winter-before", "zone": "America/New_York", "utc": "2024-01-01T04:30:00Z", "expected_date": "2023-12-31"},
            {"id": "initial-winter-after", "zone": "America/New_York", "utc": "2024-01-01T05:30:00Z", "expected_date": "2024-01-01"},
            {"id": "initial-summer-before", "zone": "America/New_York", "utc": "2024-07-01T03:30:00Z", "expected_date": "2024-06-30"},
            {"id": "initial-summer-after", "zone": "America/New_York", "utc": "2024-07-01T04:30:00Z", "expected_date": "2024-07-01"},
        ]
        replacement_probes = [
            {"id": "replacement-one", "zone": "America/New_York", "utc": "2024-01-01T04:30:00Z", "expected_date": "2024-01-01"},
            {"id": "replacement-two", "zone": "America/New_York", "utc": "2024-01-01T05:30:00Z", "expected_date": "2024-01-01"},
            {"id": "replacement-three", "zone": "America/New_York", "utc": "2024-07-01T03:30:00Z", "expected_date": "2024-07-01"},
            {"id": "replacement-four", "zone": "America/New_York", "utc": "2024-07-01T04:30:00Z", "expected_date": "2024-07-01"},
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            zones = root / "zones.txt"
            zones.write_text("America/New_York\n", encoding="utf-8")
            initial_source = root / "initial" / "America"
            replacement_source = root / "replacement" / "America"
            initial_source.mkdir(parents=True)
            replacement_source.mkdir(parents=True)
            with zipfile.ZipFile(source_archive, "r") as package:
                (initial_source / "New_York").write_bytes(package.read("zoneinfo/America/New_York"))
                (replacement_source / "New_York").write_bytes(package.read("zoneinfo/UTC"))
            release_id = tzcontract.release_id("test.1", "testa")
            archive_path = root / f"{release_id}.zip"
            built = tzcontract.build_release_archive(
                root / "initial", zones, archive_path, "test.1", "testa",
            )
            manifest_path = root / "manifest.json"
            tzcontract.write_manifest_atomic(manifest_path, {
                "schema_version": 1,
                "active_release_id": release_id,
                "rollback_release_id": None,
                "releases": {release_id: {
                    "tzdata_version": "test.1", "iana_version": "testa",
                    "archive": built["archive"], "sha256": built["sha256"],
                    "zones_count": 1, "probes": initial_probes,
                }},
            })
            probes_path = root / "replacement-probes.json"
            probes_path.write_text(json.dumps(replacement_probes), encoding="utf-8")
            manifest_before = manifest_path.read_bytes()
            archive_before = archive_path.read_bytes()
            args = types.SimpleNamespace(
                manifest=manifest_path,
                source_root=root / "replacement",
                zones_file=zones,
                tzdata_version="test.1",
                iana_version="testa",
                probes=probes_path,
            )
            with self.assertRaisesRegex(ValueError, "不能以不同内容覆盖"):
                manager.prepare(args)
            self.assertEqual(manifest_path.read_bytes(), manifest_before)
            self.assertEqual(archive_path.read_bytes(), archive_before)
            self.assertFalse(list(root.glob(".*.candidate.*")))


class FixedEvaluationSuiteTests(unittest.TestCase):
    def test_offline_suite_passes_without_running_model(self):
        result = nl2db_evaluation.run_suite(with_model=False)
        self.assertTrue(result["offline"]["all_cases_passed"], result["offline"]["error_categories"])
        self.assertEqual(result["model_nl2sql"]["status"], "not_run")


class ModelBaselineContractTests(unittest.TestCase):
    @staticmethod
    def _result(passed: list[bool], generated_at: str) -> dict:
        cases = [{
            "id": f"model-case-{index + 1}",
            "passed": value,
            "error_category": None if value else "result_rows",
            "latency_ms": 1000 + index * 100,
            "sql": f"SELECT {index + 1}",
            "error": None,
        } for index, value in enumerate(passed)]
        count = sum(passed)
        return {
            "generated_at": generated_at,
            "suite_version": "test-suite-v1",
            "dataset_sha256": "a" * 64,
            "model_nl2sql": {
                "status": "completed" if count == len(cases) else "completed_with_failures",
                "passed": count,
                "total": len(cases),
                "execution_accuracy": count / len(cases),
                "prompt_contract": "nl2sql-0123456789abcdef",
                "model_identity": {
                    "config_name": "test_config",
                    "name": "test profile",
                    "model": "test-model-v1",
                    "api_mode": "responses",
                    "endpoint_fingerprint": "0123456789abcdef",
                },
                "latency_ms": {"total": 2200, "median": 1100, "maximum": 1200},
                "cases": cases,
            },
        }

    def test_model_identity_redacts_endpoint_and_credentials(self):
        identity = nl2db_evaluation._redacted_model_identity("private_config", {
            "name": "Private", "model": "model-v1", "api_mode": "responses",
            "base_url": "https://private.example/v1", "api_key": "secret-key-value",
        })
        serialized = json.dumps(identity)
        self.assertNotIn("secret-key-value", serialized)
        self.assertNotIn("private.example", serialized)
        self.assertEqual(set(identity), model_baselines.IDENTITY_KEYS)
        self.assertRegex(identity["endpoint_fingerprint"], r"^[0-9a-f]{16}$")

    def test_append_only_model_history_compares_compatible_runs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "baselines.json"
            first = model_baselines.build_record(
                self._result([True, False], "2026-08-17T09:00:00+08:00"), "first",
            )
            second = model_baselines.build_record(
                self._result([True, True], "2026-08-17T10:00:00+08:00"), "second",
            )
            self.assertIsNone(model_baselines.append_record(path, first)["comparison"])
            appended = model_baselines.append_record(path, second)
            self.assertEqual(appended["history_count"], 2)
            self.assertEqual(appended["comparison"]["accuracy_delta"], 0.5)
            self.assertEqual(appended["comparison"]["improvements"], ["model-case-2"])
            self.assertEqual(model_baselines.latest_summary(path)["latest"]["passed"], 2)
            with self.assertRaisesRegex(ValueError, "只能追加"):
                model_baselines.append_record(path, second)


class AccessControlContractTests(unittest.TestCase):
    def test_role_tokens_are_deterministic_distinct_and_one_way_labeled(self):
        first = db_access_control.derive_role_tokens("a" * 32)
        second = db_access_control.derive_role_tokens("a" * 32)
        self.assertEqual(first, second)
        self.assertEqual(len(set(first.values())), 3)
        self.assertTrue(first["viewer"].startswith("vw_"))
        self.assertTrue(first["operator"].startswith("op_"))
        for role, token in first.items():
            self.assertEqual(db_access_control.role_for_token(token, first), role)
        self.assertIsNone(db_access_control.role_for_token("invalid", first))

    def test_http_permission_policy_is_least_privilege(self):
        self.assertEqual(db_access_control.required_role("GET", "/db/audit"), "viewer")
        self.assertEqual(db_access_control.required_role("POST", "/db/ask"), "viewer")
        self.assertEqual(
            db_access_control.required_role("POST", "/db/charts-cache-status"),
            "viewer",
        )
        self.assertEqual(db_access_control.required_role("POST", "/db/semantics"), "operator")
        self.assertEqual(db_access_control.required_role("GET", "/db/excel/export"), "viewer")
        self.assertEqual(
            db_access_control.required_role("POST", "/db/excel/import/prepare"),
            "operator",
        )
        self.assertEqual(
            db_access_control.required_role("POST", "/db/excel/import/confirm"),
            "operator",
        )
        self.assertEqual(db_access_control.required_role("POST", "/db/audit/backups"), "admin")
        self.assertEqual(
            db_access_control.required_role(
                "POST", "/db/audit/reconciliation/resolve",
            ),
            "admin",
        )
        self.assertEqual(db_access_control.required_role("GET", "/db/auth/credentials"), "admin")
        self.assertEqual(
            db_access_control.required_role("POST", "/db/auth/credentials/id/revoke"),
            "admin",
        )
        self.assertEqual(db_access_control.required_role("DELETE", "/db/databases/x"), "admin")
        self.assertEqual(db_access_control.required_role("GET", "/model-profiles"), "admin")
        self.assertTrue(db_access_control.role_allows("admin", "operator"))
        self.assertFalse(db_access_control.role_allows("viewer", "operator"))


class LocalIdentityStoreTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path_patch = mock.patch.object(
            db_identity_store, "_DB_PATH", Path(self.tmp.name) / "identities.db",
        )
        self.path_patch.start()
        db_identity_store.init_db()

    def tearDown(self):
        self.path_patch.stop()
        self.tmp.cleanup()

    def test_token_is_hashed_and_expiry_and_revocation_are_enforced(self):
        issued_at = datetime(2026, 8, 17, 4, 0, tzinfo=timezone.utc)
        database_ref = "a" * 64
        credential = db_identity_store.issue_credential(
            label="数据分析员 A", role="operator", ttl_hours=2,
            scope_mode="restricted", database_refs=[database_ref], now=issued_at,
        )
        token = credential["token"]
        self.assertTrue(token.startswith("id_"))
        self.assertNotIn("token", db_identity_store.list_credentials(now=issued_at)[0])
        self.assertNotIn(token.encode("utf-8"), db_identity_store._DB_PATH.read_bytes())
        authenticated = db_identity_store.authenticate(
            token, now=issued_at + timedelta(minutes=30),
        )
        self.assertEqual(authenticated["role"], "operator")
        self.assertEqual(authenticated["credentialRef"], credential["credentialRef"])
        self.assertEqual(
            authenticated["databaseScope"],
            {
                "mode": "restricted", "databaseRefs": [database_ref],
                "tableScopes": {}, "columnScopes": {}, "rowScopes": {},
            },
        )
        self.assertIsNone(
            db_identity_store.authenticate(token, now=issued_at + timedelta(hours=2)),
        )

        active = db_identity_store.issue_credential(
            label="临时查看", role="viewer", ttl_hours=24, now=issued_at,
        )
        revoked = db_identity_store.revoke_credential(
            active["id"], now=issued_at + timedelta(hours=1),
        )
        self.assertEqual(revoked["status"], "revoked")
        self.assertIsNone(
            db_identity_store.authenticate(
                active["token"], now=issued_at + timedelta(hours=1, minutes=1),
            ),
        )

    def test_issue_validation_rejects_unbounded_or_invalid_credentials(self):
        for kwargs in (
            {"label": "", "role": "viewer", "ttl_hours": 1},
            {"label": "name", "role": "owner", "ttl_hours": 1},
            {"label": "name", "role": "viewer", "ttl_hours": 0},
            {"label": "name", "role": "viewer", "ttl_hours": 24 * 365 + 1},
            {
                "label": "name", "role": "viewer", "ttl_hours": 1,
                "scope_mode": "restricted", "database_refs": [],
            },
            {
                "label": "name", "role": "viewer", "ttl_hours": 1,
                "scope_mode": "restricted", "database_refs": ["not-a-ref"],
            },
            {
                "label": "name", "role": "viewer", "ttl_hours": 1,
                "scope_mode": "all", "database_refs": ["b" * 64],
            },
            {
                "label": "name", "role": "viewer", "ttl_hours": 1,
                "scope_mode": "restricted", "database_refs": ["b" * 64],
                "table_scopes": {"c" * 64: ["items"]},
            },
            {
                "label": "name", "role": "viewer", "ttl_hours": 1,
                "scope_mode": "restricted", "database_refs": ["b" * 64],
                "table_scopes": {"b" * 64: []},
            },
            {
                "label": "name", "role": "viewer", "ttl_hours": 1,
                "scope_mode": "restricted", "database_refs": ["b" * 64],
                "table_scopes": {"b" * 64: ["items"]},
                "column_scopes": {"b" * 64: {"missing": ["id"]}},
            },
            {
                "label": "name", "role": "viewer", "ttl_hours": 1,
                "scope_mode": "restricted", "database_refs": ["b" * 64],
                "table_scopes": {"b" * 64: ["items"]},
                "column_scopes": {"b" * 64: {"items": []}},
            },
            {
                "label": "name", "role": "viewer", "ttl_hours": 1,
                "scope_mode": "restricted", "database_refs": ["b" * 64],
                "row_scopes": {"b" * 64: {"items": [
                    {"column": "tenant", "operator": "eq", "value": "north"},
                ]}},
            },
            {
                "label": "name", "role": "viewer", "ttl_hours": 1,
                "scope_mode": "restricted", "database_refs": ["b" * 64],
                "table_scopes": {"b" * 64: ["items"]},
                "row_scopes": {"b" * 64: {"items": [
                    {"column": "tenant", "operator": "contains", "value": "north"},
                ]}},
            },
        ):
            with self.subTest(kwargs=kwargs), self.assertRaises(ValueError):
                db_identity_store.issue_credential(**kwargs)

    def test_v1_credentials_migrate_as_explicit_legacy_all_database_scope(self):
        db_identity_store._DB_PATH.unlink()
        token = "id_" + "x" * 48
        token_hash = db_identity_store._token_hash(token)
        with closing(sqlite3.connect(db_identity_store._DB_PATH)) as conn:
            conn.execute(
                """
                CREATE TABLE local_credentials (
                    credential_id TEXT PRIMARY KEY,
                    credential_ref TEXT NOT NULL UNIQUE,
                    label TEXT NOT NULL,
                    role TEXT NOT NULL,
                    token_hash TEXT NOT NULL UNIQUE,
                    created_at TEXT NOT NULL,
                    expires_at TEXT NOT NULL,
                    revoked_at TEXT NOT NULL DEFAULT ''
                )
                """
            )
            conn.execute(
                "INSERT INTO local_credentials VALUES (?, ?, ?, ?, ?, ?, ?, '')",
                (
                    "legacy-id", db_identity_store._credential_ref(token_hash),
                    "旧凭据", "viewer", token_hash,
                    "2026-08-17T00:00:00+00:00", "2026-08-18T00:00:00+00:00",
                ),
            )
            conn.execute("PRAGMA user_version=1")
            conn.commit()
        db_identity_store.init_db()
        authenticated = db_identity_store.authenticate(
            token, now=datetime(2026, 8, 17, 1, 0, tzinfo=timezone.utc),
        )
        self.assertEqual(
            authenticated["databaseScope"],
            {
                "mode": "all", "databaseRefs": [],
                "tableScopes": {}, "columnScopes": {}, "rowScopes": {},
            },
        )
        with closing(sqlite3.connect(db_identity_store._DB_PATH)) as conn:
            self.assertEqual(conn.execute("PRAGMA user_version").fetchone()[0], 5)

    def test_table_scopes_round_trip_without_persisting_raw_token(self):
        database_ref = "d" * 64
        issued = db_identity_store.issue_credential(
            label="table viewer", role="viewer", ttl_hours=24,
            scope_mode="restricted", database_refs=[database_ref],
            table_scopes={database_ref: ["Orders", "customers", "orders"]},
            column_scopes={database_ref: {"orders": ["ID", "amount", "id"]}},
            row_scopes={database_ref: {"orders": [
                {"column": "Tenant_ID", "operator": "eq", "value": "north"},
            ]}},
        )
        self.assertEqual(
            issued["databaseScope"]["tableScopes"],
            {database_ref: ["customers", "Orders"]},
        )
        self.assertEqual(
            issued["databaseScope"]["columnScopes"],
            {database_ref: {"Orders": ["amount", "ID"]}},
        )
        self.assertEqual(
            issued["databaseScope"]["rowScopes"],
            {database_ref: {"Orders": [
                {"column": "Tenant_ID", "operator": "eq", "value": "north"},
            ]}},
        )
        authenticated = db_identity_store.authenticate(issued["token"])
        self.assertEqual(
            authenticated["databaseScope"]["tableScopes"],
            {database_ref: ["customers", "Orders"]},
        )
        self.assertEqual(
            authenticated["databaseScope"]["columnScopes"],
            {database_ref: {"Orders": ["amount", "ID"]}},
        )
        self.assertEqual(
            authenticated["databaseScope"]["rowScopes"],
            {database_ref: {"Orders": [
                {"column": "Tenant_ID", "operator": "eq", "value": "north"},
            ]}},
        )
        self.assertNotIn(issued["token"].encode("utf-8"), db_identity_store._DB_PATH.read_bytes())


class SessionAccessScopeStoreTests(unittest.TestCase):
    def test_legacy_sessions_migrate_to_all_and_new_scope_round_trips(self):
        with tempfile.TemporaryDirectory() as tmp:
            database_path = Path(tmp) / "sessions.db"
            with closing(sqlite3.connect(database_path)) as conn:
                conn.execute(
                    "CREATE TABLE sessions ("
                    "id TEXT PRIMARY KEY, title TEXT DEFAULT '', db_id TEXT DEFAULT '', "
                    "last_question TEXT DEFAULT '', count INTEGER DEFAULT 0, "
                    "created_at REAL DEFAULT 0, updated_at REAL DEFAULT 0)"
                )
                conn.execute(
                    "INSERT INTO sessions VALUES ('legacy', 'legacy', 'db1', '', 0, 1, 1)"
                )
                conn.execute(
                    "CREATE TABLE messages ("
                    "id INTEGER PRIMARY KEY AUTOINCREMENT, session_id TEXT NOT NULL, "
                    "role TEXT NOT NULL, content TEXT DEFAULT '', created_at REAL DEFAULT 0)"
                )
                conn.execute(
                    "INSERT INTO messages(session_id, role, content, created_at) "
                    "VALUES ('legacy', 'assistant', '旧文本', 1)"
                )
                conn.commit()
            with mock.patch.object(db_sessions_store, "_DATA_DIR", Path(tmp)), \
                    mock.patch.object(db_sessions_store, "_DB_PATH", database_path):
                db_sessions_store.init_db()
                self.assertEqual(
                    db_sessions_store.get_session("legacy")["accessScopeRef"], "all",
                )
                db_sessions_store.upsert_session(
                    "scoped", db_id="db1", access_scope_ref="e" * 64,
                )
                self.assertEqual(
                    db_sessions_store.get_session("scoped")["accessScopeRef"], "e" * 64,
                )
                with closing(sqlite3.connect(database_path)) as conn:
                    message_columns = {
                        row[1] for row in conn.execute(
                            "PRAGMA table_info(messages)"
                        ).fetchall()
                    }
                self.assertIn("display_payload", message_columns)
                display = {
                    "kind": "schema", "columns": ["表名"],
                    "rows": [[f"table_{index}"] for index in range(12)],
                    "row_count": 12,
                }
                db_sessions_store.append_message(
                    "legacy", "assistant", "模型只读预览", display_payload=display,
                )
                messages = db_sessions_store.get_session("legacy")["messages"]
                self.assertNotIn("display", messages[0])
                self.assertEqual(messages[1]["display"]["row_count"], 12)
                self.assertEqual(len(messages[1]["display"]["rows"]), 12)
                self.assertEqual(
                    db_sessions_store.get_history("legacy")[-1],
                    {"role": "assistant", "content": "模型只读预览"},
                )


class ExcelTransferTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.database = self.root / "target.sqlite"
        with closing(sqlite3.connect(self.database)) as connection:
            connection.execute(
                "CREATE TABLE items (id INTEGER PRIMARY KEY AUTOINCREMENT, value TEXT NOT NULL)"
            )
            connection.execute(
                "CREATE TABLE tags (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE)"
            )
            connection.execute("INSERT INTO items(value) VALUES ('=literal-formula-text')")
            connection.commit()
        self.agent = dc.DBQuillAgent(db_path=str(self.database))

    def tearDown(self):
        self.temp.cleanup()

    def _workbook(self, name: str, sheets: dict[str, list[list[object]]]) -> Path:
        from openpyxl import Workbook
        path = self.root / name
        workbook = Workbook()
        first = True
        for sheet_name, rows in sheets.items():
            worksheet = workbook.active if first else workbook.create_sheet()
            first = False
            worksheet.title = sheet_name
            for row in rows:
                worksheet.append(row)
        workbook.save(path)
        workbook.close()
        return path

    def test_prepare_then_confirm_imports_rows_in_one_transaction(self):
        source = self._workbook("merge.xlsx", {
            "items": [["value"], ["alpha"], ["beta"]],
            "tags": [["name"], ["new-tag"]],
        })
        preview = db_excel_transfer.prepare_import(
            str(source), db_id="db1", database_ref="ref1", access_scope_ref="all",
            agent=self.agent,
        )
        self.assertEqual(preview["tableCount"], 2)
        self.assertEqual(preview["rowCount"], 3)
        self.assertFalse(preview["requiresAdmin"])
        result = db_excel_transfer.execute_import(
            preview["confirmId"], db_id="db1", database_ref="ref1",
            access_scope_ref="all", agent=self.agent,
        )
        self.assertEqual(result["affected"], 3)
        with closing(sqlite3.connect(self.database)) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM items").fetchone()[0], 3)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM tags").fetchone()[0], 1)
        self.assertFalse(source.exists())

    def test_import_failure_rolls_back_every_worksheet(self):
        source = self._workbook("rollback.xlsx", {
            "items": [["value"], ["must-rollback"]],
            "tags": [["name"], ["duplicate"], ["duplicate"]],
        })
        preview = db_excel_transfer.prepare_import(
            str(source), db_id="db1", database_ref="ref1", access_scope_ref="all",
            agent=self.agent,
        )
        with self.assertRaisesRegex(db_excel_transfer.ExcelTransferError, "全部变更已回滚"):
            db_excel_transfer.execute_import(
                preview["confirmId"], db_id="db1", database_ref="ref1",
                access_scope_ref="all", agent=self.agent,
            )
        with closing(sqlite3.connect(self.database)) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM items").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM tags").fetchone()[0], 0)

    def test_formula_cells_are_rejected_before_confirmation(self):
        source = self._workbook("formula.xlsx", {"items": [["value"], ["=1+1"]]})
        with self.assertRaisesRegex(db_excel_transfer.ExcelTransferError, "包含公式"):
            db_excel_transfer.prepare_import(
                str(source), db_id="db1", database_ref="ref1", access_scope_ref="all",
                agent=self.agent,
            )

    def test_export_is_multisheet_round_trip_safe_and_formula_text_stays_text(self):
        exported = db_excel_transfer.export_workbook(str(self.root / "exports"), agent=self.agent)
        self.assertEqual(exported["tableCount"], 2)
        self.assertEqual(exported["rowCount"], 1)
        from openpyxl import load_workbook
        workbook = load_workbook(exported["path"], data_only=False)
        try:
            self.assertIn(db_excel_transfer.MANIFEST_SHEET, workbook.sheetnames)
            self.assertEqual(workbook["items"]["B2"].value, "=literal-formula-text")
            self.assertEqual(workbook["items"]["B2"].data_type, "s")
        finally:
            workbook.close()

    def test_excel_transfer_enforces_table_column_and_row_scope(self):
        scoped_agent = dc.DBQuillAgent(
            db_path=str(self.database),
            allowed_tables=["items"],
            allowed_columns={"items": ["value"]},
            row_filters={
                "items": [{
                    "column": "value", "operator": "eq", "value": "=literal-formula-text",
                }],
            },
        )
        source = self._workbook("row-scoped.xlsx", {"items": [["value"], ["blocked"]]})
        with self.assertRaisesRegex(db_excel_transfer.ExcelTransferError, "行级授权"):
            db_excel_transfer.prepare_import(
                str(source), db_id="db1", database_ref="ref1",
                access_scope_ref="row-scope", agent=scoped_agent,
            )

        exported = db_excel_transfer.export_workbook(
            str(self.root / "scoped-exports"), agent=scoped_agent,
        )
        self.assertEqual(exported["tableCount"], 1)
        self.assertEqual(exported["rowCount"], 1)
        from openpyxl import load_workbook
        workbook = load_workbook(exported["path"], data_only=False)
        try:
            self.assertNotIn("tags", workbook.sheetnames)
            self.assertEqual(
                list(workbook["items"].values),
                [("value",), ("=literal-formula-text",)],
            )
        finally:
            workbook.close()


class PublicProgressAndMemoryTests(unittest.TestCase):
    def test_progress_scope_emits_only_public_structured_events(self):
        events = []
        with dc.progress_scope(lambda stage, label, percent, detail: events.append(
            (stage, label, percent, detail),
        )):
            dc._emit_progress(
                "intent", "意图判断完成", 30,
                {"phase": "intent", "intent": "query"},
            )
        self.assertEqual(events[0][0], "intent")
        self.assertEqual(events[0][2], 30)
        self.assertEqual(events[0][3]["intent"], "query")

    def test_memory_snapshot_truthfully_marks_durable_personal_memory_disabled(self):
        agent = self.agent = dc.DBQuillAgent(db_path=str(self._database_for_memory()))
        answer = agent._attach_memory_snapshot(
            dc.DBAnswer(kind="conversation", narrative="ok"),
            question="继续", resolved_question="上一个完整问题；追问：继续",
            history=[{"role": "user", "content": "上一个完整问题"}],
        )
        layers = {item["key"]: item for item in answer.memory["layers"]}
        self.assertTrue(layers["working"]["active"])
        self.assertTrue(layers["topic"]["active"])
        self.assertFalse(layers["durable"]["active"])

    def _database_for_memory(self) -> Path:
        temporary = tempfile.NamedTemporaryFile(suffix=".sqlite", delete=False)
        temporary.close()
        path = Path(temporary.name)
        self.addCleanup(path.unlink, missing_ok=True)
        with closing(sqlite3.connect(path)) as connection:
            connection.execute("CREATE TABLE items (id INTEGER PRIMARY KEY, value TEXT)")
            connection.commit()
        return path


class LocalApiAuthTests(AioHTTPTestCase):
    async def get_application(self):
        app = web.Application(middlewares=[desktop_bridge.cors_middleware])
        app.router.add_get("/status", desktop_bridge.status_handler)
        app.router.add_get("/db/auth/context", desktop_bridge.db_auth_context_handler)
        app.router.add_get("/db/auth/credentials", desktop_bridge.db_auth_credentials_handler)
        app.router.add_post("/db/auth/credentials", desktop_bridge.db_auth_credentials_handler)
        app.router.add_post(
            "/db/auth/credentials/{credential_id}/revoke",
            desktop_bridge.db_auth_credential_revoke_handler,
        )
        app.router.add_get("/db/databases", desktop_bridge.db_databases_handler)
        app.router.add_post("/db/attach", desktop_bridge.db_attach_handler)
        app.router.add_delete(
            "/db/databases/{db_id}", desktop_bridge.db_detach_handler,
        )
        app.router.add_post("/upload", desktop_bridge.upload_handler)
        app.router.add_post(
            "/db/excel/import/prepare", desktop_bridge.db_excel_import_prepare_handler,
        )
        app.router.add_post(
            "/db/excel/import/confirm", desktop_bridge.db_excel_import_confirm_handler,
        )
        app.router.add_get("/db/excel/export", desktop_bridge.db_excel_export_handler)

        async def allowed(_request):
            return web.json_response({"ok": True})

        app.router.add_post("/db/semantics", allowed)
        app.router.add_post("/db/ask", allowed)
        app.router.add_post("/db/write/confirm", allowed)
        app.router.add_get("/db/ask/{run_id}/progress", desktop_bridge.db_progress_handler)
        app.router.add_get("/db/sessions", desktop_bridge.db_sessions_handler)
        app.router.add_get("/db/session/{sid}", desktop_bridge.db_session_detail_handler)
        app.router.add_get("/db/schedules", desktop_bridge.db_schedules_handler)
        app.router.add_post("/db/schedules/{id}/run", desktop_bridge.db_schedules_run_handler)
        app.router.add_get("/db/schedules/logs", desktop_bridge.db_schedules_logs_handler)
        app.router.add_get("/db/audit", desktop_bridge.db_audit_handler)
        app.router.add_get("/db/audit/export", desktop_bridge.db_audit_export_handler)
        app.router.add_post(
            "/db/audit/reconciliation/resolve",
            desktop_bridge.db_audit_reconciliation_resolve_handler,
        )
        app.router.add_post("/db/audit/backups", allowed)
        return app

    async def asyncSetUp(self):
        await super().asyncSetUp()
        self.audit_tmp = tempfile.TemporaryDirectory()
        self.audit_data_patch = mock.patch.object(
            db_audit_store, "_DATA_DIR", Path(self.audit_tmp.name),
        )
        self.audit_path_patch = mock.patch.object(
            db_audit_store, "_DB_PATH", Path(self.audit_tmp.name) / "audit.db",
        )
        self.audit_data_patch.start()
        self.audit_path_patch.start()
        db_audit_store.init_db()
        self.upload_tmp = tempfile.TemporaryDirectory()
        self.upload_dir_patch = mock.patch.object(
            desktop_bridge._uploads, "root", Path(self.upload_tmp.name),
        )
        self.upload_dir_patch.start()
        self.original_audit_store = desktop_bridge._audit_store
        desktop_bridge._audit_store = db_audit_store
        self.identity_tmp = tempfile.TemporaryDirectory()
        self.identity_path_patch = mock.patch.object(
            db_identity_store, "_DB_PATH", Path(self.identity_tmp.name) / "identities.db",
        )
        self.identity_path_patch.start()
        db_identity_store.init_db()
        self.original_identity_store = desktop_bridge._identity_store
        desktop_bridge._identity_store = db_identity_store
        self.original_session_store = desktop_bridge._sess_store
        desktop_bridge._sess_store = None
        desktop_bridge._DB_SESSIONS.clear()
        desktop_bridge._DB_RUNS.clear()
        self.original_scheduler = desktop_bridge._db_sched
        self.fake_scheduler = mock.Mock()
        desktop_bridge._db_sched = self.fake_scheduler
        self.role_db_id = "role-database-test"
        desktop_bridge._DB_AGENT_DBS[self.role_db_id] = {
            "id": self.role_db_id, "name": "roles.db", "path": "D:/secret/roles.db",
            "tables": ["items", "secrets"], "kind": "sqlite", "attachedAt": 0,
        }
        self.other_db_id = "other-database-test"
        desktop_bridge._DB_AGENT_DBS[self.other_db_id] = {
            "id": self.other_db_id, "name": "other.db", "path": "D:/secret/other.db",
            "tables": ["private_items"], "kind": "sqlite", "attachedAt": 0,
        }
        self.schedule_tasks = [
            {"id": "allowed-task", "dbId": self.role_db_id, "type": "sql"},
            {"id": "other-task", "dbId": self.other_db_id, "type": "sql"},
        ]
        self.fake_scheduler.list_tasks.return_value = self.schedule_tasks
        self.fake_scheduler.get_task.side_effect = lambda task_id: next(
            (item for item in self.schedule_tasks if item["id"] == task_id), None,
        )
        self.fake_scheduler.run_now.return_value = {"ok": True, "rows": []}
        self.fake_scheduler.list_logs.return_value = [{"file": "global.md"}]

    async def asyncTearDown(self):
        desktop_bridge._db_sched = self.original_scheduler
        desktop_bridge._DB_RUNS.clear()
        desktop_bridge._DB_SESSIONS.clear()
        desktop_bridge._sess_store = self.original_session_store
        desktop_bridge._identity_store = self.original_identity_store
        self.identity_path_patch.stop()
        self.identity_tmp.cleanup()
        desktop_bridge._audit_store = self.original_audit_store
        desktop_bridge._DB_AGENT_DBS.pop(self.role_db_id, None)
        desktop_bridge._DB_AGENT_DBS.pop(self.other_db_id, None)
        self.upload_dir_patch.stop()
        self.upload_tmp.cleanup()
        self.audit_path_patch.stop()
        self.audit_data_patch.stop()
        self.audit_tmp.cleanup()
        await super().asyncTearDown()

    async def _issue_restricted(self, role="viewer"):
        allowed_ref = desktop_bridge._database_scope_ref(
            desktop_bridge._DB_AGENT_DBS[self.role_db_id],
        )
        response = await self.client.post(
            "/db/auth/credentials",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
            json={
                "label": f"限定{role}", "role": role, "ttlHours": 24,
                "databaseScope": {
                    "mode": "restricted", "databaseRefs": [allowed_ref],
                },
            },
        )
        self.assertEqual(response.status, 201)
        return (await response.json())["credential"]

    async def _issue_table_restricted(self, role="viewer"):
        allowed_ref = desktop_bridge._database_scope_ref(
            desktop_bridge._DB_AGENT_DBS[self.role_db_id],
        )
        response = await self.client.post(
            "/db/auth/credentials",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
            json={
                "label": f"table-{role}", "role": role, "ttlHours": 24,
                "databaseScope": {
                    "mode": "restricted",
                    "databaseRefs": [allowed_ref],
                    "tableScopes": {allowed_ref: ["items"]},
                },
            },
        )
        self.assertEqual(response.status, 201, await response.text())
        return (await response.json())["credential"]

    async def test_missing_token_is_rejected(self):
        response = await self.client.get("/status")
        self.assertEqual(response.status, 401)
        self.assertNotIn("Access-Control-Allow-Origin", response.headers)

    async def test_valid_header_is_accepted_without_cors(self):
        response = await self.client.get(
            "/status",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
        )
        self.assertEqual(response.status, 200)
        payload = await response.json()
        self.assertTrue(payload["authRequired"])
        self.assertEqual(payload["bridgeProtocol"], desktop_bridge.BRIDGE_PROTOCOL_VERSION)
        self.assertEqual(payload["uploadProtocol"], "multipart-v1")
        self.assertNotIn("Access-Control-Allow-Origin", response.headers)

    async def test_legacy_xls_upload_is_rejected_before_file_write(self):
        form = FormData()
        form.add_field(
            "file", b"not-an-xls", filename="legacy.XLS",
            content_type="application/vnd.ms-excel",
        )
        response = await self.client.post(
            "/upload?sid=legacy-format",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
            data=form,
        )
        self.assertEqual(response.status, 415)
        payload = await response.json()
        self.assertFalse(payload["ok"])
        self.assertEqual(
            payload["error"],
            "旧版 Excel .xls 暂不支持，请先转换为 .xlsx 后再上传。",
        )
        self.assertEqual(list(Path(self.upload_tmp.name).rglob("*")), [])

    async def test_sqlite_upload_streams_attaches_and_returns_tables(self):
        source_dir = tempfile.TemporaryDirectory()
        self.addCleanup(source_dir.cleanup)
        source_path = Path(source_dir.name) / "sample.sqlite"
        _make_db(source_path, rows=2)

        form = FormData()
        form.add_field(
            "file", source_path.read_bytes(), filename="中文数据.sqlite",
            content_type="application/octet-stream",
        )
        response = await self.client.post(
            "/upload?sid=successful-sqlite-upload",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
            data=form,
        )
        self.assertEqual(response.status, 200, await response.text())
        payload = await response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["db"]["kind"], "sqlite")
        self.assertEqual(payload["db"]["name"], "中文数据.sqlite")
        self.assertEqual(payload["db"]["tables"], ["items"])
        uploaded_path = Path(payload["path"])
        self.assertTrue(uploaded_path.is_file())
        self.assertTrue(uploaded_path.is_relative_to(Path(self.upload_tmp.name)))
        uploaded_id = payload["db"]["id"]
        self.assertIn(uploaded_id, desktop_bridge._DB_AGENT_DBS)
        self.addCleanup(desktop_bridge._DB_AGENT_DBS.pop, uploaded_id, None)

    async def test_csv_upload_streams_converts_and_attaches_database(self):
        csv_bytes = "id,name\n1,alpha\n2,beta\n".encode("utf-8")
        form = FormData()
        form.add_field(
            "file", csv_bytes, filename="sample.csv", content_type="text/csv",
        )
        response = await self.client.post(
            "/upload?sid=successful-csv-upload",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
            data=form,
        )
        self.assertEqual(response.status, 200, await response.text())
        payload = await response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["db"]["kind"], "csv")
        self.assertEqual(payload["db"]["tables"], ["sample"])
        converted_path = Path(payload["db"]["path"])
        self.assertTrue(converted_path.is_file())
        with closing(sqlite3.connect(converted_path)) as connection:
            rows = connection.execute(
                'SELECT "id", "name" FROM "sample" ORDER BY "id"'
            ).fetchall()
        self.assertEqual(rows, [(1, "alpha"), (2, "beta")])
        uploaded_id = payload["db"]["id"]
        self.assertIn(uploaded_id, desktop_bridge._DB_AGENT_DBS)
        self.addCleanup(desktop_bridge._DB_AGENT_DBS.pop, uploaded_id, None)

    async def test_excel_import_confirmation_and_export_endpoints_round_trip(self):
        source_dir = tempfile.TemporaryDirectory()
        self.addCleanup(source_dir.cleanup)
        database = Path(source_dir.name) / "excel-target.sqlite"
        _make_db(database, rows=1)
        db_id = "excel-endpoint-test"
        desktop_bridge._DB_AGENT_DBS[db_id] = {
            "id": db_id, "name": "excel-target.sqlite", "path": str(database),
            "tables": ["items"], "kind": "sqlite", "attachedAt": 0,
        }
        self.addCleanup(desktop_bridge._DB_AGENT_DBS.pop, db_id, None)
        self.addCleanup(desktop_bridge._DB_AGENT_CACHE.pop, db_id, None)
        from openpyxl import Workbook
        workbook_path = Path(source_dir.name) / "merge.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "items"
        worksheet.append(["value"])
        worksheet.append(["from-excel"])
        workbook.save(workbook_path)
        workbook.close()
        form = FormData()
        form.add_field(
            "file", workbook_path.read_bytes(), filename="merge.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        prepared = await self.client.post(
            f"/db/excel/import/prepare?dbId={db_id}&sid=excel-endpoint",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
            data=form,
        )
        self.assertEqual(prepared.status, 200, await prepared.text())
        preview = (await prepared.json())["preview"]
        self.assertEqual(preview["rowCount"], 1)
        confirmed = await self.client.post(
            "/db/excel/import/confirm",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
            json={"dbId": db_id, "confirmId": preview["confirmId"], "approved": True},
        )
        self.assertEqual(confirmed.status, 200, await confirmed.text())
        self.assertEqual((await confirmed.json())["answer"]["write"]["affected"], 1)
        with closing(sqlite3.connect(database)) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM items").fetchone()[0], 2)
        exported = await self.client.get(
            f"/db/excel/export?dbId={db_id}",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
        )
        self.assertEqual(exported.status, 200)
        self.assertIn("application/vnd.openxmlformats", exported.headers["Content-Type"])
        self.assertTrue((await exported.read()).startswith(b"PK"))

    async def test_streaming_upload_enforces_limit_and_removes_partial_file(self):
        form = FormData()
        form.add_field(
            "file", b"0123456789abcdef", filename="oversized.sqlite",
            content_type="application/octet-stream",
        )
        with mock.patch.object(desktop_bridge, "_UPLOAD_MAX_BYTES", 8):
            response = await self.client.post(
                "/upload?sid=oversized",
                headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
                data=form,
            )
        self.assertEqual(response.status, 413, await response.text())
        payload = await response.json()
        self.assertFalse(payload["ok"])
        self.assertIn("too large", payload["error"])
        self.assertEqual(
            [path for path in Path(self.upload_tmp.name).rglob("*") if path.is_file()],
            [],
        )

    async def test_cross_origin_request_is_rejected(self):
        response = await self.client.get(
            "/status",
            headers={
                "X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN,
                "Origin": "https://attacker.invalid",
            },
        )
        self.assertEqual(response.status, 403)

    async def test_query_token_bootstraps_cookie_then_redirects(self):
        response = await self.client.get(
            f"/status?token={desktop_bridge.BRIDGE_TOKEN}",
            allow_redirects=False,
        )
        self.assertEqual(response.status, 302)
        self.assertEqual(response.headers["Location"], "/status")
        cookie = response.cookies[desktop_bridge._AUTH_COOKIE]
        self.assertTrue(cookie["httponly"])
        self.assertEqual(cookie["samesite"], "Strict")

    async def test_derived_role_tokens_are_distinct_and_permission_matrix_is_enforced(self):
        tokens = desktop_bridge._ROLE_TOKENS
        self.assertEqual(len(set(tokens.values())), 3)
        viewer_headers = {"X-DBQuill-Token": tokens["viewer"]}
        operator_headers = {"X-DBQuill-Token": tokens["operator"]}

        status = await self.client.get("/status", headers=viewer_headers)
        self.assertEqual(status.status, 200)
        status_payload = await status.json()
        self.assertEqual(status_payload["access"]["role"], "viewer")
        self.assertNotIn("appRoot", status_payload)
        self.assertNotIn("profilePath", status_payload)
        context = await self.client.get("/db/auth/context", headers=viewer_headers)
        context_payload = await context.json()
        self.assertEqual(context_payload["access"]["label"], "查看者")
        self.assertFalse(context_payload["access"]["capabilities"]["manage_workspace"])
        viewer_databases = await self.client.get("/db/databases", headers=viewer_headers)
        viewer_entry = next(
            item for item in (await viewer_databases.json())["databases"]
            if item["id"] == self.role_db_id
        )
        self.assertNotIn("path", viewer_entry)
        operator_databases = await self.client.get("/db/databases", headers=operator_headers)
        operator_entry = next(
            item for item in (await operator_databases.json())["databases"]
            if item["id"] == self.role_db_id
        )
        self.assertEqual(operator_entry["path"], "D:/secret/roles.db")

        viewer_mutation = await self.client.post(
            "/db/semantics", headers=viewer_headers, json={},
        )
        self.assertEqual(viewer_mutation.status, 403)
        operator_mutation = await self.client.post(
            "/db/semantics", headers=operator_headers, json={},
        )
        self.assertEqual(operator_mutation.status, 200)
        operator_backup = await self.client.post(
            "/db/audit/backups", headers=operator_headers, json={},
        )
        self.assertEqual(operator_backup.status, 403)

        denied_events = db_audit_store.list_events(category="access_control")
        self.assertEqual(len(denied_events), 2)
        self.assertEqual(denied_events[0]["actor"], "local_operator")
        self.assertEqual(denied_events[1]["actor"], "local_viewer")

    async def test_viewer_query_token_bootstraps_viewer_cookie(self):
        viewer_token = desktop_bridge._ROLE_TOKENS["viewer"]
        response = await self.client.get(
            "/status?token=" + viewer_token, allow_redirects=False,
        )
        self.assertEqual(response.status, 302)
        self.assertEqual(response.cookies[desktop_bridge._AUTH_COOKIE].value, viewer_token)
        self.assertEqual(response.headers["X-DBQuill-Role"], "viewer")

    async def test_v01_auth_header_is_accepted_and_response_is_dual_named(self):
        viewer_token = desktop_bridge._ROLE_TOKENS["viewer"]
        response = await self.client.get(
            "/db/auth/context",
            headers={"X-DBAgent-Token": viewer_token},
        )
        self.assertEqual(response.status, 200)
        self.assertEqual(response.headers["X-DBQuill-Role"], "viewer")
        self.assertEqual(response.headers["X-DBAgent-Role"], "viewer")

    async def test_expiring_credential_can_be_issued_used_audited_and_revoked(self):
        admin_headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        issued = await self.client.post(
            "/db/auth/credentials", headers=admin_headers,
            json={
                "label": "本地操作员甲", "role": "operator", "ttlHours": 24,
                "databaseScope": {
                    "mode": "restricted",
                    "databaseRefs": [desktop_bridge._database_scope_ref(
                        desktop_bridge._DB_AGENT_DBS[self.role_db_id]
                    )],
                },
            },
        )
        self.assertEqual(issued.status, 201)
        issued_payload = await issued.json()
        credential = issued_payload["credential"]
        token = credential["token"]
        credential_ref = credential["credentialRef"]
        self.assertTrue(token.startswith("id_"))

        credential_headers = {"X-DBQuill-Token": token}
        status = await self.client.get("/status", headers=credential_headers)
        status_payload = await status.json()
        self.assertEqual(status_payload["access"]["role"], "operator")
        self.assertEqual(status_payload["access"]["principalKind"], "credential")
        self.assertEqual(status_payload["access"]["credentialRef"], credential_ref)
        context = await self.client.get("/db/auth/context", headers=credential_headers)
        principal = (await context.json())["access"]["principal"]
        self.assertEqual(principal["label"], "本地操作员甲")
        self.assertEqual(principal["credentialRef"], credential_ref)

        forbidden_list = await self.client.get(
            "/db/auth/credentials", headers=credential_headers,
        )
        self.assertEqual(forbidden_list.status, 403)
        forbidden_backup = await self.client.post(
            "/db/audit/backups", headers=credential_headers, json={},
        )
        self.assertEqual(forbidden_backup.status, 403)
        denial_events = [
            event for event in db_audit_store.list_events(category="access_control")
            if event["action"] == "deny" and event["details"].get("credential_ref")
        ]
        self.assertEqual(len(denial_events), 2)
        self.assertTrue(all(event["actor"] == "local_operator" for event in denial_events))
        self.assertTrue(all(
            event["details"]["credential_ref"] == credential_ref
            for event in denial_events
        ))

        listed = await self.client.get("/db/auth/credentials", headers=admin_headers)
        listed_payload = await listed.json()
        serialized_list = json.dumps(listed_payload, ensure_ascii=False)
        self.assertNotIn(token, serialized_list)
        self.assertNotIn("token_hash", serialized_list)
        self.assertEqual(listed_payload["credentials"][0]["status"], "active")

        revoked = await self.client.post(
            f"/db/auth/credentials/{credential['id']}/revoke",
            headers=admin_headers,
        )
        self.assertEqual(revoked.status, 200)
        rejected = await self.client.get("/status", headers=credential_headers)
        self.assertEqual(rejected.status, 401)
        all_events = db_audit_store.list_events(category="access_control")
        self.assertNotIn(token, json.dumps(all_events, ensure_ascii=False))
        self.assertNotIn("本地操作员甲", json.dumps(all_events, ensure_ascii=False))
        self.assertEqual(db_audit_store.reconciliation_status()["unresolved_count"], 0)

    async def test_credential_issue_fails_closed_when_audit_is_unavailable(self):
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        with mock.patch.object(desktop_bridge, "_audit_store", None):
            response = await self.client.post(
                "/db/auth/credentials", headers=headers,
                json={
                    "label": "不得落库", "role": "viewer", "ttlHours": 24,
                    "databaseScope": {"mode": "all", "databaseRefs": []},
                },
            )
        self.assertEqual(response.status, 503)
        self.assertEqual(db_identity_store.list_credentials(), [])

    async def test_restricted_credential_filters_databases_and_blocks_body_db_id(self):
        admin_headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        allowed_ref = desktop_bridge._database_scope_ref(
            desktop_bridge._DB_AGENT_DBS[self.role_db_id],
        )
        issued = await self.client.post(
            "/db/auth/credentials", headers=admin_headers,
            json={
                "label": "仅角色库操作", "role": "operator", "ttlHours": 24,
                "databaseScope": {
                    "mode": "restricted", "databaseRefs": [allowed_ref],
                },
            },
        )
        self.assertEqual(issued.status, 201)
        token = (await issued.json())["credential"]["token"]
        headers = {"X-DBQuill-Token": token}

        context = await self.client.get("/db/auth/context", headers=headers)
        access = (await context.json())["access"]
        self.assertEqual(
            access["databaseScope"], {
                "mode": "restricted", "databaseCount": 1,
                "tableScopeDatabaseCount": 0, "tableScopeTableCount": 0,
                "columnScopeTableCount": 0, "columnScopeColumnCount": 0,
                "rowScopeTableCount": 0, "rowScopeFilterCount": 0,
            },
        )
        databases = await self.client.get("/db/databases", headers=headers)
        visible = (await databases.json())["databases"]
        self.assertEqual([item["id"] for item in visible], [self.role_db_id])
        self.assertEqual(visible[0]["databaseRef"], allowed_ref)

        allowed = await self.client.post(
            "/db/ask", headers=headers,
            json={"dbId": self.role_db_id, "question": "count"},
        )
        self.assertEqual(allowed.status, 200)
        denied = await self.client.post(
            "/db/ask", headers=headers,
            json={"dbId": self.other_db_id, "question": "count"},
        )
        self.assertEqual(denied.status, 404)
        denied_write = await self.client.post(
            "/db/write/confirm", headers=headers,
            json={"dbId": self.other_db_id, "confirmId": "private", "approved": True},
        )
        self.assertEqual(denied_write.status, 404)
        attach_path = Path(self.identity_tmp.name) / "new-scope.db"
        with closing(sqlite3.connect(attach_path)) as conn:
            conn.execute("CREATE TABLE items(id INTEGER PRIMARY KEY)")
            conn.commit()
        before_ids = set(desktop_bridge._DB_AGENT_DBS)
        denied_attach = await self.client.post(
            "/db/attach", headers=headers, json={"path": str(attach_path)},
        )
        denied_attach_payload = await denied_attach.json()
        self.assertEqual(denied_attach.status, 403, denied_attach_payload)
        self.assertEqual(set(desktop_bridge._DB_AGENT_DBS), before_ids)
        events = db_audit_store.list_events(category="access_control")
        scope_denials = [event for event in events if event["action"] == "deny_database_scope"]
        self.assertEqual(len(scope_denials), 3)
        self.assertEqual(scope_denials[0]["details"]["database_scope_count"], 1)

    async def test_credential_issue_requires_explicit_known_database_scope(self):
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        missing = await self.client.post(
            "/db/auth/credentials", headers=headers,
            json={"label": "缺少范围", "role": "viewer", "ttlHours": 24},
        )
        self.assertEqual(missing.status, 400)
        unknown = await self.client.post(
            "/db/auth/credentials", headers=headers,
            json={
                "label": "未知范围", "role": "viewer", "ttlHours": 24,
                "databaseScope": {"mode": "restricted", "databaseRefs": ["f" * 64]},
            },
        )
        self.assertEqual(unknown.status, 400)

        allowed_ref = desktop_bridge._database_scope_ref(
            desktop_bridge._DB_AGENT_DBS[self.role_db_id],
        )
        unknown_table = await self.client.post(
            "/db/auth/credentials", headers=headers,
            json={
                "label": "unknown table", "role": "viewer", "ttlHours": 24,
                "databaseScope": {
                    "mode": "restricted", "databaseRefs": [allowed_ref],
                    "tableScopes": {allowed_ref: ["missing_table"]},
                },
            },
        )
        self.assertEqual(unknown_table.status, 400)

        remote_id = "remote-table-scope-test"
        remote_entry = {
            "id": remote_id, "name": "remote", "tables": ["items"],
            "conn": {
                "dialect": "postgresql", "host": "localhost", "port": 5432,
                "database": "example", "user": "reader", "password": "secret",
            },
        }
        desktop_bridge._DB_AGENT_DBS[remote_id] = remote_entry
        try:
            remote_ref = desktop_bridge._database_scope_ref(remote_entry)
            remote_scope = await self.client.post(
                "/db/auth/credentials", headers=headers,
                json={
                    "label": "remote table", "role": "viewer", "ttlHours": 24,
                    "databaseScope": {
                        "mode": "restricted", "databaseRefs": [remote_ref],
                        "tableScopes": {remote_ref: ["items"]},
                    },
                },
            )
            self.assertEqual(remote_scope.status, 400)
        finally:
            desktop_bridge._DB_AGENT_DBS.pop(remote_id, None)

    async def test_table_scope_filters_schema_sessions_runs_schedules_and_audit(self):
        credential = await self._issue_table_restricted("admin")
        headers = {"X-DBQuill-Token": credential["token"]}
        database_ref = desktop_bridge._database_scope_ref(
            desktop_bridge._DB_AGENT_DBS[self.role_db_id],
        )
        scope_ref = desktop_bridge.hashlib.sha256(
            f"dbagent-table-access-v1:{database_ref}:items".encode("utf-8")
        ).hexdigest()
        other_scope_ref = desktop_bridge.hashlib.sha256(
            f"dbagent-table-access-v1:{database_ref}:secrets".encode("utf-8")
        ).hexdigest()

        context = await self.client.get("/db/auth/context", headers=headers)
        self.assertEqual((await context.json())["access"]["databaseScope"], {
            "mode": "restricted", "databaseCount": 1,
            "tableScopeDatabaseCount": 1, "tableScopeTableCount": 1,
            "columnScopeTableCount": 0, "columnScopeColumnCount": 0,
            "rowScopeTableCount": 0, "rowScopeFilterCount": 0,
        })
        databases = await self.client.get("/db/databases", headers=headers)
        visible_db = (await databases.json())["databases"][0]
        self.assertEqual(visible_db["tables"], ["items"])
        self.assertTrue(visible_db["tableScopeRestricted"])

        desktop_bridge._DB_SESSIONS.update({
            "matching-table-session": {
                "id": "matching-table-session", "dbId": self.role_db_id,
                "accessScopeRef": scope_ref, "messages": [], "updatedAt": 3,
            },
            "all-table-session": {
                "id": "all-table-session", "dbId": self.role_db_id,
                "accessScopeRef": "all", "messages": [], "updatedAt": 2,
            },
            "other-table-session": {
                "id": "other-table-session", "dbId": self.role_db_id,
                "accessScopeRef": other_scope_ref, "messages": [], "updatedAt": 1,
            },
        })
        desktop_bridge._DB_RUNS.update({
            "matching-table-run": {
                "id": "matching-table-run", "dbId": self.role_db_id,
                "accessScopeRef": scope_ref, "result": None,
            },
            "all-table-run": {
                "id": "all-table-run", "dbId": self.role_db_id,
                "accessScopeRef": "all", "result": None,
            },
        })
        sessions = await self.client.get("/db/sessions", headers=headers)
        self.assertEqual(
            [item["id"] for item in (await sessions.json())["sessions"]],
            ["matching-table-session"],
        )
        matching_run = await self.client.get(
            "/db/ask/matching-table-run/progress", headers=headers,
        )
        self.assertEqual(matching_run.status, 200)
        hidden_run = await self.client.get(
            "/db/ask/all-table-run/progress", headers=headers,
        )
        self.assertEqual(hidden_run.status, 404)
        schedules = await self.client.get("/db/schedules", headers=headers)
        self.assertEqual((await schedules.json())["tasks"], [])
        detach = await self.client.delete(
            f"/db/databases/{self.role_db_id}", headers=headers,
        )
        self.assertEqual(detach.status, 404)
        self.assertIn(self.role_db_id, desktop_bridge._DB_AGENT_DBS)
        audit = await self.client.get(
            f"/db/audit?dbId={self.role_db_id}", headers=headers,
        )
        self.assertEqual(audit.status, 404)

    async def test_column_scope_is_validated_and_reported_without_hidden_names(self):
        path = Path(self.identity_tmp.name) / "column-scope.db"
        with closing(sqlite3.connect(path)) as conn:
            conn.execute(
                "CREATE TABLE items(id INTEGER PRIMARY KEY, public TEXT, secret TEXT)"
            )
            conn.execute("INSERT INTO items(public, secret) VALUES ('visible', 'hidden')")
            conn.commit()
        entry = desktop_bridge._DB_AGENT_DBS[self.role_db_id]
        entry.update({"path": str(path), "tables": ["items"]})
        database_ref = desktop_bridge._database_scope_ref(entry)
        admin_headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}

        unknown = await self.client.post(
            "/db/auth/credentials", headers=admin_headers,
            json={
                "label": "unknown column", "role": "viewer", "ttlHours": 24,
                "databaseScope": {
                    "mode": "restricted", "databaseRefs": [database_ref],
                    "tableScopes": {database_ref: ["items"]},
                    "columnScopes": {database_ref: {"items": ["missing"]}},
                },
            },
        )
        self.assertEqual(unknown.status, 400)

        issued = await self.client.post(
            "/db/auth/credentials", headers=admin_headers,
            json={
                "label": "public fields", "role": "viewer", "ttlHours": 24,
                "databaseScope": {
                    "mode": "restricted", "databaseRefs": [database_ref],
                    "tableScopes": {database_ref: ["ITEMS"]},
                    "columnScopes": {
                        database_ref: {"items": ["PUBLIC", "id"]},
                    },
                },
            },
        )
        self.assertEqual(issued.status, 201, await issued.text())
        credential = (await issued.json())["credential"]
        self.assertEqual(
            credential["databaseScope"]["columnScopes"],
            {database_ref: {"items": ["id", "public"]}},
        )
        headers = {"X-DBQuill-Token": credential["token"]}
        context = await self.client.get("/db/auth/context", headers=headers)
        self.assertEqual((await context.json())["access"]["databaseScope"], {
            "mode": "restricted", "databaseCount": 1,
            "tableScopeDatabaseCount": 1, "tableScopeTableCount": 1,
            "columnScopeTableCount": 1, "columnScopeColumnCount": 2,
            "rowScopeTableCount": 0, "rowScopeFilterCount": 0,
        })
        databases = await self.client.get("/db/databases", headers=headers)
        visible = (await databases.json())["databases"][0]
        self.assertTrue(visible["tableScopeRestricted"])
        self.assertTrue(visible["columnScopeRestricted"])
        self.assertNotIn("secret", json.dumps(visible))
        principal = desktop_bridge._principal_for_token(credential["token"])
        principal_token = desktop_bridge._REQUEST_PRINCIPAL.set(principal)
        role_token = desktop_bridge._REQUEST_ROLE.set("viewer")
        try:
            agent = desktop_bridge._db_get_agent(self.role_db_id)
            self.assertEqual(
                [column.name for column in agent.schema.tables["items"].columns],
                ["id", "public"],
            )
            chart = desktop_bridge._db_chart_exec(
                self.role_db_id, "items", "public", "", "count",
            )
            self.assertEqual(chart["labels"], ["visible"])
            with self.assertRaises(ValueError):
                desktop_bridge._db_chart_exec(
                    self.role_db_id, "items", "secret", "", "count",
                )
            semantic_store = mock.Mock()
            semantic_store.list_entries.return_value = [
                {
                    "id": "visible-semantic", "kind": "column_alias",
                    "term": "公开字段", "table": "items", "column": "public",
                    "description": "visible",
                },
                {
                    "id": "hidden-semantic", "kind": "column_alias",
                    "term": "秘密字段", "table": "items", "column": "secret",
                    "description": "hidden",
                },
            ]
            with mock.patch.object(desktop_bridge, "_semantic_store", semantic_store):
                semantics = desktop_bridge._db_semantics(entry, agent.schema)
            self.assertEqual([item["id"] for item in semantics], ["visible-semantic"])
            proposal = types.SimpleNamespace(
                table="items", kind="UPDATE",
                access_scope_ref=desktop_bridge._current_access_scope_ref(entry),
            )
            self.assertTrue(desktop_bridge._write_proposal_scope_allowed(entry, proposal))
            narrower = dict(principal)
            narrower["databaseScope"] = {
                **principal["databaseScope"],
                "columnScopes": {database_ref: {"items": ["id"]}},
            }
            narrower_token = desktop_bridge._REQUEST_PRINCIPAL.set(narrower)
            try:
                self.assertFalse(
                    desktop_bridge._write_proposal_scope_allowed(entry, proposal)
                )
            finally:
                desktop_bridge._REQUEST_PRINCIPAL.reset(narrower_token)
        finally:
            desktop_bridge._REQUEST_ROLE.reset(role_token)
            desktop_bridge._REQUEST_PRINCIPAL.reset(principal_token)

    async def test_row_scope_is_validated_filters_agent_and_chart_and_is_read_only(self):
        path = Path(self.identity_tmp.name) / "row-scope.db"
        with closing(sqlite3.connect(path)) as conn:
            conn.execute(
                "CREATE TABLE items(id INTEGER PRIMARY KEY, public TEXT, tenant TEXT)"
            )
            conn.executemany(
                "INSERT INTO items(public, tenant) VALUES (?, ?)",
                [("north", "tenant-a"), ("south", "tenant-b")],
            )
            conn.commit()
        entry = desktop_bridge._DB_AGENT_DBS[self.role_db_id]
        entry.update({"path": str(path), "tables": ["items"]})
        database_ref = desktop_bridge._database_scope_ref(entry)
        admin_headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}

        unknown = await self.client.post(
            "/db/auth/credentials", headers=admin_headers,
            json={
                "label": "unknown row column", "role": "viewer", "ttlHours": 24,
                "databaseScope": {
                    "mode": "restricted", "databaseRefs": [database_ref],
                    "tableScopes": {database_ref: ["items"]},
                    "rowScopes": {database_ref: {"items": [
                        {"column": "missing", "operator": "eq", "value": "tenant-a"},
                    ]}},
                },
            },
        )
        self.assertEqual(unknown.status, 400)

        issued = await self.client.post(
            "/db/auth/credentials", headers=admin_headers,
            json={
                "label": "north rows", "role": "operator", "ttlHours": 24,
                "databaseScope": {
                    "mode": "restricted", "databaseRefs": [database_ref],
                    "tableScopes": {database_ref: ["ITEMS"]},
                    "columnScopes": {database_ref: {"items": ["id", "public"]}},
                    "rowScopes": {database_ref: {"items": [
                        {"column": "TENANT", "operator": "eq", "value": "tenant-a"},
                    ]}},
                },
            },
        )
        self.assertEqual(issued.status, 201, await issued.text())
        credential = (await issued.json())["credential"]
        self.assertEqual(
            credential["databaseScope"]["rowScopes"],
            {database_ref: {"items": [
                {"column": "tenant", "operator": "eq", "value": "tenant-a"},
            ]}},
        )
        headers = {"X-DBQuill-Token": credential["token"]}
        context = await self.client.get("/db/auth/context", headers=headers)
        access = (await context.json())["access"]
        self.assertEqual(access["databaseScope"]["rowScopeTableCount"], 1)
        self.assertEqual(access["databaseScope"]["rowScopeFilterCount"], 1)
        self.assertFalse(access["capabilities"]["approve_bounded_write"])
        self.assertFalse(access["capabilities"]["approve_high_risk_write"])

        databases = await self.client.get("/db/databases", headers=headers)
        visible = (await databases.json())["databases"][0]
        self.assertTrue(visible["rowScopeRestricted"])
        principal = desktop_bridge._principal_for_token(credential["token"])
        principal_token = desktop_bridge._REQUEST_PRINCIPAL.set(principal)
        role_token = desktop_bridge._REQUEST_ROLE.set("operator")
        try:
            agent = desktop_bridge._db_get_agent(self.role_db_id)
            self.assertEqual(agent.schema.tables["items"].row_count, 1)
            self.assertEqual(
                agent.security.execute("SELECT public FROM items").rows, [["north"]],
            )
            self.assertIsNotNone(
                agent.security.execute("SELECT * FROM main.items").error,
            )
            chart = desktop_bridge._db_chart_exec(
                self.role_db_id, "items", "public", "", "count",
            )
            self.assertEqual(chart["labels"], ["north"])
            with self.assertRaises(dc.WriteSecurityError):
                agent.write_security.validate_write(
                    "UPDATE items SET public='changed' WHERE id=1"
                )
            proposal = types.SimpleNamespace(
                table="items", kind="UPDATE",
                access_scope_ref=desktop_bridge._current_access_scope_ref(entry),
            )
            self.assertFalse(desktop_bridge._write_proposal_scope_allowed(entry, proposal))
        finally:
            desktop_bridge._REQUEST_ROLE.reset(role_token)
            desktop_bridge._REQUEST_PRINCIPAL.reset(principal_token)

    async def test_restricted_scope_hides_sessions_runs_schedules_and_global_logs(self):
        credential = await self._issue_restricted("viewer")
        headers = {"X-DBQuill-Token": credential["token"]}
        desktop_bridge._DB_SESSIONS.update({
            "allowed-session": {
                "id": "allowed-session", "dbId": self.role_db_id,
                "title": "allowed", "messages": [], "updatedAt": 2,
            },
            "other-session": {
                "id": "other-session", "dbId": self.other_db_id,
                "title": "private", "messages": [], "updatedAt": 1,
            },
        })
        desktop_bridge._DB_RUNS.update({
            "allowed-run": {
                "id": "allowed-run", "dbId": self.role_db_id,
                "question": "allowed", "result": None,
            },
            "other-run": {
                "id": "other-run", "dbId": self.other_db_id,
                "question": "private", "result": None,
            },
        })

        sessions = await self.client.get("/db/sessions", headers=headers)
        self.assertEqual(
            [item["id"] for item in (await sessions.json())["sessions"]],
            ["allowed-session"],
        )
        hidden_session = await self.client.get(
            "/db/session/other-session", headers=headers,
        )
        self.assertEqual(hidden_session.status, 404)
        hidden_session_reuse = await self.client.post(
            "/db/ask", headers=headers,
            json={
                "dbId": self.role_db_id, "sessionId": "other-session",
                "question": "continue",
            },
        )
        self.assertEqual(hidden_session_reuse.status, 404)
        allowed_run = await self.client.get(
            "/db/ask/allowed-run/progress", headers=headers,
        )
        self.assertEqual(allowed_run.status, 200)
        hidden_run = await self.client.get(
            "/db/ask/other-run/progress", headers=headers,
        )
        self.assertEqual(hidden_run.status, 404)

        schedules = await self.client.get("/db/schedules", headers=headers)
        self.assertEqual(
            [item["id"] for item in (await schedules.json())["tasks"]],
            ["allowed-task"],
        )
        hidden_run_now = await self.client.post(
            "/db/schedules/other-task/run", headers=headers, json={},
        )
        self.assertEqual(hidden_run_now.status, 404)
        self.fake_scheduler.run_now.assert_not_called()
        logs = await self.client.get("/db/schedules/logs", headers=headers)
        self.assertEqual((await logs.json())["logs"], [])

    async def test_restricted_admin_cannot_escalate_to_global_credentials_or_backups(self):
        credential = await self._issue_restricted("admin")
        headers = {"X-DBQuill-Token": credential["token"]}
        context = await self.client.get("/db/auth/context", headers=headers)
        capabilities = (await context.json())["access"]["capabilities"]
        self.assertFalse(capabilities["manage_credentials"])
        self.assertFalse(capabilities["create_audit_backup"])
        credentials = await self.client.get("/db/auth/credentials", headers=headers)
        self.assertEqual(credentials.status, 403)
        backup = await self.client.post(
            "/db/audit/backups", headers=headers,
            json={"dbId": self.role_db_id},
        )
        self.assertEqual(backup.status, 403)
        denials = [
            event for event in db_audit_store.list_events(category="access_control")
            if event["action"] == "deny_database_scope"
        ]
        self.assertEqual(len(denials), 2)

    async def test_restricted_audit_requires_and_filters_authorized_database(self):
        credential = await self._issue_restricted("viewer")
        headers = {"X-DBQuill-Token": credential["token"]}
        db_audit_store.append_event(
            category="system", action="allowed_event", outcome="succeeded",
            summary="allowed", risk="low", actor="system",
            database_key=desktop_bridge._db_semantic_key(
                desktop_bridge._DB_AGENT_DBS[self.role_db_id],
            ),
        )
        db_audit_store.append_event(
            category="system", action="private_event", outcome="succeeded",
            summary="private", risk="low", actor="system",
            database_key=desktop_bridge._db_semantic_key(
                desktop_bridge._DB_AGENT_DBS[self.other_db_id],
            ),
        )
        missing_scope = await self.client.get("/db/audit", headers=headers)
        self.assertEqual(missing_scope.status, 400)
        visible = await self.client.get(
            f"/db/audit?dbId={self.role_db_id}", headers=headers,
        )
        payload = await visible.json()
        self.assertEqual([event["action"] for event in payload["events"]], ["allowed_event"])
        self.assertEqual(payload["reconciliation"]["scope"], "database")
        self.assertTrue(payload["backups"]["scope_restricted"])
        hidden = await self.client.get(
            f"/db/audit/export?dbId={self.other_db_id}", headers=headers,
        )
        self.assertEqual(hidden.status, 404)


class SessionHandlerTests(AioHTTPTestCase):
    async def get_application(self):
        app = web.Application(middlewares=[desktop_bridge.cors_middleware])
        app.router.add_delete("/db/session/{sid}", desktop_bridge.db_session_delete_handler)
        return app

    async def asyncSetUp(self):
        await super().asyncSetUp()
        self.original_store = desktop_bridge._sess_store
        desktop_bridge._sess_store = None
        desktop_bridge._DB_SESSIONS.clear()

    async def asyncTearDown(self):
        desktop_bridge._DB_SESSIONS.clear()
        desktop_bridge._sess_store = self.original_store
        await super().asyncTearDown()

    async def test_memory_only_session_can_be_deleted(self):
        desktop_bridge._DB_SESSIONS["memory-session"] = {
            "id": "memory-session",
            "messages": [],
            "updatedAt": 0,
        }
        response = await self.client.delete(
            "/db/session/memory-session",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
        )
        self.assertEqual(response.status, 200)
        self.assertTrue((await response.json())["ok"])
        self.assertNotIn("memory-session", desktop_bridge._DB_SESSIONS)


class DatabaseSafetyTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)

    def tearDown(self):
        self.tmp.cleanup()

    def test_large_limit_is_bounded_before_materialization(self):
        path = self.root / "many.db"
        _make_db(path, rows=1000)
        security = dc.SQLSecurity(dc.DBConnector(str(path)), max_rows=10)
        result = security.execute("SELECT id, value FROM items LIMIT 1000000")
        self.assertIsNone(result.error)
        self.assertEqual(len(result.rows), 10)
        self.assertTrue(result.truncated)

    def test_read_query_releases_temporary_database_handle(self):
        path = self.root / "releasable.db"
        _make_db(path, rows=3)
        security = dc.SQLSecurity(dc.DBConnector(str(path)), max_rows=10)
        result = security.execute("SELECT id FROM items ORDER BY id")
        self.assertIsNone(result.error)
        path.unlink()
        self.assertFalse(path.exists())

    def test_sql_security_distinguishes_code_from_literals_and_comments(self):
        path = self.root / "lexical.db"
        _make_db(path, rows=2)
        read = dc.SQLSecurity(dc.DBConnector(str(path)), max_rows=10)
        literal = read.execute("SELECT 'DROP TABLE items; PRAGMA writable_schema' AS note")
        self.assertIsNone(literal.error)
        self.assertEqual(literal.rows[0][0], "DROP TABLE items; PRAGMA writable_schema")
        commented = read.execute("SELECT COUNT(*) FROM items -- DROP TABLE items;")
        self.assertIsNone(commented.error)
        self.assertEqual(commented.rows, [[2]])
        with self.assertRaises(dc.SQLSecurityError):
            read.validate("SELECT load_extension('unsafe')")
        with self.assertRaises(dc.SQLSecurityError):
            read.validate('SELECT "load_extension"(\'unsafe\')')
        with self.assertRaises(dc.SQLSecurityError):
            read.validate("SELECT 1 /*!50000 UNION SELECT 2 */")
        executable_text = read.execute("SELECT '/*!50000 UNION SELECT 2 */' AS note")
        self.assertIsNone(executable_text.error)

        write = dc.WriteSecurity()
        with self.assertRaises(dc.WriteSecurityError):
            write.validate_write("UPDATE items SET value='x' /* WHERE id=1 */")
        with self.assertRaises(dc.WriteSecurityError):
            write.validate_write("UPDATE items SET value='WHERE id=1'")
        with self.assertRaises(dc.WriteSecurityError):
            write.validate_write("UPDATE items SET value='x' # WHERE id=1")
        with self.assertRaises(dc.WriteSecurityError):
            write.validate_write("UPDATE items SET value=$$WHERE id=1$$")
        with self.assertRaises(dc.WriteSecurityError):
            write.validate_write(
                "UPDATE items SET value='x' /* outer /* WHERE id=1 */ still comment */"
            )
        allowed = write.validate_write("UPDATE items SET value='safe;verified' WHERE id=1")
        self.assertEqual(allowed["kind"], "UPDATE")

    def test_table_scope_is_enforced_by_schema_and_sqlite_authorizer(self):
        path = self.root / "table-scope.db"
        _make_db(path, rows=2)
        with closing(sqlite3.connect(path)) as conn:
            conn.execute("CREATE TABLE secrets(id INTEGER PRIMARY KEY, value TEXT)")
            conn.execute("INSERT INTO secrets(value) VALUES ('hidden')")
            conn.execute(
                "CREATE TABLE links(id INTEGER PRIMARY KEY, secret_id INTEGER "
                "REFERENCES secrets(id))"
            )
            conn.commit()

        agent = dc.DBQuillAgent(
            db_path=str(path), sample_rows=0, allowed_tables=["items", "links"],
        )
        self.assertEqual(list(agent.schema.tables), ["items", "links"])
        self.assertNotIn("secrets", agent.schema.compact())
        allowed = agent.security.execute("SELECT id, value FROM items ORDER BY id")
        self.assertIsNone(allowed.error)
        blocked = agent.security.execute(
            "WITH hidden AS (SELECT id FROM secrets) "
            "SELECT items.id FROM items JOIN hidden ON hidden.id = items.id"
        )
        self.assertIsNotNone(blocked.error)
        self.assertNotIn("hidden", json.dumps(agent.schema.l1_index()))
        # Defense in depth: even if a future regression hands RAG a broader
        # schema object, its actual SQLite reads remain table-authorized.
        agent.rag.schema = dc.SchemaDiscovery(agent.connector, sample_rows=0).discover()
        evidence = agent.rag._recall(["hidden"])
        self.assertFalse(any(item["table"] == "secrets" for item in evidence))

    def test_table_scope_blocks_ddl_cross_table_preview_and_final_write(self):
        path = self.root / "table-write-scope.db"
        _make_db(path, rows=1)
        with closing(sqlite3.connect(path)) as conn:
            conn.execute("CREATE TABLE secrets(id INTEGER PRIMARY KEY, value TEXT)")
            conn.execute("INSERT INTO secrets(value) VALUES ('hidden')")
            conn.execute(
                "CREATE TRIGGER items_secret_sync AFTER UPDATE ON items "
                "BEGIN UPDATE secrets SET value = NEW.value WHERE id = 1; END"
            )
            conn.commit()
        agent = dc.DBQuillAgent(
            db_path=str(path), sample_rows=0, allowed_tables=["items"],
        )
        with self.assertRaises(dc.WriteSecurityError):
            agent.write_security.validate_write("CREATE TABLE expanded(id INTEGER)")
        with self.assertRaises(dc.WriteSecurityError):
            agent.write_security.validate_write("UPDATE secrets SET value='x' WHERE id=1")

        cross_table_sql = (
            "UPDATE items SET value=(SELECT value FROM secrets WHERE id=1) WHERE id=1"
        )
        meta = agent.write_security.validate_write(cross_table_sql)
        with self.assertRaises(sqlite3.DatabaseError):
            agent.write_previewer.preview(cross_table_sql, meta)

        confirm_id = "table-scope-cross-read"
        dc.WRITE_REGISTRY.register(dc.WriteProposal(
            confirm_id=confirm_id, sql=cross_table_sql, kind="UPDATE", table="items",
            summary_zh="cross table", dangerous=False, preview={"affected": 1},
            db_path=str(path.resolve()),
        ))
        result = agent.confirm_write(confirm_id, approve=True)
        self.assertEqual(result.kind, "error")

        trigger_confirm_id = "table-scope-trigger-write"
        dc.WRITE_REGISTRY.register(dc.WriteProposal(
            confirm_id=trigger_confirm_id,
            sql="UPDATE items SET value='triggered' WHERE id=1",
            kind="UPDATE", table="items", summary_zh="trigger write",
            dangerous=False, preview={"affected": 1}, db_path=str(path.resolve()),
        ))
        trigger_result = agent.confirm_write(trigger_confirm_id, approve=True)
        self.assertEqual(trigger_result.kind, "error")
        with closing(sqlite3.connect(path)) as conn:
            self.assertEqual(conn.execute("SELECT value FROM items WHERE id=1").fetchone()[0], "v0")
            self.assertEqual(conn.execute("SELECT value FROM secrets WHERE id=1").fetchone()[0], "hidden")

    def test_column_scope_filters_schema_and_physically_blocks_read_and_write_bypasses(self):
        path = self.root / "column-scope.db"
        with closing(sqlite3.connect(path)) as conn:
            conn.execute(
                "CREATE TABLE records(id INTEGER PRIMARY KEY, public TEXT, secret TEXT)"
            )
            conn.execute(
                "INSERT INTO records(public, secret) VALUES ('visible', 'hidden')"
            )
            conn.commit()
        agent = dc.DBQuillAgent(
            db_path=str(path), sample_rows=2, allowed_tables=["records"],
            allowed_columns={"records": ["id", "public"]},
        )
        table = agent.schema.tables["records"]
        self.assertEqual([column.name for column in table.columns], ["id", "public"])
        self.assertNotIn("secret", agent.schema.compact())
        allowed = agent.security.execute("SELECT id, public FROM records")
        self.assertIsNone(allowed.error)
        self.assertEqual(allowed.rows, [[1, "visible"]])
        counted = agent.security.execute("SELECT COUNT(*) FROM records")
        self.assertIsNone(counted.error)
        self.assertEqual(counted.rows, [[1]])
        for sql in (
            "SELECT secret FROM records",
            "SELECT public FROM records WHERE secret = 'hidden'",
            "WITH hidden AS (SELECT secret FROM records) SELECT * FROM hidden",
            "SELECT name FROM pragma_table_info('records')",
        ):
            with self.subTest(sql=sql):
                self.assertIsNotNone(agent.security.execute(sql).error)

        # A broader schema accidentally handed to RAG still cannot make a
        # physical read of the hidden field.
        agent.rag.schema = dc.SchemaDiscovery(agent.connector, sample_rows=0).discover()
        evidence = agent.rag._recall(["hidden"])
        self.assertFalse(any("hidden" in value for item in evidence for value in item["row"]))

        with self.assertRaises(dc.WriteSecurityError):
            agent.write_security.validate_write(
                "INSERT INTO records VALUES (2, 'shown', 'hidden')"
            )
        with self.assertRaises(dc.WriteSecurityError):
            agent.write_security.validate_write(
                "INSERT INTO records(id, public, secret) VALUES (2, 'shown', 'hidden')"
            )
        with self.assertRaises(dc.WriteSecurityError):
            agent.write_security.validate_write("DELETE FROM records WHERE id=1")

        update_sql = "UPDATE records SET public='changed' WHERE id=1"
        update_meta = agent.write_security.validate_write(update_sql)
        preview = agent.write_previewer.preview(update_sql, update_meta)
        self.assertEqual(preview["before"]["columns"], ["id", "public"])
        self.assertEqual(preview["after"]["columns"], ["id", "public"])
        hidden_filter = "UPDATE records SET public='x' WHERE secret='hidden'"
        with self.assertRaises(sqlite3.DatabaseError):
            agent.write_previewer.preview(
                hidden_filter, agent.write_security.validate_write(hidden_filter),
            )
        hidden_target = "UPDATE records SET secret='exposed' WHERE id=1"
        with self.assertRaises(sqlite3.DatabaseError):
            agent.write_previewer.preview(
                hidden_target, agent.write_security.validate_write(hidden_target),
            )

        confirm_id = "column-scope-hidden-target"
        dc.WRITE_REGISTRY.register(dc.WriteProposal(
            confirm_id=confirm_id, sql=hidden_target, kind="UPDATE", table="records",
            summary_zh="hidden target", dangerous=False, preview={"affected": 1},
            db_path=str(path.resolve()),
        ))
        result = agent.confirm_write(confirm_id, approve=True)
        self.assertEqual(result.kind, "error")
        with closing(sqlite3.connect(path)) as conn:
            self.assertEqual(
                conn.execute("SELECT public, secret FROM records WHERE id=1").fetchone(),
                ("visible", "hidden"),
            )

    def test_row_scope_filters_all_read_paths_and_denies_main_table_and_writes(self):
        path = self.root / "row-scope.db"
        injection_tenant = "north' OR 1=1 --"
        with closing(sqlite3.connect(path)) as conn:
            conn.execute(
                "CREATE TABLE records(id INTEGER PRIMARY KEY, public TEXT, tenant TEXT)"
            )
            conn.executemany(
                "INSERT INTO records(public, tenant) VALUES (?, ?)",
                [
                    ("north-only", "north"),
                    ("south-only", "south"),
                    ("literal-injection-only", injection_tenant),
                ],
            )
            conn.commit()

        agent = dc.DBQuillAgent(
            db_path=str(path), sample_rows=5, allowed_tables=["records"],
            allowed_columns={"records": ["id", "public"]},
            row_filters={"records": [
                {"column": "tenant", "operator": "eq", "value": "north"},
            ]},
        )
        table = agent.schema.tables["records"]
        self.assertEqual([column.name for column in table.columns], ["id", "public"])
        self.assertEqual(table.row_count, 1)
        self.assertIn("north-only", table.columns[1].sample_values)
        self.assertNotIn("south-only", agent.schema.compact())

        allowed = agent.security.execute("SELECT id, public FROM records")
        self.assertIsNone(allowed.error)
        self.assertEqual(allowed.rows, [[1, "north-only"]])
        self.assertEqual(
            agent.security.execute("SELECT COUNT(*) FROM records").rows, [[1]],
        )
        self.assertEqual(
            agent.security.execute("SELECT public FROM temp.records").rows,
            [["north-only"]],
        )
        for sql in (
            "SELECT * FROM main.records",
            "WITH bypass AS (SELECT * FROM main.records) SELECT * FROM bypass",
            "SELECT tenant FROM records",
            "SELECT name FROM pragma_table_info('records')",
        ):
            with self.subTest(sql=sql):
                self.assertIsNotNone(agent.security.execute(sql).error)

        agent.rag.schema = dc.SchemaDiscovery(agent.connector, sample_rows=0).discover()
        evidence = agent.rag._recall(["only"])
        recalled = [value for item in evidence for value in item["row"]]
        self.assertIn("north-only", recalled)
        self.assertNotIn("south-only", recalled)
        self.assertNotIn("literal-injection-only", recalled)

        for sql in (
            "INSERT INTO records(id, public) VALUES (4, 'new')",
            "UPDATE records SET public='changed' WHERE id=1",
            "DELETE FROM records WHERE id=1",
        ):
            with self.subTest(sql=sql), self.assertRaises(dc.WriteSecurityError):
                agent.write_security.validate_write(sql)

        literal_agent = dc.DBQuillAgent(
            db_path=str(path), sample_rows=0, allowed_tables=["records"],
            row_filters={"records": [
                {"column": "tenant", "operator": "eq", "value": injection_tenant},
            ]},
        )
        literal = literal_agent.security.execute("SELECT public FROM records")
        self.assertIsNone(literal.error)
        self.assertEqual(literal.rows, [["literal-injection-only"]])

    def test_confirmation_cannot_cross_database_boundary(self):
        first = self.root / "first.db"
        second = self.root / "second.db"
        _make_db(first)
        _make_db(second)
        proposal = dc.WriteProposal(
            confirm_id="cross-db-test",
            sql="UPDATE items SET value='changed' WHERE id=1",
            kind="UPDATE",
            table="items",
            summary_zh="update one row",
            dangerous=False,
            preview={},
            db_path=str(first.resolve()),
        )
        dc.WRITE_REGISTRY.register(proposal)

        second_agent = dc.DBQuillAgent(db_path=str(second), sample_rows=0)
        rejected = second_agent.confirm_write(proposal.confirm_id, approve=True)
        self.assertEqual(rejected.kind, "error")
        with closing(sqlite3.connect(second)) as conn:
            self.assertEqual(conn.execute("SELECT value FROM items WHERE id=1").fetchone()[0], "v0")

        first_agent = dc.DBQuillAgent(db_path=str(first), sample_rows=0)
        accepted = first_agent.confirm_write(proposal.confirm_id, approve=True)
        self.assertEqual(accepted.kind, "write_result")
        with closing(sqlite3.connect(first)) as conn:
            self.assertEqual(conn.execute("SELECT value FROM items WHERE id=1").fetchone()[0], "changed")

    def test_scheduled_sql_is_physical_read_only(self):
        path = self.root / "scheduled.db"
        _make_db(path, rows=2)
        blocked = db_scheduler._execute_sql(
            str(path), "UPDATE items SET value='unsafe' WHERE id=1",
        )
        self.assertFalse(blocked["ok"])
        self.assertTrue(blocked["blocked_write"])
        self.assertTrue(blocked["requires_confirmation"])
        result = db_scheduler._execute_sql(str(path), "SELECT value FROM items WHERE id=1")
        self.assertTrue(result["ok"])
        with closing(sqlite3.connect(path)) as conn:
            self.assertEqual(conn.execute("SELECT value FROM items WHERE id=1").fetchone()[0], "v0")

        with mock.patch.object(db_scheduler, "_TASKS_DIR", self.root / "readonly-tasks"):
            with self.assertRaisesRegex(ValueError, "仅允许单条 SELECT/WITH"):
                db_scheduler.create_task({
                    "name": "must not write", "dbId": "test", "type": "sql",
                    "sql": "DELETE FROM items WHERE id=1",
                    "schedule": {"mode": "interval", "minutes": 60},
                })

    def test_scheduled_nl_write_stops_at_confirmation_boundary(self):
        calls = {"ask": 0, "confirm": 0}

        class FakeAgent:
            def __init__(self, **kwargs):
                pass

            def ask(self, prompt, history=None):
                calls["ask"] += 1
                return types.SimpleNamespace(
                    kind="write_pending",
                    narrative="pending",
                    confirm_id="confirm-once",
                    columns=[],
                    rows=[],
                    sql="UPDATE items SET value='x' WHERE id=1",
                )

            def confirm_write(self, confirm_id, approve=True):
                calls["confirm"] += 1
                return types.SimpleNamespace(
                    kind="write_result",
                    narrative="written once",
                    confirm_id=confirm_id,
                    columns=[],
                    rows=[],
                    sql="UPDATE items SET value='x' WHERE id=1",
                )

        fake_module = types.ModuleType("dbquill_core")
        fake_module.DBQuillAgent = FakeAgent
        with mock.patch.dict(sys.modules, {"dbquill_core": fake_module}):
            result = db_scheduler._run_nl(str(self.root / "unused.db"), "update it")

        self.assertFalse(result["ok"])
        self.assertTrue(result["requires_confirmation"])
        self.assertTrue(result["blocked_write"])
        self.assertEqual(calls, {"ask": 1, "confirm": 0})

    def test_scheduled_nl_accepts_legacy_frontend_content_field(self):
        tasks_dir = self.root / "tasks"
        with mock.patch.object(db_scheduler, "_TASKS_DIR", tasks_dir):
            task = db_scheduler.create_task({
                "name": "legacy nl",
                "dbId": "test-db",
                "type": "nl",
                "sql": "summarize recent orders",
                "schedule": {"mode": "interval", "minutes": 60},
            })
        self.assertEqual(task["type"], "nl")
        self.assertEqual(task["prompt"], "summarize recent orders")
        self.assertEqual(task["sql"], "")

    def test_scheduled_nl_clarification_fails_without_execution(self):
        class FakeAgent:
            def __init__(self, **kwargs):
                pass

            def ask(self, prompt, history=None):
                return types.SimpleNamespace(
                    kind="clarification",
                    narrative="请补充目标表",
                    clarification={"missing": "target_table", "missing_label": "目标表"},
                    confirm_id=None,
                )

            def confirm_write(self, confirm_id, approve=True):
                raise AssertionError("clarification must never be auto-confirmed")

        fake_module = types.ModuleType("dbquill_core")
        fake_module.DBQuillAgent = FakeAgent
        with mock.patch.dict(sys.modules, {"dbquill_core": fake_module}):
            result = db_scheduler._run_nl(str(self.root / "unused.db"), "删除记录")

        self.assertFalse(result["ok"])
        self.assertIn("缺少目标表", result["error"])

    def test_scheduled_run_emits_correlated_redacted_audit_events(self):
        path = self.root / "scheduled-audit.db"
        _make_db(path, rows=2)
        tasks_dir = self.root / "scheduled-audit-tasks"
        logs_dir = self.root / "scheduled-audit-logs"
        events = []

        def audit_sink(**event):
            events.append(event)
            return {"sequence": len(events)}

        raw_sql = "SELECT value FROM items WHERE id=1"
        with mock.patch.object(db_scheduler, "_TASKS_DIR", tasks_dir), \
             mock.patch.object(db_scheduler, "_LOGS_DIR", logs_dir), \
             mock.patch.object(
                 db_scheduler, "_resolver",
                 lambda db_id: {"path": str(path)} if db_id == "test-db" else None,
             ), \
             mock.patch.object(db_scheduler, "_audit_sink", audit_sink):
            task = db_scheduler.create_task({
                "name": "audited query", "dbId": "test-db", "type": "sql",
                "sql": raw_sql, "schedule": {"mode": "interval", "minutes": 60},
            })
            result = db_scheduler.run_task_once(task, trigger="manual")

        self.assertTrue(result["ok"])
        self.assertEqual([event["outcome"] for event in events], ["pending", "succeeded"])
        self.assertEqual(events[0]["correlation_id"], events[1]["correlation_id"])
        self.assertEqual(events[0]["actor"], "scheduler")
        self.assertNotIn(raw_sql, json.dumps(events, ensure_ascii=False))
        persisted_log = next(logs_dir.glob("*.md")).read_text(encoding="utf-8")
        self.assertNotIn(raw_sql, persisted_log)
        self.assertNotIn("audited query", persisted_log)

    def test_legacy_scheduler_logs_are_redacted_and_renamed(self):
        logs_dir = self.root / "legacy-scheduler-logs"
        logs_dir.mkdir()
        raw_sql = "UPDATE secret_table SET password='exposed' WHERE id=1"
        legacy = logs_dir / "2026-08-01_sensitive task.md"
        legacy.write_text(
            "## 执行时间：2026-08-01 10:00:00\n"
            f"- SQL：\n```sql\n{raw_sql}\n```\n- 提问：列出密码\n",
            encoding="utf-8",
        )
        with mock.patch.object(db_scheduler, "_LOGS_DIR", logs_dir):
            self.assertEqual(db_scheduler._redact_legacy_logs(), 1)
            self.assertEqual(db_scheduler._redact_legacy_logs(), 0)
        files = list(logs_dir.glob("*.md"))
        self.assertEqual(len(files), 1)
        self.assertTrue(files[0].name.startswith("legacy_"))
        self.assertNotIn("sensitive", files[0].name)
        content = files[0].read_text(encoding="utf-8")
        self.assertNotIn(raw_sql, content)
        self.assertNotIn("列出密码", content)

    def test_chart_helpers_release_sqlite_file_handles(self):
        path = self.root / "chart-lock.db"
        with closing(sqlite3.connect(path)) as conn:
            conn.execute("CREATE TABLE sales(category TEXT, amount REAL)")
            conn.executemany(
                "INSERT INTO sales VALUES (?, ?)",
                [("office", 12.5), ("software", 99.0)],
            )
            conn.commit()
        db_id = "chart-lock-test"
        desktop_bridge._DB_AGENT_DBS[db_id] = {
            "id": db_id,
            "name": path.name,
            "path": str(path),
            "tables": ["sales"],
        }
        try:
            result = desktop_bridge._db_chart_exec(
                db_id, "sales", "category", "amount", "sum"
            )
            self.assertEqual(sum(result["values"]), 111.5)
        finally:
            desktop_bridge._DB_AGENT_DBS.pop(db_id, None)
        # Windows raises PermissionError here when a connection leaked.
        path.unlink()
        self.assertFalse(path.exists())


class NaturalLanguageDatabaseTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "operations.db"
        _make_db(self.path, rows=3)
        self.agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)

    def tearDown(self):
        self.tmp.cleanup()

    def test_basic_conversation_is_local_and_has_no_database_side_effect_shape(self):
        cases = {
            "你好": "我在",
            "谢谢": "不客气",
            "你是谁": "DBQuill",
            "你能做什么": "查看数据库",
            "你好吗": "状态正常",
            "陪我聊聊": "当然可以",
            "我有点累": "不太轻松",
            "再见": "再见",
            "晚安": "再见",
        }
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            for question, expected in cases.items():
                with self.subTest(question=question):
                    answer = self.agent.ask(question)
                    self.assertEqual(answer.kind, "conversation")
                    self.assertIn(expected, answer.narrative)
                    self.assertIsNone(answer.operation)
                    self.assertIsNone(answer.sql)
                    self.assertIsNone(answer.confirm_id)
                    self.assertEqual(answer.rows, [])

    def test_social_prefix_does_not_capture_database_operations(self):
        self.assertIsNone(self.agent.conversation.answer("你好，查询 items 有多少行"))
        self.assertIsNone(self.agent.conversation.answer("谢谢，帮我删除 items 中 id=1 的记录"))
        self.assertIsNone(self.agent.conversation.answer("你能不能查一下 items"))

        with mock.patch.object(
            self.agent.nl2sql,
            "answer",
            return_value=dc.DBAnswer(kind="query", narrative="查询完成", rows=[[3]]),
        ) as query:
            answer = self.agent.ask("你好，查询 items 有多少行")
        self.assertEqual(answer.kind, "query")
        query.assert_called_once()

        routed = self.agent.router.classify(
            "谢谢，帮我删除 items 中 id=1 的记录",
            self.agent.schema.compact(),
        )
        self.assertEqual(routed.intent, "write")

    def test_pending_candidate_is_not_reclassified_as_conversation(self):
        pending = {
            "missing": "target_table",
            "missing_label": "目标表",
            "candidates": [{"label": "谢谢", "prompt": "谢谢"}],
        }
        self.assertIsNone(self.agent.conversation.answer("谢谢", clarification=pending))
        answer = self.agent.conversation.answer(
            "你好",
            clarification={**pending, "candidates": [{"label": "items", "prompt": "items"}]},
        )
        self.assertEqual(answer.kind, "conversation")
        self.assertIn("目标表", answer.narrative)
        self.assertIn("仍然保留", answer.narrative)

    def _dimension_agent(self) -> dc.DBQuillAgent:
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute(
                "CREATE TABLE customers ("
                "id INTEGER PRIMARY KEY, region TEXT, city TEXT, status TEXT)"
            )
            conn.executemany(
                "INSERT INTO customers(id, region, city, status) VALUES (?, ?, ?, ?)",
                [
                    (1, "华东", "上海", "active"),
                    (2, "华北", "北京", "active"),
                    (3, "华东", "杭州", "inactive"),
                    (4, "华南", "深圳", "active"),
                ],
            )
            conn.commit()
        return dc.DBQuillAgent(
            db_path=str(self.path),
            sample_rows=0,
            semantic_entries=[
                {
                    "kind": "dimension", "term": "客户区域", "table": "customers",
                    "column": "region", "hierarchy": {"name": "客户地域", "level": 1},
                },
                {
                    "kind": "dimension", "term": "客户城市", "table": "customers",
                    "column": "city", "hierarchy": {"name": "客户地域", "level": 2},
                },
                {
                    "kind": "dimension", "term": "活跃客户区域", "table": "customers",
                    "column": "region", "hierarchy": {"name": "活跃客户地域", "level": 1},
                    "filters": [{"column": "status", "operator": "eq", "value": "active"}],
                },
                {
                    "kind": "dimension", "term": "活跃客户城市", "table": "customers",
                    "column": "city", "hierarchy": {"name": "活跃客户地域", "level": 2},
                    "filters": [{"column": "status", "operator": "eq", "value": "active"}],
                },
                {
                    "kind": "metric", "term": "活跃客户数", "table": "customers",
                    "aggregation": "count",
                    "filters": [{"column": "status", "operator": "eq", "value": "active"}],
                },
                {
                    "kind": "metric", "term": "客户总数", "table": "customers",
                    "aggregation": "count",
                },
                {
                    "kind": "metric", "term": "非活跃客户数", "table": "customers",
                    "aggregation": "count",
                    "filters": [{"column": "status", "operator": "eq", "value": "inactive"}],
                },
            ],
        )

    def _trend_agent(self) -> dc.DBQuillAgent:
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute(
                "CREATE TABLE orders ("
                "id INTEGER PRIMARY KEY, amount REAL, status TEXT, created_at DATE)"
            )
            conn.executemany(
                "INSERT INTO orders(id, amount, status, created_at) VALUES (?, ?, ?, ?)",
                [
                    (1, 100.0, "paid", "2026-01-05"),
                    (2, 50.0, "pending", "2026-01-20"),
                    (3, 200.0, "paid", "2026-02-01"),
                    (4, 300.0, "paid", "2026-04-15"),
                ],
            )
            conn.commit()
        return dc.DBQuillAgent(
            db_path=str(self.path),
            sample_rows=0,
            semantic_entries=[
                {"kind": "table_alias", "term": "订单", "table": "orders"},
                {
                    "kind": "time_field", "term": "下单时间", "table": "orders",
                    "column": "created_at", "default_grain": "month",
                },
                {
                    "kind": "business_calendar", "term": "公司业务日历",
                    "table": "orders", "column": "created_at",
                    "calendar": {
                        "fiscal_year_start_month": 4, "fiscal_year_start_day": 1,
                        "fiscal_year_label": "start_year", "timezone": "Asia/Shanghai",
                        "week_start": 1, "weekend_days": [6, 7],
                    },
                },
                {
                    "kind": "enum_value", "term": "仅已付款订单", "table": "orders",
                    "column": "status", "value": "paid",
                },
                {
                    "kind": "metric", "term": "已付款成交额", "table": "orders",
                    "column": "amount", "aggregation": "sum",
                    "filters": [{"column": "status", "operator": "eq", "value": "paid"}],
                },
                {
                    "kind": "metric", "term": "成交额", "table": "orders",
                    "column": "amount", "aggregation": "sum",
                },
                {
                    "kind": "metric", "term": "订单笔数", "table": "orders",
                    "aggregation": "count",
                },
                {
                    "kind": "metric", "term": "已付款订单笔数", "table": "orders",
                    "aggregation": "count",
                    "filters": [{"column": "status", "operator": "eq", "value": "paid"}],
                },
            ],
        )

    def _multi_metric_agent(self) -> dc.DBQuillAgent:
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute(
                "CREATE TABLE orders ("
                "id INTEGER PRIMARY KEY, customer_id INTEGER, amount REAL, status TEXT)"
            )
            conn.executemany(
                "INSERT INTO orders(id, customer_id, amount, status) VALUES (?, ?, ?, ?)",
                [
                    (1, 1, 100.0, "paid"),
                    (2, 1, 50.0, "pending"),
                    (3, 2, 200.0, "paid"),
                    (4, 3, 300.0, "paid"),
                ],
            )
            conn.commit()
        return dc.DBQuillAgent(
            db_path=str(self.path),
            sample_rows=0,
            semantic_entries=[
                {
                    "kind": "enum_value", "term": "已付款订单", "table": "orders",
                    "column": "status", "value": "paid",
                },
                {
                    "kind": "metric", "term": "订单笔数", "table": "orders",
                    "aggregation": "count",
                },
                {
                    "kind": "metric", "term": "成交额", "table": "orders",
                    "column": "amount", "aggregation": "sum",
                },
                {
                    "kind": "metric", "term": "已付款成交额", "table": "orders",
                    "column": "amount", "aggregation": "sum",
                    "filters": [{"column": "status", "operator": "eq", "value": "paid"}],
                },
                {
                    "kind": "metric", "term": "客户数", "table": "orders",
                    "column": "customer_id", "aggregation": "count_distinct",
                },
                {
                    "kind": "metric", "term": "平均订单金额", "table": "orders",
                    "column": "amount", "aggregation": "avg",
                },
                {
                    "kind": "metric", "term": "最大订单金额", "table": "orders",
                    "column": "amount", "aggregation": "max",
                },
                {
                    "kind": "metric", "term": "最小订单金额", "table": "orders",
                    "column": "amount", "aggregation": "min",
                },
            ],
        )

    def test_schema_overview_executes_without_llm(self):
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            answer = self.agent.ask("有哪些表？每张表多少行？")
        self.assertEqual(answer.kind, "schema")
        self.assertEqual(answer.columns, ["表名", "行数", "字段数"])
        self.assertEqual(answer.rows, [["items", 3, 2]])
        self.assertEqual(answer.operation["action"], "inspect_schema")
        self.assertEqual(answer.operation["status"], "executed")
        self.assertEqual(answer.operation["engine"], "native")

    def test_multi_metric_executes_single_deterministic_query_without_model(self):
        agent = self._multi_metric_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("统计成交额和订单笔数")

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.columns, ["成交额", "订单笔数"])
        self.assertEqual(answer.rows, [[650.0, 4]])
        self.assertEqual(answer.metric_plan["engine"], "native_multi_metric")
        self.assertEqual(answer.metric_plan["version"], "1.0")
        self.assertEqual(answer.metric_plan["status"], "executed")
        self.assertEqual(answer.operation["action"], "select")

    def test_multi_metric_normalizes_single_table_respectively_to_query(self):
        agent = self._multi_metric_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ), mock.patch.object(
            agent.orchestrator, "answer", side_effect=AssertionError("graph should not run"),
        ):
            answer = agent.ask("分别统计订单笔数和成交额")

        self.assertEqual(answer.columns, ["订单笔数", "成交额"])
        self.assertEqual(answer.rows, [[4, 650.0]])
        self.assertEqual(answer.operation["action"], "select")
        self.assertEqual(answer.operation["intent"], "query")
        self.assertIsNone(answer.graph)

    def test_multi_metric_keeps_measure_filters_inside_conditional_aggregate(self):
        agent = self._multi_metric_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("统计成交额和已付款成交额")

        self.assertEqual(answer.columns, ["成交额", "已付款成交额"])
        self.assertEqual(answer.rows, [[650.0, 600.0]])
        self.assertIn("SUM(CASE WHEN", answer.sql)
        self.assertNotIn(" WHERE ", answer.sql.upper())
        self.assertEqual(answer.metric_plan["measures"][1]["filters"], [
            {"column": "status", "operator": "eq", "value": "paid"},
        ])

    def test_multi_metric_applies_one_global_enum_filter_to_all_measures(self):
        agent = self._multi_metric_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("统计已付款订单的成交额和订单笔数")

        self.assertEqual(answer.rows, [[600.0, 3]])
        self.assertEqual(answer.metric_plan["global_filters"], [
            {"column": "status", "operator": "eq", "value": "paid"},
        ])
        self.assertIn('WHERE ("orders"."status" = \'paid\')', answer.sql)

    def test_multi_metric_accepts_six_measures_in_question_order(self):
        agent = self._multi_metric_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask(
                "统计订单笔数、成交额、已付款成交额、客户数、平均订单金额和最大订单金额"
            )

        self.assertEqual(answer.columns, [
            "订单笔数", "成交额", "已付款成交额", "客户数", "平均订单金额", "最大订单金额",
        ])
        self.assertEqual(answer.rows, [[4, 650.0, 600.0, 3, 162.5, 300.0]])
        self.assertEqual(len(answer.metric_plan["measures"]), 6)

    def test_multi_metric_arithmetic_expression_falls_back_whole_question(self):
        agent = self._multi_metric_agent()
        generated = dc.DBAnswer(
            kind="query", narrative="算术表达由模型链路处理",
            sql="SELECT 1", columns=["value"], rows=[[1]],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask("计算成交额除以订单笔数")

        execute.assert_called_once()
        self.assertIsNone(answer.metric_plan)
        self.assertEqual(answer.narrative, generated.narrative)

    def test_seven_metrics_fall_back_without_partial_deterministic_execution(self):
        agent = self._multi_metric_agent()
        generated = dc.DBAnswer(
            kind="query", narrative="七指标交回模型链路",
            sql="SELECT 1", columns=["value"], rows=[[1]],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask(
                "统计订单笔数、成交额、已付款成交额、客户数、平均订单金额、"
                "最大订单金额和最小订单金额"
            )

        execute.assert_called_once()
        self.assertIsNone(answer.metric_plan)
        self.assertEqual(answer.narrative, generated.narrative)

    def test_table_structure_executes_without_llm(self):
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            answer = self.agent.ask("items 表有哪些字段？")
        self.assertEqual(answer.kind, "schema")
        self.assertEqual(answer.operation["action"], "inspect_table")
        self.assertEqual(answer.operation["target_tables"], ["items"])
        self.assertEqual([row[0] for row in answer.rows], ["id", "value"])

    def test_generic_write_capability_uses_model_and_opens_table_picker(self):
        routed = {
            "intent": "write",
            "interaction": "guided_insert",
            "target_table": "",
            "confidence": 0.97,
            "reasoning": "用户希望进入引导录入流程",
        }
        with mock.patch.object(dc, "_llm_ask_json", return_value=routed) as classify:
            answer = self.agent.ask("能不能写入")
        classify.assert_called_once()
        self.assertEqual(answer.kind, "write_form")
        self.assertEqual(answer.operation["action"], "insert")
        self.assertEqual(answer.operation["mode"], "write")
        self.assertEqual(answer.operation["status"], "needs_clarification")
        self.assertTrue(answer.operation["requires_confirmation"])
        self.assertEqual(answer.write["mode"], "insert_form")
        self.assertEqual(answer.write["selected_table"], "")
        self.assertEqual([item["name"] for item in answer.write["tables"]], ["items"])
        self.assertIsNone(answer.confirm_id)
        self.assertEqual(answer.steps[0]["source"], "model")
        self.assertEqual(answer.steps[0]["interaction"], "guided_insert")

    def test_table_write_request_uses_model_and_returns_one_real_example_row(self):
        routed = {
            "intent": "write",
            "interaction": "guided_insert",
            "target_table": "items",
            "confidence": 0.98,
            "reasoning": "用户要向指定表录入数据",
        }
        with mock.patch.object(dc, "_llm_ask_json", return_value=routed) as classify:
            answer = self.agent.ask("向 items 表写入数据")
        classify.assert_called_once()
        self.assertEqual(answer.kind, "write_form")
        self.assertEqual(answer.write["selected_table"], "items")
        self.assertEqual(
            [column["name"] for column in answer.write["columns"]],
            ["id", "value"],
        )
        self.assertTrue(answer.write["columns"][0]["automatic"])
        self.assertEqual(answer.write["example"]["columns"], ["id", "value"])
        self.assertEqual(answer.write["example"]["rows"], [[1, "v0"]])

    def test_model_can_map_natural_business_phrase_to_guided_insert_table(self):
        routed = {
            "intent": "write",
            "interaction": "guided_insert",
            "target_table": "items",
            "confidence": 0.94,
            "reasoning": "用户想录入一条商品数据",
        }
        with mock.patch.object(dc, "_llm_ask_json", return_value=routed) as classify:
            answer = self.agent.ask("帮我录一条商品数据")
        classify.assert_called_once()
        self.assertEqual(answer.kind, "write_form")
        self.assertEqual(answer.write["selected_table"], "items")
        self.assertEqual(answer.steps[0]["source"], "model")

    def test_guided_insert_keeps_safe_form_fallback_when_model_is_unavailable(self):
        with mock.patch.object(dc, "_llm_ask_json", side_effect=dc.DBQuillError("offline")) as classify:
            answer = self.agent.ask("我想往库里加点东西")
        classify.assert_called_once()
        self.assertEqual(answer.kind, "write_form")
        self.assertEqual(answer.write["selected_table"], "")
        self.assertEqual(answer.steps[0]["source"], "safety_fallback")

    def test_model_target_table_must_exist_in_authorized_schema(self):
        routed = {
            "intent": "write",
            "interaction": "guided_insert",
            "target_table": "hallucinated_table",
            "confidence": 0.92,
            "reasoning": "用户想录入数据",
        }
        with mock.patch.object(dc, "_llm_ask_json", return_value=routed):
            answer = self.agent.ask("帮我新增一条数据")
        self.assertEqual(answer.kind, "write_form")
        self.assertEqual(answer.write["selected_table"], "")
        self.assertEqual(answer.operation["target_tables"], [])

    def test_structured_insert_previews_before_explicit_confirmation(self):
        with closing(sqlite3.connect(self.path)) as conn:
            before = conn.execute("SELECT COUNT(*) FROM items").fetchone()[0]
        pending = self.agent.prepare_structured_insert("items", [
            {"column": "id", "mode": "omit", "value": ""},
            {"column": "value", "mode": "value", "value": "new row"},
        ])
        self.assertEqual(pending.kind, "write_pending")
        self.assertEqual(pending.operation["status"], "awaiting_confirmation")
        self.assertIn("new row", json.dumps(pending.write["preview"], ensure_ascii=False))
        with closing(sqlite3.connect(self.path)) as conn:
            self.assertEqual(conn.execute("SELECT COUNT(*) FROM items").fetchone()[0], before)
        result = self.agent.confirm_write(pending.confirm_id, approve=True)
        self.assertEqual(result.kind, "write_result")
        with closing(sqlite3.connect(self.path)) as conn:
            self.assertEqual(conn.execute("SELECT value FROM items ORDER BY id DESC LIMIT 1").fetchone()[0], "new row")

    def test_structured_insert_quotes_injection_payload_as_data(self):
        payload = "x'); DROP TABLE items; --"
        pending = self.agent.prepare_structured_insert("items", [
            {"column": "id", "mode": "omit"},
            {"column": "value", "mode": "value", "value": payload},
        ])
        self.assertEqual(pending.kind, "write_pending")
        result = self.agent.confirm_write(pending.confirm_id, approve=True)
        self.assertEqual(result.kind, "write_result")
        with closing(sqlite3.connect(self.path)) as conn:
            self.assertEqual(conn.execute("SELECT value FROM items ORDER BY id DESC LIMIT 1").fetchone()[0], payload)
            self.assertIsNotNone(conn.execute("SELECT name FROM sqlite_master WHERE name='items'").fetchone())

    def test_structured_insert_enforces_required_fields_and_honors_defaults(self):
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute(
                "CREATE TABLE contacts (id INTEGER PRIMARY KEY, name TEXT NOT NULL, "
                "city TEXT DEFAULT 'unknown', note TEXT)"
            )
            conn.commit()
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        rejected = agent.prepare_structured_insert("contacts", [
            {"column": "name", "mode": "omit"},
        ])
        self.assertEqual(rejected.kind, "error")
        self.assertIn("必填", rejected.error)
        pending = agent.prepare_structured_insert("contacts", [
            {"column": "id", "mode": "omit"},
            {"column": "name", "mode": "value", "value": "Alice"},
            {"column": "city", "mode": "omit"},
            {"column": "note", "mode": "null"},
        ])
        self.assertEqual(pending.kind, "write_pending")
        self.assertEqual(agent.confirm_write(pending.confirm_id, approve=True).kind, "write_result")
        with closing(sqlite3.connect(self.path)) as conn:
            self.assertEqual(
                conn.execute("SELECT name, city, note FROM contacts").fetchone(),
                ("Alice", "unknown", None),
            )

    def test_structured_create_table_previews_then_requires_explicit_confirmation(self):
        pending = self.agent.prepare_structured_create_table("customer_orders", [
            {
                "name": "id", "type": "INTEGER", "primaryKey": True,
                "autoIncrement": True, "nullable": False, "defaultMode": "none",
            },
            {
                "name": "customer_name", "type": "TEXT", "nullable": False,
                "unique": False, "defaultMode": "none",
            },
            {
                "name": "created_at", "type": "DATETIME", "nullable": False,
                "defaultMode": "current_timestamp",
            },
        ])
        self.assertEqual(pending.kind, "write_pending")
        self.assertEqual(pending.operation["action"], "create")
        self.assertEqual(pending.operation["risk"], "high")
        self.assertEqual(pending.operation["status"], "awaiting_confirmation")
        self.assertTrue(pending.write["preview"]["ddl"]["after"]["exists"])
        with closing(sqlite3.connect(self.path)) as conn:
            self.assertIsNone(conn.execute(
                "SELECT name FROM sqlite_master WHERE name='customer_orders'"
            ).fetchone())
        result = self.agent.confirm_write(pending.confirm_id, approve=True)
        self.assertEqual(result.kind, "write_result")
        with closing(sqlite3.connect(self.path)) as conn:
            self.assertIsNotNone(conn.execute(
                "SELECT name FROM sqlite_master WHERE name='customer_orders'"
            ).fetchone())
            columns = conn.execute("PRAGMA table_info(customer_orders)").fetchall()
        self.assertEqual([column[1] for column in columns], ["id", "customer_name", "created_at"])
        self.assertEqual(columns[2][4], "CURRENT_TIMESTAMP")

    def test_structured_create_table_rejects_unsafe_or_ambiguous_definitions(self):
        cases = [
            ("bad-name", [{"name": "id", "type": "INTEGER"}], "表名"),
            ("safe_table", [{"name": "bad-name", "type": "TEXT"}], "字段名"),
            ("safe_table", [{"name": "payload", "type": "TEXT); DROP TABLE items;--"}], "类型"),
            ("safe_table", [
                {"name": "id", "type": "INTEGER", "primaryKey": True},
                {"name": "other_id", "type": "INTEGER", "primaryKey": True},
            ], "一个主键"),
            ("safe_table", [{
                "name": "name", "type": "TEXT", "nullable": False,
                "defaultMode": "null",
            }], "NULL"),
        ]
        for table, columns, expected in cases:
            with self.subTest(table=table, expected=expected):
                answer = self.agent.prepare_structured_create_table(table, columns)
                self.assertEqual(answer.kind, "error")
                self.assertIn(expected, answer.error)

    def test_table_scoped_agent_cannot_prepare_structured_ddl(self):
        agent = dc.DBQuillAgent(
            db_path=str(self.path), sample_rows=0,
            allowed_tables=["items"], allowed_columns={"items": ["id", "value"]},
        )
        answer = agent.prepare_structured_create_table(
            "new_table", [{"name": "id", "type": "INTEGER"}],
        )
        self.assertEqual(answer.kind, "error")
        self.assertIn("不允许创建", answer.error)

    def test_row_scoped_credential_cannot_open_structured_write_form(self):
        agent = dc.DBQuillAgent(
            db_path=str(self.path),
            sample_rows=0,
            allowed_tables=["items"],
            allowed_columns={"items": ["id", "value"]},
            row_filters={"items": [{"column": "id", "operator": "eq", "value": 1}]},
        )
        answer = agent.write_form("items")
        self.assertEqual(answer.kind, "error")
        self.assertIn("只读", answer.error)

    def test_write_language_maps_to_confirmed_operation(self):
        result = dc.IntentResult(intent="write", confidence=0.95, reasoning="明确修改数据")
        plan = self.agent.operation_planner.from_intent(
            "把 items 中 id=1 的 value 改成 ok",
            result,
        )
        self.assertEqual(plan.action, "update")
        self.assertEqual(plan.mode, "write")
        self.assertEqual(plan.target_tables, ["items"])
        self.assertTrue(plan.requires_confirmation)
        self.assertEqual(plan.risk, "medium")

    def test_query_answer_carries_executed_operation(self):
        routed = dc.IntentResult(intent="query", confidence=0.9, reasoning="明确查数")
        generated = dc.DBAnswer(
            kind="query",
            narrative="查询完成",
            sql="SELECT id, value FROM items LIMIT 10",
            columns=["id", "value"],
            rows=[[1, "v0"]],
        )
        with mock.patch.object(self.agent.router, "classify", return_value=routed), \
             mock.patch.object(self.agent.nl2sql, "answer", return_value=generated):
            answer = self.agent.ask("查询 items 的前十条数据")
        self.assertEqual(answer.operation["action"], "select")
        self.assertEqual(answer.operation["target_tables"], ["items"])
        self.assertEqual(answer.operation["status"], "executed")
        self.assertEqual(answer.operation["sql"], generated.sql)

    def test_write_answer_waits_for_confirmation(self):
        routed = dc.IntentResult(intent="write", confidence=0.95, reasoning="明确修改数据")
        proposed = dc.DBAnswer(
            kind="write_pending",
            narrative="将修改一行",
            sql="UPDATE items SET value='ok' WHERE id=1",
            confirm_id="pending-test",
            write={"kind": "UPDATE", "table": "items", "dangerous": False, "preview": {}},
        )
        with mock.patch.object(self.agent.router, "classify", return_value=routed), \
             mock.patch.object(self.agent.write_executor, "prepare", return_value=proposed):
            answer = self.agent.ask("把 items 中 id=1 的 value 改成 ok")
        self.assertEqual(answer.operation["action"], "update")
        self.assertEqual(answer.operation["status"], "awaiting_confirmation")
        self.assertTrue(answer.operation["requires_confirmation"])

    def test_destructive_language_is_high_risk(self):
        result = dc.IntentResult(intent="write", confidence=0.95, reasoning="明确删除结构")
        plan = self.agent.operation_planner.from_intent("删除 items 表", result)
        self.assertEqual(plan.action, "drop")
        self.assertEqual(plan.risk, "high")
        self.assertTrue(plan.requires_confirmation)

    def test_row_delete_is_not_misclassified_as_drop_table(self):
        result = dc.IntentResult(intent="write", confidence=0.95, reasoning="明确删除数据")
        plan = self.agent.operation_planner.from_intent("删除 items 表中 id=1 的记录", result)
        self.assertEqual(plan.action, "delete")
        self.assertEqual(plan.target_tables, ["items"])
        self.assertEqual(plan.risk, "high")
        english = self.agent.operation_planner.from_intent("delete from items where id=1", result)
        self.assertEqual(english.action, "delete")

    def test_ambiguous_write_stops_before_llm(self):
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            answer = self.agent.ask("删除记录")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.operation["status"], "needs_clarification")
        self.assertEqual(answer.clarification["missing"], "target_table")
        self.assertEqual(answer.clarification["candidates"][0]["label"], "items")

    def test_update_without_filter_requires_clarification(self):
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            answer = self.agent.ask("把 items 的 value 改成 ok")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "filter_condition")
        self.assertIsNone(answer.confirm_id)

    def test_create_table_without_columns_requires_definition(self):
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            answer = self.agent.ask("创建 orders 表")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "object_definition")
        self.assertEqual(answer.operation["target_tables"], ["orders"])

    def test_alter_without_column_definition_requires_clarification(self):
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            answer = self.agent.ask("给 items 添加一列")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "schema_change_definition")

    def test_add_column_ddl_is_not_misclassified_as_schema_view(self):
        # 回归：中文 DDL 请求曾被 plan_schema 当作只读结构查看，写链路不可达
        plan = self.agent.operation_planner.plan_schema("给 items 表新增 extra TEXT 字段。")
        self.assertIsNone(plan)
        routed = dc.IntentResult(intent="write", confidence=0.95, reasoning="明确修改结构")
        operation = self.agent.operation_planner.from_intent(
            "给 items 表新增 extra TEXT 字段。", routed,
        )
        self.assertEqual(operation.action, "alter")
        self.assertEqual(operation.risk, "high")
        self.assertEqual(operation.target_tables, ["items"])

    def test_add_column_with_name_and_type_reaches_write_prepare(self):
        # 完整 DDL 定义（字段名+类型齐全）不应停在澄清，也不应回落为只读结构查看；
        # 应直接进入写执行器准备预览
        proposed = dc.DBAnswer(
            kind="write_pending",
            narrative="将新增字段 extra",
            sql="ALTER TABLE items ADD COLUMN extra TEXT",
            confirm_id="ddl-pending",
            write={"kind": "ALTER", "table": "items", "dangerous": True, "preview": {}},
        )
        with mock.patch.object(self.agent.write_executor, "prepare", return_value=proposed):
            answer = self.agent.ask("给 items 表新增 extra TEXT 字段。")
        self.assertEqual(answer.kind, "write_pending")
        self.assertEqual(answer.operation["action"], "alter")
        self.assertNotEqual(answer.operation["action"], "inspect_table")
        self.assertEqual(answer.operation["risk"], "high")

    def test_drop_column_is_not_misclassified_as_drop_table(self):
        routed = dc.IntentResult(intent="write", confidence=0.95, reasoning="明确删除字段")
        operation = self.agent.operation_planner.from_intent(
            "删除 items 表的 value 字段。", routed,
        )
        self.assertEqual(operation.action, "alter")
        self.assertEqual(operation.risk, "high")
        self.assertEqual(operation.target_tables, ["items"])
        proposed = dc.DBAnswer(
            kind="write_pending",
            narrative="将删除字段 value",
            sql="ALTER TABLE items DROP COLUMN value",
            confirm_id="ddl-drop",
            write={"kind": "ALTER", "table": "items", "dangerous": True, "preview": {}},
        )
        with mock.patch.object(self.agent.write_executor, "prepare", return_value=proposed):
            answer = self.agent.ask("删除 items 表的 value 字段。")
        self.assertEqual(answer.kind, "write_pending")
        self.assertNotEqual(answer.operation["action"], "drop")
        self.assertEqual(answer.operation["action"], "alter")

    def test_drop_table_still_maps_to_drop(self):
        routed = dc.IntentResult(intent="write", confidence=0.95, reasoning="明确删除表")
        operation = self.agent.operation_planner.from_intent("删除 items 表", routed)
        self.assertEqual(operation.action, "drop")
        self.assertEqual(operation.risk, "high")

    def test_multi_table_generic_count_asks_for_target(self):
        _add_table(self.path)
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            answer = agent.ask("一共多少条记录？")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "target_table")
        self.assertEqual(
            [item["label"] for item in answer.clarification["candidates"]],
            ["items", "orders"],
        )

    def test_disconnected_cross_table_query_asks_for_relationship(self):
        _add_table(self.path)
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            answer = agent.ask("统计 items 和 orders 的数量")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "table_relationship")
        self.assertEqual(answer.operation["target_tables"], ["items", "orders"])

    def test_explicit_cross_table_relationship_allows_query(self):
        _add_table(self.path)
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        generated = dc.DBAnswer(
            kind="query", narrative="查询完成", sql="SELECT COUNT(*) FROM items",
            columns=["数量"], rows=[[3]],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask("统计 items 和 orders 的数量，关联条件：items.id = orders.id")
        execute.assert_called_once()
        self.assertEqual(answer.kind, "query")

    def test_derived_metric_without_definition_requires_clarification(self):
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            answer = self.agent.ask("统计 items 的转化率")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "metric_definition")

    def test_aggregate_without_numeric_field_requires_clarification(self):
        _add_table(self.path)
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            answer = agent.ask("计算 orders 的平均值")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "aggregation_field")
        self.assertEqual(answer.clarification["candidates"][0]["label"], "orders.amount")

    def test_explicit_aggregate_field_allows_query(self):
        _add_table(self.path)
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        generated = dc.DBAnswer(
            kind="query", narrative="平均值为 12.5", sql="SELECT AVG(amount) FROM orders",
            columns=["平均值"], rows=[[12.5]],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask("计算 orders.amount 的平均值")
        execute.assert_called_once()
        self.assertEqual(answer.kind, "query")

    def test_time_clarifications_accumulate_before_query(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN created_at TEXT")
            conn.execute("ALTER TABLE orders ADD COLUMN updated_at TEXT")
            conn.commit()
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            first = agent.ask("统计 orders 最近的趋势")
        self.assertEqual(first.clarification["missing"], "time_field")
        self.assertEqual(
            [item["label"] for item in first.clarification["candidates"]],
            ["orders.created_at", "orders.updated_at"],
        )

        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            second = agent.ask(
                "orders.created_at",
                history=[{"role": "user", "content": first.clarification["original_question"]}],
                clarification=first.clarification,
            )
        self.assertEqual(second.clarification["missing"], "time_range")
        self.assertEqual(second.clarification["candidates"][1]["label"], "最近 30 天")

        generated = dc.DBAnswer(
            kind="query", narrative="趋势查询完成", sql="SELECT created_at FROM orders",
            columns=["时间"], rows=[],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            third = agent.ask(
                "最近 30 天",
                history=[{"role": "user", "content": "orders.created_at"}],
                clarification=second.clarification,
            )
        execute.assert_not_called()
        self.assertEqual(third.clarification["missing"], "time_grain")
        self.assertEqual(third.clarification["candidates"][2]["label"], "按月")

        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            fourth = agent.ask(
                "月",
                history=[{"role": "user", "content": "最近 30 天"}],
                clarification=third.clarification,
            )
        execute.assert_called_once()
        self.assertEqual(fourth.kind, "query")

    def test_explicit_time_range_does_not_trigger_clarification(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN created_at TEXT")
            conn.commit()
        agent = dc.DBQuillAgent(
            db_path=str(self.path), sample_rows=0,
            semantic_entries=[{
                "kind": "time_field", "term": "下单时间", "table": "orders",
                "column": "created_at", "default_grain": "month",
            }],
        )
        generated = dc.DBAnswer(
            kind="query", narrative="趋势查询完成", sql="SELECT created_at FROM orders",
            columns=["时间"], rows=[],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask("统计 orders.created_at 最近 30 天的趋势")
        execute.assert_called_once()
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.semantic["matches"][0]["default_grain"], "month")

    def test_trend_without_explicit_or_configured_grain_requests_clarification(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN created_at TEXT")
            conn.commit()
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        with mock.patch.object(agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run")):
            answer = agent.ask("统计 orders.created_at 最近 30 天的趋势")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "time_grain")
        self.assertEqual(
            [item["label"] for item in answer.clarification["candidates"]],
            ["按日", "按周", "按月", "按季度", "按年"],
        )

    def test_explicit_time_grain_overrides_missing_default(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN created_at TEXT")
            conn.commit()
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        generated = dc.DBAnswer(
            kind="query", narrative="按周趋势完成", sql="SELECT created_at FROM orders",
            columns=["时间"], rows=[],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask("统计 orders.created_at 最近 30 天按周的趋势")
        execute.assert_called_once()
        self.assertEqual(answer.kind, "query")

    def test_date_range_without_from_word_hits_deterministic_calendar(self):
        # 回归：日期范围“2026-08-10 到 2026-08-18”不带“从”字也应命中确定性日历通道
        entries = [{
            "kind": "metric", "term": "已支付订单数", "table": "orders", "column": "id",
            "aggregation": "count",
            "filters": [{"column": "status", "operator": "eq", "value": "paid"}],
        }, {
            "kind": "time_field", "term": "下单日期", "table": "orders", "column": "created_at",
            "default_grain": "month",
        }, {
            "kind": "business_calendar", "term": "公司业务日历", "table": "orders",
            "column": "created_at",
            "calendar": {
                "fiscal_year_start_month": 1, "fiscal_year_start_day": 1,
                "fiscal_year_label": "start_year", "timezone": "UTC",
                "storage_basis": "declared_date", "week_start": 1, "weekend_days": [6, 7],
            },
        }]
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0, semantic_entries=entries)
        # created_at 为 TEXT，确定性通道要求 DATE 声明；直接验证编译器对“无从”范围的识别
        plan = agent.calendar_query.compiler.compile(
            "统计 2026-08-10 到 2026-08-18 工作日内的已支付订单数",
            agent.semantic_catalog.resolve("统计 2026-08-10 到 2026-08-18 工作日内的已支付订单数"),
        )
        self.assertIsNone(plan)  # TEXT 字段保守回退，不静默执行
        # 但日期范围本身必须被识别（这是本轮修复点）
        matched = dc.CalendarFilterCompiler._DATE_RANGE_RE.search(
            "统计 2026-08-10 到 2026-08-18 工作日内的已支付订单数",
        )
        self.assertIsNotNone(matched)
        self.assertEqual(matched.group("start"), "2026-08-10")
        self.assertEqual(matched.group("end"), "2026-08-18")

    def test_llm_maps_chinese_table_and_columns_fail_closed(self):
        # 规则层零命中时，LLM 表映射生效；LLM 故障/发明名一律回退澄清
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE items ADD COLUMN city TEXT")
            conn.commit()
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)

        # LLM 不可用 → 不崩溃，回退澄清（fail-closed）
        with mock.patch.object(dc, "_llm_ask_json", side_effect=dc.DBQuillError("down")):
            answer = agent.ask("新增一位客户：姓名“测试”，城市“厦门”")
        self.assertEqual(answer.kind, "clarification")

        # LLM 返回真实存在的表 → 映射生效，不再要求目标表
        with mock.patch.object(
            dc, "_llm_ask_json",
            side_effect=[
                {"table": "items"},
                {"columns": [{"term": "姓名", "column": "value"}]},
                {"sql": "INSERT INTO items(value) VALUES ('测试')", "summary_zh": "新增记录"},
            ],
        ):
            answer = agent.ask("新增一条记录：姓名“测试”")
        self.assertEqual(answer.kind, "write_pending")
        self.assertEqual(answer.operation["target_tables"], ["items"])
        self.assertIn("大模型映射", answer.operation.get("reasoning", ""))

        # 引导录入模型发明不存在的表 → 丢弃表名并退回授权选表，不绑定虚构对象
        with mock.patch.object(dc, "_llm_ask_json", return_value={
            "intent": "write",
            "interaction": "guided_insert",
            "target_table": "ghost_table",
            "confidence": 0.9,
            "reasoning": "用户想新增客户",
        }):
            answer = agent.ask("新增一位客户")
        self.assertEqual(answer.kind, "write_form")
        self.assertEqual(answer.write["selected_table"], "")
        self.assertEqual(answer.operation["target_tables"], [])

        # 发明不存在的字段 → 拒绝映射，仍澄清
        with mock.patch.object(
            dc, "_llm_ask_json",
            side_effect=[{"table": "items"}, {"columns": [{"term": "外星", "column": "ufo"}]}],
        ):
            answer = agent.ask("新增一条记录：外星“x”")
        self.assertEqual(answer.kind, "clarification")

    def test_trend_default_grain_executes_exact_range_without_model(self):
        agent = self._trend_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask(
                "统计下单时间从 2026-01-01 到 2026-02-28 的订单数量趋势"
            )
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.columns, ["月份", "记录数"])
        self.assertEqual(answer.rows, [["2026-01", 2], ["2026-02", 1]])
        self.assertEqual(answer.trend_plan["engine"], "native_trend")
        self.assertEqual(answer.trend_plan["grain"], "month")
        self.assertEqual(answer.trend_plan["grain_source"], "semantic_default")
        self.assertEqual(answer.trend_plan["date_range"]["end"], "2026-02-28")

    def test_trend_relative_days_freezes_runtime_reference_date(self):
        agent = self._trend_agent()
        agent.trend_query.reference_date = date(2026, 2, 1)
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("统计下单时间最近 30 天的成交额趋势")
        self.assertEqual(answer.rows, [["2026-01", 150.0], ["2026-02", 200.0]])
        self.assertEqual(answer.trend_plan["version"], "1.3")
        self.assertEqual(answer.trend_plan["date_range"]["start"], "2026-01-03")
        self.assertEqual(answer.trend_plan["date_range"]["end"], "2026-02-01")
        self.assertEqual(answer.trend_plan["date_range"]["source"], "relative")
        self.assertEqual(answer.trend_plan["date_range"]["reference_date"], "2026-02-01")
        self.assertEqual(
            answer.trend_plan["date_range"]["reference_source"], "injected_reference",
        )

    def test_trend_relative_days_accepts_explicit_anchor(self):
        agent = self._trend_agent()
        agent.trend_query.reference_date = date(2030, 1, 1)
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask(
                "截至 2026-02-01，按日统计下单时间最近 2 天的订单数量趋势"
            )
        self.assertEqual(answer.rows, [["2026-02-01", 1]])
        self.assertEqual(answer.trend_plan["date_range"]["start"], "2026-01-31")
        self.assertEqual(answer.trend_plan["date_range"]["reference_source"], "explicit_anchor")

    def test_trend_relative_weeks_use_inclusive_day_window(self):
        agent = self._trend_agent()
        agent.trend_query.reference_date = date(2026, 2, 1)
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("统计下单时间最近 4 周的订单数量趋势")
        self.assertEqual(answer.rows, [["2026-01", 2], ["2026-02", 1]])
        self.assertEqual(answer.trend_plan["date_range"]["start"], "2026-01-05")
        self.assertEqual(answer.trend_plan["date_range"]["days"], 28)
        self.assertEqual(answer.trend_plan["date_range"]["unit"], "week")

    def test_trend_relative_months_fall_back_without_guessing(self):
        agent = self._trend_agent()
        generated = dc.DBAnswer(
            kind="query", narrative="滚动月口径交回模型", sql="SELECT created_at FROM orders",
            columns=["created_at"], rows=[],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask("统计下单时间最近 2 个月的订单数量趋势")
        execute.assert_called_once()
        self.assertIsNone(answer.trend_plan)
        self.assertEqual(answer.narrative, "滚动月口径交回模型")

    def test_trend_uses_business_calendar_fiscal_range(self):
        agent = self._trend_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("统计下单时间 2026 财年的成交额趋势")
        self.assertEqual(answer.rows, [["2026-04", 300.0]])
        self.assertEqual(answer.trend_plan["date_range"]["start"], "2026-04-01")
        self.assertEqual(answer.trend_plan["date_range"]["end"], "2027-04-01")
        self.assertFalse(answer.trend_plan["date_range"]["end_inclusive"])
        self.assertEqual(answer.trend_plan["date_range"]["source"], "business_calendar")
        self.assertEqual(answer.trend_plan["rules"]["calendar_term"], "公司业务日历")
        self.assertEqual(answer.trend_plan["rules"]["calendar_mode"], "fiscal_year")

    def test_trend_fiscal_quarter_does_not_override_explicit_grain(self):
        agent = self._trend_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按月统计下单时间 2025 财年第 4 季度的成交额趋势")
        self.assertEqual(answer.rows, [["2026-01", 150.0], ["2026-02", 200.0]])
        self.assertEqual(answer.trend_plan["grain"], "month")
        self.assertEqual(answer.trend_plan["rules"]["calendar_mode"], "fiscal_quarter")

    def test_trend_uses_business_calendar_workday_rules(self):
        agent = self._trend_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按月统计下单时间 2025 财年工作日的订单数量趋势")
        self.assertEqual(answer.rows, [["2026-01", 2]])
        self.assertEqual(answer.trend_plan["rules"]["calendar_mode"], "fiscal_business_days")
        self.assertIn("strftime('%w'", answer.sql)

    def test_trend_business_calendar_controls_week_start(self):
        agent = self._trend_agent()
        calendar = next(
            item for item in agent.semantic_catalog.entries
            if item.get("kind") == "business_calendar"
        )
        calendar["calendar"]["week_start"] = 7
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按周统计下单时间 2025 财年的订单数量趋势")
        self.assertEqual(answer.rows, [
            ["2026-01-04", 1], ["2026-01-18", 1], ["2026-02-01", 1],
        ])
        self.assertEqual(answer.trend_plan["rules"]["week_start_iso"], 7)

    def test_trend_rejects_conflicting_exact_and_fiscal_ranges(self):
        agent = self._trend_agent()
        generated = dc.DBAnswer(
            kind="query", narrative="冲突范围交回模型", sql="SELECT created_at FROM orders",
            columns=["created_at"], rows=[],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask(
                "统计下单时间 2026 财年从 2026-04-01 到 2026-04-30 的成交额趋势"
            )
        execute.assert_called_once()
        self.assertIsNone(answer.trend_plan)
        self.assertEqual(answer.narrative, "冲突范围交回模型")

    def test_trend_explicit_week_overrides_default(self):
        agent = self._trend_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按周统计下单时间的订单数量")
        self.assertEqual(answer.columns, ["周起始日", "记录数"])
        self.assertEqual(answer.rows, [
            ["2026-01-05", 1], ["2026-01-19", 1],
            ["2026-01-26", 1], ["2026-04-13", 1],
        ])
        self.assertEqual(answer.trend_plan["grain"], "week")
        self.assertEqual(answer.trend_plan["grain_source"], "explicit")
        self.assertEqual(answer.trend_plan["rules"]["week_start_iso"], 1)

    def test_trend_reuses_controlled_metric_filters(self):
        agent = self._trend_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按季度统计下单时间的已付款成交额")
        self.assertEqual(answer.columns, ["季度", "已付款成交额"])
        self.assertEqual(answer.rows, [["2026-Q1", 300.0], ["2026-Q2", 300.0]])
        self.assertEqual(answer.trend_plan["filters"], [
            {"column": "status", "operator": "eq", "value": "paid"},
        ])
        self.assertIn('"orders"."status" = \'paid\'', answer.sql)

    def test_trend_executes_multiple_metrics_with_filter_isolation(self):
        agent = self._trend_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按月统计下单时间的订单笔数和已付款订单笔数")
        self.assertEqual(answer.columns, ["月份", "订单笔数", "已付款订单笔数"])
        self.assertEqual(answer.rows, [
            ["2026-01", 2, 1], ["2026-02", 1, 1], ["2026-04", 1, 1],
        ])
        self.assertEqual(answer.trend_plan["version"], "1.3")
        self.assertEqual(
            [item["term"] for item in answer.trend_plan["measures"]],
            ["订单笔数", "已付款订单笔数"],
        )
        self.assertIn("COUNT(CASE WHEN", answer.sql)
        self.assertIn('"orders"."status" = \'paid\'', answer.sql)

    def test_trend_multiple_metrics_preserve_exact_date_range(self):
        agent = self._trend_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask(
                "统计下单时间从 2026-01-01 到 2026-02-28 的成交额和已付款成交额趋势"
            )
        self.assertEqual(answer.columns, ["月份", "成交额", "已付款成交额"])
        self.assertEqual(answer.rows, [
            ["2026-01", 150.0, 100.0], ["2026-02", 200.0, 200.0],
        ])
        self.assertEqual(answer.trend_plan["date_range"]["start"], "2026-01-01")
        self.assertEqual(answer.trend_plan["date_range"]["end"], "2026-02-28")

    def test_trend_multiple_metrics_apply_one_global_enum_filter(self):
        agent = self._trend_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask(
                "按月统计下单时间仅已付款订单的订单笔数和已付款订单笔数"
            )
        self.assertEqual(answer.rows, [
            ["2026-01", 1, 1], ["2026-02", 1, 1], ["2026-04", 1, 1],
        ])
        self.assertEqual(answer.trend_plan["global_filters"], [
            {"column": "status", "operator": "eq", "value": "paid"},
        ])
        self.assertIn(" WHERE ", answer.sql)

    def test_trend_multiple_metrics_preserve_free_condition_for_model(self):
        agent = self._trend_agent()
        generated = dc.DBAnswer(
            kind="query", narrative="多指标自由条件由模型链路处理",
            sql="SELECT created_at FROM orders", columns=["created_at"], rows=[],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask(
                "按月统计下单时间的订单笔数和已付款订单笔数，并且 amount > 100"
            )
        execute.assert_called_once()
        self.assertIsNone(answer.trend_plan)
        self.assertEqual(answer.narrative, "多指标自由条件由模型链路处理")

    def test_trend_timestamp_uses_explicit_fixed_offset_basis(self):
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute(
                "CREATE TABLE timestamp_events (id INTEGER PRIMARY KEY, occurred_at TIMESTAMP)"
            )
            conn.executemany(
                "INSERT INTO timestamp_events(id, occurred_at) VALUES (?, ?)",
                [
                    (1, "2026-08-01T15:59:59Z"),
                    (2, "2026-08-01T16:00:00Z"),
                    (3, "2026-08-02T16:00:00Z"),
                ],
            )
            conn.commit()
        agent = dc.DBQuillAgent(
            db_path=str(self.path), sample_rows=0,
            semantic_entries=[
                {
                    "kind": "time_field", "term": "事件时间", "table": "timestamp_events",
                    "column": "occurred_at", "default_grain": "day",
                },
                {
                    "kind": "business_calendar", "term": "事件日历",
                    "table": "timestamp_events", "column": "occurred_at",
                    "calendar": {
                        "fiscal_year_start_month": 1, "fiscal_year_start_day": 1,
                        "fiscal_year_label": "start_year", "timezone": "Asia/Shanghai",
                        "storage_basis": "utc_datetime", "business_utc_offset_minutes": 480,
                        "week_start": 1, "weekend_days": [],
                    },
                },
            ],
        )
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按日统计事件时间的数量")
        self.assertEqual(answer.rows, [
            ["2026-08-01", 1], ["2026-08-02", 1], ["2026-08-03", 1],
        ])
        self.assertEqual(answer.trend_plan["rules"]["storage_basis"], "utc_datetime")
        self.assertEqual(answer.trend_plan["rules"]["business_utc_offset_minutes"], 480)

    def test_trend_timestamp_uses_versioned_iana_dst_conversion(self):
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute(
                "CREATE TABLE dst_events (id INTEGER PRIMARY KEY, occurred_at TIMESTAMP)"
            )
            conn.executemany(
                "INSERT INTO dst_events(id, occurred_at) VALUES (?, ?)",
                [
                    (1, "2024-01-01T04:30:00Z"),
                    (2, "2024-01-01T05:30:00Z"),
                    (3, "2024-07-01T03:30:00Z"),
                    (4, "2024-07-01T04:30:00Z"),
                    (5, "2024-07-01T04:30:00+08:00"),
                    (6, "not-a-timestamp"),
                ],
            )
            conn.commit()
        agent = dc.DBQuillAgent(
            db_path=str(self.path), sample_rows=0,
            semantic_entries=[
                {
                    "kind": "time_field", "term": "美东事件时间", "table": "dst_events",
                    "column": "occurred_at", "default_grain": "day",
                },
                {
                    "kind": "business_calendar", "term": "美东事件日历",
                    "table": "dst_events", "column": "occurred_at",
                    "calendar": {
                        "fiscal_year_start_month": 1, "fiscal_year_start_day": 1,
                        "fiscal_year_label": "start_year", "timezone": "America/New_York",
                        "storage_basis": "utc_datetime",
                        "timezone_conversion": "iana_tzdata",
                        "week_start": 1, "weekend_days": [],
                    },
                },
            ],
        )
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按日统计美东事件时间的数量")
        self.assertEqual(answer.rows, [
            ["2023-12-31", 1], ["2024-01-01", 1],
            ["2024-06-30", 1], ["2024-07-01", 1],
        ])
        self.assertEqual(answer.trend_plan["version"], "1.3")
        self.assertEqual(answer.trend_plan["rules"]["timezone_conversion"], "iana_tzdata")
        self.assertEqual(answer.trend_plan["rules"]["tzdata_version"], "2026.3")
        self.assertEqual(answer.trend_plan["rules"]["iana_version"], "2026c")
        self.assertIn("dbquill_iana_date", answer.sql)
        self.assertIn(dc.TimezoneRuntime.VERSION_TOKEN, answer.sql)

    def test_trend_timestamp_without_storage_basis_falls_back(self):
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute(
                "CREATE TABLE events (id INTEGER PRIMARY KEY, occurred_at TIMESTAMP)"
            )
            conn.commit()
        agent = dc.DBQuillAgent(
            db_path=str(self.path), sample_rows=0,
            semantic_entries=[{
                "kind": "time_field", "term": "发生时间", "table": "events",
                "column": "occurred_at", "default_grain": "day",
            }],
        )
        generated = dc.DBAnswer(
            kind="query", narrative="未声明时间戳口径，交回模型链路",
            sql="SELECT occurred_at FROM events", columns=["occurred_at"], rows=[],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask("按日统计发生时间的数量")
        execute.assert_called_once()
        self.assertIsNone(answer.trend_plan)

    def test_trend_preserves_unsupported_free_condition_for_model(self):
        agent = self._trend_agent()
        generated = dc.DBAnswer(
            kind="query", narrative="自由条件由模型链路处理",
            sql="SELECT created_at, COUNT(*) FROM orders GROUP BY created_at",
            columns=["created_at", "count"], rows=[],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask("按月统计下单时间的订单数量，并且 status = paid")
        execute.assert_called_once()
        self.assertIsNone(answer.trend_plan)
        self.assertEqual(answer.narrative, "自由条件由模型链路处理")

    def test_dimension_group_count_executes_deterministically_without_model(self):
        agent = self._dimension_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按客户区域统计数量")
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.columns, ["客户区域", "记录数"])
        self.assertEqual(answer.rows, [["华东", 2], ["华北", 1], ["华南", 1]])
        self.assertEqual(answer.dimension_plan["engine"], "native_dimension")
        self.assertEqual(answer.dimension_plan["mode"], "group_by")
        self.assertEqual(answer.dimension_plan["status"], "executed")

    def test_dimension_group_reuses_controlled_metric_filters(self):
        agent = self._dimension_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按客户区域统计活跃客户数")
        self.assertEqual(answer.columns, ["客户区域", "活跃客户数"])
        self.assertEqual(answer.rows, [["华东", 1], ["华北", 1], ["华南", 1]])
        self.assertEqual(answer.dimension_plan["filters"], [
            {"column": "status", "operator": "eq", "value": "active"},
        ])
        self.assertIn('"customers"."status" = \'active\'', answer.sql)

    def test_dimension_group_executes_multiple_metrics_with_filter_isolation(self):
        agent = self._dimension_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按客户区域统计客户总数和活跃客户数")

        self.assertEqual(answer.columns, ["客户区域", "客户总数", "活跃客户数"])
        self.assertEqual(answer.rows, [
            ["华东", 2, 1], ["华北", 1, 1], ["华南", 1, 1],
        ])
        self.assertEqual(answer.dimension_plan["version"], "1.2")
        self.assertEqual(
            [item["term"] for item in answer.dimension_plan["measures"]],
            ["客户总数", "活跃客户数"],
        )
        self.assertIn("COUNT(CASE WHEN", answer.sql)
        self.assertNotIn(" WHERE ", answer.sql.upper())

    def test_dimension_drilldown_executes_multiple_metrics(self):
        agent = self._dimension_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask(
                "从客户区域下钻到客户城市统计活跃客户数和客户总数"
            )

        self.assertEqual(answer.dimension_plan["mode"], "drilldown")
        self.assertEqual(answer.columns, [
            "客户区域", "客户城市", "活跃客户数", "客户总数",
        ])
        self.assertEqual(answer.rows, [
            ["华东", "上海", 1, 1], ["华东", "杭州", 0, 1],
            ["华北", "北京", 1, 1], ["华南", "深圳", 1, 1],
        ])

    def test_dimension_fixed_filters_apply_to_all_measures(self):
        agent = self._dimension_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("按活跃客户区域统计客户总数和非活跃客户数")
        self.assertEqual(answer.columns, ["活跃客户区域", "客户总数", "非活跃客户数"])
        self.assertEqual(answer.rows, [
            ["华东", 1, 0], ["华北", 1, 0], ["华南", 1, 0],
        ])
        self.assertEqual(answer.dimension_plan["version"], "1.2")
        self.assertEqual(answer.dimension_plan["dimension_filters"], [{
            "dimension_term": "活跃客户区域",
            "column": "status", "operator": "eq", "value": "active",
        }])
        self.assertIn('"customers"."status" = \'active\'', answer.sql)

    def test_dimension_drilldown_deduplicates_shared_fixed_filters(self):
        agent = self._dimension_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask(
                "从活跃客户区域下钻到活跃客户城市统计客户总数"
            )
        self.assertEqual(answer.rows, [
            ["华东", "上海", 1], ["华北", "北京", 1], ["华南", "深圳", 1],
        ])
        self.assertEqual(len(answer.dimension_plan["dimension_filters"]), 1)

    def test_dimension_multi_metric_preserves_free_condition_for_model(self):
        agent = self._dimension_agent()
        generated = dc.DBAnswer(
            kind="query", narrative="多指标自由条件由模型链路处理",
            sql="SELECT region FROM customers", columns=["region"], rows=[],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask(
                "按客户区域统计客户总数和活跃客户数，并且 city = 上海"
            )

        execute.assert_called_once()
        self.assertIsNone(answer.dimension_plan)
        self.assertEqual(answer.narrative, generated.narrative)

    def test_explicit_same_table_dimension_drilldown_executes_path(self):
        agent = self._dimension_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("从客户区域下钻到客户城市统计数量")
        self.assertEqual(answer.dimension_plan["mode"], "drilldown")
        self.assertEqual(answer.columns, ["客户区域", "客户城市", "记录数"])
        self.assertEqual(answer.rows, [
            ["华东", "上海", 1], ["华东", "杭州", 1],
            ["华北", "北京", 1], ["华南", "深圳", 1],
        ])
        self.assertEqual(answer.dimension_plan["hierarchy"]["from_level"], 1)
        self.assertEqual(answer.dimension_plan["hierarchy"]["to_level"], 2)

    def test_next_dimension_level_can_be_selected_deterministically(self):
        agent = self._dimension_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask("客户区域下钻一级统计数量")
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.dimension_plan["mode"], "drilldown")
        self.assertEqual(
            [item["term"] for item in answer.dimension_plan["dimensions"]],
            ["客户区域", "客户城市"],
        )

    def test_ambiguous_drilldown_requires_level_then_accepts_followup(self):
        agent = self._dimension_agent()
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            first = agent.ask("客户区域下钻统计数量")
            self.assertEqual(first.kind, "clarification")
            self.assertEqual(first.clarification["missing"], "dimension_level")
            self.assertEqual(
                [item["label"] for item in first.clarification["candidates"]],
                ["客户城市"],
            )
            second = agent.ask(
                "客户城市",
                history=[{"role": "user", "content": "客户区域下钻统计数量"}],
                clarification=first.clarification,
            )
        self.assertEqual(second.kind, "query")
        self.assertEqual(second.dimension_plan["hierarchy"]["to_term"], "客户城市")

    def test_dimension_executor_preserves_unsupported_extra_conditions_for_model(self):
        agent = self._dimension_agent()
        generated = dc.DBAnswer(
            kind="query", narrative="复杂条件由模型链路处理",
            sql="SELECT region, COUNT(*) FROM customers WHERE status='active' GROUP BY region",
            columns=["region", "count"], rows=[],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask("按客户区域统计数量，并且 status = active")
        execute.assert_called_once()
        self.assertIsNone(answer.dimension_plan)
        self.assertEqual(answer.narrative, "复杂条件由模型链路处理")

    def test_business_calendar_requires_explicit_definition(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN created_at TEXT")
            conn.commit()
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            answer = agent.ask("统计 orders.created_at 2026 财年的订单数量")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "business_calendar")

        generated = dc.DBAnswer(
            kind="query", narrative="查询完成", sql="SELECT COUNT(*) FROM orders",
            columns=["数量"], rows=[[1]],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            completed = agent.ask(
                "统计 orders.created_at 2026 财年的订单数量；"
                "业务日历：财年从 2026-04-01 开始"
            )
        execute.assert_called_once()
        self.assertEqual(completed.kind, "query")

    def test_structured_business_calendar_resolves_only_its_bound_time_field(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN created_at TEXT")
            conn.execute("ALTER TABLE orders ADD COLUMN updated_at TEXT")
            conn.execute(
                "CREATE TABLE holidays (holiday_date DATE, name TEXT, is_working INTEGER)"
            )
            conn.commit()
        calendar_entry = {
            "kind": "business_calendar", "term": "公司业务日历", "table": "orders",
            "column": "created_at",
            "calendar": {
                "fiscal_year_start_month": 4,
                "fiscal_year_start_day": 1,
                "timezone": "Asia/Shanghai",
                "week_start": 1,
                "weekend_days": [6, 7],
                "holiday_table": "holidays",
                "holiday_date_column": "holiday_date",
                "holiday_name_column": "name",
                "working_override_column": "is_working",
            },
        }
        agent = dc.DBQuillAgent(
            db_path=str(self.path), sample_rows=0, semantic_entries=[calendar_entry],
        )
        generated = dc.DBAnswer(
            kind="query", narrative="查询完成", sql="SELECT COUNT(*) FROM orders",
            columns=["数量"], rows=[[1]],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask("统计 orders.created_at 2026 财年的订单数量")
        execute.assert_called_once()
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.semantic["matches"][0]["kind"], "business_calendar")
        self.assertIn("财年从 04-01 开始", execute.call_args.args[0])
        self.assertIn("不得猜测法定节假日", execute.call_args.args[0])

        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")):
            mismatch = agent.ask("统计 orders.updated_at 2026 财年的订单数量")
        self.assertEqual(mismatch.kind, "clarification")
        self.assertEqual(mismatch.clarification["missing"], "business_calendar")

    def test_date_only_fiscal_year_uses_deterministic_calendar_without_llm(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN created_at DATE")
            conn.execute("UPDATE orders SET amount=10, created_at='2025-04-01' WHERE id=1")
            conn.executemany(
                "INSERT INTO orders(amount, created_at) VALUES (?, ?)",
                [(20, "2026-03-31"), (30, "2026-04-01")],
            )
            conn.commit()
        entries = [
            {
                "kind": "metric", "term": "成交额", "table": "orders",
                "column": "amount", "aggregation": "sum",
            },
            {
                "kind": "business_calendar", "term": "公司业务日历", "table": "orders",
                "column": "created_at",
                "calendar": {
                    "fiscal_year_start_month": 4,
                    "fiscal_year_start_day": 1,
                    "fiscal_year_label": "end_year",
                    "timezone": "Asia/Shanghai",
                    "week_start": 1,
                    "weekend_days": [6, 7],
                },
            },
        ]
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0, semantic_entries=entries)
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")), \
             mock.patch.object(agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run")):
            answer = agent.ask("统计 orders.created_at 2026 财年的成交额")

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [[30.0]])
        self.assertEqual(answer.columns, ["成交额"])
        self.assertEqual(answer.calendar_plan["engine"], "native_calendar")
        self.assertEqual(answer.calendar_plan["date_range"]["start"], "2025-04-01")
        self.assertEqual(answer.calendar_plan["date_range"]["end"], "2026-04-01")
        self.assertFalse(answer.calendar_plan["date_range"]["end_inclusive"])
        self.assertEqual(answer.calendar_plan["status"], "executed")
        self.assertIn('date("orders"."created_at")', answer.sql)

    def test_date_only_workdays_apply_weekends_and_holiday_overrides(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN created_at DATE")
            conn.execute("UPDATE orders SET created_at='2026-08-07' WHERE id=1")
            conn.executemany(
                "INSERT INTO orders(amount, created_at) VALUES (?, ?)",
                [
                    (10, "2026-08-08"),
                    (10, "2026-08-09"),
                    (10, "2026-08-10"),
                    (10, "2026-08-11"),
                ],
            )
            conn.execute(
                "CREATE TABLE holidays (holiday_date DATE PRIMARY KEY, name TEXT, is_working INTEGER)"
            )
            conn.executemany(
                "INSERT INTO holidays(holiday_date, name, is_working) VALUES (?, ?, ?)",
                [
                    ("2026-08-08", "调休工作日", 1),
                    ("2026-08-10", "公司假日", 0),
                ],
            )
            conn.commit()
        calendar_entry = {
            "kind": "business_calendar", "term": "公司业务日历", "table": "orders",
            "column": "created_at",
            "calendar": {
                "fiscal_year_start_month": 1,
                "fiscal_year_start_day": 1,
                "fiscal_year_label": "start_year",
                "timezone": "Asia/Shanghai",
                "week_start": 1,
                "weekend_days": [6, 7],
                "holiday_table": "holidays",
                "holiday_date_column": "holiday_date",
                "holiday_name_column": "name",
                "working_override_column": "is_working",
            },
        }
        agent = dc.DBQuillAgent(
            db_path=str(self.path), sample_rows=0, semantic_entries=[calendar_entry],
        )
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")), \
             mock.patch.object(agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run")):
            answer = agent.ask(
                "统计 orders.created_at 从 2026-08-07 到 2026-08-11 的工作日订单数量"
            )

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [[3]])
        self.assertEqual(answer.calendar_plan["mode"], "business_days")
        self.assertTrue(answer.calendar_plan["date_range"]["end_inclusive"])
        self.assertEqual(answer.calendar_plan["rules"]["holiday_table"], "holidays")
        self.assertIn("strftime('%w'", answer.sql)
        self.assertIn("EXISTS", answer.sql)

    def test_utc_timestamp_calendar_applies_explicit_fixed_business_offset(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN occurred_at TIMESTAMP")
            conn.execute(
                "UPDATE orders SET occurred_at='2026-08-01T15:59:59Z' WHERE id=1"
            )
            conn.executemany(
                "INSERT INTO orders(amount, occurred_at) VALUES (?, ?)",
                [
                    (10, "2026-08-01T16:00:00Z"),
                    (10, "2026-08-02T15:59:59Z"),
                    (10, "2026-08-02T16:00:00Z"),
                ],
            )
            conn.commit()
        calendar_entry = {
            "kind": "business_calendar", "term": "北京时间日历", "table": "orders",
            "column": "occurred_at",
            "calendar": {
                "fiscal_year_start_month": 1, "fiscal_year_start_day": 1,
                "fiscal_year_label": "start_year", "timezone": "Asia/Shanghai",
                "storage_basis": "utc_datetime", "business_utc_offset_minutes": 480,
                "week_start": 1, "weekend_days": [],
            },
        }
        agent = dc.DBQuillAgent(
            db_path=str(self.path), sample_rows=0, semantic_entries=[calendar_entry],
        )
        with mock.patch.object(dc, "_llm_ask_json", side_effect=AssertionError("LLM should not run")), \
             mock.patch.object(agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run")):
            answer = agent.ask(
                "统计 orders.occurred_at 从 2026-08-02 到 2026-08-02 的工作日记录数"
            )

        self.assertEqual(answer.rows, [[2]])
        self.assertEqual(answer.calendar_plan["version"], "1.2")
        self.assertEqual(answer.calendar_plan["rules"]["storage_basis"], "utc_datetime")
        self.assertEqual(answer.calendar_plan["rules"]["business_utc_offset_minutes"], 480)
        self.assertEqual(answer.calendar_plan["rules"]["timezone_conversion"], "fixed_offset")
        self.assertIn("'+480 minutes'", answer.sql)

    def test_iana_dst_calendar_applies_dynamic_business_date(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN occurred_at TIMESTAMP")
            conn.execute("UPDATE orders SET occurred_at='2024-07-01T03:30:00Z' WHERE id=1")
            conn.execute(
                "INSERT INTO orders(amount, occurred_at) VALUES (?, ?)",
                (10, "2024-07-01T04:30:00Z"),
            )
            conn.commit()
        agent = dc.DBQuillAgent(
            db_path=str(self.path), sample_rows=0,
            semantic_entries=[{
                "kind": "business_calendar", "term": "美东业务日历", "table": "orders",
                "column": "occurred_at",
                "calendar": {
                    "fiscal_year_start_month": 1, "fiscal_year_start_day": 1,
                    "fiscal_year_label": "start_year", "timezone": "America/New_York",
                    "storage_basis": "utc_datetime", "timezone_conversion": "iana_tzdata",
                    "week_start": 1, "weekend_days": [],
                },
            }],
        )
        with mock.patch.object(
            agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run"),
        ):
            answer = agent.ask(
                "统计 orders.occurred_at 从 2024-07-01 到 2024-07-01 的工作日记录数"
            )
        self.assertEqual(answer.rows, [[1]])
        self.assertEqual(answer.calendar_plan["version"], "1.2")
        self.assertEqual(answer.calendar_plan["rules"]["timezone_conversion"], "iana_tzdata")
        self.assertEqual(answer.calendar_plan["rules"]["iana_version"], "2026c")
        self.assertIsNone(answer.calendar_plan["rules"]["business_utc_offset_minutes"])
        self.assertIn("dbquill_iana_date", answer.sql)

    def test_local_timestamp_calendar_uses_recorded_wall_date_without_conversion(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN occurred_at DATETIME")
            conn.execute("UPDATE orders SET occurred_at='2026-08-01 23:59:59' WHERE id=1")
            conn.execute(
                "INSERT INTO orders(amount, occurred_at) VALUES (?, ?)",
                (10, "2026-08-02 00:00:00"),
            )
            conn.commit()
        calendar_entry = {
            "kind": "business_calendar", "term": "本地时间日历", "table": "orders",
            "column": "occurred_at",
            "calendar": {
                "fiscal_year_start_month": 1, "fiscal_year_start_day": 1,
                "fiscal_year_label": "start_year", "timezone": "Asia/Shanghai",
                "storage_basis": "local_datetime", "week_start": 1, "weekend_days": [],
            },
        }
        agent = dc.DBQuillAgent(
            db_path=str(self.path), sample_rows=0, semantic_entries=[calendar_entry],
        )
        with mock.patch.object(agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run")):
            answer = agent.ask(
                "统计 orders.occurred_at 从 2026-08-02 到 2026-08-02 的工作日记录数"
            )

        self.assertEqual(answer.rows, [[1]])
        self.assertEqual(answer.calendar_plan["rules"]["storage_basis"], "local_datetime")
        self.assertEqual(answer.calendar_plan["rules"]["timezone_conversion"], "none")
        self.assertNotIn("minutes", answer.sql)

    def test_timestamp_without_explicit_storage_basis_stays_on_model_path(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN occurred_at TIMESTAMP")
            conn.commit()
        calendar_entry = {
            "kind": "business_calendar", "term": "旧时间戳日历", "table": "orders",
            "column": "occurred_at",
            "calendar": {
                "fiscal_year_start_month": 1, "fiscal_year_start_day": 1,
                "fiscal_year_label": "start_year", "timezone": "UTC",
                "week_start": 1, "weekend_days": [],
            },
        }
        agent = dc.DBQuillAgent(
            db_path=str(self.path), sample_rows=0, semantic_entries=[calendar_entry],
        )
        generated = dc.DBAnswer(
            kind="query", narrative="存储基准未声明，回退模型链路",
            sql="SELECT COUNT(*) FROM orders", columns=["记录数"], rows=[[1]],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask(
                "统计 orders.occurred_at 从 2026-08-02 到 2026-08-02 的工作日记录数"
            )
        execute.assert_called_once()
        self.assertIsNone(answer.calendar_plan)
        calendar = answer.semantic["matches"][0]["calendar"]
        self.assertEqual(calendar["storage_basis"], "unspecified")
        self.assertEqual(calendar["storage_basis_source"], "legacy_default")

    def test_fiscal_quarter_uses_exact_three_month_boundary(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN created_at DATE")
            conn.execute("UPDATE orders SET amount=5, created_at='2026-06-30' WHERE id=1")
            conn.executemany(
                "INSERT INTO orders(amount, created_at) VALUES (?, ?)",
                [(10, "2026-07-01"), (20, "2026-09-30"), (40, "2026-10-01")],
            )
            conn.commit()
        entries = [
            {
                "kind": "metric", "term": "成交额", "table": "orders",
                "column": "amount", "aggregation": "sum",
            },
            {
                "kind": "business_calendar", "term": "公司业务日历", "table": "orders",
                "column": "created_at",
                "calendar": {
                    "fiscal_year_start_month": 4, "fiscal_year_start_day": 1,
                    "fiscal_year_label": "start_year", "timezone": "UTC",
                    "week_start": 1, "weekend_days": [6, 7],
                },
            },
        ]
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0, semantic_entries=entries)
        with mock.patch.object(agent.nl2sql, "answer", side_effect=AssertionError("NL2SQL should not run")):
            answer = agent.ask("统计 orders.created_at 2026 财年第2季度的成交额")

        self.assertEqual(answer.rows, [[30.0]])
        self.assertEqual(answer.calendar_plan["mode"], "fiscal_quarter")
        self.assertEqual(answer.calendar_plan["date_range"]["start"], "2026-07-01")
        self.assertEqual(answer.calendar_plan["date_range"]["end"], "2026-10-01")
        self.assertEqual(answer.calendar_plan["rules"]["fiscal_quarter"], 2)

    def test_deterministic_calendar_does_not_drop_grouping_or_extra_conditions(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN created_at DATE")
            conn.commit()
        entries = [
            {
                "kind": "metric", "term": "成交额", "table": "orders",
                "column": "amount", "aggregation": "sum",
            },
            {
                "kind": "business_calendar", "term": "公司业务日历", "table": "orders",
                "column": "created_at",
                "calendar": {
                    "fiscal_year_start_month": 4, "fiscal_year_start_day": 1,
                    "fiscal_year_label": "start_year", "timezone": "UTC",
                    "week_start": 1, "weekend_days": [6, 7],
                },
            },
        ]
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0, semantic_entries=entries)
        generated = dc.DBAnswer(
            kind="query", narrative="模型链路", sql="SELECT amount FROM orders",
            columns=["amount"], rows=[],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask(
                "统计 orders.created_at 2026 财年的成交额，并按 orders.amount 分组"
            )
        execute.assert_called_once()
        self.assertIsNone(answer.calendar_plan)
        self.assertEqual(answer.narrative, "模型链路")

    def test_legacy_calendar_requires_year_label_before_deterministic_fiscal_query(self):
        _add_table(self.path)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE orders ADD COLUMN created_at DATE")
            conn.commit()
        entries = [
            {
                "kind": "metric", "term": "成交额", "table": "orders",
                "column": "amount", "aggregation": "sum",
            },
            {
                "kind": "business_calendar", "term": "旧业务日历", "table": "orders",
                "column": "created_at",
                "calendar": {
                    "fiscal_year_start_month": 4, "fiscal_year_start_day": 1,
                    "timezone": "UTC", "week_start": 1, "weekend_days": [6, 7],
                },
            },
        ]
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0, semantic_entries=entries)
        generated = dc.DBAnswer(
            kind="query", narrative="旧配置回退", sql="SELECT SUM(amount) FROM orders",
            columns=["成交额"], rows=[[12.5]],
        )
        with mock.patch.object(agent.nl2sql, "answer", return_value=generated) as execute:
            answer = agent.ask("统计 orders.created_at 2026 财年的成交额")
        execute.assert_called_once()
        self.assertIsNone(answer.calendar_plan)
        calendar = next(
            item["calendar"] for item in answer.semantic["matches"]
            if item["kind"] == "business_calendar"
        )
        self.assertEqual(calendar["fiscal_year_label_source"], "legacy_default")

    def test_single_table_generic_count_infers_only_target(self):
        generated = dc.DBAnswer(
            kind="query",
            narrative="共 3 条",
            sql="SELECT COUNT(*) FROM items",
            columns=["数量"],
            rows=[[3]],
        )
        with mock.patch.object(self.agent.nl2sql, "answer", return_value=generated) as execute:
            answer = self.agent.ask("一共多少条记录？")
        execute.assert_called_once()
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.operation["target_tables"], ["items"])

    def test_clarification_followups_accumulate_without_guessing(self):
        first = self.agent.ask("删除记录")
        second = self.agent.ask(
            "items",
            history=[
                {"role": "user", "content": "删除记录"},
                {"role": "assistant", "content": first.narrative},
            ],
            clarification=first.clarification,
        )
        self.assertEqual(second.kind, "clarification")
        self.assertEqual(second.clarification["missing"], "filter_condition")
        proposed = dc.DBAnswer(
            kind="write_pending",
            narrative="将删除一行",
            sql="DELETE FROM items WHERE id=1",
            confirm_id="delete-one",
            write={"kind": "DELETE", "table": "items", "dangerous": False, "preview": {}},
        )
        with mock.patch.object(self.agent.write_executor, "prepare", return_value=proposed) as execute:
            third = self.agent.ask(
                "id=1",
                history=[
                    {"role": "user", "content": "items"},
                    {"role": "assistant", "content": second.narrative},
                ],
                clarification=second.clarification,
            )
        execute.assert_called_once()
        self.assertEqual(third.kind, "write_pending")
        self.assertEqual(third.operation["status"], "awaiting_confirmation")

    def test_complete_new_request_replaces_pending_clarification(self):
        pending = self.agent.ask("删除记录")
        proposed = dc.DBAnswer(
            kind="write_pending",
            narrative="将更新一行",
            sql="UPDATE items SET value='ok' WHERE id=1",
            confirm_id="replace-pending",
            write={"kind": "UPDATE", "table": "items", "dangerous": False, "preview": {}},
        )
        with mock.patch.object(self.agent.write_executor, "prepare", return_value=proposed) as execute:
            answer = self.agent.ask(
                "把 items 中 id=1 的 value 改成 ok",
                history=[
                    {"role": "user", "content": "删除记录"},
                    {"role": "assistant", "content": pending.narrative},
                ],
                clarification=pending.clarification,
            )
        self.assertEqual(answer.kind, "write_pending")
        self.assertEqual(answer.operation["action"], "update")
        execute.assert_called_once_with("把 items 中 id=1 的 value 改成 ok", history=mock.ANY)

    def test_operation_type_short_reply_is_merged(self):
        pending = {
            "missing": "operation_type",
            "missing_label": "操作类型",
            "original_question": "处理 items",
            "candidates": [],
        }
        resolved = self.agent.operation_planner.resolve_followup(
            "删除",
            history=[{"role": "user", "content": "处理 items"}],
            clarification=pending,
        )
        self.assertEqual(resolved, "处理 items；操作类型：删除")

    def test_clear_read_request_replaces_pending_clarification(self):
        pending = {
            "missing": "filter_condition",
            "missing_label": "筛选条件",
            "original_question": "删除 items 的记录",
            "candidates": [],
        }
        resolved = self.agent.operation_planner.resolve_followup(
            "查询 items 一共多少条记录",
            history=[{"role": "user", "content": "删除 items 的记录"}],
            clarification=pending,
        )
        self.assertEqual(resolved, "查询 items 一共多少条记录")

    def test_compose_answer_uses_validated_operation_graph(self):
        routed = dc.IntentResult(intent="compose", confidence=0.91, reasoning="需要查数并结合内容")
        query_answer = dc.DBAnswer(
            kind="query",
            narrative="items 共 3 条",
            sql="SELECT COUNT(*) FROM items LIMIT 500",
            columns=["COUNT(*)"],
            rows=[[3]],
        )
        retrieve_answer = dc.DBAnswer(
            kind="retrieve",
            narrative="检索到相关记录",
            evidence=[{"table": "items", "row": ["1", "v0"]}],
        )
        with mock.patch.object(self.agent.router, "classify", return_value=routed), \
             mock.patch.object(self.agent.nl2sql, "answer", return_value=query_answer), \
             mock.patch.object(self.agent.rag, "answer", return_value=retrieve_answer), \
             mock.patch.object(dc, "_llm_ask_json", return_value={"answer_zh": "综合结果"}):
            answer = self.agent.ask("结合 items 的数量和内容给出结论")

        self.assertEqual(answer.kind, "compose")
        self.assertEqual(answer.narrative, "综合结果")
        self.assertEqual(answer.operation["action"], "analyze")
        self.assertEqual(answer.graph["status"], "completed")
        self.assertEqual(
            [(node["node_id"], node["depends_on"]) for node in answer.graph["nodes"]],
            [
                ("query-data", []),
                ("retrieve-context", []),
                ("synthesize-answer", ["query-data", "retrieve-context"]),
            ],
        )
        self.assertEqual(answer.evidence, retrieve_answer.evidence)

    def test_independent_multi_query_runs_through_natural_language_entry(self):
        _add_table(self.path, "orders")
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)

        def query_answer(question, history=None, allowed_tables=None):
            table = allowed_tables[0]
            total = 3 if table == "items" else 1
            return dc.DBAnswer(
                kind="query",
                narrative=f"{table} 共 {total} 条",
                sql=f"SELECT COUNT(*) AS total FROM {table}",
                columns=["total"],
                rows=[[total]],
            )

        with mock.patch.object(agent.nl2sql, "answer", side_effect=query_answer) as query, \
             mock.patch.object(dc, "_llm_ask_json", return_value={"answer_zh": "两张表已分别统计"}):
            answer = agent.ask("分别统计 items 和 orders 的数量")

        self.assertEqual(answer.kind, "compose")
        self.assertEqual(answer.operation["action"], "analyze")
        self.assertEqual(answer.operation["target_tables"], ["items", "orders"])
        self.assertEqual(answer.graph["strategy"], "deterministic-multi-query")
        self.assertEqual([item["label"] for item in answer.datasets], ["items", "orders"])
        self.assertEqual(
            [call.kwargs["allowed_tables"] for call in query.call_args_list],
            [["items"], ["orders"]],
        )

    def test_three_independent_queries_run_through_natural_language_entry(self):
        _add_table(self.path, "orders")
        _add_table(self.path, "events")
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)

        def query_answer(question, history=None, allowed_tables=None):
            table = allowed_tables[0]
            return dc.DBAnswer(
                kind="query",
                narrative=f"{table} 已统计",
                sql=f"SELECT COUNT(*) AS total FROM {table}",
                columns=["total"],
                rows=[[1]],
            )

        with mock.patch.object(agent.nl2sql, "answer", side_effect=query_answer) as query, \
             mock.patch.object(dc, "_llm_ask_json", return_value={"answer_zh": "三张表已分别统计"}):
            answer = agent.ask("分别统计 items、orders 和 events 的数量")

        self.assertEqual(answer.kind, "compose")
        self.assertEqual(answer.operation["target_tables"], ["events", "items", "orders"])
        self.assertEqual(answer.graph["strategy"], "deterministic-multi-query")
        self.assertEqual([item["label"] for item in answer.datasets], ["events", "items", "orders"])
        self.assertEqual(
            [call.kwargs["allowed_tables"] for call in query.call_args_list],
            [["events"], ["items"], ["orders"]],
        )


class BoundedReadExplorerTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "people.db"
        with closing(sqlite3.connect(self.path)) as conn:
            conn.executescript(
                """
                CREATE TABLE scholars (
                    id INTEGER PRIMARY KEY,
                    name TEXT,
                    title TEXT,
                    short_bio TEXT
                );
                CREATE TABLE audit_mentions (
                    id INTEGER PRIMARY KEY,
                    subject TEXT,
                    note TEXT
                );
                INSERT INTO scholars(name, title, short_bio) VALUES (
                    '肖仰华', '教授、博导',
                    '复旦大学教授、博士生导师，上海市数据科学重点实验室主任'
                );
                INSERT INTO audit_mentions(subject, note) VALUES ('肖仰华', '导入记录');
                """
            )
            conn.commit()
        self.agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=2)

    def tearDown(self):
        self.tmp.cleanup()

    def test_observe_replan_loop_can_inspect_retrieve_and_finish(self):
        evidence = [{
            "table": "scholars",
            "columns": ["id", "name", "title", "short_bio"],
            "row": ["1", "肖仰华", "教授、博导", "上海市数据科学重点实验室主任"],
            "matched": "肖仰华",
        }]
        actions = [
            {
                "action": "inspect_schema", "tables": ["scholars"],
                "reason_code": "schema_uncertain",
            },
            {
                "action": "retrieve", "tables": [],
                "question": "检索肖仰华的职称、导师身份和现任职务",
                "reason_code": "evidence_missing",
            },
            {
                "action": "finish", "answer_zh": "肖仰华是教授、博导，并担任上海市数据科学重点实验室主任。",
                "report": {
                    "findings": ["现任教授、博士生导师", "担任重点实验室主任"],
                    "scope": "覆盖已命中的人物档案与职务记录",
                    "limitations": "未见完整历史任职时间线",
                },
                "reason_code": "answer_grounded",
            },
        ]
        with mock.patch.object(
            self.agent.read_explorer, "_next_action", side_effect=actions,
        ) as planner, mock.patch.object(
            self.agent.rag,
            "recall_evidence",
            return_value=evidence,
        ) as retrieve:
            answer = self.agent.read_explorer.explore(
                "肖仰华的工作是什么？",
                None,
                dc.DBAnswer(kind="retrieve", narrative="未找到。", evidence=[]),
            )

        self.assertIsNotNone(answer)
        self.assertEqual(planner.call_count, 3)
        retrieve.assert_called_once()
        self.assertIn("明确实体：“肖仰华”", retrieve.call_args.args[0])
        self.assertEqual(answer.kind, "retrieve")
        self.assertIn("实验室主任", answer.narrative)
        self.assertEqual(len(answer.report["findings"]), 2)
        self.assertIn("历史任职", answer.report["limitations"])
        self.assertEqual(answer.evidence, evidence)
        actions_seen = [
            step.get("action") for step in answer.steps
            if step.get("planner_step")
        ]
        self.assertEqual(actions_seen, ["inspect_schema", "retrieve", "finish"])

    def test_synthesis_separates_concise_answer_from_bounded_report(self):
        payload = {
            "answer_zh": "肖仰华是教授和博士生导师。他还担任上海市数据科学重点实验室主任。第三句不应进入默认回答。",
            "report": {
                "findings": ["教授", "博导", "实验室主任", "期刊编委", "超出上限"],
                "scope": "覆盖人物档案和职务记录",
                "limitations": "缺少历史时间线",
                "internal": "不应透传",
            },
            "evidence_sufficient": False,
        }
        with mock.patch.object(dc, "_llm_ask_json", return_value=payload) as ask:
            output = self.agent.read_explorer._synthesize_answer(
                "肖仰华有哪些工作？",
                [{"source": "search_values", "kind": "value_search"}],
                None,
            )

        self.assertNotIn("第三句", output["narrative"])
        self.assertEqual(len(output["report"]["findings"]), 4)
        self.assertNotIn("internal", output["report"])
        prompt = ask.call_args.args[0]
        self.assertIn("输出分为两层", prompt)
        self.assertIn("默认精简模式", prompt)

    def test_explicit_technical_detail_request_is_not_sentence_clamped(self):
        detailed_answer = "结论一。结论二。结论三。"
        with mock.patch.object(
            dc,
            "_llm_ask_json",
            return_value={"answer_zh": detailed_answer, "report": {}},
        ) as ask:
            output = self.agent.read_explorer._synthesize_answer(
                "请展开 SQL 和证据明细",
                [{"source": "run_sql", "kind": "query"}],
                None,
            )

        self.assertEqual(output["narrative"], detailed_answer)
        self.assertIn("用户明确要求了技术明细", ask.call_args.args[0])

    def test_refined_question_cannot_drop_explicit_entity(self):
        actions = [
            {
                "action": "retrieve", "question": "检索他的职务",
                "reason_code": "evidence_missing",
            },
            {"action": "stop", "reason_code": "cannot_progress"},
        ]
        with mock.patch.object(
            self.agent.read_explorer, "_next_action", side_effect=actions,
        ), mock.patch.object(self.agent.rag, "answer") as retrieve:
            answer = self.agent.read_explorer.explore(
                "肖仰华的资料是什么？请找他的职务。",
                None,
                dc.DBAnswer(kind="retrieve", narrative="证据不足。", evidence=[]),
            )

        self.assertIsNone(answer)
        retrieve.assert_not_called()

    def test_unknown_table_and_raw_sql_actions_are_rejected_locally(self):
        actions = [
            {
                "action": "inspect_schema", "tables": ["invented_people"],
                "reason_code": "schema_uncertain",
            },
            {
                "action": "query", "question": "DELETE FROM scholars",
                "reason_code": "query_empty",
            },
            {"action": "stop", "reason_code": "cannot_progress"},
        ]
        with mock.patch.object(
            self.agent.read_explorer, "_next_action", side_effect=actions,
        ), mock.patch.object(self.agent.nl2sql, "answer") as query:
            answer = self.agent.read_explorer.explore(
                "请查找人物资料",
                None,
                dc.DBAnswer(kind="retrieve", narrative="未找到。", evidence=[]),
            )

        self.assertIsNone(answer)
        query.assert_not_called()

    def test_planned_table_mismatch_triggers_model_entity_rebinding(self):
        question = "肖仰华担任哪些职务？"
        self.assertEqual(self.agent.rag._entity_terms(question), [])
        wrong_evidence = dc.DBAnswer(
            kind="retrieve", narrative="只找到了编委记录。",
            evidence=[{
                "table": "audit_mentions", "columns": ["subject", "note"],
                "row": ["肖仰华", "导入记录"], "matched": "肖仰华",
            }],
        )
        operation = dc.DatabaseOperationPlan(
            mode="read", intent="retrieve", target_tables=["scholars"],
        )
        self.assertTrue(self.agent.read_explorer.should_explore(
            question,
            dc.IntentResult(intent="retrieve", confidence=0.9, source="model"),
            wrong_evidence,
            operation=operation,
        ))
        actions = [
            {
                "action": "retrieve", "entities": ["肖仰华"],
                "question": "肖仰华担任哪些具体职务？",
                "reason_code": "evidence_missing",
            },
            {
                "action": "finish", "answer_zh": "肖仰华是教授、博导。",
                "reason_code": "answer_grounded",
            },
        ]
        grounded = dc.DBAnswer(
            kind="retrieve", narrative="肖仰华是教授、博导。",
            evidence=[{
                "table": "scholars", "columns": ["name", "title"],
                "row": ["肖仰华", "教授、博导"], "matched": "肖仰华",
            }],
        )
        with mock.patch.object(
            self.agent.read_explorer, "_next_action", side_effect=actions,
        ), mock.patch.object(
            self.agent.rag, "recall_evidence", return_value=grounded.evidence,
        ) as retrieve:
            answer = self.agent.read_explorer.explore(
                question, None, wrong_evidence, target_tables=["scholars"],
            )

        self.assertIsNotNone(answer)
        self.assertIn("明确实体：“肖仰华”", retrieve.call_args.args[0])
        self.assertTrue(any(
            item.get("table") == "scholars" for item in answer.evidence
        ))
        invented = self.agent.read_explorer._normalize_action({
            "action": "retrieve", "entities": ["李四"],
            "question": "检索李四的职务",
        })
        self.assertEqual(
            self.agent.read_explorer._validate_action(
                invented, [], original_question=question,
            ),
            "entity_not_grounded",
        )

    def test_router_entities_and_read_target_are_schema_and_source_grounded(self):
        payload = {
            "intent": "query", "interaction": "auto",
            "target_table": "scholars",
            "entities": ["肖仰华", "李四", "肖仰华"],
            "confidence": 0.96,
            "reasoning": "查询明确学者的职务",
        }
        question = "肖仰华担任哪些职务？"
        with mock.patch.object(dc, "_llm_ask_json", return_value=payload):
            routed = self.agent.router.classify(
                question, self.agent.schema.compact(),
            )
        self.assertEqual(routed.entities, ["肖仰华"])
        operation = self.agent.operation_planner.from_intent(question, routed)
        self.assertEqual(operation.target_tables, ["scholars"])
        narrow = dc.DBAnswer(
            kind="query", narrative="已查询 title",
            sql="SELECT title FROM scholars", columns=["title"], rows=[["教授、博导"]],
        )
        self.assertTrue(self.agent.read_explorer.should_explore(
            question, routed, narrow, operation=operation,
        ))

    def test_database_identity_grounding_repairs_missing_router_entities(self):
        question = "肖仰华担任哪些职务？"
        keywords = self.agent.rag._keywords(question)
        self.assertIn("肖仰华", keywords)
        evidence = self.agent.rag.recall_evidence(question)
        grounded = self.agent.rag.ground_entity_context(question, evidence)
        self.assertEqual(grounded["entities"], ["肖仰华"])
        self.assertEqual(grounded["tables"], ["scholars"])

        routed = dc.IntentResult(
            intent="query", confidence=0.9, reasoning="查询职务", source="model",
        )
        narrow = dc.DBAnswer(
            kind="query", narrative="只查到 title",
            sql="SELECT title FROM scholars WHERE name='肖仰华'",
            columns=["title"], rows=[["教授、博导"]],
        )
        actions = [
            {
                "action": "retrieve", "entities": ["肖仰华"],
                "question": "肖仰华担任哪些职务？",
                "reason_code": "evidence_missing",
            },
            {
                "action": "finish",
                "answer_zh": "肖仰华是教授、博导，并担任上海市数据科学重点实验室主任。",
                "reason_code": "answer_grounded",
            },
        ]
        with mock.patch.object(self.agent.router, "classify", return_value=routed), \
             mock.patch.object(self.agent.nl2sql, "answer", return_value=narrow), \
             mock.patch.object(self.agent.read_explorer, "_next_action", side_effect=actions):
            answer = self.agent.ask(question)

        self.assertIn("实验室主任", answer.narrative)
        self.assertEqual(answer.operation["target_tables"], ["scholars"])
        self.assertTrue(any(
            step.get("tool") == "entity_grounding" for step in answer.steps
        ))
        self.assertTrue(any(
            item.get("table") == "scholars" for item in answer.evidence
        ))

    def test_repeated_action_is_blocked_and_loop_remains_bounded(self):
        repeated = {
            "action": "inspect_schema", "tables": ["scholars"],
            "reason_code": "schema_uncertain",
        }
        actions = [repeated, repeated, {"action": "stop", "reason_code": "cannot_progress"}]
        with mock.patch.object(
            self.agent.read_explorer, "_next_action", side_effect=actions,
        ) as planner:
            answer = self.agent.read_explorer.explore(
                "查找人物资料",
                None,
                dc.DBAnswer(kind="retrieve", narrative="未找到。", evidence=[]),
            )

        self.assertIsNone(answer)
        self.assertEqual(planner.call_count, 3)

    def test_grounded_primary_answer_keeps_exploration_audit_when_no_better_result(self):
        initial = dc.DBAnswer(
            kind="retrieve",
            narrative="已找到人物档案。",
            evidence=[{
                "table": "scholars", "columns": ["name", "title"],
                "row": ["肖仰华", "教授、博导"], "matched": "肖仰华",
            }],
        )
        actions = [
            {
                "action": "inspect_schema", "tables": ["scholars"],
                "reason_code": "schema_uncertain",
            },
            {"action": "stop", "reason_code": "cannot_progress"},
        ]
        with mock.patch.object(
            self.agent.read_explorer, "_next_action", side_effect=actions,
        ):
            answer = self.agent.read_explorer.explore(
                "肖仰华的资料是什么？请继续核实。",
                None,
                initial,
            )

        self.assertIsNotNone(answer)
        self.assertEqual(answer.narrative, initial.narrative)
        self.assertTrue(any(
            step.get("tool") == "bounded_read_explorer"
            and step.get("status") == "budget_complete"
            for step in answer.steps
        ))

    def test_grounded_primary_answer_can_finish_without_redundant_tool_call(self):
        initial = dc.DBAnswer(
            kind="retrieve", narrative="肖仰华是教授、博导。",
            evidence=[{
                "table": "scholars", "columns": ["name", "title"],
                "row": ["肖仰华", "教授、博导"], "matched": "肖仰华",
            }],
        )
        with mock.patch.object(
            self.agent.read_explorer,
            "_next_action",
            return_value={
                "action": "finish", "answer_zh": "证据已足够：肖仰华是教授、博导。",
                "reason_code": "answer_grounded",
            },
        ):
            answer = self.agent.read_explorer.explore(
                "肖仰华的资料是什么？请核实。",
                None,
                initial,
            )

        self.assertIsNotNone(answer)
        self.assertIn("证据已足够", answer.narrative)
        completed = [
            step for step in answer.steps
            if step.get("tool") == "bounded_read_explorer"
            and step.get("status") == "completed"
        ]
        self.assertEqual(completed[-1]["tool_actions"], 0)

    def test_real_deadline_cancels_a_blocked_planner_call(self):
        explorer = dc.BoundedReadExplorer(
            self.agent.nl2sql,
            self.agent.rag,
            self.agent.schema,
            budget=dc.ReadExplorationBudget(max_seconds=0.05),
        )

        def blocked(*args, **kwargs):
            event = dc._ACTIVE_CANCEL_EVENT.get()
            self.assertIsNotNone(event)
            self.assertTrue(event.wait(1.0))
            raise dc.LLMServiceError("cancelled")

        started = time.monotonic()
        with mock.patch.object(explorer, "_next_action", side_effect=blocked):
            answer = explorer.explore(
                "查找人物资料",
                None,
                dc.DBAnswer(kind="retrieve", narrative="未找到。", evidence=[]),
            )

        self.assertIsNone(answer)
        self.assertLess(time.monotonic() - started, 0.5)

    def test_write_and_successful_simple_query_never_enter_explorer(self):
        write_intent = dc.IntentResult(intent="write", confidence=1.0)
        write_plan = dc.DatabaseOperationPlan(mode="write", intent="write")
        self.assertFalse(self.agent.read_explorer.should_explore(
            "删除 scholars 的记录",
            write_intent,
            dc.DBAnswer(kind="write_pending", narrative="等待确认"),
            operation=write_plan,
        ))
        self.assertFalse(self.agent.read_explorer.should_explore(
            "scholars 有几条记录",
            dc.IntentResult(intent="query", confidence=1.0),
            dc.DBAnswer(
                kind="query", narrative="共 1 条", sql="SELECT COUNT(*) FROM scholars",
                columns=["COUNT(*)"], rows=[[1]],
            ),
            operation=dc.DatabaseOperationPlan(mode="read", intent="query"),
        ))
        empty_query = dc.DBAnswer(
            kind="query", narrative="未返回记录", sql="SELECT id FROM scholars",
            columns=["id"], rows=[],
        )
        self.assertFalse(self.agent.read_explorer.should_explore(
            "scholars 有哪些记录",
            dc.IntentResult(intent="query", confidence=1.0),
            empty_query,
            operation=dc.DatabaseOperationPlan(mode="read", intent="query"),
        ))
        self.assertTrue(self.agent.read_explorer.should_explore(
            "肖仰华的资料是什么？",
            dc.IntentResult(intent="query", confidence=0.8),
            empty_query,
            operation=dc.DatabaseOperationPlan(mode="read", intent="query"),
        ))

    def test_agent_entry_uses_explorer_after_empty_retrieval(self):
        actions = [
            {
                "action": "query", "tables": ["scholars"],
                "question": "查询 scholars 表中肖仰华的 title 和 short_bio",
                "reason_code": "evidence_missing",
            },
            {
                "action": "finish",
                "answer_zh": "肖仰华是教授、博导，并担任上海市数据科学重点实验室主任。",
                "reason_code": "answer_grounded",
            },
        ]
        routed = dc.IntentResult(
            intent="retrieve", confidence=0.9, reasoning="人物资料检索", source="model",
        )
        generated = dc.DBAnswer(
            kind="query", narrative="查询完成", sql="SELECT title, short_bio FROM scholars",
            columns=["title", "short_bio"],
            rows=[["教授、博导", "上海市数据科学重点实验室主任"]],
        )
        with mock.patch.object(self.agent.router, "classify", return_value=routed), \
             mock.patch.object(
                 self.agent.rag, "answer",
                 return_value=dc.DBAnswer(kind="retrieve", narrative="未找到。", evidence=[]),
             ), \
             mock.patch.object(self.agent.nl2sql, "answer", return_value=generated) as query, \
             mock.patch.object(self.agent.read_explorer, "_next_action", side_effect=actions):
            answer = self.agent.ask("查找肖仰华的资料")

        query.assert_called_once_with(
            "查询 scholars 表中肖仰华的 title 和 short_bio",
            history=None,
            allowed_tables=["scholars"],
        )
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.operation["action"], "query")
        self.assertIn("实验室主任", answer.narrative)
        self.assertTrue(any(
            step.get("tool") == "bounded_read_explorer"
            and step.get("status") == "completed"
            for step in answer.steps
        ))

    def test_model_can_direct_global_value_search_without_program_table_mapping(self):
        actions = [
            {
                "action": "search_values",
                "terms": ["肖仰华"],
                "tables": [],
                "match_mode": "exact",
                "reason_code": "evidence_missing",
            },
            {
                "action": "finish",
                "answer_zh": "肖仰华是教授、博导，并担任实验室主任。",
                "reason_code": "answer_grounded",
            },
        ]
        with mock.patch.object(
            self.agent.read_explorer, "_next_action", side_effect=actions,
        ) as planner:
            answer = self.agent.read_explorer.explore(
                "肖仰华做过什么工作？",
                None,
                dc.DBAnswer(kind="retrieve", narrative="尚未调查。", evidence=[]),
                anchor_entities=["肖仰华"],
            )

        self.assertIsNotNone(answer)
        self.assertEqual(planner.call_count, 2)
        self.assertIn("实验室主任", answer.narrative)
        self.assertTrue(any(
            item.get("table") == "scholars"
            and item.get("matched") == "肖仰华"
            for item in answer.evidence
        ))
        search_step = next(
            step for step in answer.steps
            if step.get("action") == "search_values"
        )
        self.assertGreaterEqual(search_step["queries_executed"], 2)
        self.assertNotIn("terms", search_step)

    def test_model_can_explore_opaque_schema_and_use_read_only_sql(self):
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("CREATE TABLE x_17 (c_1 TEXT, c_2 TEXT, c_3 TEXT)")
            conn.execute(
                "INSERT INTO x_17 VALUES (?, ?, ?)",
                ("肖仰华", "实验室主任", "数据科学"),
            )
            conn.commit()
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        actions = [
            {
                "action": "search_values", "terms": ["肖仰华"],
                "tables": [], "reason_code": "evidence_missing",
            },
            {
                "action": "inspect_schema", "tables": ["x_17"],
                "reason_code": "schema_uncertain",
            },
            {
                "action": "run_sql",
                "sql": "SELECT c_2, c_3 FROM x_17 WHERE c_1 = '肖仰华'",
                "reason_code": "evidence_missing",
            },
            {
                "action": "finish",
                "answer_zh": "肖仰华担任实验室主任，相关方向是数据科学。",
                "reason_code": "answer_grounded",
            },
        ]
        with mock.patch.object(
            agent.read_explorer, "_next_action", side_effect=actions,
        ):
            answer = agent.read_explorer.explore(
                "肖仰华做什么工作？",
                None,
                dc.DBAnswer(kind="retrieve", narrative="尚未调查。", evidence=[]),
                anchor_entities=["肖仰华"],
            )

        self.assertIsNotNone(answer)
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.columns, ["c_2", "c_3"])
        self.assertEqual(answer.rows, [["实验室主任", "数据科学"]])
        self.assertIn("实验室主任", answer.narrative)

    def test_model_can_follow_declared_relation_after_entity_wide_recall(self):
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute(
                "CREATE TABLE professional_engagements ("
                "person_ref INTEGER REFERENCES scholars(id), duty TEXT, context TEXT)"
            )
            conn.execute(
                "INSERT INTO professional_engagements VALUES (1, ?, ?)",
                ("期刊编委", "国际学术期刊"),
            )
            conn.commit()
        agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        actions = [
            {
                "action": "search_values", "terms": ["肖仰华"],
                "tables": [], "match_mode": "exact",
                "reason_code": "evidence_missing",
            },
            {
                "action": "find_relations",
                "tables": ["scholars", "professional_engagements"],
                "reason_code": "relation_unknown",
            },
            {
                "action": "run_sql",
                "sql": (
                    "SELECT p.duty, p.context FROM scholars AS s "
                    "JOIN professional_engagements AS p ON p.person_ref = s.id "
                    "WHERE s.name = '肖仰华'"
                ),
                "reason_code": "evidence_missing",
            },
            {
                "action": "finish",
                "answer_zh": "肖仰华还担任国际学术期刊编委。",
                "reason_code": "answer_grounded",
            },
        ]
        with mock.patch.object(
            agent.read_explorer, "_next_action", side_effect=actions,
        ):
            answer = agent.read_explorer.explore(
                "肖仰华有哪些工作？",
                None,
                dc.DBAnswer(kind="retrieve", narrative="尚未调查。", evidence=[]),
                anchor_entities=["肖仰华"],
            )

        self.assertIsNotNone(answer)
        self.assertEqual(answer.rows, [["期刊编委", "国际学术期刊"]])
        self.assertTrue(any(
            step.get("action") == "find_relations"
            and step.get("edge_count") == 1
            for step in answer.steps
        ))

    def test_generic_value_search_uses_bound_parameters(self):
        injected = self.agent.read_explorer.tools.search_values(
            ["肖仰华' OR 1=1 --"],
            tables=["scholars"],
            match_mode="exact",
        )
        self.assertEqual(injected["evidence"], [])
        normal = self.agent.read_explorer.tools.search_values(
            ["肖仰华"],
            tables=["scholars"],
            match_mode="exact",
        )
        self.assertEqual(len(normal["evidence"]), 1)
        with closing(sqlite3.connect(self.path)) as conn:
            self.assertEqual(
                conn.execute("SELECT COUNT(*) FROM scholars").fetchone()[0],
                1,
            )

    def test_value_search_reports_all_hit_tables_before_evidence_truncation(self):
        result = self.agent.read_explorer.tools.search_values(
            ["肖仰华"],
            tables=None,
            match_mode="exact",
            max_results=1,
        )
        self.assertEqual(len(result["evidence"]), 1)
        self.assertTrue(result["truncated"])
        self.assertEqual(
            {item["table"] for item in result["hit_tables"]},
            {"scholars", "audit_mentions"},
        )

    def test_model_read_strategy_controls_exploration_without_breaking_fast_path(self):
        initial = dc.DBAnswer(
            kind="retrieve",
            narrative="肖仰华是教授。",
            evidence=[{
                "table": "scholars", "columns": ["name", "title"],
                "row": ["肖仰华", "教授、博导"], "matched": "肖仰华",
            }],
        )
        operation = dc.DatabaseOperationPlan(mode="read", intent="retrieve")
        self.assertTrue(self.agent.read_explorer.should_explore(
            "肖仰华有哪些工作？",
            dc.IntentResult(
                intent="retrieve", entities=["肖仰华"],
                read_strategy="explore",
            ),
            initial,
            operation=operation,
        ))
        self.assertFalse(self.agent.read_explorer.should_explore(
            "肖仰华是谁？",
            dc.IntentResult(
                intent="retrieve", entities=["肖仰华"],
                read_strategy="fast",
            ),
            initial,
            operation=operation,
        ))

    def test_router_can_request_autonomous_exploration(self):
        payload = {
            "intent": "retrieve",
            "interaction": "auto",
            "target_table": "",
            "entities": ["肖仰华"],
            "read_strategy": "explore",
            "initial_exploration": {
                "action": "search_values", "terms": ["肖仰华"],
                "tables": [], "columns": [], "match_mode": "exact",
                "reason_code": "evidence_missing",
            },
            "confidence": 0.97,
            "reasoning": "信息可能散落在多张英文表中",
        }
        with mock.patch.object(dc, "_llm_ask_json", return_value=payload):
            routed = self.agent.router.classify(
                "肖仰华有哪些工作？",
                self.agent.schema.compact(),
            )
        self.assertEqual(routed.read_strategy, "explore")
        self.assertEqual(routed.entities, ["肖仰华"])
        self.assertEqual(
            routed.initial_exploration["action"], "search_values",
        )

    def test_router_first_exploration_action_is_reused_without_another_model_call(self):
        initial_action = {
            "action": "search_values", "terms": ["肖仰华"],
            "tables": [], "match_mode": "exact",
            "reason_code": "evidence_missing",
        }
        with mock.patch.object(
            self.agent.read_explorer,
            "_next_action",
            return_value={
                "action": "finish",
                "answer_zh": "肖仰华是教授、博导。",
                "reason_code": "answer_grounded",
            },
        ) as planner:
            answer = self.agent.read_explorer.explore(
                "肖仰华有哪些工作？",
                None,
                dc.DBAnswer(kind="retrieve", narrative="尚未调查。"),
                anchor_entities=["肖仰华"],
                initial_action=initial_action,
            )

        self.assertIsNotNone(answer)
        self.assertEqual(planner.call_count, 1)
        first_action = next(
            step for step in answer.steps
            if step.get("action") == "search_values"
        )
        self.assertTrue(first_action["reused_router_plan"])
        self.assertEqual(first_action["model_calls"], 0)

    def test_query_budget_reserves_one_final_grounded_synthesis_call(self):
        explorer = dc.AutonomousReadExplorer(
            self.agent.nl2sql,
            self.agent.rag,
            self.agent.schema,
            budget=dc.ReadExplorationBudget(
                max_planner_steps=4,
                max_tool_actions=3,
                max_query_actions=1,
                max_seconds=10.0,
                synthesis_reserve_seconds=2.0,
            ),
        )
        initial_action = {
            "action": "search_values", "terms": ["肖仰华"],
            "tables": [], "match_mode": "exact",
            "reason_code": "evidence_missing",
        }
        sql_action = {
            "action": "run_sql",
            "sql": "SELECT title FROM scholars WHERE name = '肖仰华'",
            "reason_code": "evidence_missing",
        }
        with mock.patch.object(
            explorer, "_next_action", return_value=sql_action,
        ) as planner, mock.patch.object(
            explorer,
            "_synthesize_answer",
            return_value="数据库证据显示肖仰华是教授、博导。",
        ) as synthesis:
            answer = explorer.explore(
                "肖仰华有哪些工作？",
                None,
                dc.DBAnswer(kind="retrieve", narrative="尚未调查。"),
                anchor_entities=["肖仰华"],
                initial_action=initial_action,
            )

        self.assertIsNotNone(answer)
        self.assertIn("教授、博导", answer.narrative)
        self.assertEqual(planner.call_count, 1)
        synthesis.assert_called_once()
        self.assertTrue(any(
            step.get("reserved_synthesis") is True
            and step.get("status") == "completed"
            for step in answer.steps
        ))

    def test_explore_strategy_bypasses_disposable_primary_model_answer(self):
        routed = dc.IntentResult(
            intent="retrieve",
            entities=["肖仰华"],
            read_strategy="explore",
            confidence=0.98,
            source="model",
        )
        explored = dc.DBAnswer(
            kind="retrieve",
            narrative="自主调查完成。",
            evidence=[{
                "table": "scholars", "columns": ["name", "title"],
                "row": ["肖仰华", "教授、博导"], "matched": "肖仰华",
            }],
        )
        with mock.patch.object(self.agent.router, "classify", return_value=routed), \
             mock.patch.object(self.agent.rag, "answer") as rag_answer, \
             mock.patch.object(self.agent.nl2sql, "answer") as query_answer, \
             mock.patch.object(self.agent.orchestrator, "answer") as compose_answer, \
             mock.patch.object(
                 self.agent.read_explorer, "explore", return_value=explored,
             ) as explorer:
            answer = self.agent.ask(
                "请调查肖仰华的全部任职资料，必要时查看相关表。"
            )

        self.assertEqual(answer.narrative, "自主调查完成。")
        rag_answer.assert_not_called()
        query_answer.assert_not_called()
        compose_answer.assert_not_called()
        initial = explorer.call_args.args[2]
        self.assertEqual(initial.narrative, "尚未执行数据调查。")
        self.assertTrue(any(
            step.get("tool") == "autonomous_read_dispatch"
            for step in initial.steps
        ))

    def test_catalog_and_global_search_stay_inside_authorized_scope(self):
        scoped = dc.DBQuillAgent(
            db_path=str(self.path),
            sample_rows=0,
            allowed_tables=["scholars"],
        )
        catalog = scoped.read_explorer.tools.catalog()
        self.assertIn("scholars", catalog)
        self.assertNotIn("audit_mentions", catalog)
        result = scoped.read_explorer.tools.search_values(
            ["肖仰华"], tables=None, match_mode="exact",
        )
        self.assertEqual(
            {item["table"] for item in result["evidence"]},
            {"scholars"},
        )


class FourLayerProbeFixTests(unittest.TestCase):
    """2026-08-20 四层面探测修复回归：引号召回、SQL 关系门禁、支付口径提示、
    值域路由、表关系叙述、纠正吸收、业务概览回退与流中断重试。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "probe.db"
        with closing(sqlite3.connect(self.path)) as conn:
            conn.executescript(
                """
                CREATE TABLE customers (id INTEGER PRIMARY KEY, name TEXT, region TEXT);
                CREATE TABLE orders (
                    id INTEGER PRIMARY KEY,
                    customer_id INTEGER REFERENCES customers(id),
                    region TEXT, status TEXT, paid_at DATETIME, note TEXT, total REAL);
                CREATE TABLE regional_targets (id INTEGER PRIMARY KEY, region TEXT, target REAL);
                CREATE TABLE products (id INTEGER PRIMARY KEY, name TEXT);
                CREATE TABLE inventory_snapshots (
                    product_id INTEGER REFERENCES products(id),
                    snapshot_date DATE, stock INTEGER,
                    PRIMARY KEY (product_id, snapshot_date));
                """
            )
            conn.executemany(
                "INSERT INTO customers(id, name, region) VALUES (?, ?, ?)",
                [(1, "高远", "华东"), (2, "陈晨", "华北")],
            )
            conn.executemany(
                "INSERT INTO orders(id, customer_id, region, status, paid_at, note, total)"
                " VALUES (?, ?, ?, ?, ?, ?, ?)",
                [
                    (1001, 1, "华东", "paid", "2026-08-01 10:00:00", "企业采购，加急配送", 100.0),
                    (1002, 2, "华北", "refunded", "2026-08-02 10:00:00", "普通配送", 50.0),
                ],
            )
            conn.execute("INSERT INTO regional_targets(id, region, target) VALUES (1, '华东', 1000)")
            conn.execute("INSERT INTO products(id, name) VALUES (1, '通勤双肩包')")
            conn.execute(
                "INSERT INTO inventory_snapshots(product_id, snapshot_date, stock)"
                " VALUES (1, '2026-08-17', 56)"
            )
            conn.commit()
        self.agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=5)

    def tearDown(self):
        self.tmp.cleanup()

    def test_quoted_term_prioritized_in_rag_keywords(self):
        keywords = self.agent.rag._keywords(
            "找出订单备注中提到“加急”的订单，只返回订单 id 和 note。"
        )
        self.assertEqual(keywords[0], "加急")
        evidence = self.agent.rag._recall(["加急"])
        self.assertTrue(any("加急" in str(item.get("matched") or "") for item in evidence))

    def test_nl2sql_relation_gate_blocks_invented_join(self):
        sql = (
            "WITH targets AS (SELECT region, SUM(target) AS amount FROM regional_targets GROUP BY region), "
            "actuals AS (SELECT region, SUM(total) AS amount FROM orders GROUP BY region) "
            "SELECT t.region, t.amount, a.amount FROM targets t JOIN actuals a ON t.region = a.region"
        )
        with mock.patch.object(
            dc, "_llm_ask_json", return_value={"sql": sql, "summary_zh": "区域目标对比"},
        ):
            answer = self.agent.nl2sql.answer("各区域的销售目标完成情况怎么样？")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "table_relationship")

    def test_nl2sql_relation_gate_allows_fk_and_explicit_join(self):
        fk_sql = (
            "SELECT c.name FROM orders o JOIN customers c "
            "ON o.customer_id = c.id WHERE o.status = 'paid'"
        )
        with mock.patch.object(
            dc, "_llm_ask_json", return_value={"sql": fk_sql, "summary_zh": "已支付客户"},
        ):
            answer = self.agent.nl2sql.answer("统计有已支付订单的客户姓名")
        self.assertEqual(answer.kind, "query")
        explicit_sql = (
            "SELECT o.region FROM orders o JOIN regional_targets t ON o.region = t.region"
        )
        with mock.patch.object(
            dc, "_llm_ask_json", return_value={"sql": explicit_sql, "summary_zh": "对齐区域"},
        ):
            answer = self.agent.nl2sql.answer(
                "通过 orders.region = regional_targets.region 对齐区域统计"
            )
        self.assertEqual(answer.kind, "query")

    def test_compose_graph_propagates_relation_clarification(self):
        class _StubNL2SQL:
            def answer(self, question, history=None, allowed_tables=None):
                return dc.DBAnswer(
                    kind="clarification",
                    narrative="请先补充表关联条件。",
                    clarification={
                        "missing": "table_relationship", "original_question": question,
                    },
                )

        class _StubRAG:
            def answer(self, question, history=None):
                return dc.DBAnswer(kind="retrieve", narrative="无内容", evidence=[])

        planner = dc.OperationGraphPlanner(self.agent.schema)
        executor = dc.OperationGraphExecutor(
            _StubNL2SQL(), _StubRAG(), schema=self.agent.schema,
            validator=dc.OperationGraphValidator(),
        )
        graph = planner.plan_compose("各区域的销售目标完成情况怎么样？")
        answer = executor.execute(graph)
        self.assertEqual(answer.kind, "clarification")
        node_states = {node["tool"]: node["status"] for node in answer.graph["nodes"]}
        self.assertEqual(node_states.get("query"), "failed")

    def test_payment_status_hint_scoped_to_paid_enum_schema(self):
        hints = self.agent.nl2sql._schema_semantic_hints(None)
        self.assertTrue(any("支付口径提示" in h for h in hints))
        self.assertFalse(any("支付口径提示" in h for h in self.agent.nl2sql._schema_semantic_hints(["customers"])))
        no_sample_agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=0)
        self.assertFalse(any(
            "支付口径提示" in h for h in no_sample_agent.nl2sql._schema_semantic_hints(None)
        ))

    def test_value_domain_question_not_captured_by_schema_view(self):
        plan = self.agent.operation_planner.plan_schema(
            "orders 表的 status 字段实际有哪些取值？每种取值各有多少条订单？"
        )
        self.assertIsNone(plan)
        fields_plan = self.agent.operation_planner.plan_schema("orders 表有哪些字段？")
        self.assertIsNotNone(fields_plan)
        self.assertEqual(fields_plan.action, "inspect_table")

    def test_table_relation_narration_from_fk_metadata(self):
        answer = self.agent.ask("inventory_snapshots 和 products 之间是什么关系？")
        self.assertEqual(answer.kind, "schema")
        self.assertIn("inventory_snapshots.product_id", answer.narrative)
        no_fk = self.agent.ask("orders 和 regional_targets 之间有什么关系？")
        self.assertEqual(no_fk.kind, "schema")
        self.assertIn("没有已声明的外键", no_fk.narrative)
        stats_plan = self.agent.operation_planner.plan_schema(
            "orders 和 customers 关联统计每个客户的订单总额"
        )
        self.assertNotEqual(getattr(stats_plan, "action", None), "inspect_relations")

    def test_correction_reply_merges_into_metric_definition(self):
        merged = self.agent.operation_planner.resolve_followup(
            "不对，客单价应该只统计已支付订单",
            history=[
                {"role": "user", "content": "平均客单价是多少？"},
                {"role": "assistant", "content": "计算订单总金额的平均值作为平均客单价"},
            ],
        )
        self.assertIn("指标口径：", merged)
        self.assertIn("只统计已支付订单", merged)
        self.assertIn("平均客单价", merged)
        self.assertTrue(dc.NaturalLanguageDatabasePlanner._has_metric_definition(merged))
        write_context = self.agent.operation_planner.resolve_followup(
            "不对，城市改成杭州",
            history=[
                {"role": "user", "content": "把客户城市改为厦门"},
                {"role": "assistant", "content": "写操作预览：更新 customers 表"},
            ],
        )
        self.assertEqual(write_context, "不对，城市改成杭州")

    def test_business_overview_fallback_on_empty_recall(self):
        with mock.patch.object(
            dc, "_llm_ask_json",
            return_value={"answer_zh": "这是一个包含客户、订单、商品与库存的电商演示库。"},
        ):
            answer = self.agent.rag.answer("这个数据库大概是做什么业务的？请根据表和字段说明。")
        self.assertEqual(answer.kind, "retrieve")
        self.assertIn("电商演示库", answer.narrative)
        self.assertEqual(answer.steps[0]["tool"], "schema_overview")
        with mock.patch.object(
            dc, "_llm_ask_json", side_effect=dc.DBQuillError("llm down"),
        ):
            failed = self.agent.rag.answer("这个数据库大概是做什么业务的？")
        self.assertIn("未在数据库中找到", failed.narrative)

    def test_chunked_encoding_error_retried_when_no_stream_output(self):
        import model_gateway

        calls = {"n": 0}

        def _fake_post(*args, **kwargs):
            calls["n"] += 1
            if calls["n"] == 1:
                raise requests.exceptions.ChunkedEncodingError("Response ended prematurely")
            response = mock.MagicMock()
            response.status_code = 200
            response.__enter__.return_value = response
            response.json.return_value = {
                "choices": [{"message": {"content": "ok"}}]
            }
            return response

        profile = {
            "retry_count": 1, "connect_timeout": 1, "read_timeout": 5,
            "total_timeout": 30, "verify_tls": True,
            "api_mode": "chat_completions",
        }
        with mock.patch.object(model_gateway.requests, "post", side_effect=_fake_post), \
             mock.patch.object(model_gateway, "_pause", return_value=False):
            text = "".join(model_gateway._request_text(
                profile, "http://example.invalid", {}, {"model": "m"}, use_stream=False,
            ))
        self.assertEqual(calls["n"], 2)
        self.assertEqual(text, "ok")

    def test_openai_compatible_v1_base_appends_operation_path(self):
        import model_gateway

        self.assertEqual(
            model_gateway._operation_url("https://provider.example/v1", "chat/completions"),
            "https://provider.example/v1/chat/completions",
        )
        self.assertEqual(
            model_gateway._operation_url("https://provider.example/v1/", "/responses"),
            "https://provider.example/v1/responses",
        )
        self.assertEqual(
            model_gateway._operation_url(
                "https://provider.example/v1/chat/completions", "chat/completions",
            ),
            "https://provider.example/v1/chat/completions",
        )

    def test_connection_error_retry_does_not_pollute_successful_response(self):
        import model_gateway

        calls = {"n": 0}

        def _fake_post(*args, **kwargs):
            calls["n"] += 1
            if calls["n"] == 1:
                raise requests.ConnectionError("temporary disconnect")
            response = mock.MagicMock()
            response.status_code = 200
            response.__enter__.return_value = response
            response.json.return_value = {
                "choices": [{"message": {"content": "ok"}}]
            }
            return response

        profile = {
            "retry_count": 1, "connect_timeout": 1, "read_timeout": 5,
            "total_timeout": 30, "verify_tls": True,
            "api_mode": "chat_completions",
        }
        with mock.patch.object(model_gateway.requests, "post", side_effect=_fake_post), \
             mock.patch.object(model_gateway, "_pause", return_value=False):
            text = "".join(model_gateway._request_text(
                profile, "http://example.invalid", {}, {"model": "m"}, use_stream=False,
            ))

        self.assertEqual(calls["n"], 2)
        self.assertEqual(text, "ok")

    def test_llm_retry_chain_enforces_total_deadline(self):
        import model_gateway

        response = mock.MagicMock()
        response.status_code = 200
        response.__enter__.return_value = response
        response.json.return_value = {
            "choices": [{"message": {"content": "late"}}]
        }
        profile = {
            "retry_count": 4,
            "connect_timeout": 10,
            "read_timeout": 120,
            "total_timeout": 10,
            "verify_tls": True,
            "api_mode": "chat_completions",
        }

        with mock.patch.object(model_gateway.requests, "post", return_value=response) as post:
            with mock.patch.object(
                model_gateway.time, "monotonic", side_effect=[0.0, 0.0, 11.0, 11.0],
            ):
                text = "".join(model_gateway._request_text(
                    profile, "http://example.invalid", {}, {"model": "m"}, use_stream=False,
                ))

        self.assertEqual(text, "!!!Error: TotalTimeout")
        self.assertEqual(post.call_count, 1)
        self.assertEqual(post.call_args.kwargs["timeout"], (10.0, 10.0))

    def test_llm_http_error_is_sanitized_and_not_parsed_as_model_json(self):
        raw = (
            '!!!Error: HTTP 402: {"error":{"message":"Insufficient Balance",'
            '"request_id":"secret-provider-id"}}'
        )

        with mock.patch.object(dc, "_llm_ask", return_value=raw):
            with self.assertRaises(dc.LLMServiceError) as raised:
                dc._llm_ask_json("prompt")

        self.assertEqual(str(raised.exception), "LLM 服务请求失败（HTTP 402）")
        self.assertNotIn("secret-provider-id", str(raised.exception))

    def test_nl2sql_does_not_retry_terminal_llm_service_error(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "items": dc.DBTable(name="items", columns=[dc.DBColumn("id", "INTEGER")]),
        })
        security = mock.Mock()
        executor = dc.NL2SQLExecutor(security, schema)

        with mock.patch.object(
            dc, "_llm_ask_json", side_effect=dc.LLMServiceError("LLM 服务请求失败（HTTP 402）"),
        ) as ask:
            answer = executor.answer("How many items are there?")

        self.assertEqual(answer.kind, "error")
        self.assertEqual(answer.error, "LLM 服务请求失败（HTTP 402）")
        self.assertEqual(ask.call_count, 1)
        security.execute.assert_not_called()


class SchemaContextTests(unittest.TestCase):
    def test_identifier_tokens_normalize_common_plurals_conservatively(self):
        tokens = dc.SchemaSnapshot._identifier_tokens(
            "accounts policies boxes addresses status analysis series",
        )

        self.assertTrue({"account", "policy", "box", "address"} <= tokens)
        self.assertTrue({"status", "analysis", "series"} <= tokens)
        self.assertEqual(
            dc.SchemaSnapshot._canonical_identifier_tokens("customer_accounts"),
            {"customer", "account"},
        )

    def test_prepared_schema_context_index_reuses_and_invalidates_on_mutation(self):
        schema = dc.SchemaSnapshot(
            db_path="fixture",
            tables={
                "accounts": dc.DBTable(
                    name="accounts",
                    columns=[
                        dc.DBColumn(f"field_{index}", "TEXT")
                        for index in range(20)
                    ],
                ),
            },
        )
        original = dc.SchemaSnapshot._identifier_tokens
        with mock.patch.object(
            dc.SchemaSnapshot, "_identifier_tokens", wraps=original,
        ) as tokenizer:
            first = schema.compact_for_question(
                "Show field 19.", max_total_columns=4,
            )
            cold_calls = tokenizer.call_count
            before_warm = tokenizer.call_count
            warm = schema.compact_for_question(
                "Show field 18.", max_total_columns=4,
            )
            warm_calls = tokenizer.call_count - before_warm

            schema.tables["accounts"].columns[19].name = "customer_tier"
            before_rebuild = tokenizer.call_count
            rebuilt = schema.compact_for_question(
                "Show customer tier.", max_total_columns=4,
            )
            rebuild_calls = tokenizer.call_count - before_rebuild

        self.assertIn("field_19:TEXT", first)
        self.assertIn("field_18:TEXT", warm)
        self.assertGreater(cold_calls, 10)
        self.assertLessEqual(warm_calls, 2)
        self.assertGreater(rebuild_calls, warm_calls * 3)
        self.assertIn("customer_tier:TEXT", rebuilt)
        self.assertNotIn("field_19:TEXT", rebuilt)

    def test_prepared_fk_graph_invalidates_when_declared_relation_changes(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "customers": dc.DBTable(name="customers", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
            ]),
            "orders": dc.DBTable(name="orders", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn(
                    "customer_id", "INTEGER",
                    fk_table="customers", fk_column="id",
                ),
            ]),
        })

        first = schema._prepared_fk_adjacency()
        second = schema._prepared_fk_adjacency()
        self.assertIs(first, second)
        self.assertTrue(
            schema._unique_shortest_fk_path_columns("customers", "orders"),
        )

        schema.tables["orders"].columns[1].fk_column = "missing"
        rebuilt = schema._prepared_fk_adjacency()
        self.assertIsNot(first, rebuilt)
        self.assertEqual(
            schema._unique_shortest_fk_path_columns("customers", "orders"),
            [],
        )

    def test_prepared_schema_context_index_is_consistent_during_concurrent_cold_use(self):
        schema = dc.SchemaSnapshot(
            db_path="fixture",
            tables={
                "accounts": dc.DBTable(
                    name="accounts",
                    columns=[
                        dc.DBColumn(f"field_{index}", "TEXT")
                        for index in range(40)
                    ],
                ),
            },
        )

        def render(index: int) -> tuple[int, str]:
            return index, schema.compact_for_question(
                f"Show field {index}.", max_total_columns=4,
            )

        with ThreadPoolExecutor(max_workers=8) as executor:
            rendered = list(executor.map(render, range(32, 40)))

        for index, context in rendered:
            self.assertIn(f"field_{index}:TEXT", context)

    def test_small_schema_keeps_columns_beyond_legacy_twelve_column_slice(self):
        columns = [dc.DBColumn(f"field_{index}", "TEXT") for index in range(20)]
        schema = dc.SchemaSnapshot(
            db_path="fixture",
            tables={"wide_table": dc.DBTable(name="wide_table", columns=columns)},
        )

        context = schema.compact_for_question("show field_19")

        self.assertIn("field_12:TEXT", context)
        self.assertIn("field_19:TEXT", context)
        self.assertNotIn("columns omitted", context)

    def test_large_schema_budget_prioritizes_question_and_relation_columns(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "countries": dc.DBTable(name="countries", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                *[dc.DBColumn(f"attribute_{index}", "TEXT") for index in range(12)],
                dc.DBColumn("head_of_state", "TEXT"),
            ]),
            "languages": dc.DBTable(name="languages", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn(
                    "country_id", "INTEGER", fk_table="countries", fk_column="id",
                ),
                *[dc.DBColumn(f"detail_{index}", "TEXT") for index in range(12)],
            ]),
        })

        context = schema.compact_for_question(
            "Which language belongs to the head of state?",
            max_total_columns=6,
        )

        self.assertIn("head_of_state:TEXT", context)
        self.assertIn("id:INTEGER(PK)", context)
        self.assertIn("country_id:INTEGER(FK->countries.id)", context)
        self.assertEqual(context.count("columns omitted"), 2)

    def test_three_thousand_column_schema_keeps_plural_filter_and_bounds_context(self):
        tables = {}
        for index in range(999):
            table_name = f"archive_partition_{index:04d}"
            tables[table_name] = dc.DBTable(name=table_name, columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("measure_alpha", "TEXT"),
                dc.DBColumn("measure_beta", "TEXT"),
            ])
        tables["customer_accounts"] = dc.DBTable(
            name="customer_accounts",
            columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("customer_lifetime_value", "REAL"),
                dc.DBColumn("account_segment", "TEXT"),
            ],
        )
        schema = dc.SchemaSnapshot(db_path="fixture", tables=tables)

        context = schema.compact_for_question(
            "Show customer lifetime value for enterprise accounts.",
        )

        self.assertIn("customer_lifetime_value:REAL", context)
        self.assertIn("account_segment:TEXT", context)
        self.assertIn("TABLE customer_accounts", context)
        self.assertIn("archive_partition_0000", context)
        self.assertIn("archive_partition_0998", context)
        self.assertIn("TABLE INDEX (details omitted)", context)
        self.assertLess(len(context), 60_000)

    def test_three_thousand_column_schema_reserves_unique_fk_path(self):
        tables = {}
        for index in range(247):
            table_name = f"history_shard_{index:03d}"
            tables[table_name] = dc.DBTable(
                name=table_name,
                columns=[
                    dc.DBColumn("id", "INTEGER", pk=True),
                    *[
                        dc.DBColumn(f"payload_{column}", "TEXT")
                        for column in range(11)
                    ],
                ],
            )
        tables["customers"] = dc.DBTable(name="customers", columns=[
            dc.DBColumn("id", "INTEGER", pk=True),
            dc.DBColumn("name", "TEXT"),
            *[dc.DBColumn(f"customer_detail_{index}", "TEXT") for index in range(10)],
        ])
        tables["account_links"] = dc.DBTable(name="account_links", columns=[
            dc.DBColumn("link_id", "INTEGER", pk=True),
            dc.DBColumn(
                "customer_id", "INTEGER", fk_table="customers", fk_column="id",
            ),
            dc.DBColumn(
                "order_id", "INTEGER", fk_table="orders", fk_column="id",
            ),
            *[dc.DBColumn(f"link_detail_{index}", "TEXT") for index in range(9)],
        ])
        tables["orders"] = dc.DBTable(name="orders", columns=[
            dc.DBColumn("id", "INTEGER", pk=True),
            dc.DBColumn("total_amount", "REAL"),
            *[dc.DBColumn(f"order_detail_{index}", "TEXT") for index in range(10)],
        ])
        schema = dc.SchemaSnapshot(db_path="fixture", tables=tables)

        context = schema.compact_for_question(
            "List customer names with order totals.",
            max_total_columns=8,
        )

        self.assertIn("TABLE customers", context)
        self.assertIn("id:INTEGER(PK)", context)
        self.assertIn("name:TEXT", context)
        self.assertIn("TABLE account_links", context)
        self.assertIn("customer_id:INTEGER(FK->customers.id)", context)
        self.assertIn("order_id:INTEGER(FK->orders.id)", context)
        self.assertIn("TABLE orders", context)
        self.assertIn("total_amount:REAL", context)
        self.assertIn("history_shard_000", context)
        self.assertIn("history_shard_246", context)
        self.assertEqual(
            set(schema._unique_shortest_fk_path_columns("customers", "orders")),
            {
                ("customers", 0),
                ("account_links", 1),
                ("account_links", 2),
                ("orders", 0),
            },
        )

    def test_equal_length_fk_paths_do_not_receive_unique_path_authority(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "customers": dc.DBTable(name="customers", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
            ]),
            "orders": dc.DBTable(name="orders", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
            ]),
            "customer_order_links": dc.DBTable(
                name="customer_order_links",
                columns=[
                    dc.DBColumn(
                        "customer_id", "INTEGER",
                        fk_table="customers", fk_column="id",
                    ),
                    dc.DBColumn(
                        "order_id", "INTEGER",
                        fk_table="orders", fk_column="id",
                    ),
                ],
            ),
            "customer_order_audit": dc.DBTable(
                name="customer_order_audit",
                columns=[
                    dc.DBColumn(
                        "customer_id", "INTEGER",
                        fk_table="customers", fk_column="id",
                    ),
                    dc.DBColumn(
                        "order_id", "INTEGER",
                        fk_table="orders", fk_column="id",
                    ),
                ],
            ),
        })

        self.assertEqual(
            schema._unique_shortest_fk_path_columns("customers", "orders"),
            [],
        )

    def test_question_relevant_column_dictionary_maps_business_terms_and_codes(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "schools": dc.DBTable(name="schools", columns=[
                dc.DBColumn(
                    "DOC", "INTEGER",
                    semantic_name="District Ownership Code",
                    description="Numeric code for the administrative authority category.",
                    value_description=(
                        "31 - State Special Schools\n"
                        "52 - Elementary School District\n"
                        "54 - Unified School District"
                    ),
                ),
                dc.DBColumn(
                    "StatusType", "TEXT",
                    description="Operational status of the district.",
                    value_description="Active - operating\nClosed - no longer operating",
                ),
            ]),
        })

        context = schema.compact_for_question(
            "Which state special schools have the most students?",
        )

        self.assertIn("QUESTION-RELEVANT COLUMN DICTIONARY", context)
        self.assertIn("schools.DOC", context)
        self.assertIn("31 - State Special Schools", context)
        self.assertNotIn("Closed - no longer operating", context)

    def test_grounding_echoes_only_question_matched_values_and_date_shapes(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "gasstations": dc.DBTable(name="gasstations", columns=[
                dc.DBColumn(
                    "Segment", "TEXT", sample_values=["Premium", "Other", "Discount"],
                ),
                dc.DBColumn(
                    "event_date", "TEXT", sample_values=["2019-10-08T12:00:00"],
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        context = executor._schema_context(
            'How many "discount" stations were recorded on October 8, 2019?',
        )

        self.assertIn("真实拼写为 'Discount'", context)
        self.assertIn("YYYY-MM-DDTHH:MM:SS", context)
        self.assertNotIn("Premium", context)
        self.assertNotIn("2019-10-08T12:00:00", context)

    def test_paired_source_destination_foreign_keys_add_neutral_role_hint(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "airports": dc.DBTable(name="airports", columns=[
                dc.DBColumn("AirportCode", "TEXT", pk=True),
            ]),
            "flights": dc.DBTable(name="flights", columns=[
                dc.DBColumn(
                    "SourceAirport", "TEXT", fk_table="airports", fk_column="AirportCode",
                ),
                dc.DBColumn(
                    "DestAirport", "TEXT", fk_table="airports", fk_column="AirportCode",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        context = executor._schema_context("Which airport has the most flights?")

        self.assertIn("多角色外键提示", context)
        self.assertIn("flights.SourceAirport", context)
        self.assertIn("flights.DestAirport", context)
        self.assertIn("未限定角色", context)

    def test_semantic_retry_hint_flags_extra_simple_projection(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "teacher": dc.DBTable(name="teacher", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Age", "INTEGER"),
                dc.DBColumn("Hometown", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        hint = executor._semantic_retry_hint(
            "What is the age and hometown of every teacher?",
            "SELECT Name, Age, Hometown FROM teacher",
        )

        self.assertIn("Name", hint)
        self.assertIn("未请求", hint)
        self.assertEqual(
            executor._semantic_retry_hint(
                "What are the name, age and hometown of every teacher?",
                "SELECT Name, Age, Hometown FROM teacher",
            ),
            "",
        )

    def test_semantic_retry_hint_flags_missing_coordinated_projection(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Documents": dc.DBTable(name="Documents", columns=[
                dc.DBColumn("Document_ID", "INTEGER", pk=True),
                dc.DBColumn("Document_Name", "TEXT"),
            ]),
            "Paragraphs": dc.DBTable(name="Paragraphs", columns=[
                dc.DBColumn("Paragraph_ID", "INTEGER", pk=True),
                dc.DBColumn("Document_ID", "INTEGER", fk_table="Documents", fk_column="Document_ID"),
                dc.DBColumn("Paragraph_Text", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        hint = executor._semantic_retry_hint(
            "Show all paragraph ids and texts for the document with name 'Welcome to NY'.",
            "SELECT p.Paragraph_Text FROM Paragraphs p JOIN Documents d "
            "ON p.Document_ID=d.Document_ID WHERE d.Document_Name='Welcome to NY'",
        )

        self.assertIn("遗漏", hint)
        self.assertIn("Paragraph_ID", hint)
        self.assertNotIn("Document_Name", hint)

    def test_semantic_retry_hint_flags_missing_id_alongside_maker(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "car_makers": dc.DBTable(name="car_makers", columns=[
                dc.DBColumn("Id", "INTEGER", pk=True),
                dc.DBColumn("Maker", "TEXT"),
            ]),
            "model_list": dc.DBTable(name="model_list", columns=[
                dc.DBColumn("Maker", "INTEGER", fk_table="car_makers", fk_column="Id"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        hint = executor._semantic_retry_hint(
            "What are the ids and makers of all car makers that produce at least 2 models?",
            "SELECT cm.Maker FROM car_makers cm JOIN model_list ml ON cm.Id=ml.Maker "
            "GROUP BY cm.Id, cm.Maker",
        )

        self.assertIn("遗漏", hint)
        self.assertIn("Id", hint)

    def test_projection_contract_maps_generic_id_to_entity_primary_key(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Pets": dc.DBTable(name="Pets", columns=[
                dc.DBColumn("PetID", "INTEGER", pk=True),
                dc.DBColumn("weight", "REAL"),
                dc.DBColumn("pet_age", "INTEGER"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        conflict = executor._projection_conflict(
            "Find the id and weight of all pets whose age is older than 1.",
            "SELECT weight FROM Pets WHERE pet_age > 1",
        )

        self.assertIsNotNone(conflict)
        self.assertEqual(
            conflict.constraints["required_output_columns"], ["PetID", "weight"],
        )
        self.assertEqual(conflict.constraints["required_output_bindings"], [
            {"table": "Pets", "column": "PetID"},
            {"table": "Pets", "column": "weight"},
        ])

    def test_projection_contract_maps_connector_column_and_entity_id(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "visitor": dc.DBTable(name="visitor", columns=[
                dc.DBColumn("ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Level_of_membership", "TEXT"),
            ]),
            "visit": dc.DBTable(name="visit", columns=[
                dc.DBColumn("visitor_ID", "INTEGER", fk_table="visitor", fk_column="ID"),
                dc.DBColumn("Total_spent", "REAL"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        conflict = executor._projection_conflict(
            "What are the id, name and membership level of visitors who have "
            "spent the largest amount of money in total in all museum tickets?",
            "SELECT v.Name FROM visitor v JOIN visit x ON v.ID=x.visitor_ID "
            "GROUP BY v.ID, v.Name ORDER BY SUM(x.Total_spent) DESC LIMIT 1",
        )

        self.assertIsNotNone(conflict)
        self.assertEqual(conflict.constraints["required_output_columns"], [
            "ID", "Name", "Level_of_membership",
        ])
        self.assertNotIn("visitor_ID", conflict.constraints["required_output_columns"])

    def test_projection_contract_treats_of_phrase_as_entity_filter(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Addresses": dc.DBTable(name="Addresses", columns=[
                dc.DBColumn("zip_postcode", "TEXT"),
                dc.DBColumn("city", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        self.assertEqual(
            executor._semantic_retry_hint(
                "What is the zip code of the address in the city Port Chelsea?",
                "SELECT zip_postcode FROM Addresses WHERE city = 'Port Chelsea'",
            ),
            "",
        )

    def test_projection_contract_does_not_promote_reference_entity_to_output(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Features": dc.DBTable(name="Features", columns=[
                dc.DBColumn("feature_name", "TEXT"),
                dc.DBColumn("feature_type_code", "TEXT"),
                dc.DBColumn("feature_type_name", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        self.assertEqual(
            executor._semantic_retry_hint(
                "What is the feature type name of feature AirCon?",
                "SELECT feature_type_name FROM Features WHERE feature_name = 'AirCon'",
            ),
            "",
        )
        hint = executor._semantic_retry_hint(
            "What is the feature type name of feature AirCon?",
            "SELECT feature_type_code, feature_type_name FROM Features "
            "WHERE feature_name = 'AirCon'",
        )
        self.assertIn("未请求", hint)
        self.assertNotIn("遗漏", hint)

    def test_projection_contract_matches_singular_question_to_plural_column_token(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "users": dc.DBTable(name="users", columns=[
                dc.DBColumn("Reputation", "INTEGER"),
                dc.DBColumn("UpVotes", "INTEGER"),
                dc.DBColumn(
                    "DisplayName", "TEXT", semantic_name="user display name",
                    description="the public display name of the user",
                ),
                dc.DBColumn(
                    "Views", "INTEGER", semantic_name="user profile view number",
                    description="number of times the user profile was viewed",
                ),
                dc.DBColumn(
                    "DownVotes", "INTEGER", semantic_name="user down vote number",
                    description="number of down votes cast by the user",
                ),
                dc.DBColumn(
                    "CreationDate", "TEXT",
                    semantic_name="user creation date",
                    description="the date when the user account was created",
                ),
                dc.DBColumn(
                    "LastAccessDate", "TEXT",
                    semantic_name="user last access date",
                    description="the date when the user last accessed the site",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        self.assertEqual(
            executor._semantic_retry_hint(
                "Give the user's reputation and up vote number of the user that commented.",
                "SELECT Reputation, UpVotes FROM users",
            ),
            "",
        )
        noisy_sql = "SELECT DisplayName, Reputation, Views, UpVotes, DownVotes FROM users"
        hint = executor._semantic_retry_hint(
            "Give the user's reputation and up vote number of the user that commented.",
            noisy_sql,
        )
        self.assertIn("DisplayName", hint)
        self.assertIn("Views", hint)
        self.assertIn("DownVotes", hint)

    def test_projection_contract_treats_of_year_as_reference_not_output(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "seasons": dc.DBTable(name="seasons", columns=[
                dc.DBColumn("url", "TEXT"),
                dc.DBColumn("year", "INTEGER"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "Show me the season page of year when the race No. 901 took place."
        hint = executor._semantic_retry_hint(
            question, "SELECT url, year FROM seasons",
        )

        self.assertIn("year", hint)

    def test_projection_contract_uses_explicit_tuple_evidence_for_coordinates(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "circuits": dc.DBTable(name="circuits", columns=[
                dc.DBColumn("location", "TEXT"),
                dc.DBColumn("lat", "REAL"),
                dc.DBColumn("lng", "REAL"),
                dc.DBColumn("name", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "What's the location coordinates of Silverstone Circuit?\n"
            "Relevant business evidence supplied by the user: "
            "location coordinates refers to (lat, lng); Silverstone Circuit refers to name"
        )

        self.assertEqual(
            executor._semantic_retry_hint(
                question, "SELECT lat, lng FROM circuits WHERE name='Silverstone Circuit'",
            ),
            "",
        )
        hint = executor._semantic_retry_hint(
            question, "SELECT location FROM circuits WHERE name='Silverstone Circuit'",
        )
        self.assertIn("lat", hint)
        self.assertIn("lng", hint)

    def test_projection_contract_maps_which_entity_to_declared_name_column(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "users": dc.DBTable(name="users", columns=[
                dc.DBColumn(
                    "DisplayName", "TEXT", semantic_name="user display name",
                    description="the public display name of the user",
                ),
                dc.DBColumn("Reputation", "INTEGER"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "Which user has a higher reputation, Harlan or Jarrod Dixon?"

        self.assertEqual(
            executor._semantic_retry_hint(question, "SELECT DisplayName FROM users"),
            "",
        )
        self.assertIn(
            "DisplayName",
            executor._semantic_retry_hint(question, "SELECT Reputation FROM users"),
        )

    def test_projection_contract_prefers_mid_sentence_state_output(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "legalities": dc.DBTable(name="legalities", columns=[
                dc.DBColumn("status", "TEXT"),
                dc.DBColumn("format", "TEXT"),
            ]),
            "cards": dc.DBTable(name="cards", columns=[
                dc.DBColumn("type", "TEXT"),
                dc.DBColumn("types", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "For artifact type of cards that do not have multiple faces on the same card, "
            "state its legalities status for vintage play format."
        )

        self.assertEqual(
            executor._output_request_phrase(question), "its legalities status",
        )

    def test_projection_contract_prefers_showing_clause_for_rank_output(self):
        self.assertEqual(
            dc.NL2SQLExecutor._output_request_phrase(
                "Rank schools by their average score in Writing where the score is greater "
                "than 499, showing their charter numbers."
            ),
            "their charter numbers",
        )

    def test_count_question_does_not_bind_entity_noun_as_scalar_projection(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "car_makers": dc.DBTable(name="car_makers", columns=[
                dc.DBColumn("Id", "INTEGER", pk=True),
            ]),
            "model_list": dc.DBTable(name="model_list", columns=[
                dc.DBColumn("ModelId", "INTEGER", pk=True),
                dc.DBColumn("Maker", "INTEGER", fk_table="car_makers", fk_column="Id"),
                dc.DBColumn("Model", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        contract = executor._compile_relational_contract(
            "How many car models are produced in the usa?"
        )

        self.assertEqual(contract.output_bindings, [])
        self.assertNotIn("exact_output_projection", contract.required_operators)

    def test_projection_contract_uses_business_dictionary_for_measurement_family(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Examination": dc.DBTable(name="Examination", columns=[
                dc.DBColumn(
                    "aCL IgA", "REAL",
                    semantic_name="anti-Cardiolipin antibody (IgA)",
                    description="anti-Cardiolipin antibody (IgA) concentration",
                ),
                dc.DBColumn(
                    "aCL IgG", "REAL",
                    semantic_name="anti-Cardiolipin antibody (IgG)",
                    description="anti-Cardiolipin antibody (IgG) concentration",
                ),
                dc.DBColumn(
                    "aCL IgM", "REAL",
                    semantic_name="anti-Cardiolipin antibody (IgM)",
                    description="anti-Cardiolipin antibody (IgM) concentration",
                ),
                dc.DBColumn("Examination Date", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        hint = executor._semantic_retry_hint(
            "For the patient with SLE, what was his/her anti-Cardiolipin antibody "
            "concentration status on 1993/11/12?",
            'SELECT "aCL IgM" FROM Examination WHERE "Examination Date" = \'1993-11-12\'',
        )

        self.assertIn("遗漏", hint)
        self.assertIn("aCL IgA", hint)
        self.assertIn("aCL IgG", hint)

    def test_projection_contract_normalizes_numbered_sibling_columns(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "schools": dc.DBTable(name="schools", columns=[
                dc.DBColumn("AdmEmail1", "TEXT"),
                dc.DBColumn("AdmEmail2", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        hint = executor._semantic_retry_hint(
            "What are the administrator email addresses?",
            "SELECT AdmEmail1 FROM schools",
        )

        self.assertIn("AdmEmail2", hint)

    def test_relational_contract_rejects_concatenated_dictionary_tuple(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "member": dc.DBTable(name="member", columns=[
                dc.DBColumn("first_name", "TEXT"),
                dc.DBColumn("last_name", "TEXT"),
                dc.DBColumn("phone", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "Give the full name and contact number of members.\n"
            "Relevant business evidence supplied by the user: "
            "full name refers to first_name, last_name; contact number refers to phone"
        )
        executor.last_relational_contract = executor._compile_relational_contract(question)

        hint = executor._semantic_retry_hint(
            question,
            "SELECT first_name || ' ' || last_name, phone FROM member",
        )

        self.assertEqual(
            executor.last_relational_contract.output_bundles,
            [["first_name", "last_name"]],
        )
        self.assertEqual(
            executor.last_relational_contract.output_bindings,
            [
                {"table": "member", "column": "first_name"},
                {"table": "member", "column": "last_name"},
                {"table": "member", "column": "phone"},
            ],
        )
        self.assertIn("多个独立列", hint)
        self.assertIn("first_name", hint)
        self.assertIn("last_name", hint)
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT first_name, last_name, phone FROM member",
                executor.last_relational_contract,
            ),
            "",
        )

    def test_relational_contract_keeps_qualified_output_bindings_for_local_compiler(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "drivers": dc.DBTable(name="drivers", columns=[
                dc.DBColumn("driverId", "INTEGER", pk=True),
                dc.DBColumn("forename", "TEXT"),
                dc.DBColumn("surname", "TEXT"),
                dc.DBColumn("name", "TEXT"),
            ]),
            "lapTimes": dc.DBTable(name="lapTimes", columns=[
                dc.DBColumn("driverId", "INTEGER", fk_table="drivers", fk_column="driverId"),
                dc.DBColumn("raceId", "INTEGER", fk_table="races", fk_column="raceId"),
                dc.DBColumn("milliseconds", "INTEGER"),
            ]),
            "races": dc.DBTable(name="races", columns=[
                dc.DBColumn("raceId", "INTEGER", pk=True),
                dc.DBColumn("name", "TEXT"),
            ]),
            "results": dc.DBTable(name="results", columns=[
                dc.DBColumn("milliseconds", "INTEGER"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "What is the best lap time recorded? List the driver and race.\n"
            "Relevant business evidence supplied by the user: "
            "best lap time refers to min(milliseconds); "
            "driver refers to drivers.forename, drivers.surname; "
            "race refers to races.name"
        )

        contract = executor._compile_relational_contract(question)
        conflict = dc.QuerySemanticConflict(
            code="relational_algebra_contract",
            message="projection mismatch",
            constraints={"relational_contract": contract.as_dict()},
        )
        bindings = executor._projection_lock_bindings(conflict)
        sql = executor._compile_projection_locked_sql(
            bindings,
            "FROM drivers d JOIN lapTimes l ON d.driverId=l.driverId "
            "JOIN races r ON l.raceId=r.raceId ORDER BY l.milliseconds LIMIT 1",
        )

        self.assertEqual(contract.version, "1.10")
        self.assertEqual(contract.output_columns, [
            "milliseconds", "forename", "surname", "name",
        ])
        self.assertEqual(bindings, [
            {
                "table": "",
                "column": "milliseconds",
                "table_candidates": ["lapTimes", "results"],
            },
            {"table": "drivers", "column": "forename"},
            {"table": "drivers", "column": "surname"},
            {"table": "races", "column": "name"},
        ])
        self.assertEqual(
            sql,
            'SELECT l."milliseconds", d."forename", d."surname", '
            'r."name" FROM drivers d JOIN lapTimes l ON d.driverId=l.driverId '
            'JOIN races r ON l.raceId=r.raceId ORDER BY l.milliseconds LIMIT 1',
        )
        executor.last_relational_contract = contract
        self.assertIsNotNone(executor._projection_conflict(question, sql))
        self.assertIsNone(executor._semantic_conflict(
            question,
            sql,
            locked_projection_columns=[
                "milliseconds", "forename", "surname", "name",
            ],
        ))

    def test_relational_contract_compiles_any_all_extreme_direction(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "country": dc.DBTable(name="country", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("SurfaceArea", "REAL"),
                dc.DBColumn("Continent", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "What countries have greater surface area than any country in Europe?"
        contract = executor._compile_relational_contract(question)

        hint = executor._relational_algebra_retry_hint(
            question,
            "SELECT Name FROM country WHERE SurfaceArea > "
            "(SELECT MAX(SurfaceArea) FROM country WHERE Continent='Europe')",
            contract,
        )

        self.assertEqual(contract.comparison_quantifier, "existential_any")
        self.assertEqual(contract.comparison_direction, "greater")
        self.assertIn("MIN", hint)
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT Name FROM country WHERE SurfaceArea > "
                "(SELECT MIN(SurfaceArea) FROM country WHERE Continent='Europe')",
                contract,
            ),
            "",
        )
        self.assertIn(
            "比较符",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT Name FROM country WHERE SurfaceArea < "
                "(SELECT MIN(SurfaceArea) FROM country WHERE Continent='Europe')",
                contract,
            ),
        )

    def test_relational_contract_distinguishes_rank_output_from_sorting(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "schools": dc.DBTable(name="schools", columns=[
                dc.DBColumn("CharterNum", "TEXT"),
                dc.DBColumn("AvgScrWrite", "REAL"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "Rank schools by average writing score, showing their charter numbers."
        contract = executor._compile_relational_contract(question)

        self.assertIn("rank_projection", contract.required_operators)
        self.assertIn(
            "排名列",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT CharterNum FROM schools ORDER BY AvgScrWrite DESC",
                contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT CharterNum, AvgScrWrite, "
                "RANK() OVER (ORDER BY AvgScrWrite DESC) FROM schools",
                contract,
            ),
            "",
        )

    def test_relational_contract_treats_adjective_flag_as_filter(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "countrylanguage": dc.DBTable(name="countrylanguage", columns=[
                dc.DBColumn("CountryCode", "TEXT"),
                dc.DBColumn("Language", "TEXT"),
                dc.DBColumn("IsOfficial", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "What is the official language spoken in the country?"
        contract = executor._compile_relational_contract(question)

        self.assertEqual(contract.modifier_filters, ["countrylanguage.IsOfficial"])
        self.assertEqual(contract.output_columns, ["Language"])
        self.assertEqual(contract.output_bindings, [
            {"table": "countrylanguage", "column": "Language"},
        ])
        self.assertIn(
            "筛选标志",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT IsOfficial, Language FROM countrylanguage WHERE IsOfficial='T'",
                contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT Language FROM countrylanguage WHERE IsOfficial='T'",
                contract,
            ),
            "",
        )
        self.assertIn(
            "没有使用该标志",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT Language FROM countrylanguage",
                contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT Language AS official_language FROM countrylanguage "
                "WHERE IsOfficial='T'",
                contract,
            ),
            "",
        )
        executor.last_relational_contract = contract
        self.assertIsNone(executor._semantic_conflict(
            question,
            "SELECT Language FROM countrylanguage WHERE IsOfficial='T'",
        ))

    def test_relational_contract_requires_having_for_relationship_threshold(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "maker": dc.DBTable(name="maker", columns=[dc.DBColumn("id", "INTEGER")]),
            "model": dc.DBTable(name="model", columns=[dc.DBColumn("maker_id", "INTEGER")]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "Which makers produce at least 2 models?"
        contract = executor._compile_relational_contract(question)

        self.assertIn("having", contract.required_operators)
        self.assertIn(
            "HAVING",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT m.id FROM maker m JOIN model x ON x.maker_id=m.id "
                "GROUP BY m.id",
                contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT m.id FROM maker m JOIN model x ON x.maker_id=m.id "
                "GROUP BY m.id HAVING COUNT(*) >= 2",
                contract,
            ),
            "",
        )

    def test_relational_contract_binds_entity_key_aggregate_input_and_all_thresholds(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "players": dc.DBTable(name="players", columns=[
                dc.DBColumn("player_id", "INTEGER", pk=True),
                dc.DBColumn("first_name", "TEXT"),
            ]),
            "rankings": dc.DBTable(name="rankings", columns=[
                dc.DBColumn(
                    "player_id", "INTEGER", fk_table="players", fk_column="player_id",
                ),
                dc.DBColumn("ranking", "REAL"),
                dc.DBColumn("score", "REAL"),
            ]),
            "models": dc.DBTable(name="models", columns=[
                dc.DBColumn("model_id", "INTEGER"),
            ]),
            "cars": dc.DBTable(name="cars", columns=[
                dc.DBColumn("car_id", "INTEGER"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        aggregate_question = "Find the average ranking for each player and their first name."
        aggregate_contract = executor._compile_relational_contract(aggregate_question)

        self.assertEqual(aggregate_contract.grouping_keys, ["players.player_id"])
        self.assertEqual(aggregate_contract.aggregate_requirements, [{
            "function": "AVG", "column": "rankings.ranking",
        }])
        self.assertIn(
            "实体主键",
            executor._relational_algebra_retry_hint(
                aggregate_question,
                "SELECT p.first_name, AVG(r.ranking) FROM players p "
                "JOIN rankings r ON r.player_id=p.player_id GROUP BY p.first_name",
                aggregate_contract,
            ),
        )
        self.assertIn(
            "聚合函数/输入列不匹配",
            executor._relational_algebra_retry_hint(
                aggregate_question,
                "SELECT p.first_name, AVG(r.score) FROM players p "
                "JOIN rankings r ON r.player_id=p.player_id "
                "GROUP BY p.player_id, p.first_name",
                aggregate_contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                aggregate_question,
                "SELECT p.first_name, AVG(r.ranking) FROM players p "
                "JOIN rankings r ON r.player_id=p.player_id "
                "GROUP BY p.player_id, p.first_name",
                aggregate_contract,
            ),
            "",
        )

        threshold_question = (
            "Which makers produce at least 2 models and make more than 3 cars?"
        )
        threshold_contract = executor._compile_relational_contract(threshold_question)
        self.assertEqual(threshold_contract.relationship_thresholds, [
            {"operator": ">=", "value": 2, "subject": "models"},
            {"operator": ">", "value": 3, "subject": "cars"},
        ])
        self.assertIn(
            "cars > 3",
            executor._relational_algebra_retry_hint(
                threshold_question,
                "SELECT maker_id FROM models GROUP BY maker_id HAVING COUNT(*) >= 2",
                threshold_contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                threshold_question,
                "SELECT maker_id FROM models GROUP BY maker_id "
                "HAVING COUNT(DISTINCT model_id) >= 2 AND COUNT(car_id) > 3",
                threshold_contract,
            ),
            "",
        )

    def test_projection_contract_understands_alias_compounds_and_specific_metadata(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "TV_Channel": dc.DBTable(name="TV_Channel", columns=[
                dc.DBColumn("Package_Option", "TEXT"),
                dc.DBColumn("series_name", "TEXT"),
                dc.DBColumn("Hight_definition_TV", "TEXT"),
            ]),
            "country": dc.DBTable(name="country", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Population", "INTEGER"),
                dc.DBColumn("LifeExpectancy", "REAL"),
                dc.DBColumn("SurfaceArea", "REAL"),
            ]),
            "races": dc.DBTable(name="races", columns=[
                dc.DBColumn("year", "INTEGER"),
                dc.DBColumn("time", "TEXT"),
            ]),
            "results": dc.DBTable(name="results", columns=[
                dc.DBColumn("milliseconds", "INTEGER"),
            ]),
            "Examination": dc.DBTable(name="Examination", columns=[
                dc.DBColumn(
                    "aCL IgA", "REAL",
                    semantic_name="anti-Cardiolipin antibody (IgA)",
                    description="anti-Cardiolipin antibody (IgA) concentration",
                ),
                dc.DBColumn(
                    "aCL IgG", "REAL",
                    semantic_name="anti-Cardiolipin antibody (IgG)",
                    description="anti-Cardiolipin antibody (IgG) concentration",
                ),
                dc.DBColumn(
                    "aCL IgM", "REAL",
                    semantic_name="anti-Cardiolipin antibody (IgM)",
                    description="anti-Cardiolipin antibody (IgM) concentration",
                ),
                dc.DBColumn(
                    "ANA", "REAL", semantic_name="anti-nucleus antibody",
                    description="anti-nucleus antibody concentration",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        self.assertEqual(
            executor._projection_retry_hint(
                "What are the package options and the name of the series for the TV "
                "Channel that supports high definition TV?",
                "SELECT Package_Option, series_name FROM TV_Channel "
                "WHERE Hight_definition_TV='yes'",
            ),
            "",
        )
        self.assertEqual(
            executor._projection_retry_hint(
                "Find the name, population and expected life length of the asian "
                "country with the largest area.",
                "SELECT Name, Population, LifeExpectancy FROM country "
                "ORDER BY SurfaceArea DESC LIMIT 1",
            ),
            "",
        )
        self.assertEqual(
            executor._projection_retry_hint(
                "What is the average time in seconds for each year?",
                "SELECT races.year, AVG(results.milliseconds) / 1000.0 AS time "
                "FROM races JOIN results ON results.raceId=races.raceId "
                "GROUP BY races.year",
            ),
            "",
        )
        self.assertIn(
            "ANA",
            executor._projection_retry_hint(
                "What was the anti-Cardiolipin antibody concentration status?",
                "SELECT `aCL IgA`, `aCL IgG`, `aCL IgM`, ANA FROM Examination",
            ),
        )

    def test_join_path_contract_checks_actual_on_columns_not_only_table_connectivity(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "parents": dc.DBTable(name="parents", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("code", "TEXT"),
            ]),
            "children": dc.DBTable(name="children", columns=[
                dc.DBColumn(
                    "parent_id", "INTEGER", fk_table="parents", fk_column="id",
                ),
                dc.DBColumn("wrong_code", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        sql = (
            "SELECT p.id FROM parents p JOIN children c "
            "ON c.wrong_code = p.code"
        )

        self.assertTrue(
            dc.SchemaRelationAnalyzer(schema).analyze(
                ["parents", "children"], "",
            )["connected"],
        )
        self.assertIn("实际 JOIN 边", executor._join_path_retry_hint("", sql))
        self.assertEqual(
            executor._join_path_retry_hint(
                "children.wrong_code = parents.code", sql,
            ),
            "",
        )
        self.assertEqual(
            executor._join_path_retry_hint(
                "", "SELECT p.id FROM parents p JOIN children c ON c.parent_id = p.id",
            ),
            "",
        )

    def test_relational_ir_requires_unique_physical_relation_path(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Player": dc.DBTable(name="Player", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("player_api_id", "INTEGER"),
                dc.DBColumn("player_fifa_api_id", "INTEGER"),
                dc.DBColumn("birthday", "TEXT"),
            ]),
            "Player_Attributes": dc.DBTable(name="Player_Attributes", columns=[
                dc.DBColumn(
                    "player_api_id", "INTEGER", fk_table="Player",
                    fk_column="player_api_id",
                ),
                dc.DBColumn("player_fifa_api_id", "INTEGER"),
                dc.DBColumn("preferred_foot", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "Calculate the percentage of players who prefer left foot, who were "
            "born between 1987 and 1992.\n"
            "Relevant business evidence supplied by the user: "
            "players who prefer left foot refers to preferred_foot = 'left'; "
            "percentage of players who prefer left foot refers to "
            "SUM(preferred_foot = 'left') * 100 / COUNT(player_fifa_api_id); "
            "born between 1987 and 1992 refers to birthday"
        )
        contract = executor._compile_relational_contract(question)

        self.assertEqual(contract.relation_paths[0]["edges"], [{
            "from": "Player_Attributes.player_api_id",
            "to": "Player.player_api_id",
            "source": "foreign_key",
        }])
        wrong = (
            "SELECT 100.0 * COUNT(CASE WHEN pa.preferred_foot='left' THEN p.id END) "
            "/ COUNT(p.id) FROM Player p JOIN Player_Attributes pa "
            "ON pa.player_fifa_api_id=p.player_fifa_api_id "
            "WHERE p.birthday BETWEEN '1987' AND '1992'"
        )
        self.assertIn(
            "唯一已声明外键路径",
            executor._relational_algebra_retry_hint(question, wrong, contract),
        )
        correct = wrong.replace(
            "pa.player_fifa_api_id=p.player_fifa_api_id",
            "pa.player_api_id=p.player_api_id",
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(question, correct, contract), "",
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT COUNT(CASE WHEN Player_Attributes.preferred_foot='left' "
                "THEN Player.id END) * 100.0 / COUNT(Player.id) "
                "FROM Player JOIN Player_Attributes ON "
                "Player.player_api_id=Player_Attributes.player_api_id",
                contract,
            ),
            "",
        )

    def test_relational_ir_compiles_path_from_exact_question_table_mentions(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "parents": dc.DBTable(name="parents", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("name", "TEXT"),
            ]),
            "children": dc.DBTable(name="children", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn(
                    "parent_id", "INTEGER", fk_table="parents", fk_column="id",
                ),
            ]),
            "show": dc.DBTable(name="show", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "Show parent names for children registered this year."
        contract = executor._compile_relational_contract(question)

        self.assertEqual(contract.relation_paths, [{
            "tables": ["parents", "children"],
            "edges": [{
                "from": "children.parent_id",
                "to": "parents.id",
                "source": "foreign_key",
            }],
            "source": "question_named_unique_shortest_declared_fk_path",
            "enforcement": "when_both_tables_referenced",
        }])
        self.assertNotIn("show", contract.relation_paths[0]["tables"])
        self.assertIn(
            "唯一已声明外键路径",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT p.name FROM parents p JOIN children c ON c.id=p.id",
                contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question, "SELECT name FROM parents", contract,
            ),
            "",
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT name FROM parents WHERE id IN "
                "(SELECT parent_id FROM children)",
                contract,
            ),
            "",
        )

    def test_relational_ir_compiles_counted_superlative_phrase(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "conductor": dc.DBTable(name="conductor", columns=[
                dc.DBColumn("Conductor_ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
            ]),
            "orchestra": dc.DBTable(name="orchestra", columns=[
                dc.DBColumn("Orchestra_ID", "INTEGER", pk=True),
                dc.DBColumn(
                    "Conductor_ID", "INTEGER", fk_table="conductor",
                    fk_column="Conductor_ID",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "Show the name of the conductor that has conducted the most number "
            "of orchestras."
        )
        contract = executor._compile_relational_contract(question)

        self.assertEqual(
            [stage["kind"] for stage in contract.aggregation_stages],
            ["group_aggregate", "rank"],
        )
        self.assertEqual(contract.aggregation_stages[0]["group_keys"], [
            "conductor.Conductor_ID",
        ])
        self.assertEqual(contract.aggregation_stages[1]["direction"], "DESC")
        self.assertEqual(contract.output_bindings, [{
            "table": "conductor", "column": "Name",
        }])
        self.assertEqual(contract.relation_paths[0]["edges"], [{
            "from": "orchestra.Conductor_ID",
            "to": "conductor.Conductor_ID",
            "source": "foreign_key",
        }])
        self.assertIn(
            "按 DESC 排名",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT c.Name FROM conductor c JOIN orchestra o "
                "ON o.Conductor_ID=c.Conductor_ID GROUP BY c.Conductor_ID "
                "ORDER BY COUNT(*) ASC LIMIT 1",
                contract,
            ),
        )
        ties_question = (
            "Show all conductors with the most number of orchestras, including "
            "all ties."
        )
        ties_contract = executor._compile_relational_contract(ties_question)
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                ties_question,
                "WITH counts AS (SELECT c.Conductor_ID, COUNT(*) AS cnt "
                "FROM conductor c JOIN orchestra o ON "
                "o.Conductor_ID=c.Conductor_ID GROUP BY c.Conductor_ID) "
                "SELECT Conductor_ID FROM counts WHERE cnt="
                "(SELECT MAX(cnt) FROM counts) ORDER BY Conductor_ID",
                ties_contract,
            ),
            "",
        )
        self.assertIn(
            "\u6309 DESC \u6392\u540d",
            executor._relational_algebra_retry_hint(
                ties_question,
                "WITH counts AS (SELECT c.Conductor_ID, COUNT(*) AS cnt, "
                "MAX(o.Orchestra_ID) AS max_id FROM conductor c JOIN orchestra o "
                "ON o.Conductor_ID=c.Conductor_ID GROUP BY c.Conductor_ID) "
                "SELECT Conductor_ID FROM counts WHERE max_id="
                "(SELECT MAX(max_id) FROM counts)",
                ties_contract,
            ),
        )

    def test_relational_ir_uses_explicit_superlative_attribute_grain(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Templates": dc.DBTable(name="Templates", columns=[
                dc.DBColumn("Template_ID", "INTEGER", pk=True),
                dc.DBColumn("Template_Type_Code", "TEXT"),
            ]),
            "Documents": dc.DBTable(name="Documents", columns=[
                dc.DBColumn("Document_ID", "INTEGER", pk=True),
                dc.DBColumn(
                    "Template_ID", "INTEGER", fk_table="Templates",
                    fk_column="Template_ID",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        contract = executor._compile_relational_contract(
            "Which template type code is used by most number of documents?"
        )

        self.assertEqual(contract.aggregation_stages[0]["group_keys"], [
            "Templates.Template_Type_Code",
        ])
        self.assertEqual(contract.output_bindings, [{
            "table": "Templates", "column": "Template_Type_Code",
        }])
        security = mock.Mock()
        security.connector = dc.DBConnector("fixture")
        native_executor = dc.NL2SQLExecutor(security, schema)
        self.assertIsNotNone(native_executor._compile_native_relational_plan(
            "Which template type code is used by most number of documents?",
            contract,
        ))

    def test_result_grain_qualifies_generic_outputs_and_anti_relation_owner(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "visitor": dc.DBTable(name="visitor", columns=[
                dc.DBColumn("ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
            ]),
            "museum": dc.DBTable(name="museum", columns=[
                dc.DBColumn("Museum_ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
            ]),
            "visit": dc.DBTable(name="visit", columns=[
                dc.DBColumn(
                    "visitor_ID", "INTEGER", fk_table="visitor", fk_column="ID",
                ),
                dc.DBColumn(
                    "Museum_ID", "INTEGER", fk_table="museum",
                    fk_column="Museum_ID",
                ),
            ]),
            "conductor": dc.DBTable(name="conductor", columns=[
                dc.DBColumn("Conductor_ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
            ]),
            "orchestra": dc.DBTable(name="orchestra", columns=[
                dc.DBColumn("Orchestra_ID", "INTEGER", pk=True),
                dc.DBColumn("Orchestra", "TEXT"),
                dc.DBColumn(
                    "Conductor_ID", "INTEGER", fk_table="conductor",
                    fk_column="Conductor_ID",
                ),
            ]),
            "performance": dc.DBTable(name="performance", columns=[
                dc.DBColumn("Performance_ID", "INTEGER", pk=True),
                dc.DBColumn(
                    "Orchestra_ID", "INTEGER", fk_table="orchestra",
                    fk_column="Orchestra_ID",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        museum_question = "What are the id and name of the museum visited most times?"
        museum_contract = executor._compile_relational_contract(museum_question)
        self.assertEqual(museum_contract.output_bindings, [
            {"table": "museum", "column": "Museum_ID"},
            {"table": "museum", "column": "Name"},
        ])
        self.assertEqual(museum_contract.result_grain, {
            "kind": "entity",
            "owner_table": "museum",
            "identity_columns": ["museum.Museum_ID"],
            "visible_columns": ["museum.Museum_ID", "museum.Name"],
            "cardinality": "single_row",
            "multiplicity": "one_row_per_entity",
            "source": "relationship_superlative_entity_identity",
        })
        self.assertEqual(museum_contract.aggregate_subjects, [{
            "role": "ranking_measure",
            "function": "COUNT",
            "source_table": "visit",
            "column": "*",
            "multiplicity": "fact_rows",
            "group_grain": ["museum.Museum_ID"],
            "source": "relationship_superlative_fact_rows",
        }])

        anti_question = "List the names of orchestras that have no performance."
        anti_contract = executor._compile_relational_contract(anti_question)
        self.assertEqual(anti_contract.output_bindings, [{
            "table": "orchestra", "column": "Orchestra",
        }])
        self.assertEqual(anti_contract.correlation_requirements[0]["outer_table"],
                         "orchestra")

    def test_entity_identity_not_visible_name_owns_counted_aggregation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "documents.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE Documents ("
                    "Document_ID INTEGER PRIMARY KEY, Document_Name TEXT);"
                    "CREATE TABLE Paragraphs ("
                    "Paragraph_ID INTEGER PRIMARY KEY, Document_ID INTEGER, "
                    "FOREIGN KEY (Document_ID) REFERENCES Documents(Document_ID));"
                    "INSERT INTO Documents VALUES "
                    "(1, 'Same'), (2, 'Same'), (3, 'Other');"
                    "INSERT INTO Paragraphs VALUES "
                    "(10, 1), (20, 2), (21, 2), (22, 2), (30, 3), (31, 3);"
                )
                conn.commit()
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            question = (
                "Return the id and name of the document with the most paragraphs."
            )
            contract = executor._compile_relational_contract(question)
            plan = executor._compile_native_relational_plan(question, contract)
            with mock.patch.object(
                dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
            ) as ask:
                answer = executor.answer(question)

        self.assertEqual(contract.output_bindings, [
            {"table": "Documents", "column": "Document_ID"},
            {"table": "Documents", "column": "Document_Name"},
        ])
        self.assertEqual(contract.grouping_keys, ["Documents.Document_ID"])
        self.assertEqual(contract.result_grain["identity_columns"], [
            "Documents.Document_ID",
        ])
        self.assertEqual(contract.result_grain["visible_columns"], [
            "Documents.Document_ID", "Documents.Document_Name",
        ])
        self.assertEqual(contract.aggregate_subjects[0]["multiplicity"], "fact_rows")
        self.assertIsNotNone(plan)
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [[2, "Same"]])
        self.assertIn('GROUP BY t0."Document_ID"', answer.sql)
        self.assertEqual(ask.call_count, 0)
        self.assertIn(
            "实体主键",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT d.Document_ID, d.Document_Name FROM Documents d "
                "JOIN Paragraphs p ON p.Document_ID=d.Document_ID "
                "GROUP BY d.Document_Name ORDER BY COUNT(*) DESC LIMIT 1",
                contract,
            ),
        )

    def test_registered_relationship_keeps_semester_identity_and_all_outputs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "semesters.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE Semesters ("
                    "semester_id INTEGER PRIMARY KEY, semester_name TEXT);"
                    "CREATE TABLE Students (student_id INTEGER PRIMARY KEY);"
                    "CREATE TABLE Student_Enrolment ("
                    "student_enrolment_id INTEGER PRIMARY KEY, "
                    "semester_id INTEGER, student_id INTEGER, "
                    "FOREIGN KEY (semester_id) REFERENCES Semesters(semester_id), "
                    "FOREIGN KEY (student_id) REFERENCES Students(student_id));"
                    "INSERT INTO Semesters VALUES (1, 'Fall'), (2, 'Spring');"
                    "INSERT INTO Students VALUES (10), (11), (12);"
                    "INSERT INTO Student_Enrolment VALUES "
                    "(100, 1, 10), (101, 2, 11), (102, 2, 12);"
                )
                conn.commit()
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            question = (
                "For each semester, what is the name and id of the one with "
                "the most students registered?"
            )
            contract = executor._compile_relational_contract(question)
            plan = executor._compile_native_relational_plan(question, contract)
            with mock.patch.object(
                dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
            ) as ask:
                answer = executor.answer(question)

        self.assertEqual(contract.output_columns, ["semester_name", "semester_id"])
        self.assertEqual(contract.grouping_keys, ["Semesters.semester_id"])
        self.assertEqual(contract.result_grain["owner_table"], "Semesters")
        self.assertIsNotNone(plan)
        self.assertEqual(answer.rows, [["Spring", 2]])
        self.assertIn('GROUP BY t0."semester_id"', answer.sql)
        self.assertEqual(ask.call_count, 0)

    def test_cte_projection_validation_uses_only_outer_select(self):
        schema = dc.SchemaSnapshot(db_path="x", generated_at=0, tables={
            "airports": dc.DBTable(name="airports", columns=[
                dc.DBColumn("AirportCode", "TEXT", pk=True),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        contract = dc.RelationalAlgebraContract(
            required_operators=["exact_output_projection"],
            output_columns=["AirportCode"],
            output_bindings=[{"table": "airports", "column": "AirportCode"}],
        )
        sql = (
            "WITH airport_flights AS ("
            "SELECT SourceAirport AS AirportCode FROM flights "
            "UNION ALL SELECT DestAirport FROM flights) "
            "SELECT AirportCode FROM airport_flights "
            "GROUP BY AirportCode ORDER BY COUNT(*) DESC LIMIT 1"
        )

        self.assertEqual(executor._top_level_projection(sql), "AirportCode")
        self.assertEqual(executor._simple_projection_columns(sql), ["airportcode"])
        self.assertEqual(
            executor._relational_algebra_retry_hint("question", sql, contract), "",
        )

    def test_output_contract_merges_explicit_followup_sentence(self):
        schema = dc.SchemaSnapshot(db_path="fixture", generated_at=0, tables={
            "schools": dc.DBTable(name="schools", columns=[
                dc.DBColumn("School", "TEXT"),
                dc.DBColumn("Street", "TEXT"),
                dc.DBColumn("County", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "List the names of schools with more than 30 difference in "
            "enrollements between K-12 and ages 5-17? Please also give the "
            "full street adress of the schools."
        )
        contract = executor._compile_relational_contract(question)

        self.assertEqual(
            executor._output_request_phrase(question),
            "schools names the full street adress",
        )
        self.assertEqual(contract.output_bindings, [
            {"table": "schools", "column": "School"},
            {"table": "schools", "column": "Street"},
        ])
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT s.School, s.Street FROM schools s WHERE 1=1",
                contract,
            ),
            "",
        )

    def test_native_relational_plan_executes_without_model_sql_generation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "orchestra.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE conductor ("
                    "Conductor_ID INTEGER PRIMARY KEY, Name TEXT);"
                    "CREATE TABLE orchestra ("
                    "Orchestra_ID INTEGER PRIMARY KEY, Conductor_ID INTEGER, "
                    "FOREIGN KEY (Conductor_ID) REFERENCES conductor(Conductor_ID));"
                    "INSERT INTO conductor VALUES (1, 'Ada'), (2, 'Bob');"
                    "INSERT INTO orchestra VALUES (10, 1), (11, 2), (12, 2);"
                )
                conn.commit()
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            question = (
                "Show the name of the conductor that has conducted the most "
                "number of orchestras."
            )

            with mock.patch.object(
                dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
            ) as ask:
                answer = executor.answer(question)

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [["Bob"]])
        self.assertEqual(ask.call_count, 0)
        self.assertIsNotNone(answer.relational_plan)
        self.assertEqual(answer.relational_plan["ranking"], {
            "direction": "DESC", "tie_policy": "single_row", "limit": 1,
        })
        self.assertIn('JOIN "orchestra"', answer.sql)
        self.assertEqual(answer.steps[0]["tool"], "native_relational_planner")
        self.assertEqual(answer.steps[0]["model_calls"], 0)

    def test_native_grouped_relationship_count_returns_label_count_and_zero(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "teacher_courses.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE teacher ("
                    "Teacher_ID INTEGER PRIMARY KEY, Name TEXT);"
                    "CREATE TABLE course ("
                    "Course_ID INTEGER PRIMARY KEY, Course TEXT);"
                    "CREATE TABLE course_arrange ("
                    "Course_ID INTEGER PRIMARY KEY REFERENCES course(Course_ID), "
                    "Teacher_ID INTEGER REFERENCES teacher(Teacher_ID));"
                    "INSERT INTO teacher VALUES (1, 'Ada'), (2, 'Bob'), (3, 'Cid');"
                    "INSERT INTO course VALUES (10, 'Math'), (11, 'SQL'), (12, 'Art');"
                    "INSERT INTO course_arrange VALUES (10, 1), (11, 1), (12, 2);"
                )
                conn.commit()
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            question = (
                "What are the names of the teachers and how many courses "
                "do they teach?"
            )
            contract = executor._compile_relational_contract(question)
            with mock.patch.object(
                dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
            ) as ask:
                answer = executor.answer(question)

        self.assertEqual(contract.version, "1.10")
        self.assertEqual(contract.output_bindings, [{
            "table": "teacher", "column": "Name",
        }])
        self.assertEqual(contract.output_layout, [
            {"kind": "column", "table": "teacher", "column": "Name"},
            {
                "kind": "aggregate", "function": "COUNT",
                "source_table": "course", "column": "*",
                "alias": "course_count",
            },
        ])
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.columns[-1], "course_count")
        self.assertEqual(answer.rows, [["Ada", 2], ["Bob", 1], ["Cid", 0]])
        self.assertEqual(answer.relational_plan["kind"], "grouped_aggregate")
        self.assertTrue(answer.relational_plan["include_zero"])
        self.assertIn('LEFT JOIN "course_arrange"', answer.sql)
        self.assertIn('LEFT JOIN "course"', answer.sql)
        self.assertIn('COUNT(t2."Course_ID")', answer.sql)
        self.assertEqual(ask.call_count, 0)

    def test_question_order_and_relationship_threshold_survive_local_repair(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "transcripts.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE Transcripts ("
                    "transcript_id INTEGER PRIMARY KEY, transcript_date TEXT);"
                    "CREATE TABLE Transcript_Contents ("
                    "content_id INTEGER PRIMARY KEY, transcript_id INTEGER "
                    "REFERENCES Transcripts(transcript_id));"
                    "INSERT INTO Transcripts VALUES (1, '2026-01-01'), "
                    "(2, '2026-02-01');"
                    "INSERT INTO Transcript_Contents VALUES "
                    "(10, 1), (11, 1), (12, 2);"
                )
                conn.commit()
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            question = (
                "Show the date and id of the transcript with at least 2 "
                "course results."
            )
            contract = executor._compile_relational_contract(question)
            reversed_sql = (
                "SELECT t.transcript_id, t.transcript_date FROM Transcripts t "
                "WHERE t.transcript_id IN (SELECT tc.transcript_id FROM "
                "Transcript_Contents tc GROUP BY tc.transcript_id "
                "HAVING COUNT(*) >= 2)"
            )
            ordered_sql = reversed_sql.replace(
                "t.transcript_id, t.transcript_date",
                "t.transcript_date, t.transcript_id",
            )
            self.assertIn(
                "重排了输出表达式",
                executor._relational_algebra_retry_hint(
                    question, reversed_sql, contract,
                ),
            )
            self.assertEqual(
                executor._relational_algebra_retry_hint(
                    question, ordered_sql, contract,
                ),
                "",
            )
            payload = {
                "intent": {
                    "outputs": ["transcript id", "transcript date"],
                    "row_grain": "detail_rows",
                    "filters": ["at least 2 course results"],
                    "grouping": [],
                    "relations": ["transcript contents"],
                    "ordering_limit": "",
                },
                "sql": reversed_sql,
                "summary_zh": "",
            }
            with mock.patch.object(dc, "_llm_ask_json", return_value=payload) as ask:
                answer = executor.answer(question)

        self.assertEqual(contract.output_columns, [
            "transcript_date", "transcript_id",
        ])
        self.assertEqual(contract.relationship_thresholds[0]["operator"], ">=")
        self.assertEqual(contract.relationship_thresholds[0]["value"], 2)
        self.assertEqual(
            contract.relationship_thresholds[0]["fact_table"],
            "Transcript_Contents",
        )
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.columns, ["transcript_date", "transcript_id"])
        self.assertEqual(answer.rows, [["2026-01-01", 1]])
        self.assertEqual(
            executor.last_candidate_search["status"],
            "local_projection_compiled",
        )
        self.assertEqual(ask.call_count, 1)

    def test_question_order_overrides_schema_order_for_three_outputs(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "battle": dc.DBTable(name="battle", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("name", "TEXT"),
                dc.DBColumn("bulgarian_commander", "TEXT"),
                dc.DBColumn("result", "TEXT"),
            ]),
            "ship": dc.DBTable(name="ship", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn(
                    "lost_in_battle", "INTEGER", fk_table="battle", fk_column="id",
                ),
                dc.DBColumn("location", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "Show names, results and bulgarian commanders of the battles "
            "with no ships lost in the 'English Channel'."
        )
        contract = executor._compile_relational_contract(question)

        self.assertEqual(contract.output_columns, [
            "name", "result", "bulgarian_commander",
        ])
        self.assertEqual(contract.output_bindings, [
            {"table": "battle", "column": "name"},
            {"table": "battle", "column": "result"},
            {"table": "battle", "column": "bulgarian_commander"},
        ])
        self.assertIn("question_ordered_output_projection", contract.evidence)
        correct = (
            "SELECT b.name, b.result, b.bulgarian_commander FROM battle b "
            "WHERE NOT EXISTS (SELECT 1 FROM ship s WHERE "
            "s.lost_in_battle = b.id AND s.location = 'English Channel')"
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(question, correct, contract),
            "",
        )

    def test_native_relational_plan_represents_typed_filter_and_zero_fact_fewest(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "stadium": dc.DBTable(name="stadium", columns=[
                dc.DBColumn("Stadium_ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Capacity", "INTEGER"),
            ]),
            "concert": dc.DBTable(name="concert", columns=[
                dc.DBColumn("concert_ID", "INTEGER", pk=True),
                dc.DBColumn(
                    "Stadium_ID", "INTEGER", fk_table="stadium",
                    fk_column="Stadium_ID",
                ),
                dc.DBColumn("Year", "INTEGER"),
            ]),
            "conductor": dc.DBTable(name="conductor", columns=[
                dc.DBColumn("Conductor_ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
            ]),
            "orchestra": dc.DBTable(name="orchestra", columns=[
                dc.DBColumn("Orchestra_ID", "INTEGER", pk=True),
                dc.DBColumn(
                    "Conductor_ID", "INTEGER", fk_table="conductor",
                    fk_column="Conductor_ID",
                ),
            ]),
        })
        security = mock.Mock()
        security.connector = dc.DBConnector("fixture")
        executor = dc.NL2SQLExecutor(security, schema)
        base = (
            "Show the stadium name and capacity of the stadium with the most "
            "number of concerts."
        )
        filtered = base[:-1] + " in year 2014 or after."
        fewest = (
            "Show the name of the conductor that has conducted the fewest "
            "number of orchestras."
        )

        base_contract = executor._compile_relational_contract(base)
        filtered_contract = executor._compile_relational_contract(filtered)
        fewest_contract = executor._compile_relational_contract(fewest)

        self.assertIsNotNone(
            executor._compile_native_relational_plan(base, base_contract)
        )
        filtered_plan = executor._compile_native_relational_plan(
            filtered, filtered_contract,
        )
        fewest_plan = executor._compile_native_relational_plan(
            fewest, fewest_contract,
        )
        self.assertIsNotNone(filtered_plan)
        self.assertEqual(filtered_plan.filters[0].as_dict(), {
            "column": {"table": "concert", "column": "Year"},
            "operator": ">=", "value": 2014, "value_type": "number",
        })
        self.assertIsNotNone(fewest_plan)
        self.assertEqual(fewest_plan.ranking.direction, "ASC")
        self.assertEqual(fewest_plan.joins[0].join_type, "LEFT")
        self.assertEqual(fewest_plan.aggregate.column, "Conductor_ID")

    def test_native_relational_renderer_rejects_undeclared_join_edge(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "parents": dc.DBTable(name="parents", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("name", "TEXT"),
            ]),
            "children": dc.DBTable(name="children", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("parent_id", "INTEGER"),
            ]),
        })
        plan = dc.RelationalQueryPlan(
            sources=["parents", "children"],
            joins=[dc.RelationalJoinEdge(
                left=dc.RelationalColumnRef("children", "parent_id"),
                right=dc.RelationalColumnRef("parents", "id"),
            )],
            projections=[dc.RelationalColumnRef("parents", "name")],
            group_keys=[dc.RelationalColumnRef("parents", "id")],
            aggregate=dc.RelationalAggregate("COUNT", "children"),
            ranking=dc.RelationalRanking("DESC", "single_row", 1),
            contract_version="1.3",
        )

        with self.assertRaisesRegex(ValueError, "真实外键"):
            dc.SQLiteRelationalPlanRenderer(schema).render(plan)

    def test_native_relational_renderer_preserves_all_maximum_ties(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "ties.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE parents (id INTEGER PRIMARY KEY, name TEXT);"
                    "CREATE TABLE children (id INTEGER PRIMARY KEY, parent_id INTEGER, "
                    "FOREIGN KEY (parent_id) REFERENCES parents(id));"
                    "INSERT INTO parents VALUES (1, 'Ada'), (2, 'Bob'), (3, 'Cid');"
                    "INSERT INTO children VALUES (10, 1), (11, 1), (12, 2), (13, 2), (14, 3);"
                )
                conn.commit()
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            plan = dc.RelationalQueryPlan(
                sources=["parents", "children"],
                joins=[dc.RelationalJoinEdge(
                    left=dc.RelationalColumnRef("children", "parent_id"),
                    right=dc.RelationalColumnRef("parents", "id"),
                )],
                projections=[dc.RelationalColumnRef("parents", "name")],
                group_keys=[dc.RelationalColumnRef("parents", "id")],
                aggregate=dc.RelationalAggregate("COUNT", "children"),
                ranking=dc.RelationalRanking("DESC", "all_ties"),
                contract_version="1.3",
            )
            sql = dc.SQLiteRelationalPlanRenderer(schema).render(plan)
            result = dc.SQLSecurity(connector).execute(sql)

        self.assertIsNone(result.error)
        self.assertEqual(result.rows, [["Ada"], ["Bob"]])
        self.assertIn("SELECT MAX", sql)

    def test_relational_ir_binds_ratio_denominator_population(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "customers": dc.DBTable(name="customers", columns=[
                dc.DBColumn("CustomerID", "INTEGER", pk=True),
            ]),
            "yearmonth": dc.DBTable(name="yearmonth", columns=[
                dc.DBColumn(
                    "CustomerID", "INTEGER", pk=True, fk_table="customers",
                    fk_column="CustomerID",
                ),
                dc.DBColumn("Date", "TEXT", pk=True),
                dc.DBColumn("Consumption", "REAL"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "In February 2012, what percentage of customers consumed more than 528.3?\n"
            "Relevant business evidence supplied by the user: "
            "February 2012 refers to '201202' in yearmonth.Date"
        )
        contract = executor._compile_relational_contract(question)

        self.assertEqual(contract.ratio_requirements, [{
            "kind": "conditional_share",
            "population_tables": ["yearmonth"],
            "denominator_grain": "yearmonth.CustomerID",
            "base_filter_columns": ["yearmonth.Date"],
            "shared_scope": "same_relations_and_base_filters",
            "scale": 100,
        }])
        self.assertIn(
            "分母改变了总体关系范围",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT (SELECT COUNT(*) FROM yearmonth WHERE Date='201202' "
                "AND Consumption>528.3) * 100.0 / "
                "(SELECT COUNT(*) FROM customers)",
                contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT SUM(CASE WHEN Consumption>528.3 THEN 1 ELSE 0 END) * "
                "100.0 / COUNT(*) FROM yearmonth WHERE Date='201202'",
                contract,
            ),
            "",
        )
        self.assertIn(
            "没有继承总体的基础筛选列",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT (SELECT COUNT(*) FROM yearmonth WHERE Date='201202' "
                "AND Consumption>528.3) * 100.0 / "
                "(SELECT COUNT(*) FROM yearmonth)",
                contract,
            ),
        )

    def test_ratio_scope_uses_question_clauses_not_formula_function_names(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "comments": dc.DBTable(name="comments", columns=[
                dc.DBColumn("UserId", "INTEGER", fk_table="users", fk_column="Id"),
                dc.DBColumn("Score", "INTEGER"),
            ]),
            "users": dc.DBTable(name="users", columns=[
                dc.DBColumn("Id", "INTEGER", pk=True),
                dc.DBColumn("UpVotes", "INTEGER"),
            ]),
            "tags": dc.DBTable(name="tags", columns=[
                dc.DBColumn("Count", "INTEGER"),
            ]),
            "votes": dc.DBTable(name="votes", columns=[
                dc.DBColumn("Id", "INTEGER", pk=True),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "Among the comments with scores between 5 to 10, what is the "
            "percentage of the users with 0 up votes?\n"
            "Relevant business evidence supplied by the user: percentage = "
            "DIVIDE(COUNT(UserId where UpVotes = 0 and Score BETWEEN 5 and 10))"
            "*100, (COUNT(UserId where Score BETWEEN 5 and 10));"
        )

        contract = executor._compile_relational_contract(question)

        self.assertEqual(contract.ratio_requirements, [{
            "kind": "conditional_share",
            "population_tables": ["comments", "users"],
            "denominator_grain": "comments.UserId",
            "base_filter_columns": ["comments.Score"],
            "shared_scope": "same_relations_and_base_filters",
            "scale": 100,
        }])
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT COUNT(CASE WHEN u.UpVotes=0 THEN 1 END)*100.0/"
                "COUNT(c.UserId) FROM comments c JOIN users u "
                "ON c.UserId=u.Id WHERE c.Score BETWEEN 5 AND 10",
                contract,
            ),
            "",
        )

    def test_ratio_category_condition_is_not_a_shared_denominator_filter(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "loan": dc.DBTable(name="loan", columns=[
                dc.DBColumn("loan_id", "INTEGER", pk=True),
                dc.DBColumn("amount", "REAL"),
                dc.DBColumn("status", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "What is the percentage of loan amount that has been fully paid "
            "with no issue?\nRelevant business evidence supplied by the user: "
            "status = 'A' means fully paid; percentage = "
            "SUM(amount where status = 'A') / SUM(amount) * 100"
        )

        contract = executor._compile_relational_contract(question)

        self.assertEqual(
            contract.ratio_requirements[0]["base_filter_columns"], [],
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT SUM(CASE WHEN status='A' THEN amount ELSE 0 END)*100.0/"
                "SUM(amount) FROM loan",
                contract,
            ),
            "",
        )

    def test_relational_ir_orders_entity_aggregation_before_rank(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "museum": dc.DBTable(name="museum", columns=[
                dc.DBColumn("Museum_ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
            ]),
            "visit": dc.DBTable(name="visit", columns=[
                dc.DBColumn(
                    "Museum_ID", "INTEGER", fk_table="museum", fk_column="Museum_ID",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "What are the id and name of the museum visited most times?"
        contract = executor._compile_relational_contract(question)

        self.assertEqual(
            [stage["kind"] for stage in contract.aggregation_stages],
            ["group_aggregate", "rank"],
        )
        self.assertIn(
            "按 DESC 排名",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT m.Museum_ID, m.Name FROM museum m JOIN visit v "
                "ON v.Museum_ID=m.Museum_ID GROUP BY m.Museum_ID, m.Name "
                "ORDER BY COUNT(*) ASC LIMIT 1",
                contract,
            ),
        )

    def test_relational_ir_requires_correlated_anti_relationship_binding(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "student": dc.DBTable(name="student", columns=[
                dc.DBColumn("StuID", "INTEGER", pk=True),
                dc.DBColumn("Age", "INTEGER"),
            ]),
            "has_pet": dc.DBTable(name="has_pet", columns=[
                dc.DBColumn(
                    "StuID", "INTEGER", fk_table="student", fk_column="StuID",
                ),
                dc.DBColumn(
                    "PetID", "INTEGER", fk_table="pets", fk_column="PetID",
                ),
            ]),
            "pets": dc.DBTable(name="pets", columns=[
                dc.DBColumn("PetID", "INTEGER", pk=True),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "What is the average age for all students who do not own any pets ?"
        contract = executor._compile_relational_contract(question)

        self.assertEqual(len(contract.correlation_requirements), 1)
        self.assertIn(
            "没有证明反关联绑定",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT AVG(s.Age) FROM student s WHERE NOT EXISTS "
                "(SELECT 1 FROM has_pet hp)",
                contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT AVG(s.Age) FROM student s WHERE NOT EXISTS "
                "(SELECT 1 FROM has_pet hp WHERE hp.StuID=s.StuID)",
                contract,
            ),
            "",
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT AVG(Age) FROM student WHERE StuID NOT IN "
                "(SELECT StuID FROM has_pet)",
                contract,
            ),
            "",
        )

    def test_relational_contract_requires_all_ties_when_explicit(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "shop": dc.DBTable(name="shop", columns=[dc.DBColumn("name", "TEXT")]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "Show all shops tied for the most employees, including ties."
        contract = executor._compile_relational_contract(question)

        self.assertEqual(contract.tie_policy, "all_ties")
        self.assertIn(
            "并列",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT name FROM shop ORDER BY employee_count DESC LIMIT 1",
                contract,
            ),
        )

    def test_relational_contract_leaves_ambiguous_who_cardinality_undeclared(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "customers": dc.DBTable(name="customers", columns=[
                dc.DBColumn("CustomerID", "INTEGER", pk=True),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        contract = executor._compile_relational_contract(
            "Who had the least consumption in the segment?"
        )

        self.assertEqual(contract.tie_policy, "")

    def test_relational_contract_plural_which_preserves_extremum_ties(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "schools": dc.DBTable(name="schools", columns=[
                dc.DBColumn("School", "TEXT"),
                dc.DBColumn("Enrollment", "INTEGER"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        contract = executor._compile_relational_contract(
            "Which state special schools have the highest enrollment?"
        )

        self.assertEqual(contract.tie_policy, "all_ties")

    def test_relational_contract_requires_single_superlative_entity(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "museum": dc.DBTable(name="museum", columns=[
                dc.DBColumn("Museum_ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
            ]),
            "visit": dc.DBTable(name="visit", columns=[
                dc.DBColumn(
                    "Museum_ID", "INTEGER", fk_table="museum", fk_column="Museum_ID",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "What are the id and name of the museum visited most times?"
        contract = executor._compile_relational_contract(question)

        self.assertEqual(contract.tie_policy, "single_row")
        self.assertIn(
            "单数实体",
            executor._relational_algebra_retry_hint(
                question,
                "SELECT m.Museum_ID, m.Name FROM museum m JOIN visit v "
                "ON m.Museum_ID=v.Museum_ID GROUP BY m.Museum_ID, m.Name "
                "HAVING COUNT(*)=(SELECT MAX(n) FROM counts)",
                contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT m.Museum_ID, m.Name FROM museum m JOIN visit v "
                "ON m.Museum_ID=v.Museum_ID GROUP BY m.Museum_ID, m.Name "
                "ORDER BY COUNT(*) DESC, m.Museum_ID ASC LIMIT 1",
                contract,
            ),
            "",
        )

    def test_initial_singular_which_entity_requires_one_superlative_row(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "employee": dc.DBTable(name="employee", columns=[
                dc.DBColumn("Employee_ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
            ]),
            "evaluation": dc.DBTable(name="evaluation", columns=[
                dc.DBColumn(
                    "Employee_ID", "INTEGER", fk_table="employee",
                    fk_column="Employee_ID",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        contract = executor._compile_relational_contract(
            "Which employee received the most awards in evaluations? "
            "Give me the employee name."
        )

        self.assertEqual(contract.tie_policy, "single_row")
        self.assertEqual(contract.aggregation_stages[-1]["limit"], 1)

    def test_numeric_rounding_contract_rejects_text_formatting(self):
        hint = dc.NL2SQLExecutor._numeric_result_type_retry_hint(
            "Provide the percentage with five decimal places.",
            "SELECT printf('%.5f', 100.0 * SUM(flag) / COUNT(*)) FROM items",
        )

        self.assertIn("ROUND", hint)

    def test_numeric_value_domain_separates_business_type_from_storage_type(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Dogs": dc.DBTable(name="Dogs", columns=[
                dc.DBColumn(
                    "age", "VARCHAR(20)", sample_values=["2", "10", "bad"],
                ),
            ]),
            "measures": dc.DBTable(name="measures", columns=[
                dc.DBColumn("Score", "REAL", sample_values=["1.5", "2.5"]),
            ]),
            "Students": dc.DBTable(name="Students", columns=[
                dc.DBColumn("cell_mobile_number", "VARCHAR(40)"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        text_question = "Count the number of dogs of an age below the average."
        text_contract = executor._compile_relational_contract(text_question)

        self.assertEqual(text_contract.value_domain_requirements, [{
            "column": "Dogs.age",
            "semantic_type": "numeric",
            "physical_type": "VARCHAR(20)",
            "coercion": "controlled_numeric_parse",
            "invalid_value_policy": "exclude",
            "source": "numeric_operator_bound_to_physical_measure",
        }])
        unsafe = executor._relational_algebra_retry_hint(
            text_question,
            "SELECT COUNT(*) FROM Dogs WHERE CAST(age AS REAL) < "
            "(SELECT AVG(CAST(age AS REAL)) FROM Dogs)",
            text_contract,
        )
        guarded = executor._relational_algebra_retry_hint(
            text_question,
            "SELECT COUNT(*) FROM Dogs WHERE age GLOB '[0-9]*' "
            "AND CAST(age AS REAL) < (SELECT AVG(CAST(age AS REAL)) "
            "FROM Dogs WHERE age GLOB '[0-9]*')",
            text_contract,
        )
        self.assertIn("无效文本", unsafe)
        self.assertEqual(guarded, "")

        native_question = "What is the maximum score?"
        native_contract = executor._compile_relational_contract(native_question)
        self.assertEqual(
            native_contract.value_domain_requirements[0]["coercion"],
            "native_numeric",
        )
        self.assertIn(
            "已经是数值",
            executor._relational_algebra_retry_hint(
                native_question,
                "SELECT MAX(CAST(Score AS REAL)) FROM measures",
                native_contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                native_question, "SELECT MAX(Score) FROM measures", native_contract,
            ),
            "",
        )
        count_question = (
            "What is the maximum number of times that a course shows up in "
            "different transcripts?"
        )
        self.assertEqual(
            executor._compile_relational_contract(
                count_question
            ).value_domain_requirements,
            [],
        )

    def test_parallel_measure_contract_rejects_one_merged_count(self):
        hint = dc.NL2SQLExecutor._parallel_measures_retry_hint(
            "How many atoms with iodine and with sulfur type elements are there?",
            "SELECT COUNT(*) FROM atom WHERE element IN ('i', 's')",
        )

        self.assertIn("两个独立", hint)

    def test_row_grain_contract_rejects_collapsed_entity_lists(self):
        hint = dc.NL2SQLExecutor._row_grain_retry_hint(
            "List and group all patients by sex for abnormal bilirubin.",
            "SELECT sex, GROUP_CONCAT(id) FROM patient GROUP BY sex",
        )

        self.assertIn("一行一个实体", hint)

    def test_semantic_retry_hint_excludes_filter_and_order_columns_from_output_phrase(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "airlines": dc.DBTable(name="airlines", columns=[
                dc.DBColumn("Airline", "TEXT"),
                dc.DBColumn("Country", "TEXT"),
            ]),
            "employee": dc.DBTable(name="employee", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Age", "INTEGER"),
            ]),
            "airports": dc.DBTable(name="airports", columns=[
                dc.DBColumn("AirportCode", "TEXT"),
                dc.DBColumn("AirportName", "TEXT"),
                dc.DBColumn("City", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        airline_hint = executor._semantic_retry_hint(
            'Which country does Airline "JetBlue Airways" belong to?',
            "SELECT Country, Airline FROM airlines WHERE Airline='JetBlue Airways'",
        )
        employee_hint = executor._semantic_retry_hint(
            "Sort employee names by their age in ascending order.",
            "SELECT Name, Age FROM employee ORDER BY Age ASC",
        )
        airport_hint = executor._semantic_retry_hint(
            "Give the airport code and airport name corresonding to the city Anthony.",
            "SELECT AirportCode, AirportName, City FROM airports WHERE City='Anthony'",
        )

        self.assertIn("Airline", airline_hint)
        self.assertIn("Age", employee_hint)
        self.assertIn("City", airport_hint)

    def test_semantic_retry_hint_distinguishes_and_from_or_for_in(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "cartoon": dc.DBTable(name="cartoon", columns=[
                dc.DBColumn("Directed_by", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        sql = "SELECT Channel FROM cartoon WHERE Directed_by IN ('Ben Jones', 'Michael Chang')"

        self.assertIn(
            "裸 IN",
            executor._semantic_retry_hint(
                "Find channels playing cartoons directed by Ben Jones and Michael Chang", sql,
            ),
        )
        self.assertEqual(
            executor._semantic_retry_hint(
                "Find channels playing cartoons directed by Ben Jones or Michael Chang", sql,
            ),
            "",
        )

    def test_semantic_retry_hint_normalizes_iso_date_positions_and_literals(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "expense": dc.DBTable(name="expense", columns=[
                dc.DBColumn(
                    "expense_date", "TEXT", sample_values=["2019-09-01", "2019-10-03"],
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        month_hint = executor._semantic_retry_hint(
            "What is the average expense in September?",
            "SELECT AVG(amount) FROM expense e WHERE SUBSTR(e.expense_date, 5, 2)='09'",
        )
        literal_hint = executor._semantic_retry_hint(
            "Show expenses on 2019/9/1.",
            "SELECT amount FROM expense e WHERE e.expense_date='2019/9/1'",
        )

        self.assertIn("第 6-7 个字符", month_hint)
        self.assertIn("2019-09-01", literal_hint)

    def test_date_storage_retry_requires_iso_schema_evidence(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "expense": dc.DBTable(name="expense", columns=[
                dc.DBColumn("expense_date", "TEXT", sample_values=["09/01/2019"]),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        self.assertEqual(
            executor._semantic_retry_hint(
                "What is the average expense in September?",
                "SELECT AVG(amount) FROM expense WHERE SUBSTR(expense_date, 5, 2)='09'",
            ),
            "",
        )

    def test_semantic_retry_hint_matches_date_prefix_for_datetime_storage(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "event": dc.DBTable(name="event", columns=[
                dc.DBColumn(
                    "event_date",
                    "TEXT",
                    sample_values=["2020-03-10T12:00:00", "2019-10-05T09:30:00"],
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        hint = executor._semantic_retry_hint(
            "Show events on October 8, 2019.",
            "SELECT event_id FROM event e WHERE e.event_date = '2019-10-08'",
        )

        self.assertIn("LIKE '2019-10-08%'", hint)

    def test_date_prefix_retry_requires_datetime_schema_evidence(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "event": dc.DBTable(name="event", columns=[
                dc.DBColumn(
                    "event_date", "TEXT", sample_values=["2019-10-08", "2019-10-09"],
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        self.assertEqual(
            executor._semantic_retry_hint(
                "Show events on October 8, 2019.",
                "SELECT event_id FROM event WHERE event_date = '2019-10-08'",
            ),
            "",
        )

    def test_semantic_retry_hint_uses_exact_sampled_enum_case(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "gasstations": dc.DBTable(name="gasstations", columns=[
                dc.DBColumn(
                    "Segment",
                    "TEXT",
                    sample_values=["Premium", "Other", "Discount"],
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        hint = executor._semantic_retry_hint(
            'How many "discount" gas stations are there?',
            "SELECT COUNT(*) FROM gasstations g WHERE g.Segment = 'discount'",
        )

        self.assertIn("'Discount'", hint)
        self.assertIn("'discount'", hint)

    def test_enum_case_retry_requires_sampled_casefold_match(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "gasstations": dc.DBTable(name="gasstations", columns=[
                dc.DBColumn("Segment", "TEXT", sample_values=["Premium", "Other"]),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        self.assertEqual(
            executor._semantic_retry_hint(
                'How many "discount" gas stations are there?',
                "SELECT COUNT(*) FROM gasstations WHERE Segment = 'discount'",
            ),
            "",
        )

    def test_answer_locally_compiles_exact_sampled_enum_case(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "gasstations": dc.DBTable(name="gasstations", columns=[
                dc.DBColumn(
                    "Segment", "TEXT",
                    sample_values=["Premium", "Other", "Discount"],
                ),
            ]),
        })
        generated_sql = (
            "SELECT COUNT(*) FROM gasstations g WHERE g.Segment = 'discount'"
        )
        repaired_sql = generated_sql.replace("'discount'", "'Discount'")
        security = mock.Mock()
        security.execute.return_value = dc.SQLResult(
            sql=repaired_sql,
            columns=["COUNT(*)"],
            rows=[[2]],
            row_count=1,
        )
        executor = dc.NL2SQLExecutor(security, schema)

        with mock.patch.object(dc, "_llm_ask_json", return_value={
            "sql": generated_sql,
            "summary_zh": "",
        }) as ask:
            answer = executor.answer('How many "discount" gas stations are there?')

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.sql, repaired_sql)
        self.assertEqual(answer.rows, [[2]])
        self.assertEqual(ask.call_count, 1)
        security.execute.assert_called_once_with(repaired_sql)
        self.assertEqual(executor.semantic_repair_count, 0)
        self.assertEqual(
            executor.last_candidate_search["status"],
            "local_exact_enum_compiled",
        )
        self.assertEqual(executor.last_candidate_search["model_calls"], 0)

    def test_answer_locally_repairs_extra_projection_before_execution(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "teacher": dc.DBTable(name="teacher", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Age", "INTEGER"),
                dc.DBColumn("Hometown", "TEXT"),
            ]),
        })
        security = mock.Mock()
        repaired_sql = 'SELECT teacher."Age", teacher."Hometown" FROM teacher'
        security.execute.return_value = dc.SQLResult(
            sql=repaired_sql,
            columns=["Age", "Hometown"],
            rows=[],
            row_count=0,
        )
        executor = dc.NL2SQLExecutor(security, schema)
        intent = {
            "outputs": ["teacher age", "teacher hometown"],
            "row_grain": "one_row_per_entity",
            "filters": [],
            "grouping": [],
            "relations": [],
            "ordering_limit": "",
        }
        generated = {
            "sql": "SELECT Name, Age, Hometown FROM teacher",
            "summary_zh": "",
            "intent": intent,
        }

        with mock.patch.object(dc, "_llm_ask_json", return_value=generated) as ask:
            answer = executor.answer("What is the age and hometown of every teacher?")

        self.assertEqual(answer.kind, "query")
        self.assertEqual(ask.call_count, 1)
        security.execute.assert_called_once_with(repaired_sql)
        self.assertEqual(executor.semantic_repair_count, 0)
        self.assertEqual(executor.last_candidate_sql, repaired_sql)
        self.assertEqual(executor.last_semantic_hint, "")
        self.assertEqual(executor.last_query_intent.row_grain, "one_row_per_entity")
        self.assertEqual(answer.steps[0]["tool"], "query_intent_contract")
        self.assertEqual(answer.steps[1]["status"], "local_projection_compiled")
        self.assertEqual(answer.steps[1]["model_calls"], 0)

    def test_answer_injects_independent_contract_before_first_generation(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "countrylanguage": dc.DBTable(name="countrylanguage", columns=[
                dc.DBColumn("Language", "TEXT"),
                dc.DBColumn("IsOfficial", "TEXT"),
            ]),
        })
        sql = "SELECT Language FROM countrylanguage WHERE IsOfficial='T'"
        security = mock.Mock()
        security.execute.return_value = dc.SQLResult(
            sql=sql, columns=["Language"], rows=[["Dutch"]], row_count=1,
        )
        executor = dc.NL2SQLExecutor(security, schema)

        with mock.patch.object(dc, "_llm_ask_json", return_value={
            "sql": sql,
            "summary_zh": "官方语言",
        }) as ask:
            answer = executor.answer("What is the official language spoken in the country?")

        prompt = ask.call_args.args[0]
        self.assertEqual(answer.kind, "query")
        self.assertEqual(ask.call_count, 1)
        self.assertIn("本地独立关系代数合同", prompt)
        self.assertIn('"output_columns":["Language"]', prompt)
        self.assertEqual(answer.steps[-1]["tool"], "relational_algebra_contract")

    def test_relational_contract_repair_precedes_execution_of_wrong_sql(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "customers": dc.DBTable(name="customers", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("name", "TEXT"),
            ]),
            "orders": dc.DBTable(name="orders", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("customer_id", "INTEGER", fk_table="customers", fk_column="id"),
            ]),
        })
        first_sql = (
            "SELECT c.name, COUNT(*) FROM customers c JOIN orders o "
            "ON c.id=o.customer_id GROUP BY c.id, c.name"
        )
        revised_sql = (
            "SELECT c.name FROM customers c JOIN orders o ON c.id=o.customer_id "
            "GROUP BY c.id, c.name ORDER BY COUNT(*) DESC, c.id ASC LIMIT 1"
        )
        security = mock.Mock()
        security.validate.side_effect = lambda sql: sql
        security.execute.return_value = dc.SQLResult(
            sql=revised_sql, columns=["name"], rows=[["A"]], row_count=1,
        )
        executor = dc.NL2SQLExecutor(security, schema)
        generated = [
            {"sql": first_sql, "summary_zh": "候选"},
            {
                "intent": {
                    "outputs": ["customer name"],
                    "row_grain": "top_k",
                    "filters": [],
                    "grouping": ["customer"],
                    "relations": ["orders.customer_id = customers.id"],
                    "ordering_limit": "order by order count desc, limit 1",
                },
                "sql": revised_sql,
                "summary_zh": "订单最多的客户",
                "reason_code": "row_grain_and_projection",
            },
        ]

        with mock.patch.object(dc, "_llm_ask_json", side_effect=generated) as ask:
            answer = executor.answer("Which customer has the most orders?")

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [["A"]])
        self.assertEqual(executor.last_generated_sql, revised_sql)
        self.assertEqual(ask.call_count, 2)
        security.execute.assert_called_once_with(revised_sql)
        self.assertEqual(executor.last_candidate_search["status"], "primary_accepted")
        self.assertEqual(answer.steps[0]["intent"]["row_grain"], "top_k")
        self.assertEqual(answer.steps[1]["tool"], "bounded_candidate_search")

    def test_query_contract_review_failure_preserves_successful_readonly_candidate(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "orders": dc.DBTable(name="orders", columns=[
                dc.DBColumn("customer_id", "INTEGER"),
                dc.DBColumn("amount", "REAL"),
            ]),
        })
        sql = "SELECT customer_id, SUM(amount) FROM orders GROUP BY customer_id"
        security = mock.Mock()
        security.execute.return_value = dc.SQLResult(
            sql=sql, columns=["customer_id", "sum"], rows=[], row_count=0,
        )
        executor = dc.NL2SQLExecutor(security, schema)

        with mock.patch.object(dc, "_llm_ask_json", side_effect=[
            {"sql": sql, "summary_zh": "按客户汇总"},
            dc.LLMServiceError("LLM 服务请求失败（HTTP 503）"),
        ]):
            answer = executor.answer("Show totals by customer.")

        self.assertEqual(answer.kind, "query")
        self.assertEqual(executor.last_generated_sql, sql)
        self.assertEqual(security.execute.call_count, 1)
        self.assertEqual(answer.steps[0]["status"], "unavailable")

    def test_relation_gate_precedes_semantic_projection_retry(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "customers": dc.DBTable(name="customers", columns=[
                dc.DBColumn("Name", "TEXT"),
            ]),
            "orders": dc.DBTable(name="orders", columns=[
                dc.DBColumn("Status", "TEXT"),
            ]),
        })
        security = mock.Mock()
        executor = dc.NL2SQLExecutor(security, schema)
        generated = {
            "sql": "SELECT customers.Name, orders.Status FROM customers JOIN orders ON 1=1",
            "summary_zh": "",
        }

        with mock.patch.object(dc, "_llm_ask_json", return_value=generated) as ask:
            answer = executor.answer("Show customer names.")

        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "table_relationship")
        self.assertEqual(ask.call_count, 1)
        security.execute.assert_not_called()

    def test_missing_projection_is_locally_repaired_without_second_model_call(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "teacher": dc.DBTable(name="teacher", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Age", "INTEGER"),
            ]),
        })
        security = mock.Mock()
        repaired_sql = 'SELECT teacher."Name", teacher."Age" FROM teacher'
        security.execute.return_value = dc.SQLResult(
            sql=repaired_sql, columns=["Name", "Age"], rows=[["Ada", 36]],
            row_count=1,
        )
        executor = dc.NL2SQLExecutor(security, schema)
        generated = {
            "sql": "SELECT Name FROM teacher",
            "summary_zh": "",
        }

        with mock.patch.object(dc, "_llm_ask_json", return_value=generated) as ask:
            answer = executor.answer("Show teacher names and ages.")

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [["Ada", 36]])
        self.assertEqual(ask.call_count, 1)
        self.assertEqual(executor.semantic_repair_count, 0)
        self.assertEqual(executor.last_candidate_sql, repaired_sql)
        self.assertEqual(executor.last_semantic_hint, "")
        security.execute.assert_called_once_with(repaired_sql)
        self.assertEqual(executor.last_candidate_search["status"], "local_projection_compiled")

    def test_bounded_candidate_search_salvages_unique_safe_alternative(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "teacher": dc.DBTable(name="teacher", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Age", "INTEGER"),
            ]),
        })
        security = mock.Mock()

        def validate(sql):
            statement, code = dc._normalize_single_sql_statement(
                sql, dc.SQLSecurityError, "禁止多语句：只允许单条查询",
            )
            if not code.lstrip().upper().startswith(("SELECT", "WITH")):
                raise dc.SQLSecurityError("只允许 SELECT 查询（含 WITH CTE）")
            return statement

        security.validate.side_effect = validate
        security.execute.return_value = dc.SQLResult(
            sql="SELECT Name, Age FROM teacher",
            columns=["Name", "Age"],
            rows=[["Ada", 36]],
            row_count=1,
        )
        executor = dc.NL2SQLExecutor(security, schema)
        generated = {"candidates": [
                {
                    "candidate_id": "primary",
                    "strategy": "保留原投影",
                    "sql": "SELECT Name FROM teacher",
                    "summary_zh": "主修复",
                },
                {
                    "candidate_id": "unsafe",
                    "strategy": "写入语句",
                    "sql_tail": "FROM teacher; DELETE FROM teacher",
                    "summary_zh": "不应执行",
                },
                {
                    "candidate_id": "complete_projection",
                    "strategy": "补齐问句输出原子",
                    "intent": {
                        "outputs": ["teacher name", "teacher age"],
                        "row_grain": "detail_rows",
                    },
                    "sql_tail": "FROM teacher",
                    "summary_zh": "教师姓名和年龄",
                },
            ]}

        question = "Show teacher names and ages."
        executor.last_relational_contract = executor._compile_relational_contract(question)
        conflict = executor._projection_conflict(question, "SELECT Name FROM teacher")
        self.assertIsNotNone(conflict)
        with mock.patch.object(dc, "_llm_ask_json", return_value=generated) as ask:
            search = executor._search_semantic_repair_candidates(
                question=question,
                schema_txt="fixture schema",
                bad_sql="SELECT Name FROM teacher",
                semantic_conflict=conflict,
                history=None,
                allowed_tables=None,
            )

        self.assertEqual(ask.call_count, 1)
        self.assertEqual(
            search["selected"]["sql"],
            'SELECT teacher."Name", teacher."Age" FROM teacher',
        )
        self.assertEqual(
            executor.last_candidate_search["status"],
            "unique_alternative_accepted",
        )
        self.assertEqual(
            executor.last_candidate_search["selected_candidate_id"],
            "complete_projection",
        )
        self.assertEqual(
            [item["reason_code"] for item in executor.last_candidate_search["assessments"]],
            ["candidate_plan_compile", "sql_security", ""],
        )
        search_step = search["diagnostic"]
        self.assertEqual(search_step["selection_basis"], "unique_independently_validated_alternative")
        repair_prompt = ask.call_args.args[0]
        self.assertIn('"code":"projection"', repair_prompt)
        self.assertIn('"required_output_columns":["Name","Age"]', repair_prompt)
        self.assertIn('"missing_output_columns":["Age"]', repair_prompt)
        self.assertIn('"mode":"projection_locked_sql_tail"', repair_prompt)
        security.execute.assert_not_called()

    def test_projection_conflict_exposes_required_and_forbidden_columns(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "teacher": dc.DBTable(name="teacher", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Age", "INTEGER"),
                dc.DBColumn("Hometown", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        conflict = executor._projection_conflict(
            "Show teacher names and ages.",
            "SELECT Name, Age, Hometown FROM teacher",
        )

        self.assertIsNotNone(conflict)
        self.assertEqual(conflict.code, "projection")
        self.assertEqual(
            conflict.constraints["required_output_columns"], ["Name", "Age"],
        )
        self.assertEqual(
            conflict.constraints["forbidden_output_columns"], ["Hometown"],
        )

    def test_projection_lock_removes_forbidden_antibody_column_before_revalidation(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Patient": dc.DBTable(name="Patient", columns=[
                dc.DBColumn("ID", "INTEGER", pk=True),
                dc.DBColumn("Diagnosis", "TEXT"),
                dc.DBColumn("Description", "TEXT"),
            ]),
            "Examination": dc.DBTable(name="Examination", columns=[
                dc.DBColumn("ID", "INTEGER", fk_table="Patient", fk_column="ID"),
                dc.DBColumn(
                    "aCL IgA", "TEXT",
                    semantic_name="anti-Cardiolipin antibody (IgA)",
                    description="anti-Cardiolipin antibody (IgA) concentration",
                ),
                dc.DBColumn(
                    "aCL IgG", "TEXT",
                    semantic_name="anti-Cardiolipin antibody (IgG)",
                    description="anti-Cardiolipin antibody (IgG) concentration",
                ),
                dc.DBColumn(
                    "aCL IgM", "TEXT",
                    semantic_name="anti-Cardiolipin antibody (IgM)",
                    description="anti-Cardiolipin antibody (IgM) concentration",
                ),
                dc.DBColumn("ANA", "TEXT", description="antinuclear antibody concentration"),
                dc.DBColumn("Examination Date", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "For the patient with SLE, what was his/her anti-Cardiolipin antibody "
            "concentration status on 1993/11/12?"
        )
        bad_sql = (
            'SELECT e."aCL IgM", e."aCL IgG", e.ANA, e."aCL IgA" '
            "FROM Patient p JOIN Examination e ON p.ID=e.ID"
        )

        conflict = executor._projection_conflict(question, bad_sql)
        bindings = executor._projection_lock_bindings(conflict)
        compiled = executor._compile_projection_locked_sql(
            bindings,
            "FROM Patient p JOIN Examination e ON p.ID=e.ID "
            "WHERE p.Diagnosis='SLE' AND e.\"Examination Date\"='1993-11-12'",
        )

        self.assertEqual(executor._simple_projection_columns(compiled), [
            "acl iga", "acl igg", "acl igm",
        ])
        self.assertNotIn("ANA", compiled)
        self.assertIsNone(executor._semantic_conflict(
            question,
            compiled,
            locked_projection_columns=["aCL IgA", "aCL IgG", "aCL IgM"],
        ))

    def test_bounded_candidate_search_rejects_ambiguous_valid_alternatives(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "teacher": dc.DBTable(name="teacher", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Age", "INTEGER"),
            ]),
        })
        security = mock.Mock()
        security.validate.side_effect = lambda sql: sql
        executor = dc.NL2SQLExecutor(security, schema)
        generated = {"candidates": [
                {"candidate_id": "primary", "sql": "SELECT Name FROM teacher"},
                {
                    "candidate_id": "plain",
                    "sql_tail": "FROM teacher",
                },
                {
                    "candidate_id": "qualified",
                    "sql_tail": "FROM teacher WHERE Age IS NOT NULL",
                },
            ]}

        question = "Show teacher names and ages."
        executor.last_relational_contract = executor._compile_relational_contract(question)
        conflict = executor._projection_conflict(question, "SELECT Name FROM teacher")
        self.assertIsNotNone(conflict)
        with mock.patch.object(dc, "_llm_ask_json", return_value=generated) as ask:
            search = executor._search_semantic_repair_candidates(
                question=question,
                schema_txt="fixture schema",
                bad_sql="SELECT Name FROM teacher",
                semantic_conflict=conflict,
                history=None,
                allowed_tables=None,
            )

        self.assertIsNone(search["selected"])
        self.assertEqual(ask.call_count, 1)
        self.assertEqual(
            executor.last_candidate_search["status"], "ambiguous_alternatives",
        )
        self.assertEqual(executor.last_candidate_search["eligible_count"], 2)
        self.assertEqual(
            executor.last_candidate_search["selection_basis"],
            "fail_closed_without_independent_tie_breaker",
        )
        security.execute.assert_not_called()

    def test_bounded_candidate_search_applies_relation_gate_to_each_alternative(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "teacher": dc.DBTable(name="teacher", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Age", "INTEGER"),
            ]),
            "department": dc.DBTable(name="department", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
            ]),
        })
        security = mock.Mock()
        security.validate.side_effect = lambda sql: sql
        security.execute.return_value = dc.SQLResult(
            sql="SELECT Name, Age FROM teacher",
            columns=["Name", "Age"],
            rows=[["Ada", 36]],
            row_count=1,
        )
        executor = dc.NL2SQLExecutor(security, schema)
        generated = {"candidates": [
                {"candidate_id": "primary", "sql": "SELECT Name FROM teacher"},
                {
                    "candidate_id": "invented_relation",
                    "sql_tail": "FROM teacher JOIN department ON teacher.Age=department.id",
                },
                {
                    "candidate_id": "single_table",
                    "sql_tail": "FROM teacher",
                },
            ]}

        question = "Show teacher names and ages."
        executor.last_relational_contract = executor._compile_relational_contract(question)
        conflict = executor._projection_conflict(question, "SELECT Name FROM teacher")
        self.assertIsNotNone(conflict)
        with mock.patch.object(dc, "_llm_ask_json", return_value=generated):
            search = executor._search_semantic_repair_candidates(
                question=question,
                schema_txt="fixture schema",
                bad_sql="SELECT Name FROM teacher",
                semantic_conflict=conflict,
                history=None,
                allowed_tables=None,
            )

        reasons = {
            item["candidate_id"]: item["reason_code"]
            for item in executor.last_candidate_search["assessments"]
        }
        self.assertEqual(reasons["invented_relation"], "table_relationship")
        self.assertEqual(
            executor.last_candidate_search["selected_candidate_id"], "single_table",
        )
        self.assertEqual(search["selected"]["candidate_id"], "single_table")
        security.execute.assert_not_called()

    def test_bounded_candidate_search_enforces_independent_branch_scope(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "teacher": dc.DBTable(name="teacher", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Age", "INTEGER"),
            ]),
            "department": dc.DBTable(name="department", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
            ]),
        })
        security = mock.Mock()
        security.validate.side_effect = lambda sql: sql
        security.execute.return_value = dc.SQLResult(
            sql="SELECT Name, Age FROM teacher",
            columns=["Name", "Age"],
            rows=[["Ada", 36]],
            row_count=1,
        )
        executor = dc.NL2SQLExecutor(security, schema)
        generated = {"candidates": [
                {"candidate_id": "primary", "sql": "SELECT Name FROM teacher"},
                {
                    "candidate_id": "wrong_branch",
                    "sql_tail": "FROM teacher JOIN department ON teacher.Age=department.id",
                },
                {"candidate_id": "right_branch", "sql_tail": "FROM teacher"},
            ]}

        question = "Show teacher names and ages."
        executor.last_relational_contract = executor._compile_relational_contract(question)
        conflict = executor._projection_conflict(question, "SELECT Name FROM teacher")
        self.assertIsNotNone(conflict)
        with mock.patch.object(dc, "_llm_ask_json", return_value=generated):
            search = executor._search_semantic_repair_candidates(
                question=question,
                schema_txt="fixture schema",
                bad_sql="SELECT Name FROM teacher",
                semantic_conflict=conflict,
                history=None,
                allowed_tables=["teacher"],
            )

        reasons = {
            item["candidate_id"]: item["reason_code"]
            for item in executor.last_candidate_search["assessments"]
        }
        self.assertEqual(reasons["wrong_branch"], "branch_scope")
        self.assertEqual(
            search["selected"]["sql"],
            'SELECT teacher."Name", teacher."Age" FROM teacher',
        )
        security.execute.assert_not_called()

    def test_scalar_attribute_threshold_is_not_rewritten_as_relationship_count(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "shop": dc.DBTable(name="shop", columns=[
                dc.DBColumn("Shop_ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("District", "TEXT"),
                dc.DBColumn("Number_products", "INTEGER"),
            ]),
            "hiring": dc.DBTable(name="hiring", columns=[
                dc.DBColumn(
                    "Shop_ID", "INTEGER", fk_table="shop", fk_column="Shop_ID",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "List shop IDs with fewer than 3000 products."

        contract = executor._compile_relational_contract(question)

        self.assertEqual(contract.relationship_thresholds, [])
        self.assertEqual(contract.filter_requirements, [{
            "column": "shop.Number_products",
            "operator": "<",
            "value": 3000,
            "value_type": "number",
            "scope": "row_predicate",
        }])
        self.assertNotIn("having", contract.required_operators)
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT Shop_ID FROM shop WHERE Number_products < 3000",
                contract,
            ),
            "",
        )
        both_question = (
            "Which district has both stores with less than 3000 products and "
            "stores with more than 10000 products?"
        )
        both_contract = executor._compile_relational_contract(both_question)
        self.assertEqual(both_contract.relationship_thresholds, [])
        self.assertEqual(both_contract.relation_paths, [])
        self.assertNotIn("having", both_contract.required_operators)
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                both_question,
                "SELECT District FROM shop WHERE Number_products < 3000 "
                "INTERSECT SELECT District FROM shop WHERE Number_products > 10000",
                both_contract,
            ),
            "",
        )

    def test_output_aliases_and_each_detail_cardinality_are_schema_bound(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "singer": dc.DBTable(name="singer", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Country", "TEXT"),
            ]),
            "student": dc.DBTable(name="student", columns=[
                dc.DBColumn("Major", "TEXT"),
                dc.DBColumn("Age", "INTEGER"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        singer = executor._projection_conflict(
            "What are the names and nations of all singers?",
            "SELECT Name FROM singer",
        )
        student = executor._projection_conflict(
            "List the major of each student, and also how old are they?",
            "SELECT Major FROM student",
        )

        self.assertEqual(singer.constraints["required_output_columns"], ["Name", "Country"])
        self.assertEqual(singer.constraints["required_output_bindings"], [
            {"table": "singer", "column": "Name"},
            {"table": "singer", "column": "Country"},
        ])
        self.assertEqual(student.constraints["required_output_columns"], ["Major", "Age"])
        student_contract = executor._compile_relational_contract(
            "List the major of each student, and also how old are they?",
        )
        self.assertNotEqual(student_contract.tie_policy, "single_row")
        self.assertIn(
            "一行一条关联明细",
            executor._row_grain_retry_hint(
                "List the first name of each patient and the description of "
                "every treatment they received.",
                "SELECT first_name, description FROM treatment "
                "ORDER BY treatment_date DESC LIMIT 1",
            ),
        )

    def test_relation_grounding_requires_physical_table_phrase(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Courses": dc.DBTable(name="Courses", columns=[
                dc.DBColumn("Course_ID", "INTEGER", pk=True),
            ]),
            "Student_Course_Registrations": dc.DBTable(
                name="Student_Course_Registrations", columns=[
                    dc.DBColumn("Student_ID", "INTEGER"),
                    dc.DBColumn(
                        "Course_ID", "INTEGER", fk_table="Courses",
                        fk_column="Course_ID",
                    ),
                ],
            ),
            "TV_series": dc.DBTable(name="TV_series", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("series_name", "TEXT"),
            ]),
            "TV_Channel": dc.DBTable(name="TV_Channel", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("channel", "TEXT"),
            ]),
            "Cartoon": dc.DBTable(name="Cartoon", columns=[
                dc.DBColumn(
                    "channel_id", "INTEGER", fk_table="TV_Channel", fk_column="id",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)

        course_contract = executor._compile_relational_contract(
            "List student IDs from course results.",
        )
        television_contract = executor._compile_relational_contract(
            "List the series name and TV channel used by cartoons.",
        )

        self.assertFalse(any(
            "Courses" in path.get("tables", [])
            for path in course_contract.relation_paths
        ))
        self.assertFalse(any(
            "TV_series" in path.get("tables", [])
            for path in television_contract.relation_paths
        ))

    def test_complete_schema_output_contract_searches_but_still_fails_closed(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "teacher": dc.DBTable(name="teacher", columns=[
                dc.DBColumn("Name", "TEXT"),
                dc.DBColumn("Age", "INTEGER"),
            ]),
        })
        security = mock.Mock()
        executor = dc.NL2SQLExecutor(security, schema)
        generated = {
            "sql": "SELECT Name FROM teacher UNION SELECT Name FROM teacher",
            "summary_zh": "",
        }

        with mock.patch.object(dc, "_llm_ask_json", return_value=generated) as ask:
            answer = executor.answer("Show teacher names and ages.")

        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(ask.call_count, 2)
        self.assertEqual(
            executor.last_candidate_search["status"],
            "no_eligible_candidate",
        )
        self.assertEqual(executor.last_candidate_search["eligible_count"], 0)
        security.execute.assert_not_called()

    def test_native_filtered_ranking_executes_typed_predicate_without_model(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "filtered.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE stadium (Stadium_ID INTEGER PRIMARY KEY, "
                    "Name TEXT, Capacity INTEGER);"
                    "CREATE TABLE concert (concert_ID INTEGER PRIMARY KEY, "
                    "Stadium_ID INTEGER REFERENCES stadium(Stadium_ID), Year INTEGER);"
                    "INSERT INTO stadium VALUES (1, 'Old', 100), (2, 'New', 200);"
                    "INSERT INTO concert VALUES (1, 1, 2013), (2, 1, 2013), "
                    "(3, 2, 2014), (4, 2, 2015);"
                )
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            with mock.patch.object(
                dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
            ) as ask:
                answer = executor.answer(
                    "Show the stadium name and capacity of the stadium with the most "
                    "number of concerts in year 2014 or after.",
                )

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [["New", 200]])
        self.assertEqual(ask.call_count, 0)
        self.assertIn('WHERE t1."Year" >= 2014', answer.sql)

    def test_native_scalar_argmin_binds_schema_label_not_neighbor_attribute(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "scalar_argmin.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE cars_data (Id INTEGER PRIMARY KEY, "
                    "Cylinders INTEGER, Accelerate REAL);"
                    "CREATE TABLE car_names (MakeId INTEGER PRIMARY KEY "
                    "REFERENCES cars_data(Id), Model TEXT, Make TEXT);"
                    "INSERT INTO cars_data VALUES "
                    "(1, 4, 12.0), (2, 6, 8.0), (3, 8, 3.0);"
                    "INSERT INTO car_names VALUES "
                    "(1, 'volvo', 'other'), (2, 'volvo', 'other'), "
                    "(3, 'other', 'volvo');"
                )
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            question = (
                "For a volvo model, how many cylinders does the version "
                "with least accelerate have?"
            )
            contract = executor._compile_relational_contract(question)
            with mock.patch.object(
                dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
            ) as ask:
                answer = executor.answer(question)

        self.assertEqual(contract.version, "1.10")
        self.assertEqual(contract.output_columns, ["Cylinders"])
        self.assertEqual(contract.filter_requirements, [{
            "column": "car_names.Model",
            "operator": "=",
            "value": "volvo",
            "value_type": "text",
            "scope": "row_predicate",
        }])
        self.assertEqual(contract.ordering_requirements, [{
            "column": "cars_data.Accelerate",
            "direction": "ASC",
            "limit": 1,
            "tie_policy": "single_row",
        }])
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [[6]])
        self.assertEqual(answer.relational_plan["kind"], "scalar_ranking")
        self.assertEqual(ask.call_count, 0)
        self.assertIn('t1."Model" = \'volvo\'', answer.sql)
        self.assertNotIn('t1."Make" =', answer.sql)
        self.assertIn('ORDER BY t0."Accelerate" ASC, t0."Id" ASC', answer.sql)

    def test_native_category_usage_groups_requested_value_across_entities(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "category_usage.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE Ref_Template_Types ("
                    "Template_Type_Code TEXT PRIMARY KEY);"
                    "CREATE TABLE Templates (Template_ID INTEGER PRIMARY KEY, "
                    "Template_Type_Code TEXT REFERENCES "
                    "Ref_Template_Types(Template_Type_Code));"
                    "CREATE TABLE Documents (Document_ID INTEGER PRIMARY KEY, "
                    "Template_ID INTEGER REFERENCES Templates(Template_ID));"
                    "INSERT INTO Ref_Template_Types VALUES ('A'), ('B');"
                    "INSERT INTO Templates VALUES "
                    "(1, 'A'), (2, 'A'), (3, 'A'), (4, 'B');"
                    "INSERT INTO Documents VALUES "
                    "(1, 1), (2, 2), (3, 3), (4, 4), (5, 4);"
                )
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            question = (
                "Return the code of the template type that is most commonly "
                "used in documents."
            )
            contract = executor._compile_relational_contract(question)
            with mock.patch.object(
                dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
            ) as ask:
                answer = executor.answer(question)

        self.assertEqual(contract.output_bindings, [{
            "table": "Templates", "column": "Template_Type_Code",
        }])
        self.assertEqual(contract.grouping_keys, ["Templates.Template_Type_Code"])
        self.assertEqual(
            contract.aggregation_stages[0]["aggregates"][0]["source_table"],
            "Documents",
        )
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [["A"]])
        self.assertEqual(ask.call_count, 0)
        self.assertEqual(answer.relational_plan["sources"], ["Templates", "Documents"])
        self.assertIn('GROUP BY t0."Template_Type_Code"', answer.sql)
        self.assertNotIn("Ref_Template_Types", answer.sql)

    def test_native_set_intersection_executes_independent_branches_without_model(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "sets.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE Owners (owner_id INTEGER PRIMARY KEY, state TEXT);"
                    "CREATE TABLE Professionals (professional_id INTEGER PRIMARY KEY, state TEXT);"
                    "INSERT INTO Owners VALUES (1, 'CA'), (2, 'NY');"
                    "INSERT INTO Professionals VALUES (1, 'CA'), (2, 'TX');"
                )
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            with mock.patch.object(
                dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
            ) as ask:
                answer = executor.answer(
                    "Which states are where both owners and professionals live?",
                )

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [["CA"]])
        self.assertEqual(ask.call_count, 0)
        self.assertIn("\nINTERSECT\n", answer.sql)
        self.assertNotIn(" AS t0", answer.sql)

    def test_native_fewest_relationship_preserves_entities_with_zero_facts(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "outer.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE parents (id INTEGER PRIMARY KEY, name TEXT);"
                    "CREATE TABLE children (id INTEGER PRIMARY KEY, "
                    "parent_id INTEGER REFERENCES parents(id));"
                    "INSERT INTO parents VALUES (1, 'Zero'), (2, 'Two');"
                    "INSERT INTO children VALUES (1, 2), (2, 2);"
                )
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            with mock.patch.object(
                dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
            ) as ask:
                answer = executor.answer(
                    "Show the name of the parent that has the fewest number of children.",
                )

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [["Zero"]])
        self.assertEqual(ask.call_count, 0)
        self.assertIn('LEFT JOIN "children"', answer.sql)
        self.assertIn('COUNT(t1."parent_id")', answer.sql)

    def test_native_distinct_entity_count_uses_declared_fact_foreign_key(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "distinct.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE Professionals (professional_id INTEGER PRIMARY KEY);"
                    "CREATE TABLE Treatments (treatment_id INTEGER PRIMARY KEY, "
                    "professional_id INTEGER REFERENCES Professionals(professional_id));"
                    "INSERT INTO Professionals VALUES (1), (2), (3);"
                    "INSERT INTO Treatments VALUES (1, 1), (2, 1), (3, 2);"
                )
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            with mock.patch.object(
                dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
            ) as ask:
                explicit_answer = executor.answer(
                    "How many distinct professionals have treatments?",
                )
                implicit_answer = executor.answer(
                    "Find the number of professionals who have ever performed treatments.",
                )

        self.assertEqual(explicit_answer.kind, "query")
        self.assertEqual(explicit_answer.rows, [[2]])
        self.assertEqual(implicit_answer.kind, "query")
        self.assertEqual(implicit_answer.rows, [[2]])
        self.assertEqual(ask.call_count, 0)
        self.assertIn('COUNT(DISTINCT "professional_id")', explicit_answer.sql)
        self.assertIn('COUNT(DISTINCT "professional_id")', implicit_answer.sql)
        self.assertNotIn(" AS t0", explicit_answer.sql)

    def test_table_routing_prefers_complete_names_and_ignores_imperative_show(self):
        orchestra_schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "conductor": dc.DBTable(name="conductor", columns=[
                dc.DBColumn("Conductor_ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
            ]),
            "orchestra": dc.DBTable(name="orchestra", columns=[
                dc.DBColumn("Orchestra_ID", "INTEGER", pk=True),
                dc.DBColumn(
                    "Conductor_ID", "INTEGER", fk_table="conductor",
                    fk_column="Conductor_ID",
                ),
            ]),
            "performance": dc.DBTable(name="performance", columns=[
                dc.DBColumn("Performance_ID", "INTEGER", pk=True),
                dc.DBColumn(
                    "Orchestra_ID", "INTEGER", fk_table="orchestra",
                    fk_column="Orchestra_ID",
                ),
            ]),
            "show": dc.DBTable(name="show", columns=[
                dc.DBColumn("Show_ID", "INTEGER", pk=True),
                dc.DBColumn(
                    "Performance_ID", "INTEGER", fk_table="performance",
                    fk_column="Performance_ID",
                ),
            ]),
        })
        security = mock.Mock()
        security.connector = dc.DBConnector("fixture")
        executor = dc.NL2SQLExecutor(security, orchestra_schema)
        question = (
            "Show the name of the conductor that has conducted the most "
            "number of orchestras."
        )

        contract = executor._compile_relational_contract(question)
        plan = executor._compile_native_relational_plan(question, contract)

        self.assertEqual(contract.output_columns, ["Name"])
        self.assertEqual(
            [stage["kind"] for stage in contract.aggregation_stages],
            ["group_aggregate", "rank"],
        )
        self.assertIsNotNone(plan)
        self.assertEqual(plan.sources, ["conductor", "orchestra"])

        concert_schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "stadium": dc.DBTable(name="stadium", columns=[
                dc.DBColumn("Stadium_ID", "INTEGER", pk=True),
                dc.DBColumn("Name", "TEXT"),
            ]),
            "concert": dc.DBTable(name="concert", columns=[
                dc.DBColumn("concert_ID", "INTEGER", pk=True),
                dc.DBColumn(
                    "Stadium_ID", "INTEGER", fk_table="stadium",
                    fk_column="Stadium_ID",
                ),
            ]),
            "singer_in_concert": dc.DBTable(name="singer_in_concert", columns=[
                dc.DBColumn(
                    "concert_ID", "INTEGER", fk_table="concert",
                    fk_column="concert_ID",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(security, concert_schema)
        question = "Show the stadium name with most number of concerts."
        contract = executor._compile_relational_contract(question)
        plan = executor._compile_native_relational_plan(question, contract)
        self.assertIsNotNone(plan)
        self.assertEqual(plan.aggregate.source_table, "concert")
        self.assertNotIn("singer_in_concert", plan.sources)

    def test_measure_unit_per_does_not_create_entity_grouping_contract(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "cars_data": dc.DBTable(name="cars_data", columns=[
                dc.DBColumn("Id", "INTEGER", pk=True),
                dc.DBColumn("MPG", "REAL"),
                dc.DBColumn("Cylinders", "INTEGER"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "What is the average miles per gallon of all cars with 4 cylinders?"

        contract = executor._compile_relational_contract(question)

        self.assertNotIn("group_by", contract.required_operators)
        self.assertNotIn("per_entity_measure", contract.evidence)
        self.assertEqual(contract.aggregate_requirements, [{
            "function": "AVG", "column": "cars_data.MPG",
        }])
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT AVG(MPG) FROM cars_data WHERE Cylinders = 4",
                contract,
            ),
            "",
        )

    def test_schema_bound_outputs_enable_projection_lock_and_all_values_search(self):
        orchestra_schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "orchestra": dc.DBTable(name="orchestra", columns=[
                dc.DBColumn("Major_Record_Format", "TEXT"),
            ]),
        })
        security = mock.Mock()
        orchestra = dc.NL2SQLExecutor(security, orchestra_schema)
        order_question = (
            "Please show the record formats of orchestras in ascending order of count."
        )
        order_contract = orchestra._compile_relational_contract(order_question)
        orchestra.last_relational_contract = order_contract
        bad_order_sql = (
            "SELECT Major_Record_Format, COUNT(*) FROM orchestra "
            "GROUP BY Major_Record_Format ORDER BY COUNT(*) ASC"
        )
        conflict = orchestra._semantic_conflict(order_question, bad_order_sql)
        repair = orchestra._try_local_projection_repair(
            question=order_question,
            bad_sql=bad_order_sql,
            conflict=conflict,
            allowed_tables=None,
        )

        self.assertEqual(order_contract.output_columns, ["Major_Record_Format"])
        self.assertIsNotNone(repair)
        self.assertNotIn("COUNT(*) FROM", repair["selected"]["sql"])
        self.assertIn("ORDER BY COUNT(*) ASC", repair["selected"]["sql"])

        tv_schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "TV_Channel": dc.DBTable(name="TV_Channel", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("series_name", "TEXT"),
                dc.DBColumn("Country", "TEXT"),
            ]),
            "Cartoon": dc.DBTable(name="Cartoon", columns=[
                dc.DBColumn(
                    "Channel", "INTEGER", fk_table="TV_Channel", fk_column="id",
                ),
                dc.DBColumn("Directed_by", "TEXT"),
            ]),
        })
        tv = dc.NL2SQLExecutor(mock.Mock(), tv_schema)
        all_values_question = (
            "Find the series name and country of the tv channel that is playing "
            "some cartoons directed by Ben Jones and Michael Chang?"
        )
        all_values_contract = tv._compile_relational_contract(all_values_question)
        tv.last_relational_contract = all_values_contract
        bad_set_sql = (
            "SELECT c.series_name, c.Country FROM TV_Channel c JOIN Cartoon ca "
            "ON ca.Channel=c.id WHERE ca.Directed_by IN "
            "('Ben Jones', 'Michael Chang')"
        )
        conflict = tv._semantic_conflict(all_values_question, bad_set_sql)
        good_set_sql = (
            "SELECT DISTINCT c.series_name, c.Country FROM TV_Channel c WHERE EXISTS "
            "(SELECT 1 FROM Cartoon a WHERE a.Channel=c.id AND "
            "a.Directed_by='Ben Jones') AND EXISTS "
            "(SELECT 1 FROM Cartoon b WHERE b.Channel=c.id AND "
            "b.Directed_by='Michael Chang')"
        )

        self.assertEqual(all_values_contract.output_columns, ["series_name", "Country"])
        self.assertEqual(all_values_contract.set_requirements[0]["operator"], "ALL_VALUES")
        self.assertEqual(
            all_values_contract.distinct_row_requirements[0]["columns"],
            ["TV_Channel.series_name", "TV_Channel.Country"],
        )
        self.assertIsNotNone(conflict)
        self.assertTrue(tv._candidate_search_readiness(
            all_values_question, conflict,
        )["ready"])
        self.assertEqual(
            tv._relational_algebra_retry_hint(
                all_values_question, good_set_sql, all_values_contract,
            ),
            "",
        )
        self.assertIn(
            "SELECT DISTINCT",
            tv._relational_algebra_retry_hint(
                all_values_question,
                good_set_sql.replace("SELECT DISTINCT", "SELECT"),
                all_values_contract,
            ),
        )
        repeated_relation_question = (
            "What is the series name and country of all TV channels that are "
            "playing cartoons directed by Ben Jones and cartoons directed by "
            "Michael Chang?"
        )
        repeated_contract = tv._compile_relational_contract(
            repeated_relation_question,
        )
        repeated_sql = (
            "SELECT c.series_name, c.Country FROM TV_Channel c WHERE EXISTS "
            "(SELECT 1 FROM Cartoon a WHERE a.Channel=c.id AND "
            "a.Directed_by='Ben Jones') AND EXISTS "
            "(SELECT 1 FROM Cartoon b WHERE b.Channel=c.id AND "
            "b.Directed_by='Michael Chang')"
        )
        self.assertEqual(repeated_contract.set_requirements[0]["operator"], "ALL_VALUES")
        self.assertEqual(
            repeated_contract.distinct_row_requirements[0]["columns"],
            ["TV_Channel.series_name", "TV_Channel.Country"],
        )
        self.assertIn(
            "SELECT DISTINCT",
            tv._relational_algebra_retry_hint(
                repeated_relation_question, repeated_sql, repeated_contract,
            ),
        )
        tv.last_relational_contract = repeated_contract
        distinct_conflict = tv._semantic_conflict(
            repeated_relation_question, repeated_sql,
        )
        distinct_repair = tv._try_local_contract_repair(
            question=repeated_relation_question,
            bad_sql=repeated_sql,
            conflict=distinct_conflict,
            allowed_tables=None,
        )
        self.assertIsNotNone(distinct_repair)
        self.assertEqual(
            distinct_repair["diagnostic"]["status"],
            "local_distinct_tuple_compiled",
        )
        self.assertTrue(distinct_repair["selected"]["sql"].startswith(
            "SELECT DISTINCT c.series_name, c.Country",
        ))
        self.assertEqual(
            tv._semantic_conflict(
                repeated_relation_question, distinct_repair["selected"]["sql"],
            ),
            None,
        )
        self.assertIsNone(tv._try_local_distinct_tuple_repair(
            question=repeated_relation_question,
            bad_sql=repeated_sql + " LIMIT 1",
            conflict=distinct_conflict,
            allowed_tables=None,
        ))
        self.assertIsNone(tv._try_local_distinct_tuple_repair(
            question=repeated_relation_question,
            bad_sql=repeated_sql.replace("SELECT", "SELECT ALL", 1),
            conflict=distinct_conflict,
            allowed_tables=None,
        ))
        self.assertIsNone(tv._try_local_distinct_tuple_repair(
            question=repeated_relation_question,
            bad_sql=repeated_sql.replace(
                " FROM ", ", 1 AS extra_output FROM ", 1,
            ),
            conflict=distinct_conflict,
            allowed_tables=None,
        ))
        self.assertIsNone(tv._try_local_distinct_tuple_repair(
            question=repeated_relation_question,
            bad_sql=(
                "SELECT c.series_name, c.Country, COUNT(*) FROM TV_Channel c "
                "GROUP BY c.series_name, c.Country"
            ),
            conflict=distinct_conflict,
            allowed_tables=None,
        ))
        self.assertEqual(
            tv._relational_algebra_retry_hint(
                repeated_relation_question,
                repeated_sql.replace("SELECT", "SELECT DISTINCT", 1),
                repeated_contract,
            ),
            "",
        )

    def test_relational_ir_closes_what_is_superlative_to_one_row(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "airport": dc.DBTable(name="airport", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("code", "TEXT"),
            ]),
            "flight": dc.DBTable(name="flight", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn(
                    "airport_id", "INTEGER", fk_table="airport", fk_column="id",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "What is the code of airport that has fewest number of flights?"
        contract = executor._compile_relational_contract(question)
        sql = (
            "SELECT a.code FROM airport a LEFT JOIN flight f "
            "ON f.airport_id=a.id GROUP BY a.id ORDER BY COUNT(f.id) ASC"
        )

        self.assertEqual(contract.version, "1.10")
        self.assertEqual(contract.tie_policy, "single_row")
        self.assertEqual(contract.tie_breaker_columns, ["airport.id"])
        self.assertIn(
            "单数实体",
            executor._relational_algebra_retry_hint(question, sql, contract),
        )
        self.assertIn(
            "稳定二级排序",
            executor._relational_algebra_retry_hint(
                question, sql + " LIMIT 1", contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question, sql + ", a.id ASC LIMIT 1", contract,
            ),
            "",
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question, sql + ", f.airport_id ASC LIMIT 1", contract,
            ),
            "",
        )
        executor.last_relational_contract = contract
        unstable_sql = sql + " LIMIT 1"
        conflict = executor._semantic_conflict(question, unstable_sql)
        repair = executor._try_local_deterministic_tie_repair(
            question=question,
            bad_sql=unstable_sql,
            conflict=conflict,
            allowed_tables=None,
        )
        self.assertIsNotNone(repair)
        self.assertEqual(
            repair["diagnostic"]["status"],
            "local_deterministic_tie_compiled",
        )
        self.assertIn(', "a"."id" ASC LIMIT 1', repair["selected"]["sql"])
        self.assertEqual(
            executor._semantic_conflict(question, repair["selected"]["sql"]),
            None,
        )
        self.assertIsNone(executor._try_local_deterministic_tie_repair(
            question=question,
            bad_sql=(
                "SELECT a.code FROM airport a JOIN airport b ON b.id=a.id "
                "LEFT JOIN flight f ON f.airport_id=a.id GROUP BY a.id "
                "ORDER BY COUNT(f.id) ASC LIMIT 1"
            ),
            conflict=conflict,
            allowed_tables=None,
        ))
        self.assertIsNone(executor._try_local_deterministic_tie_repair(
            question=question,
            bad_sql=unstable_sql.replace("LIMIT 1", "LIMIT 2"),
            conflict=conflict,
            allowed_tables=None,
        ))
        self.assertIsNone(executor._try_local_deterministic_tie_repair(
            question=question,
            bad_sql=unstable_sql + " UNION SELECT code FROM airport",
            conflict=conflict,
            allowed_tables=None,
        ))

    def test_answer_locally_compiles_visible_tuple_distinctness(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "TV_Channel": dc.DBTable(name="TV_Channel", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("series_name", "TEXT"),
                dc.DBColumn("Country", "TEXT"),
            ]),
            "Cartoon": dc.DBTable(name="Cartoon", columns=[
                dc.DBColumn(
                    "Channel", "INTEGER", fk_table="TV_Channel", fk_column="id",
                ),
                dc.DBColumn("Directed_by", "TEXT"),
            ]),
        })
        question = (
            "What is the series name and country of all TV channels that are "
            "playing cartoons directed by Ben Jones and cartoons directed by "
            "Michael Chang?"
        )
        generated_sql = (
            "SELECT c.series_name, c.Country FROM TV_Channel c WHERE EXISTS "
            "(SELECT 1 FROM Cartoon a WHERE a.Channel=c.id AND "
            "a.Directed_by='Ben Jones') AND EXISTS "
            "(SELECT 1 FROM Cartoon b WHERE b.Channel=c.id AND "
            "b.Directed_by='Michael Chang')"
        )
        repaired_sql = generated_sql.replace("SELECT", "SELECT DISTINCT", 1)
        security = mock.Mock()
        security.execute.return_value = dc.SQLResult(
            sql=repaired_sql,
            columns=["series_name", "Country"],
            rows=[["Animation", "US"]],
            row_count=1,
        )
        executor = dc.NL2SQLExecutor(security, schema)

        with mock.patch.object(dc, "_llm_ask_json", return_value={
            "sql": generated_sql,
            "summary_zh": "",
        }) as ask:
            answer = executor.answer(question)

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.sql, repaired_sql)
        self.assertEqual(answer.rows, [["Animation", "US"]])
        self.assertEqual(ask.call_count, 1)
        security.execute.assert_called_once_with(repaired_sql)
        self.assertEqual(executor.semantic_repair_count, 0)
        self.assertEqual(
            executor.last_candidate_search["status"],
            "local_distinct_tuple_compiled",
        )
        self.assertEqual(executor.last_candidate_search["model_calls"], 0)

    def test_local_contract_pipeline_composes_projection_and_tie_repairs(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Transcripts": dc.DBTable(name="Transcripts", columns=[
                dc.DBColumn("transcript_id", "INTEGER", pk=True),
                dc.DBColumn("transcript_date", "TEXT"),
            ]),
            "Transcript_Contents": dc.DBTable(
                name="Transcript_Contents",
                columns=[
                    dc.DBColumn(
                        "transcript_id", "INTEGER", fk_table="Transcripts",
                        fk_column="transcript_id",
                    ),
                    dc.DBColumn("student_course_id", "INTEGER"),
                ],
            ),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "What is the date and id of the transcript with the least number "
            "of results?"
        )
        sql = (
            "SELECT t.transcript_id, t.transcript_date FROM Transcripts t "
            "LEFT JOIN Transcript_Contents tc "
            "ON t.transcript_id = tc.transcript_id "
            "GROUP BY t.transcript_id, t.transcript_date "
            "ORDER BY COUNT(tc.student_course_id) ASC LIMIT 1"
        )
        contract = executor._compile_relational_contract(question)
        executor.last_relational_contract = contract
        conflict = executor._semantic_conflict(question, sql)

        self.assertIsNotNone(conflict)
        repair = executor._try_local_contract_repair(
            question=question,
            bad_sql=sql,
            conflict=conflict,
            allowed_tables=None,
        )

        self.assertIsNotNone(repair)
        self.assertEqual(
            repair["diagnostic"]["status"],
            "local_contract_pipeline_compiled",
        )
        self.assertEqual(repair["diagnostic"]["pipeline_stages"], [
            "local_projection_compiled",
            "local_deterministic_tie_compiled",
        ])
        repaired_sql = repair["selected"]["sql"]
        self.assertTrue(repaired_sql.startswith(
            'SELECT t."transcript_date", t."transcript_id" FROM',
        ))
        self.assertIn(', "t"."transcript_id" ASC LIMIT 1', repaired_sql)
        self.assertIsNone(executor._semantic_conflict(question, repaired_sql))
        self.assertEqual(repair["diagnostic"]["model_calls"], 0)

    def test_schema_bound_projection_authority_suppresses_weaker_heuristic(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "country": dc.DBTable(name="country", columns=[
                dc.DBColumn("Code", "TEXT", pk=True),
                dc.DBColumn("Code2", "TEXT"),
            ]),
            "countrylanguage": dc.DBTable(name="countrylanguage", columns=[
                dc.DBColumn(
                    "CountryCode", "TEXT", fk_table="country",
                    fk_column="Code",
                ),
                dc.DBColumn("Language", "TEXT", sample_values=["Spanish"]),
                dc.DBColumn("Percentage", "REAL"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "Return the codes of countries for which Spanish is the "
            "predominantly spoken language."
        )
        sql = (
            "SELECT country.Code FROM country JOIN countrylanguage sp "
            "ON country.Code = sp.CountryCode AND sp.Language = 'Spanish' "
            "WHERE NOT EXISTS (SELECT 1 FROM countrylanguage other "
            "WHERE other.CountryCode = country.Code "
            "AND other.Percentage > sp.Percentage)"
        )
        executor.last_relational_contract = executor._compile_relational_contract(
            question,
        )
        conflict = executor._semantic_conflict(question, sql)

        self.assertIsNone(conflict)
        self.assertEqual(executor.last_relational_contract.output_bindings, [{
            "table": "country", "column": "Code",
        }])
        self.assertIsNone(executor.last_candidate_search)

    def test_relational_ir_compiles_qualified_anti_relationship_filter(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "battle": dc.DBTable(name="battle", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("name", "TEXT"),
                dc.DBColumn("result", "TEXT"),
                dc.DBColumn("bulgarian_commander", "TEXT"),
            ]),
            "ship": dc.DBTable(name="ship", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn(
                    "lost_in_battle", "INTEGER", fk_table="battle", fk_column="id",
                ),
                dc.DBColumn(
                    "location", "TEXT", sample_values=["English Channel", "Atlantic"],
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "Show names, results and bulgarian commanders of the battles with no "
            "ships lost in the 'English Channel'."
        )
        contract = executor._compile_relational_contract(question)
        good = (
            "SELECT b.name, b.result, b.bulgarian_commander FROM battle b "
            "WHERE NOT EXISTS (SELECT 1 FROM ship s WHERE "
            "s.lost_in_battle=b.id AND s.location='English Channel')"
        )
        missing_filter = good.replace(" AND s.location='English Channel'", "")
        uncorrelated = good.replace("s.lost_in_battle=b.id AND ", "")

        self.assertEqual(contract.version, "1.10")
        self.assertEqual(contract.correlation_requirements[0]["outer_table"], "battle")
        self.assertEqual(contract.correlation_requirements[0]["inner_table"], "ship")
        self.assertEqual(contract.filter_requirements, [{
            "column": "ship.location",
            "operator": "=",
            "value": "English Channel",
            "value_type": "text",
            "scope": "row_predicate",
        }])
        self.assertEqual(
            executor._relational_algebra_retry_hint(question, good, contract), "",
        )
        self.assertIn(
            "类型化过滤",
            executor._relational_algebra_retry_hint(
                question, missing_filter, contract,
            ),
        )
        executor.last_relational_contract = contract
        missing_filter_conflict = executor._semantic_conflict(
            question, missing_filter,
        )
        self.assertIsNotNone(missing_filter_conflict)
        self.assertTrue(executor._candidate_search_readiness(
            question, missing_filter_conflict,
        )["ready"])
        self.assertIn(
            "关联",
            executor._relational_algebra_retry_hint(
                question, uncorrelated, contract,
            ),
        )

    def test_semantic_gate_rejects_unrequested_wildcard_broadening(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Students": dc.DBTable(name="Students", columns=[
                dc.DBColumn("student_id", "INTEGER", pk=True),
                dc.DBColumn("first_name", "TEXT"),
                dc.DBColumn("middle_name", "TEXT"),
                dc.DBColumn("last_name", "TEXT"),
            ]),
            "Student_Enrolment": dc.DBTable(name="Student_Enrolment", columns=[
                dc.DBColumn(
                    "student_id", "INTEGER", fk_table="Students",
                    fk_column="student_id",
                ),
                dc.DBColumn(
                    "degree_program_id", "INTEGER", fk_table="Degree_Programs",
                    fk_column="degree_program_id",
                ),
            ]),
            "Degree_Programs": dc.DBTable(name="Degree_Programs", columns=[
                dc.DBColumn("degree_program_id", "INTEGER", pk=True),
                dc.DBColumn(
                    "degree_summary_name", "TEXT",
                    sample_values=["Bachelor", "Master"],
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "What are the first, middle, and last names for everybody enrolled "
            "in a Bachelors program?"
        )
        sql = (
            "SELECT s.first_name, s.middle_name, s.last_name FROM Students s "
            "JOIN Student_Enrolment e ON e.student_id=s.student_id "
            "JOIN Degree_Programs d ON d.degree_program_id=e.degree_program_id "
            "WHERE d.degree_summary_name LIKE '%Bachelor%'"
        )
        executor.last_relational_contract = executor._compile_relational_contract(question)
        conflict = executor._semantic_conflict(question, sql)

        self.assertIsNotNone(conflict)
        self.assertEqual(conflict.code, "wildcard_literal_broadening")
        repair = executor._try_local_contract_repair(
            question=question,
            bad_sql=sql,
            conflict=conflict,
            allowed_tables=None,
        )
        self.assertIsNotNone(repair)
        self.assertEqual(
            repair["diagnostic"]["status"], "local_exact_enum_compiled",
        )
        self.assertIn("d.degree_summary_name = 'Bachelor'", repair["selected"]["sql"])
        self.assertNotIn("LIKE", repair["selected"]["sql"])
        self.assertIsNone(executor._semantic_conflict(
            question, repair["selected"]["sql"],
        ))
        self.assertIsNone(executor._try_local_exact_enum_repair(
            question="List programs containing Bachelor in the description.",
            bad_sql=sql,
            conflict=conflict,
            allowed_tables=None,
        ))
        self.assertIsNone(executor._semantic_conflict(
            question,
            sql.replace("LIKE '%Bachelor%'", "= 'Bachelor'"),
        ))
        self.assertEqual(
            executor._wildcard_literal_retry_hint(
                "List programs containing Bachelor in the description.", sql,
            ),
            "",
        )
        self.assertEqual(
            executor._compile_relational_contract(
                "List degree summary names containing 'Bachelor'."
            ).filter_requirements,
            [],
        )
        self.assertEqual(
            executor._compile_relational_contract(
                "List degree summary names that are not 'Bachelor'."
            ).filter_requirements,
            [],
        )
        self.assertEqual(
            executor._compile_relational_contract(
                "List degree summary names not named 'Bachelor'."
            ).filter_requirements,
            [],
        )
        noun_match_contract = executor._compile_relational_contract(
            "List matching enrolment records in the 'Bachelor' category."
        )
        self.assertEqual(noun_match_contract.filter_requirements, [{
            "column": "Degree_Programs.degree_summary_name",
            "operator": "=",
            "value": "Bachelor",
            "value_type": "text",
            "scope": "row_predicate",
        }])

    def test_relational_ir_rejects_ungrounded_predicate_literal(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "players": dc.DBTable(name="players", columns=[
                dc.DBColumn("player_id", "INTEGER", pk=True),
                dc.DBColumn("first_name", "TEXT"),
                dc.DBColumn("country_code", "TEXT"),
            ]),
            "matches": dc.DBTable(name="matches", columns=[
                dc.DBColumn(
                    "winner_id", "INTEGER", fk_table="players",
                    fk_column="player_id",
                ),
                dc.DBColumn(
                    "loser_id", "INTEGER", fk_table="players",
                    fk_column="player_id",
                ),
                dc.DBColumn("round", "TEXT", sample_values=["F", "SF"]),
                dc.DBColumn(
                    "tourney_name", "TEXT",
                    sample_values=["WTA Championships", "Auckland"],
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "What are the first names and country codes for players who won both "
            "the WTA Championships and the Australian Open?"
        )
        contract = executor._compile_relational_contract(question)
        good = (
            "SELECT DISTINCT p.first_name, p.country_code FROM players p WHERE EXISTS "
            "(SELECT 1 FROM matches m WHERE m.winner_id=p.player_id AND "
            "m.tourney_name='WTA Championships') AND EXISTS "
            "(SELECT 1 FROM matches m WHERE m.winner_id=p.player_id AND "
            "m.tourney_name='Australian Open')"
        )
        bad = good.replace(
            "m.tourney_name='WTA Championships')",
            "m.tourney_name='WTA Championships' AND m.round='F')",
        )

        self.assertEqual(contract.version, "1.10")
        self.assertEqual(contract.set_requirements[0]["operator"], "ALL_VALUES")
        self.assertEqual(
            contract.relation_paths[0]["source"], "role_bound_declared_fk_path",
        )
        self.assertIn(
            "无来源谓词",
            executor._relational_algebra_retry_hint(question, bad, contract),
        )
        self.assertIn(
            "无来源谓词",
            executor._relational_algebra_retry_hint(
                question,
                good.replace(
                    "m.tourney_name='WTA Championships')",
                    "m.tourney_name='WTA Championships' AND m.round IN ('F'))",
                ),
                contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(question, good, contract), "",
        )

    def test_relational_ir_all_values_locks_output_entity_grain(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "AREA_CODE_STATE": dc.DBTable(name="AREA_CODE_STATE", columns=[
                dc.DBColumn("area_code", "INTEGER", pk=True),
                dc.DBColumn("state", "TEXT"),
            ]),
            "CONTESTANTS": dc.DBTable(name="CONTESTANTS", columns=[
                dc.DBColumn("contestant_number", "INTEGER", pk=True),
                dc.DBColumn(
                    "contestant_name", "TEXT",
                    sample_values=["Tabatha Gehling", "Kelly Clauss"],
                ),
            ]),
            "VOTES": dc.DBTable(name="VOTES", columns=[
                dc.DBColumn("vote_id", "INTEGER", pk=True),
                dc.DBColumn("phone_number", "INTEGER"),
                dc.DBColumn(
                    "state", "TEXT", fk_table="AREA_CODE_STATE",
                    fk_column="state",
                ),
                dc.DBColumn(
                    "contestant_number", "INTEGER", fk_table="CONTESTANTS",
                    fk_column="contestant_number",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "List the area codes in which voters voted both for the contestant "
            "'Tabatha Gehling' and the contestant 'Kelly Clauss'."
        )
        contract = executor._compile_relational_contract(question)
        wrong_grain = (
            "SELECT DISTINCT a.area_code FROM VOTES v JOIN CONTESTANTS c ON "
            "v.contestant_number=c.contestant_number JOIN AREA_CODE_STATE a ON "
            "v.state=a.state WHERE c.contestant_name IN "
            "('Tabatha Gehling','Kelly Clauss') GROUP BY v.phone_number,a.area_code "
            "HAVING COUNT(DISTINCT v.contestant_number)=2"
        )
        correct = (
            "SELECT a.area_code FROM AREA_CODE_STATE a JOIN VOTES v ON a.state=v.state "
            "JOIN CONTESTANTS c ON v.contestant_number=c.contestant_number "
            "WHERE c.contestant_name='Tabatha Gehling' INTERSECT "
            "SELECT a.area_code FROM AREA_CODE_STATE a JOIN VOTES v ON a.state=v.state "
            "JOIN CONTESTANTS c ON v.contestant_number=c.contestant_number "
            "WHERE c.contestant_name='Kelly Clauss'"
        )

        self.assertEqual(contract.output_columns, ["area_code"])
        self.assertEqual(
            contract.set_requirements[0]["row_grain"],
            "AREA_CODE_STATE.area_code",
        )
        self.assertIn(
            "同一父实体",
            executor._relational_algebra_retry_hint(
                question, wrong_grain, contract,
            ),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(question, correct, contract),
            "",
        )

    def test_relational_ir_spending_superlative_binds_sum_measure(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Owners": dc.DBTable(name="Owners", columns=[
                dc.DBColumn("owner_id", "INTEGER", pk=True),
                dc.DBColumn("last_name", "TEXT"),
            ]),
            "Dogs": dc.DBTable(name="Dogs", columns=[
                dc.DBColumn("dog_id", "INTEGER", pk=True),
                dc.DBColumn(
                    "owner_id", "INTEGER", fk_table="Owners",
                    fk_column="owner_id",
                ),
            ]),
            "Treatments": dc.DBTable(name="Treatments", columns=[
                dc.DBColumn(
                    "dog_id", "INTEGER", fk_table="Dogs", fk_column="dog_id",
                ),
                dc.DBColumn("cost_of_treatment", "REAL"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "Tell me the owner id and last name of the owner who spent the most "
            "on treatments of his or her dogs."
        )
        contract = executor._compile_relational_contract(question)
        count_sql = (
            "SELECT o.owner_id,o.last_name FROM Owners o JOIN Dogs d ON "
            "d.owner_id=o.owner_id JOIN Treatments t ON t.dog_id=d.dog_id "
            "GROUP BY o.owner_id,o.last_name ORDER BY COUNT(*) DESC, "
            "o.owner_id ASC LIMIT 1"
        )
        sum_sql = count_sql.replace("COUNT(*)", "SUM(t.cost_of_treatment)")

        self.assertEqual(contract.aggregate_requirements, [{
            "function": "SUM", "column": "Treatments.cost_of_treatment",
        }])
        self.assertEqual(
            [stage["kind"] for stage in contract.aggregation_stages],
            ["group_aggregate", "rank"],
        )
        count_hint = executor._relational_algebra_retry_hint(
            question, count_sql, contract,
        )
        self.assertIn("SUM", count_hint)
        self.assertIn("聚合", count_hint)
        self.assertEqual(
            executor._relational_algebra_retry_hint(question, sum_sql, contract),
            "",
        )

    def test_relational_ir_requires_distinct_descriptive_tuple(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Professionals": dc.DBTable(name="Professionals", columns=[
                dc.DBColumn("professional_id", "INTEGER", pk=True),
                dc.DBColumn("first_name", "TEXT"),
            ]),
            "Treatments": dc.DBTable(name="Treatments", columns=[
                dc.DBColumn(
                    "professional_id", "INTEGER", fk_table="Professionals",
                    fk_column="professional_id",
                ),
                dc.DBColumn(
                    "treatment_type_code", "TEXT", fk_table="Treatment_Types",
                    fk_column="treatment_type_code",
                ),
            ]),
            "Treatment_Types": dc.DBTable(name="Treatment_Types", columns=[
                dc.DBColumn("treatment_type_code", "TEXT", pk=True),
                dc.DBColumn("treatment_type_description", "TEXT"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "What are each professional's first name and description of the "
            "treatment they have performed?"
        )
        contract = executor._compile_relational_contract(question)
        sql = (
            "SELECT p.first_name,tt.treatment_type_description FROM Professionals p "
            "JOIN Treatments t ON p.professional_id=t.professional_id JOIN "
            "Treatment_Types tt ON t.treatment_type_code=tt.treatment_type_code"
        )

        self.assertEqual(contract.distinct_row_requirements[0]["row_grain"],
                         "unique_output_tuple")
        self.assertIn(
            "SELECT DISTINCT",
            executor._relational_algebra_retry_hint(question, sql, contract),
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question, sql.replace("SELECT ", "SELECT DISTINCT ", 1), contract,
            ),
            "",
        )

    def test_relational_ir_clarifies_ambiguous_boolean_modifier_scope(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Properties": dc.DBTable(name="Properties", columns=[
                dc.DBColumn("property_id", "INTEGER", pk=True),
                dc.DBColumn("property_name", "TEXT"),
                dc.DBColumn(
                    "property_type_code", "TEXT",
                    sample_values=["House", "Apartment"],
                ),
                dc.DBColumn("room_count", "INTEGER"),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = (
            "What are the names of properties that are either houses or "
            "apartments with more than 1 room?"
        )
        contract = executor._compile_relational_contract(question)

        self.assertEqual(contract.ambiguities[0]["kind"],
                         "boolean_modifier_scope")
        with mock.patch.object(
            dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
        ):
            answer = executor.answer(question)
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(
            answer.clarification["missing"], "boolean_filter_scope",
        )
        self.assertEqual(answer.steps[0]["status"], "needs_clarification")

        planner = dc.NaturalLanguageDatabasePlanner(schema)
        resolved = planner.resolve_followup(
            "修饰条件只作用于 or 后面的类别",
            history=[],
            clarification=answer.clarification,
        )
        self.assertIn("布尔筛选作用域", resolved)
        resolved_contract = executor._compile_relational_contract(resolved)
        self.assertEqual(resolved_contract.ambiguities, [])
        self.assertEqual(
            resolved_contract.boolean_filter_requirements[0]["scope"],
            "right_category_only",
        )
        right_only = (
            'SELECT property_name FROM Properties WHERE property_type_code="House" '
            'UNION SELECT property_name FROM Properties WHERE '
            'property_type_code="Apartment" AND room_count>1'
        )
        both_categories = (
            "SELECT property_name FROM Properties WHERE property_type_code IN "
            "('House','Apartment') AND room_count>1"
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                resolved, right_only, resolved_contract,
            ),
            "",
        )
        self.assertIn(
            "只作用于后一类别",
            executor._relational_algebra_retry_hint(
                resolved, both_categories, resolved_contract,
            ),
        )

        resolved_both = planner.resolve_followup(
            "修饰条件同时作用于两个类别",
            history=[],
            clarification=answer.clarification,
        )
        both_contract = executor._compile_relational_contract(resolved_both)
        self.assertEqual(
            both_contract.boolean_filter_requirements[0]["scope"],
            "both_categories",
        )
        self.assertEqual(
            executor._relational_algebra_retry_hint(
                resolved_both, both_categories, both_contract,
            ),
            "",
        )

    def test_native_fk_set_difference_executes_without_model(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "anti_set.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.executescript(
                    "CREATE TABLE Templates (Template_ID INTEGER PRIMARY KEY);"
                    "CREATE TABLE Documents (Document_ID INTEGER PRIMARY KEY, "
                    "Template_ID INTEGER REFERENCES Templates(Template_ID));"
                    "INSERT INTO Templates VALUES (1), (2), (3);"
                    "INSERT INTO Documents VALUES (1, 2), (2, 3);"
                )
            connector = dc.DBConnector(str(db_path))
            schema = dc.SchemaDiscovery(connector).discover()
            executor = dc.NL2SQLExecutor(dc.SQLSecurity(connector), schema)
            with mock.patch.object(
                dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
            ) as ask:
                answer = executor.answer(
                    "Show IDs for all templates not used by any document.",
                )

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [[1]])
        self.assertEqual(ask.call_count, 0)
        self.assertIn("\nEXCEPT\n", answer.sql)
        self.assertNotIn(" AS t0", answer.sql)
        self.assertEqual(answer.relational_plan["operator"], "EXCEPT")
        self.assertEqual(len(answer.relational_plan["proof_edges"]), 1)

    def test_relation_contract_accepts_unqualified_inner_fk_in_correlated_not_exists(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "Templates": dc.DBTable(name="Templates", columns=[
                dc.DBColumn("Template_ID", "INTEGER", pk=True),
            ]),
            "Documents": dc.DBTable(name="Documents", columns=[
                dc.DBColumn(
                    "Template_ID", "INTEGER", fk_table="Templates",
                    fk_column="Template_ID",
                ),
            ]),
        })
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        question = "Show IDs for all templates not used by any document."
        contract = executor._compile_relational_contract(question)

        self.assertEqual(
            executor._relational_algebra_retry_hint(
                question,
                "SELECT Template_ID FROM Templates WHERE NOT EXISTS "
                "(SELECT 1 FROM Documents WHERE Template_ID = Templates.Template_ID)",
                contract,
            ),
            "",
        )

    def test_spider_adapter_removes_only_projection_aliases(self):
        benchmark_path = Path(__file__).resolve().parents[3] / "scripts/run_spider_benchmark.py"
        spec = importlib.util.spec_from_file_location("spider_benchmark_for_test", benchmark_path)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        benchmark = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(benchmark)
        schema_record = {
            "table_names_original": ["singer", "concert"],
            "column_names_original": [
                [-1, "*"], [0, "id"], [0, "Name"], [1, "id"],
            ],
        }

        normalized = benchmark._official_parser_compatible_sql(
            "SELECT COUNT(*) AS total FROM singer s JOIN concert c ON c.id=s.id",
            schema_record,
        )

        self.assertEqual(
            normalized,
            "SELECT COUNT(*) FROM singer AS s JOIN concert AS c ON c.id=s.id",
        )
        quoted_alias = benchmark._official_parser_compatible_sql(
            'SELECT COUNT(*) AS "total" FROM singer s',
            schema_record,
        )
        self.assertEqual(quoted_alias, "SELECT COUNT(*) FROM singer AS s")

        quoted = benchmark._official_parser_compatible_sql(
            'SELECT s."Name" FROM singer s WHERE s."Name" = \'Ada\'',
            schema_record,
        )
        self.assertEqual(
            quoted,
            "SELECT s.Name FROM singer AS s WHERE s.Name = 'Ada'",
        )
        quoted_sources = benchmark._official_parser_compatible_sql(
            'SELECT s."Name" FROM "singer" AS s JOIN "concert" AS c '
            'ON c."id"=s."id"',
            schema_record,
        )
        self.assertEqual(
            quoted_sources,
            "SELECT s.Name FROM singer AS s JOIN concert AS c ON c.id=s.id",
        )
        unqualified = benchmark._official_parser_compatible_sql(
            'SELECT "Name" FROM "singer"',
            schema_record,
        )
        self.assertEqual(unqualified, "SELECT Name FROM singer")
        outer_join = benchmark._official_parser_compatible_sql(
            'SELECT s."Name" FROM "singer" AS s LEFT JOIN "concert" AS c '
            'ON c."id"=s."id"',
            schema_record,
        )
        self.assertEqual(
            outer_join,
            "SELECT s.Name FROM singer AS s JOIN concert AS c ON c.id=s.id",
        )

    def test_spider_summary_excludes_llm_infrastructure_and_rescores_old_checkpoint(self):
        benchmark_path = Path(__file__).resolve().parents[3] / "scripts/run_spider_benchmark.py"
        spec = importlib.util.spec_from_file_location("spider_infrastructure_for_test", benchmark_path)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        benchmark = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(benchmark)
        results = [
            {
                "id": "pass", "index": 0, "hardness": "easy",
                "exact_match": True, "answer_kind": "query",
                "predicted_sql": "SELECT 1", "error_category": None,
                "latency_ms": 10,
            },
            {
                "id": "infra", "index": 1, "hardness": "medium",
                "exact_match": False, "answer_kind": "error",
                "predicted_sql": "", "execution_error": "LLM 服务请求失败（HTTP 402）",
                "error_category": benchmark.INFRASTRUCTURE_ERROR_CATEGORY,
                "component_gaps": [benchmark.INFRASTRUCTURE_ERROR_CATEGORY],
                "latency_ms": 5,
            },
        ]

        summary = benchmark._summary(results, target_total=2)

        self.assertEqual(summary["exact_match"], {
            "passed": 1, "total": 1, "rate": 1.0,
        })
        self.assertEqual(summary["raw_lower_bound_exact_match"]["total"], 2)
        self.assertEqual(summary["coverage"]["infrastructure_failures"], 1)
        self.assertFalse(summary["coverage"]["complete"])
        self.assertEqual(summary["latency_ms"]["median"], 10.0)
        self.assertTrue(benchmark._is_llm_infrastructure_error(
            "LLM 服务请求失败（HTTP 402）",
        ))
        rescored = benchmark._rescore_result(
            {
                **results[1],
                "gold_sql": "SELECT 1",
                "error_category": "generation_or_execution_error",
            },
            Path("unused.sqlite"),
            {},
            None,
            None,
        )
        self.assertEqual(
            rescored["error_category"], benchmark.INFRASTRUCTURE_ERROR_CATEGORY,
        )
        self.assertEqual(benchmark.SCORING_CONTRACT, "spider-exact-set-match-adapter-v5")

    def test_spider_execution_comparator_preserves_bags_order_and_column_permutations(self):
        scoring_path = Path(__file__).resolve().parents[3] / "scripts/spider_execution_scoring.py"
        spec = importlib.util.spec_from_file_location("spider_execution_scoring_for_test", scoring_path)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        scoring = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(scoring)

        equivalent, detail = scoring.result_equivalent(
            [(1, "Ada"), (2, "Bob")],
            [("Ada", 1), ("Bob", 2)],
            order_matters=False,
        )
        self.assertTrue(equivalent)
        self.assertEqual(detail["reason"], "equivalent")
        duplicate_mismatch, detail = scoring.result_equivalent(
            [(1,), (1,)], [(1,), (2,)], order_matters=False,
        )
        self.assertFalse(duplicate_mismatch)
        self.assertIn(detail["reason"], {"column_value_mismatch", "value_or_row_association"})
        unordered, _ = scoring.result_equivalent(
            [(1,), (2,)], [(2,), (1,)], order_matters=False,
        )
        ordered, _ = scoring.result_equivalent(
            [(1,), (2,)], [(2,), (1,)], order_matters=True,
        )
        self.assertTrue(unordered)
        self.assertFalse(ordered)
        self.assertTrue(scoring.gold_order_matters(
            "SELECT id FROM t ORDER BY id",
        ))
        self.assertFalse(scoring.gold_order_matters(
            "SELECT 'ORDER BY is data' FROM t",
        ))

    def test_spider_execution_scoring_is_read_only_and_reports_empty_evidence(self):
        scoring_path = Path(__file__).resolve().parents[3] / "scripts/spider_execution_scoring.py"
        spec = importlib.util.spec_from_file_location("spider_execution_readonly_for_test", scoring_path)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        scoring = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(scoring)
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "fixture.sqlite"
            with closing(sqlite3.connect(db_path)) as connection:
                connection.executescript(
                    "CREATE TABLE people(id INTEGER PRIMARY KEY, name TEXT);"
                    "INSERT INTO people VALUES (1, 'Ada'), (2, 'Bob');"
                )
            result = scoring.score_execution(
                db_path,
                "SELECT name, id FROM people ORDER BY id",
                "SELECT id, name FROM people ORDER BY id",
            )
            blocked = scoring.execute_readonly(
                db_path, "DELETE FROM people WHERE id=1",
            )
            empty = scoring.score_execution(
                db_path,
                "SELECT name FROM people WHERE id=999",
                "SELECT id FROM people WHERE id=999",
            )
            with closing(sqlite3.connect(db_path)) as connection:
                row_count = connection.execute("SELECT COUNT(*) FROM people").fetchone()[0]

        self.assertTrue(result["agreement"])
        self.assertTrue(result["order_matters"])
        self.assertIsNotNone(blocked["error"])
        self.assertEqual(row_count, 2)
        self.assertTrue(empty["agreement"])
        self.assertTrue(empty["empty_result_match"])
        self.assertTrue(empty["gold_empty"])

    def test_spider_test_suite_catches_single_database_false_positive(self):
        scoring_path = Path(__file__).resolve().parents[3] / "scripts/spider_execution_scoring.py"
        spec = importlib.util.spec_from_file_location("spider_test_suite_scoring_for_test", scoring_path)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        scoring = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(scoring)
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            first = root / "first.sqlite"
            second = root / "second.sqlite"
            for path, rows in (
                (first, [(1, "same"), (2, "same")]),
                (second, [(1, "left"), (2, "right")]),
            ):
                with closing(sqlite3.connect(path)) as connection:
                    connection.execute("CREATE TABLE items (id INTEGER, label TEXT)")
                    connection.executemany("INSERT INTO items VALUES (?, ?)", rows)
                    connection.commit()
            single = scoring.score_execution(
                first,
                "SELECT label FROM items WHERE id = 1",
                "SELECT label FROM items WHERE id = 2",
            )
            suite = scoring.score_test_suite(
                [first, second],
                "SELECT label FROM items WHERE id = 1",
                "SELECT label FROM items WHERE id = 2",
                keep_distinct=True,
            )
            self.assertTrue(single["agreement"])
            self.assertFalse(suite["agreement"])
            self.assertEqual(suite["databases_checked"], 2)
            self.assertEqual(suite["failure_database"], "second.sqlite")

    def test_spider_test_suite_reports_distinct_sensitive_semantics(self):
        scoring_path = Path(__file__).resolve().parents[3] / "scripts/spider_execution_scoring.py"
        spec = importlib.util.spec_from_file_location("spider_distinct_scoring_for_test", scoring_path)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        scoring = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(scoring)
        sql = "SELECT DISTINCT value, 'DISTINCT', \"DISTINCT\" FROM items -- DISTINCT\n"
        normalized = scoring.remove_distinct_keywords(sql)
        self.assertNotIn("SELECT DISTINCT", normalized)
        self.assertIn("'DISTINCT'", normalized)
        self.assertIn('\"DISTINCT\"', normalized)
        self.assertTrue(normalized.rstrip().endswith("-- DISTINCT"))
        with tempfile.TemporaryDirectory() as temporary:
            database = Path(temporary) / "duplicates.sqlite"
            with closing(sqlite3.connect(database)) as connection:
                connection.execute("CREATE TABLE items (value TEXT)")
                connection.executemany("INSERT INTO items VALUES (?)", [("x",), ("x",)])
                connection.commit()
            strict = scoring.score_test_suite(
                [database], "SELECT value FROM items", "SELECT DISTINCT value FROM items",
                keep_distinct=True,
            )
            upstream = scoring.score_test_suite(
                [database], "SELECT value FROM items", "SELECT DISTINCT value FROM items",
                keep_distinct=False,
            )
            self.assertFalse(strict["agreement"])
            self.assertTrue(upstream["agreement"])

    def test_bird_summary_excludes_llm_infrastructure_from_semantic_denominator(self):
        benchmark_path = Path(__file__).resolve().parents[3] / "scripts/run_bird_benchmark.py"
        spec = importlib.util.spec_from_file_location("bird_benchmark_for_test", benchmark_path)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        benchmark = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(benchmark)
        results = [
            {
                "id": "pass", "index": 0, "difficulty": "simple", "db_id": "db",
                "execution_exact": True, "strict_sequence_match": True,
                "answer_kind": "query", "predicted_sql": "SELECT 1",
                "execution_error": None, "error_category": None,
                "generation_latency_ms": 10,
            },
            {
                "id": "infra", "index": 1, "difficulty": "moderate", "db_id": "db",
                "execution_exact": False, "strict_sequence_match": False,
                "answer_kind": "error", "predicted_sql": "",
                "execution_error": "LLM 服务请求失败（HTTP 402）",
                "error_category": benchmark.INFRASTRUCTURE_ERROR_CATEGORY,
                "generation_latency_ms": 5,
            },
        ]

        summary = benchmark._summary(results)

        self.assertEqual(summary["execution_accuracy"], {
            "passed": 1, "total": 1, "rate": 1.0,
        })
        self.assertEqual(summary["raw_lower_bound_execution_accuracy"]["total"], 2)
        self.assertEqual(summary["coverage"]["infrastructure_failures"], 1)
        self.assertFalse(summary["coverage"]["complete"])

    def test_bird_schema_adapter_loads_dictionary_and_declared_relationships(self):
        benchmark_path = Path(__file__).resolve().parents[3] / "scripts/run_bird_benchmark.py"
        spec = importlib.util.spec_from_file_location("bird_schema_adapter_for_test", benchmark_path)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        benchmark = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(benchmark)
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "users": dc.DBTable(name="users", columns=[
                dc.DBColumn("id", "INTEGER", pk=True),
                dc.DBColumn("status", "TEXT"),
            ]),
            "orders": dc.DBTable(name="orders", columns=[
                dc.DBColumn("user_id", "INTEGER"),
            ]),
        })
        with tempfile.TemporaryDirectory() as temp_dir:
            database_dir = Path(temp_dir) / "db"
            description_dir = database_dir / "database_description"
            description_dir.mkdir(parents=True)
            db_path = database_dir / "db.sqlite"
            db_path.touch()
            # Exercise the official package's deterministic CP1252 fallback.
            (description_dir / "users.csv").write_bytes(
                b"original_column_name,column_name,column_description,value_description\r\n"
                b"status,Account Status,Current user status,Active \x95 enabled\r\n"
            )
            stats = benchmark._apply_column_descriptions(db_path, schema)

        relation_count = benchmark._apply_declared_foreign_keys({
            "table_names_original": ["users", "orders"],
            "column_names_original": [[-1, "*"], [0, "id"], [1, "user_id"]],
            "foreign_keys": [[2, 1]],
        }, schema)

        self.assertEqual(stats, {"files": 1, "columns": 1})
        status = schema.tables["users"].columns[1]
        self.assertEqual(status.semantic_name, "Account Status")
        self.assertIn("enabled", status.value_description)
        self.assertEqual(relation_count, 1)
        user_id = schema.tables["orders"].columns[0]
        self.assertEqual((user_id.fk_table, user_id.fk_column), ("users", "id"))


class OperationGraphTests(unittest.TestCase):
    @staticmethod
    def _independent_schema():
        return dc.SchemaSnapshot(db_path="fixture", tables={
            "customers": dc.DBTable(
                name="customers",
                columns=[dc.DBColumn("id", "INTEGER", pk=True)],
            ),
            "events": dc.DBTable(
                name="events",
                columns=[dc.DBColumn("id", "INTEGER", pk=True)],
            ),
            "audits": dc.DBTable(
                name="audits",
                columns=[dc.DBColumn("id", "INTEGER", pk=True)],
            ),
            "snapshots": dc.DBTable(
                name="snapshots",
                columns=[dc.DBColumn("id", "INTEGER", pk=True)],
            ),
            "alerts": dc.DBTable(
                name="alerts",
                columns=[dc.DBColumn("id", "INTEGER", pk=True)],
            ),
            "jobs": dc.DBTable(
                name="jobs",
                columns=[dc.DBColumn("id", "INTEGER", pk=True)],
            ),
            "archives": dc.DBTable(
                name="archives",
                columns=[dc.DBColumn("id", "INTEGER", pk=True)],
            ),
        })

    def test_planner_builds_read_only_dag_in_stable_topological_order(self):
        graph = dc.OperationGraphPlanner().plan_compose("结合数量和记录内容")
        ordered = dc.OperationGraphValidator().validate(graph)

        self.assertEqual([node.node_id for node in ordered], [
            "query-data", "retrieve-context", "synthesize-answer",
        ])
        self.assertEqual({node.tool for node in graph.nodes}, {"query", "retrieve", "synthesize"})
        self.assertEqual(graph.version, "3.1")
        self.assertTrue(all(node.input_contract and node.output_contract for node in graph.nodes))

    def test_planner_selects_only_query_for_quantitative_question(self):
        graph = dc.OperationGraphPlanner().plan_compose("统计 items 的数量和平均值")

        self.assertEqual([node.tool for node in graph.nodes], ["query", "synthesize"])
        self.assertEqual(graph.nodes[-1].depends_on, ["query-data"])

    def test_planner_selects_only_retrieval_for_context_question(self):
        graph = dc.OperationGraphPlanner().plan_compose("根据备注内容说明主要原因")

        self.assertEqual([node.tool for node in graph.nodes], ["retrieve", "synthesize"])
        self.assertEqual(graph.nodes[-1].depends_on, ["retrieve-context"])

    def test_cross_table_query_adds_relation_preflight(self):
        schema = dc.SchemaSnapshot(db_path="fixture", tables={
            "customers": dc.DBTable(
                name="customers",
                columns=[dc.DBColumn("id", "INTEGER", pk=True)],
            ),
            "orders": dc.DBTable(
                name="orders",
                columns=[
                    dc.DBColumn("id", "INTEGER", pk=True),
                    dc.DBColumn(
                        "customer_id", "INTEGER",
                        fk_table="customers", fk_column="id",
                    ),
                ],
            ),
        })
        graph = dc.OperationGraphPlanner(schema).plan_compose(
            "统计 customers 和 orders 的数量对比",
        )
        ordered = dc.OperationGraphValidator().validate(graph)

        self.assertEqual([node.tool for node in ordered], [
            "inspect_relations", "query", "synthesize",
        ])
        self.assertEqual(graph.nodes[1].depends_on, ["inspect-relations"])

    def test_relation_preflight_blocks_disconnected_cross_table_query(self):
        schema = self._independent_schema()
        nl2sql = mock.Mock()
        rag = mock.Mock()
        graph = dc.OperationGraphPlanner(schema).plan_compose(
            "统计 customers 和 events 的数量对比",
        )
        answer = dc.OperationGraphExecutor(nl2sql, rag, schema=schema).execute(graph)

        self.assertEqual(answer.kind, "error")
        statuses = {node["node_id"]: node["status"] for node in answer.graph["nodes"]}
        self.assertEqual(statuses["inspect-relations"], "failed")
        self.assertEqual(statuses["query-data"], "skipped")
        nl2sql.answer.assert_not_called()

    def test_planner_splits_two_independent_queries_without_relation_preflight(self):
        graph = dc.OperationGraphPlanner(self._independent_schema()).plan_compose(
            "分别统计 customers 和 events 的数量",
        )
        ordered = dc.OperationGraphValidator().validate(graph)

        self.assertEqual(graph.strategy, "deterministic-multi-query")
        self.assertEqual([node.node_id for node in ordered], [
            "query-1", "query-2", "synthesize-answer",
        ])
        self.assertEqual(graph.nodes[0].parameters["allowed_tables"], ["customers"])
        self.assertEqual(graph.nodes[1].parameters["allowed_tables"], ["events"])
        self.assertNotIn("inspect_relations", {node.tool for node in graph.nodes})

    def test_planner_splits_three_independent_queries_in_schema_order(self):
        graph = dc.OperationGraphPlanner(self._independent_schema()).plan_compose(
            "分别统计 customers、events 和 audits 的数量",
        )
        ordered = dc.OperationGraphValidator().validate(graph)

        self.assertEqual(graph.strategy, "deterministic-multi-query")
        self.assertEqual([node.node_id for node in ordered], [
            "query-1", "query-2", "query-3", "synthesize-answer",
        ])
        self.assertEqual(
            [node.parameters["allowed_tables"] for node in graph.nodes if node.tool == "query"],
            [["customers"], ["events"], ["audits"]],
        )
        self.assertNotIn("inspect_relations", {node.tool for node in graph.nodes})

    def test_executor_returns_all_scoped_query_datasets_and_caps_rows(self):
        schema = self._independent_schema()
        nl2sql = mock.Mock()

        def query_answer(question, history=None, allowed_tables=None):
            table = allowed_tables[0]
            return dc.DBAnswer(
                kind="query",
                narrative=f"{table} 查询完成",
                sql=f"SELECT id FROM {table}",
                columns=["id"],
                rows=[[index] for index in range(130)],
            )

        nl2sql.answer.side_effect = query_answer
        graph = dc.OperationGraphPlanner(schema).plan_compose(
            "分别统计 customers 和 events 的数量",
        )
        with mock.patch.object(dc, "_llm_ask_json", return_value={"answer_zh": "两组结果已完成"}):
            answer = dc.OperationGraphExecutor(nl2sql, mock.Mock(), schema=schema).execute(graph)

        self.assertEqual(answer.kind, "compose")
        self.assertEqual([item["label"] for item in answer.datasets], ["customers", "events"])
        self.assertEqual([len(item["rows"]) for item in answer.datasets], [100, 100])
        self.assertEqual(
            [call.kwargs["allowed_tables"] for call in nl2sql.answer.call_args_list],
            [["customers"], ["events"]],
        )

    def test_scoped_nl2sql_rejects_cross_branch_table_reference(self):
        schema = self._independent_schema()
        security = mock.Mock()
        executor = dc.NL2SQLExecutor(security, schema)
        generated = {"sql": "SELECT COUNT(*) FROM events", "summary_zh": "错误越界"}

        with mock.patch.object(dc, "_llm_ask_json", return_value=generated) as generate:
            answer = executor.answer("统计 customers", allowed_tables=["customers"])

        self.assertEqual(answer.kind, "error")
        self.assertIn("越界引用表", answer.error)
        self.assertEqual(generate.call_count, 3)
        security.execute.assert_not_called()

    def test_scoped_nl2sql_accepts_assigned_table_reference(self):
        schema = self._independent_schema()
        security = mock.Mock()
        security.execute.return_value = dc.SQLResult(
            sql="SELECT COUNT(*) AS total FROM customers LIMIT 500",
            columns=["total"],
            rows=[[2]],
            row_count=1,
        )
        executor = dc.NL2SQLExecutor(security, schema)
        generated = {
            "sql": "SELECT COUNT(*) AS total FROM customers",
            "summary_zh": "customers 共 2 条",
        }

        with mock.patch.object(dc, "_llm_ask_json", return_value=generated):
            answer = executor.answer("统计 customers", allowed_tables=["customers"])

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.rows, [[2]])
        security.execute.assert_called_once_with(generated["sql"])

    def test_scoped_nl2sql_rejects_join_cte_and_comma_sources(self):
        schema = self._independent_schema()
        executor = dc.NL2SQLExecutor(mock.Mock(), schema)
        invalid = (
            "SELECT customers.id FROM customers JOIN events ON events.id = customers.id",
            "WITH scoped AS (SELECT id FROM customers) SELECT * FROM scoped",
            "SELECT customers.id FROM customers, events",
        )

        for sql in invalid:
            with self.subTest(sql=sql), self.assertRaises(dc.NL2SQLError):
                executor._validate_allowed_tables(sql, ["customers"])

    def test_conditional_retrieval_skips_when_all_queries_are_empty(self):
        schema = self._independent_schema()
        nl2sql = mock.Mock()
        nl2sql.answer.return_value = dc.DBAnswer(
            kind="query", narrative="没有数据", sql="SELECT id", columns=["id"], rows=[],
        )
        rag = mock.Mock()
        graph = dc.OperationGraphPlanner(schema).plan_compose(
            "分别统计 customers 和 events 的数量，如果有数据再查看记录内容",
        )

        with mock.patch.object(dc, "_llm_ask_json", return_value={"answer_zh": "没有可展开的记录"}):
            answer = dc.OperationGraphExecutor(nl2sql, rag, schema=schema).execute(graph)

        statuses = {node["node_id"]: node["status"] for node in answer.graph["nodes"]}
        self.assertEqual(answer.graph["status"], "completed")
        self.assertEqual(graph.strategy, "deterministic-multi-query-conditional")
        self.assertEqual(statuses["retrieve-context"], "skipped")
        self.assertIn("条件未满足", next(
            node["error"] for node in answer.graph["nodes"] if node["node_id"] == "retrieve-context"
        ))
        rag.answer.assert_not_called()

    def test_conditional_retrieval_treats_zero_count_as_no_data(self):
        schema = self._independent_schema()
        nl2sql = mock.Mock()
        nl2sql.answer.return_value = dc.DBAnswer(
            kind="query", narrative="共 0 条", sql="SELECT COUNT(*)", columns=["count"], rows=[[0]],
        )
        rag = mock.Mock()
        graph = dc.OperationGraphPlanner(schema).plan_compose(
            "分别统计 customers 和 events 的数量，如果有数据再查看记录内容",
        )

        with mock.patch.object(dc, "_llm_ask_json", return_value={"answer_zh": "没有数据"}):
            answer = dc.OperationGraphExecutor(nl2sql, rag, schema=schema).execute(graph)

        self.assertEqual(answer.graph["status"], "completed")
        rag.answer.assert_not_called()

    def test_conditional_retrieval_runs_after_nonempty_query(self):
        schema = self._independent_schema()
        nl2sql = mock.Mock()
        nl2sql.answer.side_effect = [
            dc.DBAnswer(kind="query", narrative="有数据", sql="SELECT 1", columns=["n"], rows=[[1]]),
            dc.DBAnswer(kind="query", narrative="无数据", sql="SELECT 1", columns=["n"], rows=[]),
        ]
        rag = mock.Mock()
        rag.answer.return_value = dc.DBAnswer(
            kind="retrieve", narrative="补充记录", evidence=[{"table": "customers"}],
        )
        graph = dc.OperationGraphPlanner(schema).plan_compose(
            "分别统计 customers 和 events 的数量，如果有数据再查看记录内容",
        )

        with mock.patch.object(dc, "_llm_ask_json", return_value={"answer_zh": "完成条件分析"}):
            answer = dc.OperationGraphExecutor(nl2sql, rag, schema=schema).execute(graph)

        self.assertEqual(answer.graph["status"], "completed")
        rag.answer.assert_called_once()

    def test_independent_word_does_not_bypass_explicit_relation_preflight(self):
        graph = dc.OperationGraphPlanner(self._independent_schema()).plan_compose(
            "分别统计 customers 和 events 的关联数量",
        )

        self.assertEqual(graph.strategy, "deterministic")
        self.assertEqual(graph.nodes[0].tool, "inspect_relations")

    def test_validator_rejects_unknown_branch_condition(self):
        graph = dc.OperationGraphPlanner().plan_compose("统计数量")
        graph.nodes[0].parameters["condition"] = "model_decides"

        with self.assertRaisesRegex(dc.OrchestratorError, "分支条件非法"):
            dc.OperationGraphValidator().validate(graph)

    def test_validator_rejects_scoped_tables_outside_graph_targets(self):
        graph = dc.OperationGraphPlanner(self._independent_schema()).plan_compose(
            "分别统计 customers 和 events 的数量",
        )
        graph.nodes[0].parameters["allowed_tables"] = ["unknown"]

        with self.assertRaisesRegex(dc.OrchestratorError, "表范围越界"):
            dc.OperationGraphValidator().validate(graph)

    def test_validator_rejects_multi_table_or_duplicate_scoped_branches(self):
        for mutation, message in (
            (lambda graph: graph.nodes[0].parameters.update(
                {"allowed_tables": ["customers", "events"]}
            ), "只能绑定一张表"),
            (lambda graph: graph.nodes[1].parameters.update(
                {"allowed_tables": ["customers"], "branch_label": "customers"}
            ), "重复目标表"),
        ):
            with self.subTest(message=message):
                graph = dc.OperationGraphPlanner(self._independent_schema()).plan_compose(
                    "分别统计 customers 和 events 的数量",
                )
                mutation(graph)
                with self.assertRaisesRegex(dc.OrchestratorError, message):
                    dc.OperationGraphValidator().validate(graph)

    def test_executor_rejects_seven_independent_branches_before_tools(self):
        schema = self._independent_schema()
        graph = dc.OperationGraphPlanner(schema).plan_compose(
            "分别统计 customers、events、audits、snapshots、alerts、jobs 和 archives 的数量",
        )
        nl2sql = mock.Mock()
        rag = mock.Mock()

        answer = dc.OperationGraphExecutor(nl2sql, rag, schema=schema).execute(graph)

        self.assertEqual(answer.kind, "error")
        self.assertIn("独立查询分支数量必须在 2 到 6 之间", answer.error)
        nl2sql.answer.assert_not_called()
        rag.answer.assert_not_called()

    def test_validator_rejects_tampered_output_contract(self):
        graph = dc.OperationGraphPlanner().plan_compose("统计数量")
        graph.nodes[0].output_contract = {
            "type": "arbitrary_payload",
            "required": ["summary"],
        }
        with self.assertRaisesRegex(dc.OrchestratorError, "契约非法"):
            dc.OperationGraphValidator().validate(graph)

    def test_validator_rejects_write_tool(self):
        graph = dc.OperationGraph(
            objective="修改数据后总结",
            nodes=[
                dc.OperationGraphNode(node_id="write-data", tool="write"),
                dc.OperationGraphNode(
                    node_id="synthesize-answer",
                    tool="synthesize",
                    depends_on=["write-data"],
                ),
            ],
        )
        with self.assertRaisesRegex(dc.OrchestratorError, "只读白名单"):
            dc.OperationGraphValidator().validate(graph)

    def test_validator_rejects_cycle(self):
        graph = dc.OperationGraph(
            objective="循环依赖测试",
            nodes=[
                dc.OperationGraphNode(
                    node_id="query-data",
                    tool="query",
                    depends_on=["synthesize-answer"],
                ),
                dc.OperationGraphNode(
                    node_id="synthesize-answer",
                    tool="synthesize",
                    depends_on=["query-data"],
                ),
            ],
        )
        with self.assertRaisesRegex(dc.OrchestratorError, "循环依赖"):
            dc.OperationGraphValidator().validate(graph)

    def test_executor_degrades_to_available_branch_and_marks_partial(self):
        nl2sql = mock.Mock()
        nl2sql.answer.return_value = dc.DBAnswer(
            kind="error",
            narrative="查询失败",
            error="query unavailable",
        )
        rag = mock.Mock()
        rag.answer.return_value = dc.DBAnswer(
            kind="retrieve",
            narrative="检索分支可用",
            evidence=[{"table": "items"}],
        )
        graph = dc.OperationGraphPlanner().plan_compose("结合数量和记录内容说明情况")
        executor = dc.OperationGraphExecutor(nl2sql, rag)

        with mock.patch.object(dc, "_llm_ask_json", return_value={"answer_zh": "使用可用分支回答"}):
            answer = executor.execute(graph)

        self.assertEqual(answer.kind, "compose")
        self.assertEqual(answer.graph["status"], "partial")
        statuses = {node["node_id"]: node["status"] for node in answer.graph["nodes"]}
        self.assertEqual(statuses["query-data"], "failed")
        self.assertEqual(statuses["retrieve-context"], "completed")
        self.assertEqual(statuses["synthesize-answer"], "completed")
        self.assertEqual(answer.evidence, [{"table": "items"}])

    def test_three_query_executor_keeps_successes_when_one_branch_fails(self):
        schema = self._independent_schema()
        nl2sql = mock.Mock()
        nl2sql.answer.side_effect = [
            dc.DBAnswer(
                kind="query", narrative="customers 完成", sql="SELECT 1",
                columns=["total"], rows=[[2]],
            ),
            dc.DBAnswer(kind="error", narrative="events 失败", error="query unavailable"),
            dc.DBAnswer(
                kind="query", narrative="audits 完成", sql="SELECT 1",
                columns=["total"], rows=[[4]],
            ),
        ]
        graph = dc.OperationGraphPlanner(schema).plan_compose(
            "分别统计 customers、events 和 audits 的数量",
        )

        with mock.patch.object(dc, "_llm_ask_json", return_value={"answer_zh": "已使用可用分支"}):
            answer = dc.OperationGraphExecutor(nl2sql, mock.Mock(), schema=schema).execute(graph)

        self.assertEqual(answer.kind, "compose")
        self.assertEqual(answer.graph["status"], "partial")
        self.assertEqual([item["label"] for item in answer.datasets], ["customers", "audits"])
        statuses = {node["node_id"]: node["status"] for node in answer.graph["nodes"]}
        self.assertEqual(statuses["query-1"], "completed")
        self.assertEqual(statuses["query-2"], "failed")
        self.assertEqual(statuses["query-3"], "completed")
        self.assertEqual(statuses["synthesize-answer"], "completed")

    def test_invalid_graph_fails_before_any_tool_runs(self):
        nl2sql = mock.Mock()
        rag = mock.Mock()
        graph = dc.OperationGraph(
            objective="非法操作图",
            nodes=[
                dc.OperationGraphNode(node_id="delete-data", tool="write"),
                dc.OperationGraphNode(
                    node_id="synthesize-answer",
                    tool="synthesize",
                    depends_on=["delete-data"],
                ),
            ],
        )
        answer = dc.OperationGraphExecutor(nl2sql, rag).execute(graph)

        self.assertEqual(answer.kind, "error")
        self.assertEqual(answer.graph["status"], "failed")
        nl2sql.answer.assert_not_called()
        rag.answer.assert_not_called()


class SemanticCatalogTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "semantic.db"
        _make_db(self.path, rows=3)
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE items ADD COLUMN created_at TEXT")
            conn.execute("ALTER TABLE items ADD COLUMN updated_at TEXT")
            conn.execute("ALTER TABLE items ADD COLUMN amount REAL DEFAULT 10")
            conn.execute(
                "CREATE TABLE holidays (holiday_date DATE, name TEXT, is_working INTEGER)"
            )
            conn.commit()

    def tearDown(self):
        self.tmp.cleanup()

    def test_catalog_validates_and_resolves_all_supported_semantic_kinds(self):
        schema = dc.SchemaDiscovery(dc.DBConnector(str(self.path)), sample_rows=0).discover()
        catalog = dc.SemanticCatalog(schema, [
            {"kind": "table_alias", "term": "商品表", "table": "items"},
            {"kind": "column_alias", "term": "商品名", "table": "items", "column": "value"},
            {
                "kind": "dimension", "term": "商品分类", "table": "items", "column": "value",
                "filters": [{"column": "value", "operator": "neq", "value": "archived"}],
            },
            {"kind": "time_field", "term": "入库时间", "table": "items", "column": "created_at"},
            {"kind": "enum_value", "term": "默认商品", "table": "items", "column": "value", "value": "v0"},
            {
                "kind": "metric", "term": "商品数", "table": "items", "aggregation": "count",
                "filters": [{"column": "value", "operator": "eq", "value": "v0"}],
            },
            {
                "kind": "ratio_metric", "term": "平均商品金额", "table": "items",
                "formula": {
                    "operator": "divide",
                    "numerator": {
                        "aggregation": "sum", "column": "amount",
                        "filters": [{"column": "amount", "operator": "gte", "value": 5}],
                    },
                    "denominator": {"aggregation": "count", "column": ""},
                    "scale": 1,
                    "zero_division": "null",
                },
            },
            {
                "kind": "business_calendar", "term": "公司日历", "table": "items",
                "column": "created_at",
                "calendar": {
                    "fiscal_year_start_month": 4,
                    "fiscal_year_start_day": 1,
                    "fiscal_year_label": "start_year",
                    "timezone": "Asia/Shanghai",
                    "week_start": 1,
                    "weekend_days": [6, 7],
                    "holiday_table": "holidays",
                    "holiday_date_column": "holiday_date",
                    "holiday_name_column": "name",
                    "working_override_column": "is_working",
                },
            },
        ], strict=True)

        resolution = catalog.resolve("按公司日历的 2026 财年、商品分类和入库时间统计商品数及平均商品金额，并找到默认商品的商品名，数据来自商品表")

        self.assertEqual(len(resolution.matches), 8)
        self.assertEqual(resolution.as_dict()["version"], "2.8")
        self.assertIn("count(表 items)", resolution.resolved_question)
        self.assertIn("items.value 的值 \"v0\"", resolution.resolved_question)
        self.assertIn("字段 items.value", resolution.resolved_question)
        self.assertIn("分组字段 items.value", resolution.resolved_question)
        self.assertIn('维度固定过滤仅按 AND 组合：items.value 不等于 "archived"', resolution.resolved_question)
        self.assertIn("时间字段 items.created_at", resolution.resolved_question)
        self.assertIn("sum(items.amount) [过滤：items.amount 大于等于 5]", resolution.resolved_question)
        self.assertIn("/ (count(表 items))", resolution.resolved_question)
        self.assertIn('items.value 等于 "v0"', resolution.resolved_question)
        self.assertIn("items.amount 大于等于 5", resolution.resolved_question)
        self.assertIn("财年从 04-01 开始", resolution.resolved_question)
        self.assertIn("holidays.holiday_date", resolution.resolved_question)
        self.assertIn("未绑定节假日表时不得猜测法定节假日", resolution.resolved_question)
        self.assertIn("FILTER 规则", catalog.prompt_context())
        self.assertIn('DIMENSION items.value; filters=AND(items.value 不等于 "archived")', catalog.prompt_context())
        self.assertIn("BUSINESS_CALENDAR target=items.created_at", catalog.prompt_context())
        self.assertIn("zero_division=NULL", catalog.prompt_context())

    def test_catalog_rejects_unknown_schema_targets_and_freeform_aggregations(self):
        schema = dc.SchemaDiscovery(dc.DBConnector(str(self.path)), sample_rows=0).discover()
        with self.assertRaisesRegex(ValueError, "目标表不存在"):
            dc.SemanticCatalog(schema, [
                {"kind": "table_alias", "term": "订单", "table": "orders"},
            ], strict=True)
        with self.assertRaisesRegex(ValueError, "指标聚合只支持"):
            dc.SemanticCatalog(schema, [
                {"kind": "metric", "term": "危险指标", "table": "items", "aggregation": "select *"},
            ], strict=True)
        with self.assertRaisesRegex(ValueError, "时间字段必须"):
            dc.SemanticCatalog(schema, [
                {"kind": "time_field", "term": "错误时间", "table": "items", "column": "value"},
            ], strict=True)
        with self.assertRaisesRegex(ValueError, "只支持分子除以分母"):
            dc.SemanticCatalog(schema, [{
                "kind": "ratio_metric", "term": "危险公式", "table": "items",
                "formula": {
                    "operator": "sql", "numerator": {"aggregation": "sum", "column": "amount"},
                    "denominator": {"aggregation": "count"}, "scale": 1,
                },
            }], strict=True)
        with self.assertRaisesRegex(ValueError, "必须是数值类型"):
            dc.SemanticCatalog(schema, [{
                "kind": "ratio_metric", "term": "错误比率", "table": "items",
                "formula": {
                    "operator": "divide", "numerator": {"aggregation": "sum", "column": "value"},
                    "denominator": {"aggregation": "count"}, "scale": 1,
                },
            }], strict=True)

    def test_dimension_hierarchy_exposes_ordered_same_table_path(self):
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE items ADD COLUMN category TEXT")
            conn.commit()
        schema = dc.SchemaDiscovery(dc.DBConnector(str(self.path)), sample_rows=0).discover()
        catalog = dc.SemanticCatalog(schema, [
            {
                "kind": "dimension", "term": "商品大类", "table": "items",
                "column": "value", "hierarchy": {"name": "商品层级", "level": 1},
            },
            {
                "kind": "dimension", "term": "商品明细", "table": "items",
                "column": "category", "hierarchy": {"name": "商品层级", "level": 2},
            },
        ], strict=True)

        resolution = catalog.resolve("按商品明细统计商品数量")
        hierarchy = resolution.matches[0]
        self.assertEqual(hierarchy["hierarchy"], {"name": "商品层级", "level": 2})
        self.assertEqual(
            [(item["level"], item["term"], item["column"]) for item in hierarchy["hierarchy_path"]],
            [(1, "商品大类", "value"), (2, "商品明细", "category")],
        )
        self.assertIn("完整路径 L1 商品大类(items.value) > L2 商品明细(items.category)", resolution.resolved_question)
        self.assertIn("hierarchy=商品层级; level=1", catalog.prompt_context())
        self.assertIn("不得据此发明跨表 JOIN", catalog.prompt_context())

    def test_hierarchy_and_default_grain_reject_conflicting_definitions(self):
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE items ADD COLUMN category TEXT")
            conn.commit()
        schema = dc.SchemaDiscovery(dc.DBConnector(str(self.path)), sample_rows=0).discover()
        with self.assertRaisesRegex(ValueError, "第 1 级已经"):
            dc.SemanticCatalog(schema, [
                {
                    "kind": "dimension", "term": "商品大类", "table": "items",
                    "column": "value", "hierarchy": {"name": "商品层级", "level": 1},
                },
                {
                    "kind": "dimension", "term": "商品明细", "table": "items",
                    "column": "category", "hierarchy": {"name": "商品层级", "level": 1},
                },
            ], strict=True)
        with self.assertRaisesRegex(ValueError, "只能绑定同一张表"):
            dc.SemanticCatalog(schema, [
                {
                    "kind": "dimension", "term": "商品大类", "table": "items",
                    "column": "value", "hierarchy": {"name": "公共层级", "level": 1},
                },
                {
                    "kind": "dimension", "term": "节假日名称", "table": "holidays",
                    "column": "name", "hierarchy": {"name": "公共层级", "level": 2},
                },
            ], strict=True)
        with self.assertRaisesRegex(ValueError, "默认粒度"):
            dc.SemanticCatalog(schema, [{
                "kind": "time_field", "term": "错误粒度", "table": "items",
                "column": "created_at", "default_grain": "hour",
            }], strict=True)
        with self.assertRaisesRegex(ValueError, "冲突的默认粒度"):
            dc.SemanticCatalog(schema, [
                {
                    "kind": "time_field", "term": "入库时间", "table": "items",
                    "column": "created_at", "default_grain": "month",
                },
                {
                    "kind": "time_field", "term": "创建日期", "table": "items",
                    "column": "created_at", "default_grain": "week",
                },
            ], strict=True)

    def test_business_calendar_rejects_invalid_dates_timezone_and_schema_bindings(self):
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE items ADD COLUMN occurred_at TIMESTAMP")
            conn.commit()
        schema = dc.SchemaDiscovery(dc.DBConnector(str(self.path)), sample_rows=0).discover()
        base = {
            "kind": "business_calendar", "term": "公司日历", "table": "items",
            "column": "created_at",
        }
        with self.assertRaisesRegex(ValueError, "fiscal_year_start_day"):
            dc.SemanticCatalog(schema, [{
                **base,
                "calendar": {
                    "fiscal_year_start_month": 2, "fiscal_year_start_day": 30,
                    "timezone": "Asia/Shanghai", "week_start": 1, "weekend_days": [6, 7],
                },
            }], strict=True)
        with self.assertRaisesRegex(ValueError, "IANA"):
            dc.SemanticCatalog(schema, [{
                **base,
                "calendar": {"timezone": "UTC+8; DROP", "weekend_days": [6, 7]},
            }], strict=True)
        with self.assertRaisesRegex(ValueError, "fiscal_year_label"):
            dc.SemanticCatalog(schema, [{
                **base,
                "calendar": {
                    "timezone": "UTC", "weekend_days": [6, 7],
                    "fiscal_year_label": "guess",
                },
            }], strict=True)
        with self.assertRaisesRegex(ValueError, "节假日日期字段"):
            dc.SemanticCatalog(schema, [{
                **base,
                "calendar": {
                    "timezone": "UTC", "weekend_days": [6, 7],
                    "holiday_table": "holidays", "holiday_date_column": "name",
                },
            }], strict=True)
        with self.assertRaisesRegex(ValueError, "工作日覆盖字段"):
            dc.SemanticCatalog(schema, [{
                **base,
                "calendar": {
                    "timezone": "UTC", "weekend_days": [6, 7],
                    "holiday_table": "holidays", "holiday_date_column": "holiday_date",
                    "working_override_column": "name",
                },
            }], strict=True)
        with self.assertRaisesRegex(ValueError, "declared_date"):
            dc.SemanticCatalog(schema, [{
                **base,
                "calendar": {
                    "timezone": "UTC", "weekend_days": [6, 7],
                    "storage_basis": "declared_date",
                },
            }], strict=True)
        timestamp_base = {
            **base, "term": "UTC 日历", "column": "occurred_at",
        }
        with self.assertRaisesRegex(ValueError, "business_utc_offset_minutes"):
            dc.SemanticCatalog(schema, [{
                **timestamp_base,
                "calendar": {
                    "timezone": "Asia/Shanghai", "weekend_days": [6, 7],
                    "storage_basis": "utc_datetime",
                },
            }], strict=True)
        with self.assertRaisesRegex(ValueError, "固定偏移必须为 0"):
            dc.SemanticCatalog(schema, [{
                **timestamp_base,
                "calendar": {
                    "timezone": "UTC", "weekend_days": [6, 7],
                    "storage_basis": "utc_datetime", "business_utc_offset_minutes": 60,
                },
            }], strict=True)

    def test_business_calendar_rejects_invalid_iana_runtime_contracts(self):
        with closing(sqlite3.connect(self.path)) as conn:
            conn.execute("ALTER TABLE items ADD COLUMN occurred_at TIMESTAMP")
            conn.commit()
        schema = dc.SchemaDiscovery(dc.DBConnector(str(self.path)), sample_rows=0).discover()
        base = {
            "kind": "business_calendar", "term": "动态日历", "table": "items",
            "column": "occurred_at",
        }

        def calendar(**overrides):
            value = {
                "timezone": "America/New_York", "weekend_days": [6, 7],
                "storage_basis": "utc_datetime", "timezone_conversion": "iana_tzdata",
            }
            value.update(overrides)
            return value

        with self.assertRaisesRegex(ValueError, "IANA 时区不存在"):
            dc.SemanticCatalog(schema, [{**base, "calendar": calendar(timezone="America/Not_Real")}], strict=True)
        with self.assertRaisesRegex(ValueError, "tzdata_version"):
            dc.SemanticCatalog(schema, [{**base, "calendar": calendar(tzdata_version="2025.2")}], strict=True)
        with self.assertRaisesRegex(ValueError, "不能同时设置"):
            dc.SemanticCatalog(schema, [{
                **base, "calendar": calendar(business_utc_offset_minutes=-300),
            }], strict=True)

    def test_metric_filters_reject_freeform_or_type_unsafe_conditions(self):
        schema = dc.SchemaDiscovery(dc.DBConnector(str(self.path)), sample_rows=0).discover()
        with self.assertRaisesRegex(ValueError, "操作符只支持"):
            dc.SemanticCatalog(schema, [{
                "kind": "metric", "term": "危险过滤", "table": "items",
                "aggregation": "count",
                "filters": [{"column": "value", "operator": "sql", "value": "1=1"}],
            }], strict=True)

        with self.assertRaisesRegex(ValueError, "范围比较只支持"):
            dc.SemanticCatalog(schema, [{
                "kind": "metric", "term": "文本范围", "table": "items",
                "aggregation": "count",
                "filters": [{"column": "value", "operator": "gt", "value": "v0"}],
            }], strict=True)
        with self.assertRaisesRegex(ValueError, "过滤值必须是数字"):
            dc.SemanticCatalog(schema, [{
                "kind": "metric", "term": "金额过滤", "table": "items",
                "aggregation": "sum", "column": "amount",
                "filters": [{"column": "amount", "operator": "in", "value": ["10"]}],
            }], strict=True)
        with self.assertRaisesRegex(ValueError, "不能超过 4 条"):
            dc.SemanticCatalog(schema, [{
                "kind": "ratio_metric", "term": "过滤过多", "table": "items",
                "formula": {
                    "operator": "divide",
                    "numerator": {
                        "aggregation": "count",
                        "filters": [
                            {"column": "value", "operator": "eq", "value": f"v{i}"}
                            for i in range(5)
                        ],
                    },
                    "denominator": {"aggregation": "count"},
                    "scale": 1,
                },
            }], strict=True)

    def test_dimension_filters_reject_cross_table_or_unbounded_conditions(self):
        schema = dc.SchemaDiscovery(dc.DBConnector(str(self.path)), sample_rows=0).discover()
        with self.assertRaisesRegex(ValueError, "不存在于表 items"):
            dc.SemanticCatalog(schema, [{
                "kind": "dimension", "term": "错误维度", "table": "items",
                "column": "value",
                "filters": [{"column": "holiday_date", "operator": "eq", "value": "2026-01-01"}],
            }], strict=True)
        with self.assertRaisesRegex(ValueError, "维度过滤条件不能超过 4 条"):
            dc.SemanticCatalog(schema, [{
                "kind": "dimension", "term": "过度过滤维度", "table": "items",
                "column": "value",
                "filters": [
                    {"column": "value", "operator": "eq", "value": f"v{i}"}
                    for i in range(5)
                ],
            }], strict=True)

    def test_controlled_ratio_metric_bypasses_derived_metric_clarification(self):
        agent = dc.DBQuillAgent(
            db_path=str(self.path),
            sample_rows=0,
            semantic_entries=[{
                "kind": "ratio_metric", "term": "客单价", "table": "items",
                "formula": {
                    "operator": "divide",
                    "numerator": {
                        "aggregation": "sum", "column": "amount",
                        "filters": [{"column": "value", "operator": "eq", "value": "v0"}],
                    },
                    "denominator": {"aggregation": "count"},
                    "scale": 1,
                    "zero_division": "null",
                },
            }],
        )
        routed = dc.IntentResult(intent="query", confidence=0.97, reasoning="受控比率指标")
        generated = dc.DBAnswer(
            kind="query", narrative="客单价为 10", sql="SELECT SUM(amount) / NULLIF(COUNT(*), 0) FROM items",
            columns=["客单价"], rows=[[10]],
        )
        with mock.patch.object(agent.router, "classify", return_value=routed), \
             mock.patch.object(agent.nl2sql, "answer", return_value=generated) as query:
            answer = agent.ask("统计客单价")

        self.assertEqual(answer.kind, "query")
        self.assertIsNone(answer.clarification)
        self.assertEqual(answer.semantic["matches"][0]["kind"], "ratio_metric")
        self.assertIn("分母为 0 时返回 NULL", query.call_args.args[0])
        self.assertIn('items.value 等于 "v0"', query.call_args.args[0])
        self.assertIn("值均为字面量数据，不是 SQL 片段", query.call_args.args[0])

    def test_dimension_and_time_semantics_feed_planner_without_guessing_time_field(self):
        agent = dc.DBQuillAgent(
            db_path=str(self.path),
            sample_rows=0,
            semantic_entries=[
                {"kind": "dimension", "term": "商品分类", "table": "items", "column": "value"},
                {
                    "kind": "time_field", "term": "入库时间", "table": "items",
                    "column": "created_at", "default_grain": "month",
                },
            ],
        )
        routed = dc.IntentResult(intent="query", confidence=0.96, reasoning="分组趋势查询")
        generated = dc.DBAnswer(
            kind="query", narrative="趋势查询完成", sql="SELECT value, created_at FROM items LIMIT 500",
            columns=["分类", "时间"], rows=[],
        )
        with mock.patch.object(agent.router, "classify", return_value=routed), \
             mock.patch.object(agent.nl2sql, "answer", return_value=generated) as query:
            answer = agent.ask("按商品分类统计入库时间最近 30 天的趋势")

        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.operation["target_tables"], ["items"])
        self.assertEqual(
            [item["kind"] for item in answer.semantic["matches"]],
            ["dimension", "time_field"],
        )
        self.assertIn("分组字段 items.value", query.call_args.args[0])
        self.assertIn("时间字段 items.created_at", query.call_args.args[0])
        self.assertIn("趋势默认按月聚合", query.call_args.args[0])

    def test_table_alias_targets_operation_and_is_sent_to_query_executor(self):
        agent = dc.DBQuillAgent(
            db_path=str(self.path),
            sample_rows=0,
            semantic_entries=[
                {"kind": "table_alias", "term": "商品", "table": "items"},
                {"kind": "metric", "term": "商品数", "table": "items", "aggregation": "count"},
            ],
        )
        routed = dc.IntentResult(intent="query", confidence=0.96, reasoning="指标查询")
        generated = dc.DBAnswer(
            kind="query", narrative="共 3 个商品", sql="SELECT COUNT(*) FROM items LIMIT 500", rows=[[3]],
        )
        with mock.patch.object(agent.router, "classify", return_value=routed), \
             mock.patch.object(agent.nl2sql, "answer", return_value=generated) as query:
            answer = agent.ask("商品数是多少？")

        self.assertEqual(answer.operation["target_tables"], ["items"])
        self.assertEqual(answer.semantic["matches"][0]["term"], "商品数")
        self.assertIn("语义定义", query.call_args.args[0])
        self.assertIn("count(表 items)", query.call_args.args[0])

    def test_write_alias_still_requires_filter_and_confirmation(self):
        agent = dc.DBQuillAgent(
            db_path=str(self.path),
            sample_rows=0,
            semantic_entries=[
                {"kind": "table_alias", "term": "商品", "table": "items"},
                {"kind": "column_alias", "term": "商品名", "table": "items", "column": "value"},
            ],
        )
        routed = dc.IntentResult(intent="write", confidence=0.97, reasoning="修改数据")
        with mock.patch.object(agent.router, "classify", return_value=routed), \
             mock.patch.object(agent.write_executor, "prepare", side_effect=AssertionError("should clarify")):
            clarification = agent.ask("把商品的商品名改为 ok")
        self.assertEqual(clarification.kind, "clarification")
        self.assertEqual(clarification.clarification["missing"], "filter_condition")
        self.assertEqual([item["term"] for item in clarification.semantic["matches"]], ["商品名", "商品"])

        pending = dc.DBAnswer(
            kind="write_pending",
            narrative="将修改一条记录",
            sql="UPDATE items SET value='ok' WHERE id=1",
            confirm_id="semantic-write",
            write={"kind": "UPDATE", "table": "items", "preview": {}},
        )
        with mock.patch.object(agent.router, "classify", return_value=routed), \
             mock.patch.object(agent.write_executor, "prepare", return_value=pending):
            answer = agent.ask("把商品中 id=1 的商品名改为 ok")
        self.assertEqual(answer.kind, "write_pending")
        self.assertTrue(answer.operation["requires_confirmation"])
        self.assertEqual(answer.operation["target_tables"], ["items"])


class SemanticStoreTests(unittest.TestCase):
    def test_semantic_entry_persists_by_database_key_and_releases_file(self):
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(db_semantic_store, "_DATA_DIR", Path(tmp)), \
             mock.patch.object(db_semantic_store, "_DB_PATH", Path(tmp) / "semantics.db"):
            db_semantic_store.init_db()
            saved = db_semantic_store.upsert_entry("sqlite:test", {
                "kind": "table_alias", "term": "商品", "table": "items",
                "column": "", "value": None, "aggregation": "", "description": "商品主表",
            })
            self.assertEqual(db_semantic_store.list_entries("sqlite:test")[0]["term"], "商品")
            self.assertEqual(db_semantic_store.list_entries("sqlite:other"), [])
            self.assertTrue(db_semantic_store.delete_entry("sqlite:test", saved["id"]))
            path = Path(tmp) / "semantics.db"
            path.unlink()
            self.assertFalse(path.exists())

    def test_ratio_metric_formula_round_trips_as_structured_json(self):
        formula = {
            "operator": "divide",
            "numerator": {
                "aggregation": "sum", "column": "amount",
                "filters": [{"column": "status", "operator": "eq", "value": "paid"}],
            },
            "denominator": {"aggregation": "count", "column": ""},
            "scale": 100,
            "zero_division": "null",
        }
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(db_semantic_store, "_DATA_DIR", Path(tmp)), \
             mock.patch.object(db_semantic_store, "_DB_PATH", Path(tmp) / "semantics.db"):
            db_semantic_store.init_db()
            saved = db_semantic_store.upsert_entry("sqlite:ratio", {
                "kind": "ratio_metric", "term": "金额率", "table": "items",
                "column": "", "value": None, "aggregation": "", "formula": formula,
                "description": "受控比率",
            })
            self.assertEqual(saved["formula"], formula)
            self.assertIsNone(saved["value"])

    def test_metric_filters_round_trip_without_becoming_sql(self):
        filters = [
            {"column": "status", "operator": "in", "value": ["paid", "settled"]},
            {"column": "deleted_at", "operator": "is_null", "value": None},
        ]
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(db_semantic_store, "_DATA_DIR", Path(tmp)), \
             mock.patch.object(db_semantic_store, "_DB_PATH", Path(tmp) / "semantics.db"):
            db_semantic_store.init_db()
            saved = db_semantic_store.upsert_entry("sqlite:metric-filter", {
                "kind": "metric", "term": "有效成交额", "table": "orders",
                "column": "amount", "aggregation": "sum", "filters": filters,
            })
            self.assertEqual(saved["filters"], filters)
            self.assertIsNone(saved["formula"])
            self.assertIsNone(saved["value"])

    def test_business_calendar_round_trips_as_structured_json(self):
        calendar = {
            "fiscal_year_start_month": 4, "fiscal_year_start_day": 1,
            "timezone": "Asia/Shanghai", "week_start": 1, "weekend_days": [6, 7],
            "holiday_table": "holidays", "holiday_date_column": "holiday_date",
            "holiday_name_column": "name", "working_override_column": "is_working",
        }
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(db_semantic_store, "_DATA_DIR", Path(tmp)), \
             mock.patch.object(db_semantic_store, "_DB_PATH", Path(tmp) / "semantics.db"):
            db_semantic_store.init_db()
            saved = db_semantic_store.upsert_entry("sqlite:calendar", {
                "kind": "business_calendar", "term": "公司日历", "table": "orders",
                "column": "created_at", "calendar": calendar,
            })
            self.assertEqual(saved["calendar"], calendar)
            self.assertIsNone(saved["formula"])
            self.assertEqual(saved["filters"], [])

    def test_dimension_hierarchy_and_time_default_grain_round_trip(self):
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(db_semantic_store, "_DATA_DIR", Path(tmp)), \
             mock.patch.object(db_semantic_store, "_DB_PATH", Path(tmp) / "semantics.db"):
            db_semantic_store.init_db()
            dimension = db_semantic_store.upsert_entry("sqlite:hierarchy", {
                "kind": "dimension", "term": "客户区域", "table": "customers",
                "column": "region", "hierarchy": {"name": "客户地域", "level": 1},
                "filters": [{"column": "status", "operator": "eq", "value": "active"}],
            })
            time_field = db_semantic_store.upsert_entry("sqlite:hierarchy", {
                "kind": "time_field", "term": "下单时间", "table": "orders",
                "column": "created_at", "default_grain": "month",
            })
            self.assertEqual(
                dimension["hierarchy"], {"name": "客户地域", "level": 1},
            )
            self.assertEqual(dimension["filters"], [
                {"column": "status", "operator": "eq", "value": "active"},
            ])
            self.assertEqual(time_field["default_grain"], "month")
            by_term = {
                item["term"]: item
                for item in db_semantic_store.list_entries("sqlite:hierarchy")
            }
            self.assertEqual(by_term["客户区域"]["hierarchy"]["level"], 1)
            self.assertEqual(by_term["客户区域"]["filters"][0]["value"], "active")
            self.assertEqual(by_term["下单时间"]["default_grain"], "month")

    def test_legacy_dimension_hierarchy_payload_remains_readable(self):
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(db_semantic_store, "_DATA_DIR", Path(tmp)), \
             mock.patch.object(db_semantic_store, "_DB_PATH", Path(tmp) / "semantics.db"):
            db_semantic_store.init_db()
            with closing(sqlite3.connect(db_semantic_store._DB_PATH)) as conn:
                conn.execute(
                    "INSERT INTO semantic_entries ("
                    "id, database_key, kind, term, term_key, table_name, column_name, "
                    "value_json, aggregation, description, created_at, updated_at"
                    ") VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        "legacy-dimension", "sqlite:legacy", "dimension", "客户区域", "客户区域",
                        "customers", "region", '{"level":1,"name":"客户地域"}', "", "", 1, 1,
                    ),
                )
                conn.commit()
            entry = db_semantic_store.list_entries("sqlite:legacy")[0]
            self.assertEqual(entry["hierarchy"], {"level": 1, "name": "客户地域"})
            self.assertEqual(entry["filters"], [])

    def test_atomic_import_preserves_identity_and_rejects_stale_revision(self):
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(db_semantic_store, "_DATA_DIR", Path(tmp)), \
             mock.patch.object(db_semantic_store, "_DB_PATH", Path(tmp) / "semantics.db"):
            db_semantic_store.init_db()
            original = db_semantic_store.upsert_entry("sqlite:import", {
                "kind": "table_alias", "term": "商品", "table": "items",
                "description": "旧定义",
            })
            _, revision = db_semantic_store.list_entries_with_revision("sqlite:import")
            saved = db_semantic_store.import_entries("sqlite:import", [
                {
                    "kind": "table_alias", "term": "商品", "table": "items",
                    "description": "新定义",
                },
                {
                    "kind": "dimension", "term": "商品分类", "table": "items",
                    "column": "value", "description": "新增定义",
                },
            ], revision)
            self.assertEqual(saved[0]["id"], original["id"])
            self.assertEqual(saved[0]["description"], "新定义")
            self.assertEqual(len(db_semantic_store.list_entries("sqlite:import")), 2)

            _, stale_revision = db_semantic_store.list_entries_with_revision("sqlite:import")
            db_semantic_store.upsert_entry("sqlite:import", {
                "kind": "column_alias", "term": "中途变更", "table": "items",
                "column": "value",
            })
            with self.assertRaisesRegex(ValueError, "预检后已发生变化"):
                db_semantic_store.import_entries("sqlite:import", [{
                    "kind": "table_alias", "term": "不得写入", "table": "items",
                }], stale_revision)
            terms = {item["term"] for item in db_semantic_store.list_entries("sqlite:import")}
            self.assertNotIn("不得写入", terms)


class AuditLedgerTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.patch_data = mock.patch.object(
            db_audit_store, "_DATA_DIR", Path(self.tmp.name),
        )
        self.patch_path = mock.patch.object(
            db_audit_store, "_DB_PATH", Path(self.tmp.name) / "audit.db",
        )
        self.patch_data.start()
        self.patch_path.start()
        db_audit_store.init_db()

    def tearDown(self):
        self.patch_path.stop()
        self.patch_data.stop()
        self.tmp.cleanup()

    def test_brand_formats_emit_dbquill_and_accept_v01_identifiers(self):
        self.assertEqual(db_audit_store.EXPORT_FORMAT, "dbquill-audit-ledger")
        self.assertTrue(db_audit_store._format_supported(
            "dbagent-audit-backup", db_audit_store.BACKUP_FORMAT,
        ))
        self.assertFalse(db_audit_store._format_supported(
            "unrelated-audit-backup", db_audit_store.BACKUP_FORMAT,
        ))

    def test_append_filter_and_verify_hash_chain(self):
        first = db_audit_store.append_event(
            category="nl_operation", action="query", outcome="succeeded",
            summary="自然语言数据库操作", database_key="sqlite:first",
            details={
                "question_sha256": db_audit_store.sha256_text("问题一"),
                "question_length": 3,
                "result_rows": 2,
            },
        )
        second = db_audit_store.append_event(
            category="semantic_change", action="upsert", outcome="succeeded",
            summary="语义定义已保存", database_key="sqlite:second",
            details={"semantic_kind": "table_alias", "semantic_ref": "a" * 16},
        )
        self.assertEqual(second["previous_hash"], first["event_hash"])
        self.assertTrue(db_audit_store.verify_chain()["ok"])
        filtered = db_audit_store.list_events(
            database_key="sqlite:first", category="nl_operation",
        )
        self.assertEqual([event["sequence"] for event in filtered], [1])

    def test_payload_whitelist_prevents_raw_question_sql_and_credentials(self):
        raw_question = "列出用户，password=top-secret"
        raw_sql = "SELECT * FROM users"
        db_audit_store.append_event(
            category="nl_operation", action="query", outcome="succeeded",
            summary="自然语言数据库操作", database_key="sqlite:safe",
            details={
                "question_sha256": db_audit_store.sha256_text(raw_question),
                "question_length": len(raw_question),
                "sql_sha256": db_audit_store.sha256_text(raw_sql),
            },
        )
        persisted = Path(db_audit_store._DB_PATH).read_bytes()
        self.assertNotIn(raw_question.encode("utf-8"), persisted)
        self.assertNotIn(raw_sql.encode("utf-8"), persisted)
        self.assertNotIn(b"top-secret", persisted)
        with self.assertRaisesRegex(ValueError, "未允许字段"):
            db_audit_store.append_event(
                category="nl_operation", action="query", outcome="succeeded",
                summary="自然语言数据库操作", details={"question": raw_question},
            )

    def test_tampering_is_detected(self):
        db_audit_store.append_event(
            category="system", action="startup", outcome="succeeded",
            summary="系统启动",
        )
        with closing(sqlite3.connect(db_audit_store._DB_PATH)) as conn:
            conn.execute("UPDATE audit_events SET summary = ? WHERE sequence = 1", ("被修改",))
            conn.commit()
        integrity = db_audit_store.verify_chain()
        self.assertFalse(integrity["ok"])
        self.assertEqual(integrity["error_sequence"], 1)

    def test_reconciliation_reports_and_closes_unresolved_operations(self):
        correlation = "operation-correlation"
        db_audit_store.append_event(
            category="semantic_change", action="upsert", outcome="approved",
            summary="语义定义变更已通过校验", correlation_id=correlation,
            details={"semantic_ref": "a" * 16},
        )
        unresolved = db_audit_store.reconciliation_status()
        self.assertTrue(unresolved["ok"])
        self.assertEqual(unresolved["unresolved_count"], 1)
        self.assertEqual(unresolved["by_category"], {"semantic_change": 1})

        db_audit_store.append_event(
            category="semantic_change", action="upsert", outcome="succeeded",
            summary="语义定义已保存", correlation_id=correlation,
            details={"semantic_ref": "a" * 16},
        )
        self.assertEqual(db_audit_store.reconciliation_status()["unresolved_count"], 0)

        run_id = "scheduled-run-123456"
        db_audit_store.append_event(
            category="schedule_execution", action="run", outcome="pending",
            summary="定时只读任务开始执行", actor="scheduler",
            run_id=run_id, correlation_id=run_id,
            details={"schedule_ref": "b" * 16, "task_type": "sql", "trigger": "scheduled"},
        )
        self.assertEqual(db_audit_store.reconciliation_status()["unresolved_count"], 1)
        db_audit_store.append_event(
            category="schedule_execution", action="run", outcome="failed",
            summary="定时只读任务执行完成", actor="scheduler",
            run_id=run_id, correlation_id=run_id,
            details={"schedule_ref": "b" * 16, "task_type": "sql", "trigger": "scheduled"},
        )
        self.assertEqual(db_audit_store.reconciliation_status()["unresolved_count"], 0)

    def test_manual_pending_disposition_is_append_only_scoped_and_not_atomic_fact(self):
        pending = db_audit_store.append_event(
            category="semantic_change", action="upsert", outcome="approved",
            summary="semantic change approved", correlation_id="manual-resolution",
            database_key="sqlite:first",
            details={"semantic_ref": "a" * 16},
        )
        evidence_text = "ticket-123 contains operator notes"
        evidence_hash = db_audit_store.sha256_text(evidence_text)

        db_audit_store.append_event(
            category="audit_control", action="resolve_pending", outcome="succeeded",
            summary="invalid manual disposition", actor="local_admin", risk="high",
            correlation_id="manual-resolution",
            details={
                "pending_sequence": pending["sequence"],
                "disposition": "not_allowed",
                "evidence_sha256": evidence_hash,
            },
            _database_ref_override=pending["database_ref"],
        )
        self.assertEqual(db_audit_store.reconciliation_status()["unresolved_count"], 1)

        with self.assertRaises(KeyError):
            db_audit_store.resolve_pending_event(
                pending["sequence"], disposition="verified_no_change",
                evidence_sha256=evidence_hash,
                expected_database_key="sqlite:other",
            )
        resolution = db_audit_store.resolve_pending_event(
            pending["sequence"], disposition="verified_no_change",
            evidence_sha256=evidence_hash,
            expected_database_key="sqlite:first",
        )
        self.assertEqual(resolution["pending_sequence"], pending["sequence"])
        self.assertIn("不等于", resolution["warning"])
        status = db_audit_store.reconciliation_status()
        self.assertEqual(status["unresolved_count"], 0)
        self.assertEqual(status["manually_resolved_count"], 1)
        events = db_audit_store.list_events(limit=10)
        self.assertEqual(events[0]["category"], "audit_control")
        self.assertEqual(events[0]["database_ref"], pending["database_ref"])
        self.assertEqual(events[0]["details"]["evidence_sha256"], evidence_hash)
        self.assertNotIn(evidence_text.encode("utf-8"), db_audit_store._DB_PATH.read_bytes())
        with self.assertRaises(KeyError):
            db_audit_store.resolve_pending_event(
                pending["sequence"], disposition="verified_no_change",
                evidence_sha256=evidence_hash,
            )
        with self.assertRaises(ValueError):
            db_audit_store.resolve_pending_event(
                999, disposition="accepted_risk", evidence_sha256=evidence_hash,
            )

    def test_external_archive_is_non_destructive_and_detects_tamper_and_prefix_drift(self):
        for index in range(1, 4):
            db_audit_store.append_event(
                category="system", action=f"event_{index}", outcome="succeeded",
                summary=f"event {index}",
            )
        with tempfile.TemporaryDirectory() as external:
            archive = db_audit_store.create_external_archive(
                external, through_sequence=2,
            )
            archive_path = Path(archive["archive_file"])
            self.assertTrue(archive["valid"])
            self.assertEqual(archive["event_count"], 2)
            self.assertEqual(archive["events_since_archive"], 1)
            self.assertFalse(archive["destructive_action"])
            self.assertEqual(db_audit_store.verify_chain()["count"], 3)
            self.assertEqual(
                db_audit_store.verify_external_archive(archive_path)["head_hash"],
                archive["head_hash"],
            )

            original = json.loads(archive_path.read_text(encoding="utf-8"))
            tampered = json.loads(json.dumps(original))
            tampered["events"][0]["summary"] = "tampered"
            archive_path.write_text(json.dumps(tampered), encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "载荷哈希"):
                db_audit_store.verify_external_archive(archive_path)
            archive_path.write_text(
                json.dumps(original, ensure_ascii=False, indent=2), encoding="utf-8",
            )

            with closing(sqlite3.connect(db_audit_store._DB_PATH)) as conn:
                conn.row_factory = sqlite3.Row
                rows = conn.execute(
                    "SELECT * FROM audit_events ORDER BY sequence ASC"
                ).fetchall()
                previous_hash = db_audit_store._GENESIS_HASH
                for row in rows:
                    event = db_audit_store._row_to_event(row)
                    if event["sequence"] == 1:
                        event["summary"] = "rewritten history"
                    event["previous_hash"] = previous_hash
                    event_hash = db_audit_store.sha256_text(
                        db_audit_store._canonical_payload(event)
                    )
                    conn.execute(
                        "UPDATE audit_events SET summary=?, previous_hash=?, event_hash=? "
                        "WHERE sequence=?",
                        (event["summary"], previous_hash, event_hash, event["sequence"]),
                    )
                    previous_hash = event_hash
                conn.commit()
            self.assertTrue(db_audit_store.verify_chain()["ok"])
            with self.assertRaisesRegex(RuntimeError, "历史前缀"):
                db_audit_store.verify_external_archive(archive_path)

        with self.assertRaisesRegex(ValueError, "受管目录"):
            db_audit_store.create_external_archive(
                Path(self.tmp.name) / "archives", through_sequence=1,
            )

    def test_retention_status_only_recommends_a_contiguous_archive_prefix(self):
        db_audit_store.append_event(
            category="system", action="retention", outcome="succeeded",
            summary="retention event",
        )
        future = datetime.now(timezone.utc).astimezone() + timedelta(days=400)
        due = db_audit_store.retention_status(365, now=future)
        self.assertTrue(due["ok"])
        self.assertEqual(due["due_count"], 1)
        self.assertEqual(due["recommended_through_sequence"], 1)
        self.assertFalse(due["destructive_action"])
        not_due = db_audit_store.retention_status(
            365, now=datetime.now(timezone.utc).astimezone(),
        )
        self.assertEqual(not_due["due_count"], 0)
        with self.assertRaises(ValueError):
            db_audit_store.retention_status(29)

    def test_backup_verification_tamper_detection_and_offline_restore(self):
        db_audit_store.append_event(
            category="system", action="first", outcome="succeeded", summary="第一条事件",
        )
        backup = db_audit_store.create_backup(reason="manual")
        verified = db_audit_store.verify_backup(backup["backup_id"])
        self.assertTrue(verified["valid"])
        self.assertEqual(verified["count"], 1)

        db_audit_store.append_event(
            category="system", action="second", outcome="succeeded", summary="第二条事件",
        )
        current_head = db_audit_store.verify_chain()["head_hash"]
        with self.assertRaisesRegex(RuntimeError, "已变化"):
            db_audit_store.restore_backup(
                backup["backup_id"], expected_current_head="0" * 64,
                confirmation="RESTORE_AUDIT_LEDGER",
            )
        restored = db_audit_store.restore_backup(
            backup["backup_id"], expected_current_head=current_head,
            confirmation="RESTORE_AUDIT_LEDGER",
        )
        self.assertEqual(restored["integrity"]["count"], 1)
        self.assertEqual(restored["integrity"]["head_hash"], backup["head_hash"])
        self.assertTrue(db_audit_store.verify_backup(restored["safety_backup_id"])["valid"])

        backup_path, _ = db_audit_store._backup_paths(backup["backup_id"])
        with closing(sqlite3.connect(backup_path)) as conn:
            conn.execute("UPDATE audit_events SET summary = ? WHERE sequence = 1", ("篡改",))
            conn.commit()
        with self.assertRaisesRegex(RuntimeError, "文件哈希不匹配"):
            db_audit_store.verify_backup(backup["backup_id"])

    def test_external_backup_is_independent_drillable_and_restorable(self):
        db_audit_store.append_event(
            category="system", action="external_source", outcome="succeeded",
            summary="外部备份源事件",
        )
        backup = db_audit_store.create_backup(reason="manual")
        with tempfile.TemporaryDirectory() as external_dir, \
                tempfile.TemporaryDirectory() as drill_dir:
            exported = db_audit_store.create_external_backup(
                backup["backup_id"], external_dir,
            )
            bundle_path = Path(exported["bundle_file"])
            self.assertTrue(exported["valid"])
            self.assertTrue(exported["independent_of_current_ledger"])
            self.assertFalse(exported["destructive_action"])
            with zipfile.ZipFile(bundle_path, "r") as archive:
                self.assertEqual(
                    set(archive.namelist()), {"manifest.json", "audit.db"},
                )
                self.assertTrue(all(
                    info.compress_type == zipfile.ZIP_STORED
                    for info in archive.infolist()
                ))

            db_audit_store.append_event(
                category="system", action="after_external", outcome="succeeded",
                summary="外部备份后的当前事件",
            )
            live_before = db_audit_store.verify_chain()
            local_database, local_manifest = db_audit_store._backup_paths(
                backup["backup_id"]
            )
            local_database.unlink()
            local_manifest.unlink()
            self.assertTrue(db_audit_store.verify_external_backup(bundle_path)["valid"])

            drill = db_audit_store.run_external_restore_drill(bundle_path, drill_dir)
            self.assertEqual(drill["source_kind"], "external_backup")
            self.assertEqual(drill["source_bundle_id"], exported["bundle_id"])
            self.assertEqual(
                db_audit_store.verify_chain()["head_hash"], live_before["head_hash"],
            )
            with self.assertRaisesRegex(ValueError, "external_backup_file"):
                db_audit_store.verify_restore_drill(drill["report_file"])
            verified_drill = db_audit_store.verify_restore_drill(
                drill["report_file"], external_backup_file=bundle_path,
            )
            self.assertTrue(verified_drill["valid"])

            with self.assertRaisesRegex(ValueError, "确认短语"):
                db_audit_store.restore_external_backup(
                    bundle_path,
                    expected_current_head=live_before["head_hash"],
                    confirmation="WRONG",
                )
            restored = db_audit_store.restore_external_backup(
                bundle_path,
                expected_current_head=live_before["head_hash"],
                confirmation="RESTORE_EXTERNAL_AUDIT_BACKUP",
            )
            self.assertEqual(restored["restored_external_bundle_id"], exported["bundle_id"])
            self.assertEqual(restored["integrity"]["count"], 1)
            self.assertEqual(restored["integrity"]["head_hash"], backup["head_hash"])
            self.assertTrue(
                db_audit_store.verify_backup(restored["safety_backup_id"])["valid"]
            )

            with zipfile.ZipFile(bundle_path, "r") as source:
                manifest_bytes = source.read("manifest.json")
                database_bytes = source.read("audit.db")
            tampered = Path(external_dir) / "tampered.zip"
            with zipfile.ZipFile(tampered, "w", compression=zipfile.ZIP_STORED) as archive:
                archive.writestr("manifest.json", manifest_bytes)
                archive.writestr("audit.db", database_bytes + b"x")
            with self.assertRaisesRegex(RuntimeError, "数据库哈希"):
                db_audit_store.verify_external_backup(tampered)

            compressed = Path(external_dir) / "compressed.zip"
            with zipfile.ZipFile(compressed, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("manifest.json", manifest_bytes)
                archive.writestr("audit.db", database_bytes)
            with self.assertRaisesRegex(ValueError, "未压缩"):
                db_audit_store.verify_external_backup(compressed)

            extra_entry = Path(external_dir) / "extra-entry.zip"
            with zipfile.ZipFile(extra_entry, "w", compression=zipfile.ZIP_STORED) as archive:
                archive.writestr("manifest.json", manifest_bytes)
                archive.writestr("audit.db", database_bytes)
                archive.writestr("../unexpected.txt", b"unexpected")
            with self.assertRaisesRegex(ValueError, "条目结构"):
                db_audit_store.verify_external_backup(extra_entry)

        with self.assertRaisesRegex(ValueError, "受管目录"):
            db_audit_store.create_external_backup(
                restored["safety_backup_id"], Path(self.tmp.name) / "external",
            )

    def test_external_backup_recovers_a_missing_live_ledger(self):
        db_audit_store.append_event(
            category="system", action="disaster_source", outcome="succeeded",
            summary="灾备恢复源事件",
        )
        backup = db_audit_store.create_backup(reason="manual")
        with tempfile.TemporaryDirectory() as external_dir:
            exported = db_audit_store.create_external_backup(
                backup["backup_id"], external_dir,
            )
            for candidate in (
                Path(db_audit_store._DB_PATH),
                Path(str(db_audit_store._DB_PATH) + "-wal"),
                Path(str(db_audit_store._DB_PATH) + "-shm"),
            ):
                if candidate.exists():
                    candidate.unlink()
            restored = db_audit_store.restore_external_backup(
                exported["bundle_file"],
                expected_current_head="0" * 64,
                confirmation="RESTORE_EXTERNAL_AUDIT_BACKUP",
            )
            self.assertEqual(restored["integrity"]["count"], 1)
            self.assertEqual(restored["integrity"]["head_hash"], backup["head_hash"])
            safety = db_audit_store.verify_backup(restored["safety_backup_id"])
            self.assertEqual(safety["count"], 0)

    def test_corrupt_ledger_requires_stable_assessment_evidence_and_confirmation(self):
        db_audit_store.append_event(
            category="system", action="corrupt_source", outcome="succeeded",
            summary="损坏恢复源事件",
        )
        backup = db_audit_store.create_backup(reason="manual")
        with tempfile.TemporaryDirectory() as external_dir, \
                tempfile.TemporaryDirectory() as evidence_dir:
            exported = db_audit_store.create_external_backup(
                backup["backup_id"], external_dir,
            )
            db_audit_store.append_event(
                category="system", action="corrupt_later", outcome="succeeded",
                summary="备份后的事件",
            )
            with closing(sqlite3.connect(db_audit_store._DB_PATH)) as conn:
                conn.execute(
                    "UPDATE audit_events SET summary = ? WHERE sequence = 2",
                    ("已损坏但未重算哈希",),
                )
                conn.commit()

            with self.assertRaisesRegex(RuntimeError, "完整性异常"):
                db_audit_store.restore_external_backup(
                    exported["bundle_file"],
                    expected_current_head="0" * 64,
                    confirmation="RESTORE_EXTERNAL_AUDIT_BACKUP",
                )
            assessment = db_audit_store.assess_current_ledger()
            self.assertFalse(assessment["integrity_ok"])
            self.assertTrue(assessment["requires_evidence_quarantine"])
            self.assertRegex(assessment["assessment_token"], r"^[0-9a-f]{64}$")

            with self.assertRaisesRegex(ValueError, "确认短语"):
                db_audit_store.create_corrupt_ledger_evidence(
                    evidence_dir,
                    expected_assessment_token=assessment["assessment_token"],
                    confirmation="WRONG",
                )
            with self.assertRaisesRegex(ValueError, "受管目录"):
                db_audit_store.create_corrupt_ledger_evidence(
                    Path(self.tmp.name) / "evidence",
                    expected_assessment_token=assessment["assessment_token"],
                    confirmation="PRESERVE_CORRUPT_AUDIT_LEDGER",
                )
            preserved = db_audit_store.create_corrupt_ledger_evidence(
                evidence_dir,
                expected_assessment_token=assessment["assessment_token"],
                confirmation="PRESERVE_CORRUPT_AUDIT_LEDGER",
            )
            self.assertTrue(preserved["valid"])
            self.assertTrue(preserved["independent_of_current_ledger"])
            self.assertEqual(
                preserved["assessment_token"], assessment["assessment_token"],
            )
            evidence_path = Path(preserved["evidence_file"])
            with zipfile.ZipFile(evidence_path, "r") as archive:
                manifest = json.loads(archive.read("manifest.json"))
                entries = {info.filename for info in archive.infolist()}
                self.assertEqual(
                    entries,
                    {"manifest.json"} | {
                        item["archive_file"] for item in manifest["files"]
                    },
                )
                self.assertTrue(all(
                    info.compress_type == zipfile.ZIP_STORED
                    for info in archive.infolist()
                ))

            with closing(sqlite3.connect(db_audit_store._DB_PATH)) as conn:
                conn.execute(
                    "UPDATE audit_events SET summary = ? WHERE sequence = 2",
                    ("现场再次变化",),
                )
                conn.commit()
            with self.assertRaisesRegex(RuntimeError, "现场已变化"):
                db_audit_store.restore_external_backup_over_corrupt_ledger(
                    exported["bundle_file"],
                    expected_assessment_token=assessment["assessment_token"],
                    evidence_output_dir=evidence_dir,
                    confirmation="RESTORE_CORRUPT_AUDIT_LEDGER",
                )
            current_assessment = db_audit_store.assess_current_ledger()
            with self.assertRaisesRegex(ValueError, "确认短语"):
                db_audit_store.restore_external_backup_over_corrupt_ledger(
                    exported["bundle_file"],
                    expected_assessment_token=current_assessment["assessment_token"],
                    evidence_output_dir=evidence_dir,
                    confirmation="WRONG",
                )
            restored = db_audit_store.restore_external_backup_over_corrupt_ledger(
                exported["bundle_file"],
                expected_assessment_token=current_assessment["assessment_token"],
                evidence_output_dir=evidence_dir,
                confirmation="RESTORE_CORRUPT_AUDIT_LEDGER",
            )
            self.assertEqual(restored["recovery_mode"], "corrupt_ledger_evidence_preserved")
            self.assertIsNone(restored["safety_backup_id"])
            self.assertEqual(restored["integrity"]["count"], 1)
            self.assertEqual(restored["integrity"]["head_hash"], backup["head_hash"])
            recovery_evidence = db_audit_store.verify_corrupt_ledger_evidence(
                restored["corrupt_evidence_file"]
            )
            self.assertEqual(
                recovery_evidence["assessment_token"],
                current_assessment["assessment_token"],
            )

            with zipfile.ZipFile(evidence_path, "r") as source:
                source_entries = {
                    info.filename: source.read(info.filename)
                    for info in source.infolist()
                }
            tampered = Path(evidence_dir) / "tampered-evidence.zip"
            with zipfile.ZipFile(tampered, "w", compression=zipfile.ZIP_STORED) as archive:
                for name, data in source_entries.items():
                    archive.writestr(
                        name, data + (b"x" if name == "ledger.db" else b""),
                    )
            with self.assertRaisesRegex(RuntimeError, "大小不匹配"):
                db_audit_store.verify_corrupt_ledger_evidence(tampered)

            healthy_assessment = db_audit_store.assess_current_ledger()
            self.assertTrue(healthy_assessment["integrity_ok"])
            with self.assertRaisesRegex(RuntimeError, "完整性正常"):
                db_audit_store.restore_external_backup_over_corrupt_ledger(
                    exported["bundle_file"],
                    expected_assessment_token=healthy_assessment["assessment_token"],
                    evidence_output_dir=evidence_dir,
                    confirmation="RESTORE_CORRUPT_AUDIT_LEDGER",
                )

    def test_corrupt_ledger_recovery_rolls_back_live_files_on_postcheck_failure(self):
        db_audit_store.append_event(
            category="system", action="rollback_source", outcome="succeeded",
            summary="回滚恢复源事件",
        )
        backup = db_audit_store.create_backup(reason="manual")
        with tempfile.TemporaryDirectory() as external_dir, \
                tempfile.TemporaryDirectory() as evidence_dir:
            exported = db_audit_store.create_external_backup(
                backup["backup_id"], external_dir,
            )
            with closing(sqlite3.connect(db_audit_store._DB_PATH)) as conn:
                conn.execute(
                    "UPDATE audit_events SET summary = ? WHERE sequence = 1",
                    ("故障现场",),
                )
                conn.commit()
            assessment = db_audit_store.assess_current_ledger()
            with mock.patch.object(
                db_audit_store, "verify_chain",
                side_effect=RuntimeError("模拟恢复后校验失败"),
            ):
                with self.assertRaisesRegex(RuntimeError, "模拟恢复后校验失败"):
                    db_audit_store.restore_external_backup_over_corrupt_ledger(
                        exported["bundle_file"],
                        expected_assessment_token=assessment["assessment_token"],
                        evidence_output_dir=evidence_dir,
                        confirmation="RESTORE_CORRUPT_AUDIT_LEDGER",
                    )
            after = db_audit_store.assess_current_ledger()
            self.assertEqual(after["assessment_token"], assessment["assessment_token"])
            self.assertFalse(after["integrity_ok"])
            self.assertTrue(any(Path(evidence_dir).glob(
                "dbquill-audit-corrupt-evidence-*.zip"
            )))

    def test_configured_external_target_probes_syncs_verifies_and_never_deletes(self):
        self.assertFalse(db_audit_store.external_backup_target_status()["configured"])
        with tempfile.TemporaryDirectory() as target_dir, \
                tempfile.TemporaryDirectory() as replacement_dir:
            with self.assertRaisesRegex(ValueError, "确认短语"):
                db_audit_store.configure_external_backup_target(
                    target_dir, confirmation="WRONG",
                )
            with self.assertRaisesRegex(ValueError, "受管目录之外"):
                db_audit_store.configure_external_backup_target(
                    Path(self.tmp.name) / "target",
                    confirmation="CONFIGURE_AUDIT_BACKUP_TARGET",
                )
            configured = db_audit_store.configure_external_backup_target(
                target_dir, confirmation="CONFIGURE_AUDIT_BACKUP_TARGET",
            )
            self.assertRegex(configured["target_id"], r"^[0-9a-f]{32}$")
            self.assertIsNone(configured["replaced_target_id"])
            with self.assertRaisesRegex(RuntimeError, "必须绑定当前 target ID"):
                db_audit_store.configure_external_backup_target(
                    target_dir, confirmation="CONFIGURE_AUDIT_BACKUP_TARGET",
                )
            with self.assertRaisesRegex(RuntimeError, "已变化"):
                db_audit_store.configure_external_backup_target(
                    replacement_dir,
                    expected_current_target_id="f" * 32,
                    confirmation="CONFIGURE_AUDIT_BACKUP_TARGET",
                )
            status = db_audit_store.external_backup_target_status()
            self.assertTrue(status["configured"])
            self.assertTrue(status["available"])
            self.assertIsNone(status["last_success"])

            with self.assertRaisesRegex(ValueError, "确认短语"):
                db_audit_store.probe_external_backup_target(confirmation="WRONG")
            probe = db_audit_store.probe_external_backup_target(
                confirmation="PROBE_AUDIT_BACKUP_TARGET",
            )
            self.assertTrue(probe["write_read"])
            self.assertTrue(probe["atomic_replace"])
            self.assertTrue(probe["temporary_cleanup"])
            self.assertFalse(list(Path(target_dir).glob(".dbquill-audit-probe-*")))

            db_audit_store.append_event(
                category="system", action="target_first", outcome="succeeded",
                summary="首次目标同步",
            )
            with self.assertRaisesRegex(ValueError, "确认短语"):
                db_audit_store.synchronize_external_backup_target(
                    confirmation="WRONG",
                )
            first = db_audit_store.synchronize_external_backup_target(
                confirmation="SYNC_AUDIT_BACKUP_TARGET",
            )
            first_path = Path(first["backup"]["bundle_file"])
            self.assertTrue(first_path.is_file())
            self.assertFalse(first["deleted_existing_files"])
            verified_first = db_audit_store.verify_latest_external_target_backup()
            self.assertTrue(verified_first["valid"])
            self.assertEqual(
                verified_first["backup"]["bundle_id"],
                first["backup"]["bundle_id"],
            )

            db_audit_store.append_event(
                category="system", action="target_second", outcome="succeeded",
                summary="第二次目标同步",
            )
            second = db_audit_store.synchronize_external_backup_target(
                confirmation="SYNC_AUDIT_BACKUP_TARGET",
            )
            self.assertNotEqual(
                first["backup"]["bundle_id"], second["backup"]["bundle_id"],
            )
            self.assertTrue(first_path.is_file())
            self.assertEqual(
                len(list(Path(target_dir).glob("dbquill-audit-external-backup-*.zip"))),
                2,
            )
            self.assertEqual(
                db_audit_store.verify_latest_external_target_backup()["backup"]["bundle_id"],
                second["backup"]["bundle_id"],
            )

            replaced = db_audit_store.configure_external_backup_target(
                replacement_dir,
                expected_current_target_id=configured["target_id"],
                confirmation="CONFIGURE_AUDIT_BACKUP_TARGET",
            )
            self.assertEqual(replaced["replaced_target_id"], configured["target_id"])
            replacement_status = db_audit_store.external_backup_target_status()
            self.assertTrue(replacement_status["previous_target_state_retained"])
            self.assertIsNone(replacement_status["last_success"])
            with self.assertRaisesRegex(FileNotFoundError, "尚无成功同步"):
                db_audit_store.verify_latest_external_target_backup()
            self.assertTrue(first_path.is_file())

    def test_external_target_detects_config_and_bundle_tamper_and_keeps_last_success(self):
        db_audit_store.append_event(
            category="system", action="target_tamper", outcome="succeeded",
            summary="目标篡改检测",
        )
        with tempfile.TemporaryDirectory() as target_dir, \
                tempfile.TemporaryDirectory() as other_dir:
            db_audit_store.configure_external_backup_target(
                target_dir, confirmation="CONFIGURE_AUDIT_BACKUP_TARGET",
            )
            synced = db_audit_store.synchronize_external_backup_target(
                confirmation="SYNC_AUDIT_BACKUP_TARGET",
            )
            bundle = Path(synced["backup"]["bundle_file"])
            original_bundle = bundle.read_bytes()
            state_path = db_audit_store._external_target_state_path()
            original_state = state_path.read_bytes()

            with open(bundle, "ab") as handle:
                handle.write(b"tamper")
            with self.assertRaisesRegex(RuntimeError, "最新副本.*不一致"):
                db_audit_store.verify_latest_external_target_backup()
            bundle.write_bytes(original_bundle)
            self.assertTrue(db_audit_store.verify_latest_external_target_backup()["valid"])

            config_path = db_audit_store._external_target_config_path()
            original_config = config_path.read_bytes()
            config = json.loads(original_config.decode("utf-8"))
            config["directory"] = str(Path(other_dir).resolve())
            config_path.write_text(json.dumps(config), encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "配置载荷哈希"):
                db_audit_store.external_backup_target_status()
            config_path.write_bytes(original_config)

            with mock.patch.object(
                db_audit_store, "create_external_backup",
                side_effect=RuntimeError("模拟目标写入失败"),
            ):
                with self.assertRaisesRegex(RuntimeError, "模拟目标写入失败"):
                    db_audit_store.synchronize_external_backup_target(
                        confirmation="SYNC_AUDIT_BACKUP_TARGET",
                    )
            self.assertEqual(state_path.read_bytes(), original_state)
            self.assertEqual(
                len(list(Path(target_dir).glob("dbquill-audit-external-backup-*.zip"))),
                1,
            )
            self.assertTrue(db_audit_store.verify_latest_external_target_backup()["valid"])

    def test_external_target_history_is_redacted_tamper_evident_and_health_gated(self):
        db_audit_store.append_event(
            category="system", action="target_health", outcome="succeeded",
            summary="目标健康历史",
        )
        with tempfile.TemporaryDirectory() as target_dir:
            db_audit_store.configure_external_backup_target(
                target_dir, confirmation="CONFIGURE_AUDIT_BACKUP_TARGET",
            )
            first = db_audit_store.synchronize_external_backup_target(
                confirmation="SYNC_AUDIT_BACKUP_TARGET",
            )
            history = db_audit_store.external_backup_target_history()
            self.assertEqual(history["total_matching"], 1)
            self.assertEqual(history["consecutive_failures"], 0)
            self.assertEqual(history["items"][0]["outcome"], "succeeded")
            self.assertEqual(
                history["items"][0]["bundle_id"], first["backup"]["bundle_id"],
            )
            self.assertTrue(
                db_audit_store.check_external_backup_target_health(
                    max_age_hours=1,
                )["healthy"]
            )

            state_path = db_audit_store._external_target_state_path()
            successful_state = state_path.read_bytes()
            sensitive_error = "高度敏感失败细节 / secret-path"
            with mock.patch.object(
                db_audit_store, "create_external_backup",
                side_effect=RuntimeError(sensitive_error),
            ):
                with self.assertRaisesRegex(RuntimeError, "高度敏感失败细节"):
                    db_audit_store.synchronize_external_backup_target(
                        confirmation="SYNC_AUDIT_BACKUP_TARGET",
                    )
            self.assertEqual(state_path.read_bytes(), successful_state)
            failed_history = db_audit_store.external_backup_target_history()
            self.assertEqual(failed_history["total_matching"], 2)
            self.assertEqual(failed_history["consecutive_failures"], 1)
            self.assertEqual(failed_history["items"][0]["outcome"], "failed")
            self.assertEqual(failed_history["items"][0]["error_type"], "RuntimeError")
            history_path = db_audit_store._external_target_history_path()
            raw_history = history_path.read_bytes()
            self.assertNotIn(sensitive_error.encode("utf-8"), raw_history)
            self.assertNotIn(str(Path(target_dir).resolve()).encode("utf-8"), raw_history)
            self.assertEqual(
                db_audit_store.external_backup_target_status()["consecutive_failures"],
                1,
            )
            with self.assertRaisesRegex(RuntimeError, "最近一次同步失败"):
                db_audit_store.check_external_backup_target_health(max_age_hours=1)

            db_audit_store.synchronize_external_backup_target(
                confirmation="SYNC_AUDIT_BACKUP_TARGET",
            )
            recovered_history = db_audit_store.external_backup_target_history()
            self.assertEqual(recovered_history["total_matching"], 3)
            self.assertEqual(recovered_history["consecutive_failures"], 0)
            self.assertTrue(
                db_audit_store.check_external_backup_target_health(
                    max_age_hours=1,
                )["healthy"]
            )

            current_state = state_path.read_bytes()
            stale = json.loads(current_state.decode("utf-8"))
            stale["synchronized_at"] = (
                datetime.now(timezone.utc).astimezone() - timedelta(hours=2)
            ).isoformat(timespec="seconds")
            stale["payload_sha256"] = db_audit_store._target_payload_sha256(
                stale, db_audit_store._TARGET_STATE_FIELDS,
            )
            state_path.write_text(json.dumps(stale), encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "超过健康时限"):
                db_audit_store.check_external_backup_target_health(max_age_hours=1)
            state_path.write_bytes(current_state)

            history_lines = raw_history.decode("utf-8").splitlines()
            tampered = json.loads(history_lines[0])
            tampered["attempt_id"] = "f" * 32
            history_lines[0] = json.dumps(tampered, separators=(",", ":"))
            history_path.write_text("\n".join(history_lines) + "\n", encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "同步尝试载荷哈希"):
                db_audit_store.external_backup_target_history()

    def test_restore_drill_is_isolated_and_report_detects_tamper(self):
        db_audit_store.append_event(
            category="system", action="drill_source", outcome="succeeded",
            summary="恢复演练源事件",
        )
        backup = db_audit_store.create_backup(reason="manual")
        db_audit_store.append_event(
            category="system", action="after_backup", outcome="succeeded",
            summary="备份后的当前事件",
        )
        live_before = db_audit_store.verify_chain()
        with tempfile.TemporaryDirectory() as external_dir:
            drill = db_audit_store.run_restore_drill(backup["backup_id"], external_dir)
            live_after = db_audit_store.verify_chain()
            self.assertEqual(live_after["count"], live_before["count"])
            self.assertEqual(live_after["head_hash"], live_before["head_hash"])
            self.assertTrue(drill["live_ledger_unchanged"])
            self.assertFalse(drill["restored_artifact_retained"])
            self.assertFalse(drill["destructive_action"])
            report_path = Path(drill["report_file"])
            self.assertTrue(report_path.is_file())
            self.assertFalse(any(Path(external_dir).glob("*.db")))
            verified = db_audit_store.verify_restore_drill(report_path)
            self.assertTrue(verified["valid"])
            legacy_path = Path(external_dir) / "legacy-drill.json"
            legacy_payload = json.loads(report_path.read_text(encoding="utf-8"))
            for key in ("report_version", "source_kind", "source_bundle_id"):
                legacy_payload.pop(key)
            legacy_payload["payload_sha256"] = (
                db_audit_store._restore_drill_payload_sha256(legacy_payload)
            )
            legacy_path.write_text(
                json.dumps(legacy_payload, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            self.assertTrue(db_audit_store.verify_restore_drill(legacy_path)["valid"])
            payload = json.loads(report_path.read_text(encoding="utf-8"))
            payload["event_count"] += 1
            report_path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8",
            )
            with self.assertRaisesRegex(RuntimeError, "载荷哈希"):
                db_audit_store.verify_restore_drill(report_path)

        with self.assertRaisesRegex(ValueError, "受管目录"):
            db_audit_store.run_restore_drill(
                backup["backup_id"], Path(self.tmp.name) / "drills",
            )

    def test_external_anchor_verifies_ledger_prefix_after_growth_and_detects_tampering(self):
        first = db_audit_store.append_event(
            category="system", action="first", outcome="succeeded", summary="第一条事件",
        )
        with tempfile.TemporaryDirectory() as external_dir:
            anchor = db_audit_store.create_external_anchor(external_dir)
            anchor_path = Path(anchor["anchor_file"])
            self.assertEqual(anchor["event_count"], 1)
            self.assertEqual(anchor["head_hash"], first["event_hash"])
            self.assertNotEqual(anchor_path.parent, Path(self.tmp.name).resolve())

            db_audit_store.append_event(
                category="system", action="second", outcome="succeeded", summary="第二条事件",
            )
            verified = db_audit_store.verify_external_anchor(anchor_path)
            self.assertTrue(verified["valid"])
            self.assertEqual(verified["events_since_anchor"], 1)
            self.assertNotEqual(verified["current_head_hash"], verified["head_hash"])

            # Simulate an attacker who can rewrite and re-hash the entire local chain.
            with closing(sqlite3.connect(db_audit_store._DB_PATH)) as conn:
                conn.row_factory = sqlite3.Row
                rows = conn.execute(
                    "SELECT * FROM audit_events ORDER BY sequence ASC"
                ).fetchall()
                rewritten_first = db_audit_store._row_to_event(rows[0])
                rewritten_first["summary"] = "被整体重写的第一条事件"
                rewritten_first_hash = db_audit_store.sha256_text(
                    db_audit_store._canonical_payload(rewritten_first)
                )
                rewritten_second = db_audit_store._row_to_event(rows[1])
                rewritten_second["previous_hash"] = rewritten_first_hash
                rewritten_second_hash = db_audit_store.sha256_text(
                    db_audit_store._canonical_payload(rewritten_second)
                )
                conn.execute(
                    "UPDATE audit_events SET summary = ?, event_hash = ? WHERE sequence = 1",
                    (rewritten_first["summary"], rewritten_first_hash),
                )
                conn.execute(
                    "UPDATE audit_events SET previous_hash = ?, event_hash = ? WHERE sequence = 2",
                    (rewritten_first_hash, rewritten_second_hash),
                )
                conn.commit()
            self.assertTrue(db_audit_store.verify_chain()["ok"])
            with self.assertRaisesRegex(RuntimeError, "历史前缀"):
                db_audit_store.verify_external_anchor(anchor_path)

            raw = json.loads(anchor_path.read_text(encoding="utf-8"))
            raw["head_hash"] = "f" * 64
            anchor_path.write_text(json.dumps(raw), encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "载荷哈希"):
                db_audit_store.verify_external_anchor(anchor_path)

        with self.assertRaisesRegex(ValueError, "受管目录"):
            db_audit_store.create_external_anchor(Path(self.tmp.name) / "anchors")


class AuditApiTests(AioHTTPTestCase):
    async def get_application(self):
        app = web.Application(middlewares=[desktop_bridge.cors_middleware])
        app.router.add_get("/db/audit", desktop_bridge.db_audit_handler)
        app.router.add_get("/db/audit/verify", desktop_bridge.db_audit_verify_handler)
        app.router.add_get("/db/audit/export", desktop_bridge.db_audit_export_handler)
        app.router.add_post(
            "/db/audit/reconciliation/resolve",
            desktop_bridge.db_audit_reconciliation_resolve_handler,
        )
        app.router.add_get("/db/audit/backups", desktop_bridge.db_audit_backups_handler)
        app.router.add_post("/db/audit/backups", desktop_bridge.db_audit_backup_create_handler)
        app.router.add_get("/db/write/form", desktop_bridge.db_write_form_handler)
        app.router.add_post(
            "/db/write/prepare-insert",
            desktop_bridge.db_write_prepare_insert_handler,
        )
        app.router.add_post(
            "/db/write/prepare-create-table",
            desktop_bridge.db_write_prepare_create_table_handler,
        )
        app.router.add_post("/db/write/confirm", desktop_bridge.db_write_confirm_handler)
        return app

    async def asyncSetUp(self):
        await super().asyncSetUp()
        self.tmp = tempfile.TemporaryDirectory()
        self.audit_patch_data = mock.patch.object(
            db_audit_store, "_DATA_DIR", Path(self.tmp.name),
        )
        self.audit_patch_path = mock.patch.object(
            db_audit_store, "_DB_PATH", Path(self.tmp.name) / "audit.db",
        )
        self.audit_patch_data.start()
        self.audit_patch_path.start()
        db_audit_store.init_db()
        self.original_audit_store = desktop_bridge._audit_store
        desktop_bridge._audit_store = db_audit_store
        path = Path(self.tmp.name) / "api.db"
        _make_db(path)
        self.db_path = path
        self.db_id = "audit-api-test"
        desktop_bridge._DB_AGENT_DBS[self.db_id] = {
            "id": self.db_id, "name": path.name, "path": str(path),
            "tables": ["items"], "attachedAt": 0,
        }

    async def asyncTearDown(self):
        desktop_bridge._DB_AGENT_CACHE.pop(self.db_id, None)
        desktop_bridge._DB_AGENT_DBS.pop(self.db_id, None)
        desktop_bridge._audit_store = self.original_audit_store
        self.audit_patch_path.stop()
        self.audit_patch_data.stop()
        self.tmp.cleanup()
        await super().asyncTearDown()

    async def test_audit_query_and_export_return_verified_redacted_events(self):
        db_key = desktop_bridge._db_semantic_key(desktop_bridge._DB_AGENT_DBS[self.db_id])
        db_audit_store.append_event(
            category="nl_operation", action="query", outcome="succeeded",
            summary="自然语言数据库操作", database_key=db_key,
            details={
                "question_sha256": db_audit_store.sha256_text("敏感问题"),
                "question_length": 4,
            },
        )
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        response = await self.client.get(
            f"/db/audit?dbId={self.db_id}&limit=20", headers=headers,
        )
        self.assertEqual(response.status, 200)
        payload = await response.json()
        self.assertTrue(payload["integrity"]["ok"])
        self.assertFalse(payload["retention"]["destructive_action"])
        self.assertEqual(payload["retention"]["retention_days"], 365)
        self.assertEqual(len(payload["events"]), 1)
        self.assertNotIn("敏感问题", json.dumps(payload, ensure_ascii=False))
        exported = await self.client.get(
            f"/db/audit/export?dbId={self.db_id}", headers=headers,
        )
        self.assertEqual(exported.status, 200)
        self.assertEqual((await exported.json())["ledger"]["format"], "dbquill-audit-ledger")

    async def test_structured_insert_api_requires_operator_and_waits_for_confirmation(self):
        viewer_headers = {
            "X-DBQuill-Token": desktop_bridge._ROLE_TOKENS["viewer"],
        }
        operator_headers = {
            "X-DBQuill-Token": desktop_bridge._ROLE_TOKENS["operator"],
        }
        form = await self.client.get(
            f"/db/write/form?dbId={self.db_id}&table=items",
            headers=viewer_headers,
        )
        self.assertEqual(form.status, 200, await form.text())
        form_payload = await form.json()
        self.assertEqual(form_payload["answer"]["kind"], "write_form")
        self.assertEqual(form_payload["answer"]["write"]["example"]["rows"], [[1, "v0"]])

        body = {
            "dbId": self.db_id,
            "table": "items",
            "fields": [
                {"column": "id", "mode": "omit", "value": ""},
                {"column": "value", "mode": "value", "value": "api-row-secret"},
            ],
        }
        denied = await self.client.post(
            "/db/write/prepare-insert", headers=viewer_headers, json=body,
        )
        self.assertEqual(denied.status, 403)
        prepared = await self.client.post(
            "/db/write/prepare-insert", headers=operator_headers, json=body,
        )
        self.assertEqual(prepared.status, 200, await prepared.text())
        pending = await prepared.json()
        self.assertEqual(pending["answer"]["kind"], "write_pending")
        with closing(sqlite3.connect(self.db_path)) as conn:
            self.assertEqual(conn.execute("SELECT COUNT(*) FROM items").fetchone()[0], 1)

        confirmed = await self.client.post(
            "/db/write/confirm",
            headers=operator_headers,
            json={
                "dbId": self.db_id,
                "confirmId": pending["answer"]["confirm_id"],
                "approved": True,
            },
        )
        self.assertEqual(confirmed.status, 200, await confirmed.text())
        with closing(sqlite3.connect(self.db_path)) as conn:
            self.assertEqual(
                conn.execute("SELECT value FROM items ORDER BY id DESC LIMIT 1").fetchone()[0],
                "api-row-secret",
            )
        db_key = desktop_bridge._db_semantic_key(desktop_bridge._DB_AGENT_DBS[self.db_id])
        events = db_audit_store.list_events(database_key=db_key)
        self.assertNotIn("api-row-secret", json.dumps(events, ensure_ascii=False))

    async def test_structured_create_table_api_requires_admin_and_refreshes_schema(self):
        operator_headers = {
            "X-DBQuill-Token": desktop_bridge._ROLE_TOKENS["operator"],
        }
        admin_headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        body = {
            "dbId": self.db_id,
            "table": "api_orders",
            "columns": [
                {
                    "name": "id", "type": "INTEGER", "primaryKey": True,
                    "autoIncrement": True, "nullable": False, "defaultMode": "none",
                },
                {
                    "name": "amount", "type": "REAL", "nullable": False,
                    "defaultMode": "value", "defaultValue": "0",
                },
            ],
        }
        denied = await self.client.post(
            "/db/write/prepare-create-table", headers=operator_headers, json=body,
        )
        self.assertEqual(denied.status, 403)
        prepared = await self.client.post(
            "/db/write/prepare-create-table", headers=admin_headers, json=body,
        )
        self.assertEqual(prepared.status, 200, await prepared.text())
        pending = await prepared.json()
        self.assertEqual(pending["answer"]["kind"], "write_pending")
        with closing(sqlite3.connect(self.db_path)) as conn:
            self.assertIsNone(conn.execute(
                "SELECT name FROM sqlite_master WHERE name='api_orders'"
            ).fetchone())

        operator_confirm = await self.client.post(
            "/db/write/confirm", headers=operator_headers,
            json={
                "dbId": self.db_id,
                "confirmId": pending["answer"]["confirm_id"],
                "approved": True,
            },
        )
        self.assertEqual(operator_confirm.status, 403)
        confirmed = await self.client.post(
            "/db/write/confirm", headers=admin_headers,
            json={
                "dbId": self.db_id,
                "confirmId": pending["answer"]["confirm_id"],
                "approved": True,
            },
        )
        self.assertEqual(confirmed.status, 200, await confirmed.text())
        result = await confirmed.json()
        self.assertEqual(result["answer"]["kind"], "write_result")
        with closing(sqlite3.connect(self.db_path)) as conn:
            self.assertIsNotNone(conn.execute(
                "SELECT name FROM sqlite_master WHERE name='api_orders'"
            ).fetchone())
        self.assertIn("api_orders", desktop_bridge._DB_AGENT_DBS[self.db_id]["tables"])
        self.assertNotIn(self.db_id, desktop_bridge._DB_AGENT_CACHE)

    async def test_pending_resolution_requires_admin_evidence_and_is_append_only(self):
        db_key = desktop_bridge._db_semantic_key(
            desktop_bridge._DB_AGENT_DBS[self.db_id]
        )
        pending = db_audit_store.append_event(
            category="semantic_change", action="upsert", outcome="approved",
            summary="semantic approved", correlation_id="api-manual-resolution",
            database_key=db_key, details={"semantic_ref": "a" * 16},
        )
        body = {
            "dbId": self.db_id,
            "sequence": pending["sequence"],
            "disposition": "verified_no_change",
            "evidenceRef": "ticket-456",
        }
        operator = await self.client.post(
            "/db/audit/reconciliation/resolve",
            headers={"X-DBQuill-Token": desktop_bridge._ROLE_TOKENS["operator"]},
            json=body,
        )
        self.assertEqual(operator.status, 403)
        missing_evidence = await self.client.post(
            "/db/audit/reconciliation/resolve",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
            json={**body, "evidenceRef": ""},
        )
        self.assertEqual(missing_evidence.status, 400)

        resolved = await self.client.post(
            "/db/audit/reconciliation/resolve",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
            json=body,
        )
        self.assertEqual(resolved.status, 200, await resolved.text())
        payload = await resolved.json()
        self.assertEqual(payload["reconciliation"]["unresolved_count"], 0)
        self.assertEqual(payload["reconciliation"]["manually_resolved_count"], 1)
        self.assertEqual(
            payload["resolution"]["evidence_sha256"],
            db_audit_store.sha256_text("ticket-456"),
        )
        duplicate = await self.client.post(
            "/db/audit/reconciliation/resolve",
            headers={"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN},
            json=body,
        )
        self.assertEqual(duplicate.status, 404)
        self.assertNotIn(b"ticket-456", db_audit_store._DB_PATH.read_bytes())

    async def test_write_confirmation_fails_closed_when_audit_is_unavailable(self):
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        with mock.patch.object(desktop_bridge, "_audit_store", None), \
             mock.patch.object(desktop_bridge, "_db_get_agent") as get_agent:
            response = await self.client.post(
                "/db/write/confirm", headers=headers,
                json={"dbId": self.db_id, "confirmId": "abc123", "approved": True},
            )
        self.assertEqual(response.status, 503)
        get_agent.assert_not_called()

    async def test_audit_backup_api_creates_verified_backup_and_reports_status(self):
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        response = await self.client.post(
            "/db/audit/backups", headers=headers, json={"dbId": self.db_id},
        )
        self.assertEqual(response.status, 200)
        payload = await response.json()
        self.assertTrue(payload["backup"]["valid"])
        self.assertGreaterEqual(payload["backup"]["count"], 1)

        listed = await self.client.get("/db/audit/backups", headers=headers)
        self.assertEqual(listed.status, 200)
        listed_payload = await listed.json()
        self.assertEqual(listed_payload["status"]["valid_count"], 1)
        self.assertEqual(listed_payload["status"]["invalid_count"], 0)

        audit = await self.client.get(
            f"/db/audit?dbId={self.db_id}", headers=headers,
        )
        audit_payload = await audit.json()
        self.assertEqual(audit_payload["reconciliation"]["unresolved_count"], 0)
        self.assertEqual(audit_payload["backups"]["valid_count"], 1)
        self.assertEqual(
            [event["outcome"] for event in audit_payload["events"][:2]],
            ["succeeded", "approved"],
        )

    async def test_write_confirmation_fails_closed_when_chain_is_tampered(self):
        db_audit_store.append_event(
            category="system", action="startup", outcome="succeeded", summary="系统启动",
        )
        with closing(sqlite3.connect(db_audit_store._DB_PATH)) as conn:
            conn.execute("UPDATE audit_events SET summary = ? WHERE sequence = 1", ("被修改",))
            conn.commit()
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        with mock.patch.object(desktop_bridge, "_db_get_agent") as get_agent:
            response = await self.client.post(
                "/db/write/confirm", headers=headers,
                json={"dbId": self.db_id, "confirmId": "abc123", "approved": True},
            )
        self.assertEqual(response.status, 503)
        get_agent.assert_not_called()

    async def test_approved_write_records_intent_and_result_without_raw_sql(self):
        confirm_id = "audit-write-success"
        raw_sql = "UPDATE items SET value='audited' WHERE id=1"
        dc.WRITE_REGISTRY.register(dc.WriteProposal(
            confirm_id=confirm_id,
            sql=raw_sql,
            kind="UPDATE",
            table="items",
            summary_zh="update one row",
            dangerous=False,
            preview={"affected": 1},
            db_path=str(self.db_path.resolve()),
        ))
        agent = dc.DBQuillAgent(db_path=str(self.db_path), sample_rows=0)
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        with mock.patch.object(desktop_bridge, "_db_get_agent", return_value=agent):
            response = await self.client.post(
                "/db/write/confirm", headers=headers,
                json={"dbId": self.db_id, "confirmId": confirm_id, "approved": True},
            )
        self.assertEqual(response.status, 200)
        payload = await response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["audit"], {"intent_recorded": True, "result_recorded": True})
        with closing(sqlite3.connect(self.db_path)) as conn:
            self.assertEqual(
                conn.execute("SELECT value FROM items WHERE id=1").fetchone()[0], "audited",
            )
        db_key = desktop_bridge._db_semantic_key(desktop_bridge._DB_AGENT_DBS[self.db_id])
        events = db_audit_store.list_events(database_key=db_key)
        self.assertEqual([event["outcome"] for event in events], ["succeeded", "approved"])
        self.assertNotIn(raw_sql, json.dumps(events, ensure_ascii=False))

    async def test_cancelled_write_is_consumed_without_database_change(self):
        confirm_id = "audit-write-cancel"
        dc.WRITE_REGISTRY.register(dc.WriteProposal(
            confirm_id=confirm_id,
            sql="UPDATE items SET value='must-not-write' WHERE id=1",
            kind="UPDATE",
            table="items",
            summary_zh="cancel one update",
            dangerous=False,
            preview={"affected": 1},
            db_path=str(self.db_path.resolve()),
        ))
        agent = dc.DBQuillAgent(db_path=str(self.db_path), sample_rows=0)
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        with mock.patch.object(desktop_bridge, "_db_get_agent", return_value=agent):
            response = await self.client.post(
                "/db/write/confirm", headers=headers,
                json={"dbId": self.db_id, "confirmId": confirm_id, "approved": False},
            )
        self.assertEqual(response.status, 200, await response.text())
        payload = await response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["answer"]["operation"]["status"], "cancelled")
        self.assertEqual(payload["answer"]["narrative"], "已取消写操作")
        with closing(sqlite3.connect(self.db_path)) as conn:
            self.assertEqual(
                conn.execute("SELECT value FROM items WHERE id=1").fetchone()[0], "v0",
            )
        db_key = desktop_bridge._db_semantic_key(desktop_bridge._DB_AGENT_DBS[self.db_id])
        events = db_audit_store.list_events(database_key=db_key)
        self.assertEqual([event["outcome"] for event in events], ["cancelled", "rejected"])
        with mock.patch.object(desktop_bridge, "_db_get_agent", return_value=agent):
            repeated = await self.client.post(
                "/db/write/confirm", headers=headers,
                json={"dbId": self.db_id, "confirmId": confirm_id, "approved": True},
            )
        self.assertEqual(repeated.status, 200)
        self.assertFalse((await repeated.json())["ok"])

    async def test_operator_can_approve_bounded_update_and_actor_is_recorded(self):
        confirm_id = "operator-bounded-update"
        dc.WRITE_REGISTRY.register(dc.WriteProposal(
            confirm_id=confirm_id,
            sql="UPDATE items SET value='operator' WHERE id=1",
            kind="UPDATE", table="items", summary_zh="update one row",
            dangerous=False, preview={"affected": 1},
            db_path=str(self.db_path.resolve()),
        ))
        agent = dc.DBQuillAgent(db_path=str(self.db_path), sample_rows=0)
        headers = {"X-DBQuill-Token": desktop_bridge._ROLE_TOKENS["operator"]}
        with mock.patch.object(desktop_bridge, "_db_get_agent", return_value=agent):
            response = await self.client.post(
                "/db/write/confirm", headers=headers,
                json={"dbId": self.db_id, "confirmId": confirm_id, "approved": True},
            )
        self.assertEqual(response.status, 200)
        with closing(sqlite3.connect(self.db_path)) as conn:
            self.assertEqual(conn.execute("SELECT value FROM items WHERE id=1").fetchone()[0], "operator")
        db_key = desktop_bridge._db_semantic_key(desktop_bridge._DB_AGENT_DBS[self.db_id])
        events = db_audit_store.list_events(database_key=db_key)
        self.assertEqual(events[0]["actor"], "local_operator")
        self.assertEqual(events[1]["actor"], "local_operator")
        self.assertEqual(events[1]["details"]["approval_policy"], "bounded_dml")

    async def test_operator_cannot_approve_delete_but_admin_can_use_same_proposal(self):
        confirm_id = "operator-delete-denied"
        dc.WRITE_REGISTRY.register(dc.WriteProposal(
            confirm_id=confirm_id,
            sql="DELETE FROM items WHERE id=1",
            kind="DELETE", table="items", summary_zh="delete one row",
            dangerous=False, preview={"affected": 1},
            db_path=str(self.db_path.resolve()),
        ))
        operator_headers = {"X-DBQuill-Token": desktop_bridge._ROLE_TOKENS["operator"]}
        denied = await self.client.post(
            "/db/write/confirm", headers=operator_headers,
            json={"dbId": self.db_id, "confirmId": confirm_id, "approved": True},
        )
        self.assertEqual(denied.status, 403)
        denied_payload = await denied.json()
        self.assertEqual(denied_payload["requiredRole"], "admin")
        self.assertEqual(denied_payload["approvalPolicy"], "high_risk")
        self.assertIsNotNone(dc.WRITE_REGISTRY.get(confirm_id))
        with closing(sqlite3.connect(self.db_path)) as conn:
            self.assertEqual(conn.execute("SELECT COUNT(*) FROM items").fetchone()[0], 1)

        agent = dc.DBQuillAgent(db_path=str(self.db_path), sample_rows=0)
        admin_headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        with mock.patch.object(desktop_bridge, "_db_get_agent", return_value=agent):
            approved = await self.client.post(
                "/db/write/confirm", headers=admin_headers,
                json={"dbId": self.db_id, "confirmId": confirm_id, "approved": True},
            )
        self.assertEqual(approved.status, 200)
        with closing(sqlite3.connect(self.db_path)) as conn:
            self.assertEqual(conn.execute("SELECT COUNT(*) FROM items").fetchone()[0], 0)
        access_events = db_audit_store.list_events(category="access_control")
        self.assertEqual(access_events[0]["actor"], "local_operator")
        self.assertEqual(access_events[0]["details"]["approval_policy"], "high_risk")


class ScheduleAuditApiTests(AioHTTPTestCase):
    async def get_application(self):
        app = web.Application(middlewares=[desktop_bridge.cors_middleware])
        app.router.add_get("/db/schedules", desktop_bridge.db_schedules_handler)
        app.router.add_post("/db/schedules", desktop_bridge.db_schedules_create_handler)
        app.router.add_patch("/db/schedules/{id}", desktop_bridge.db_schedules_update_handler)
        app.router.add_delete("/db/schedules/{id}", desktop_bridge.db_schedules_delete_handler)
        app.router.add_post("/db/schedules/{id}/run", desktop_bridge.db_schedules_run_handler)
        return app

    async def asyncSetUp(self):
        await super().asyncSetUp()
        self.tmp = tempfile.TemporaryDirectory()
        root = Path(self.tmp.name)
        self.tasks_patch = mock.patch.object(db_scheduler, "_TASKS_DIR", root / "tasks")
        self.logs_patch = mock.patch.object(db_scheduler, "_LOGS_DIR", root / "logs")
        self.tasks_patch.start()
        self.logs_patch.start()
        self.audit_tmp = tempfile.TemporaryDirectory()
        self.audit_data_patch = mock.patch.object(
            db_audit_store, "_DATA_DIR", Path(self.audit_tmp.name),
        )
        self.audit_path_patch = mock.patch.object(
            db_audit_store, "_DB_PATH", Path(self.audit_tmp.name) / "audit.db",
        )
        self.audit_data_patch.start()
        self.audit_path_patch.start()
        db_audit_store.init_db()
        self.original_store = desktop_bridge._audit_store
        desktop_bridge._audit_store = db_audit_store
        self.original_resolver = db_scheduler._resolver
        self.original_sink = db_scheduler._audit_sink
        self.db_path = root / "schedule.db"
        _make_db(self.db_path, rows=2)
        self.db_id = "schedule-api-test"
        desktop_bridge._DB_AGENT_DBS[self.db_id] = {
            "id": self.db_id, "name": self.db_path.name, "path": str(self.db_path),
            "tables": ["items"], "attachedAt": 0,
        }
        db_scheduler._resolver = desktop_bridge._db_sched_resolver
        db_scheduler._audit_sink = desktop_bridge._db_sched_audit_sink

    async def asyncTearDown(self):
        desktop_bridge._DB_AGENT_DBS.pop(self.db_id, None)
        desktop_bridge._audit_store = self.original_store
        db_scheduler._resolver = self.original_resolver
        db_scheduler._audit_sink = self.original_sink
        self.audit_path_patch.stop()
        self.audit_data_patch.stop()
        self.logs_patch.stop()
        self.tasks_patch.stop()
        self.audit_tmp.cleanup()
        self.tmp.cleanup()
        await super().asyncTearDown()

    async def test_read_only_schedule_mutation_and_run_are_fully_reconciled(self):
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        raw_sql = "SELECT value FROM items WHERE id=1"
        created = await self.client.post(
            "/db/schedules", headers=headers,
            json={
                "name": "daily query", "dbId": self.db_id, "type": "sql",
                "sql": raw_sql, "schedule": {"mode": "interval", "minutes": 60},
            },
        )
        self.assertEqual(created.status, 200)
        task = (await created.json())["task"]
        run = await self.client.post(
            f"/db/schedules/{task['id']}/run", headers=headers, json={},
        )
        self.assertEqual(run.status, 200)
        self.assertTrue((await run.json())["result"]["ok"])

        events = db_audit_store.list_events(
            database_key=desktop_bridge._db_semantic_key(
                desktop_bridge._DB_AGENT_DBS[self.db_id]
            ),
        )
        self.assertEqual(
            [(event["category"], event["outcome"]) for event in events],
            [
                ("schedule_execution", "succeeded"),
                ("schedule_execution", "pending"),
                ("schedule_change", "succeeded"),
                ("schedule_change", "approved"),
            ],
        )
        self.assertEqual(db_audit_store.reconciliation_status()["unresolved_count"], 0)
        self.assertNotIn(raw_sql, json.dumps(events, ensure_ascii=False))

    async def test_schedule_write_is_rejected_and_audit_unavailable_fails_closed(self):
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        write_sql = "UPDATE items SET value='forbidden' WHERE id=1"
        rejected = await self.client.post(
            "/db/schedules", headers=headers,
            json={
                "name": "must reject", "dbId": self.db_id, "type": "sql",
                "sql": write_sql, "schedule": {"mode": "interval", "minutes": 60},
            },
        )
        self.assertEqual(rejected.status, 400)
        self.assertEqual(db_scheduler.list_tasks(), [])
        with closing(sqlite3.connect(self.db_path)) as conn:
            self.assertEqual(conn.execute("SELECT value FROM items WHERE id=1").fetchone()[0], "v0")

        with mock.patch.object(desktop_bridge, "_audit_store", None):
            unavailable = await self.client.post(
                "/db/schedules", headers=headers,
                json={
                    "name": "audit required", "dbId": self.db_id, "type": "sql",
                    "sql": "SELECT * FROM items",
                    "schedule": {"mode": "interval", "minutes": 60},
                },
            )
        self.assertEqual(unavailable.status, 503)
        self.assertEqual(db_scheduler.list_tasks(), [])


class NaturalLanguageApiTests(AioHTTPTestCase):
    async def get_application(self):
        app = web.Application(middlewares=[desktop_bridge.cors_middleware])
        app.router.add_post("/db/ask", desktop_bridge.db_ask_handler)
        app.router.add_get("/db/ask/{run_id}/progress", desktop_bridge.db_progress_handler)
        return app

    async def asyncSetUp(self):
        await super().asyncSetUp()
        self.audit_tmp = tempfile.TemporaryDirectory()
        self.audit_patch_data = mock.patch.object(
            db_audit_store, "_DATA_DIR", Path(self.audit_tmp.name),
        )
        self.audit_patch_path = mock.patch.object(
            db_audit_store, "_DB_PATH", Path(self.audit_tmp.name) / "audit.db",
        )
        self.audit_patch_data.start()
        self.audit_patch_path.start()
        db_audit_store.init_db()
        self.original_audit_store = desktop_bridge._audit_store
        desktop_bridge._audit_store = db_audit_store
        self.identity_tmp = tempfile.TemporaryDirectory()
        self.identity_patch_path = mock.patch.object(
            db_identity_store, "_DB_PATH", Path(self.identity_tmp.name) / "identities.db",
        )
        self.identity_patch_path.start()
        db_identity_store.init_db()
        self.original_identity_store = desktop_bridge._identity_store
        desktop_bridge._identity_store = db_identity_store
        self.tmp = tempfile.TemporaryDirectory()
        path = Path(self.tmp.name) / "api.db"
        _make_db(path, rows=2)
        self.db_id = "nl-api-test"
        desktop_bridge._DB_AGENT_DBS[self.db_id] = {
            "id": self.db_id,
            "name": path.name,
            "path": str(path),
            "tables": ["items"],
            "attachedAt": 0,
        }

    async def asyncTearDown(self):
        desktop_bridge._identity_store = self.original_identity_store
        self.identity_patch_path.stop()
        self.identity_tmp.cleanup()
        desktop_bridge._DB_AGENT_CACHE.pop(self.db_id, None)
        desktop_bridge._DB_AGENT_DBS.pop(self.db_id, None)
        desktop_bridge._DB_RUNS.clear()
        desktop_bridge._audit_store = self.original_audit_store
        self.audit_patch_path.stop()
        self.audit_patch_data.stop()
        self.audit_tmp.cleanup()
        self.tmp.cleanup()
        await super().asyncTearDown()

    async def test_operation_plan_survives_full_http_workflow(self):
        credential = db_identity_store.issue_credential(
            label="只读分析员", role="viewer", ttl_hours=24,
        )
        headers = {"X-DBQuill-Token": credential["token"]}
        response = await self.client.post(
            "/db/ask",
            headers=headers,
            json={"dbId": self.db_id, "question": "有哪些表？每张表多少行？"},
        )
        self.assertEqual(response.status, 202)
        run_id = (await response.json())["run"]["id"]
        run = None
        for _ in range(50):
            progress = await self.client.get(f"/db/ask/{run_id}/progress", headers=headers)
            run = (await progress.json())["run"]
            if run["status"] != "running":
                break
            await asyncio.sleep(0.01)
        self.assertEqual(run["status"], "done")
        self.assertEqual(run["result"]["kind"], "schema")
        self.assertEqual(run["result"]["operation"]["action"], "inspect_schema")
        self.assertEqual(run["result"]["operation"]["status"], "executed")
        events = db_audit_store.list_events(database_key=desktop_bridge._db_semantic_key(
            desktop_bridge._DB_AGENT_DBS[self.db_id]
        ))
        self.assertEqual(events[0]["category"], "nl_operation")
        self.assertEqual(events[0]["outcome"], "succeeded")
        self.assertEqual(events[0]["actor"], "local_viewer")
        self.assertEqual(
            events[0]["details"]["credential_ref"], credential["credentialRef"],
        )
        self.assertNotIn("有哪些表", json.dumps(events[0], ensure_ascii=False))

    async def test_short_clarification_reply_uses_structured_session_state(self):
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}

        async def ask_and_wait(question: str, session_id: str = "") -> dict:
            response = await self.client.post(
                "/db/ask",
                headers=headers,
                json={"dbId": self.db_id, "question": question, "sessionId": session_id},
            )
            self.assertEqual(response.status, 202)
            started = (await response.json())["run"]
            for _ in range(100):
                progress = await self.client.get(
                    f"/db/ask/{started['id']}/progress",
                    headers=headers,
                )
                run = (await progress.json())["run"]
                if run["status"] != "running":
                    return run
                await asyncio.sleep(0.01)
            self.fail("clarification workflow timed out")

        # 2026-08-20 起写澄清链会先做 LLM 表映射与结构化写解析；用确定性的
        # 离线 mock（均表示“无法确定”）保证本用例不依赖真实模型时序
        def _llm_unsure(prompt, cfg, history=None):
            if "应映射到数据库哪张表" in prompt:
                return {"table": None}
            return {"result": None}

        with mock.patch.object(dc, "_llm_ask_json", side_effect=_llm_unsure):
            first = await ask_and_wait("删除记录")
            second = await ask_and_wait("items", first["sessionId"])
        self.assertEqual(first["result"]["kind"], "clarification")
        self.assertEqual(first["result"]["clarification"]["missing"], "target_table")

        self.assertEqual(second["result"]["kind"], "clarification")
        self.assertEqual(second["result"]["clarification"]["missing"], "filter_condition")
        self.assertIn("目标表：items", second["result"]["clarification"]["original_question"])

    async def test_basic_conversation_preserves_pending_database_clarification(self):
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}

        async def ask_and_wait(question: str, session_id: str = "") -> dict:
            response = await self.client.post(
                "/db/ask",
                headers=headers,
                json={"dbId": self.db_id, "question": question, "sessionId": session_id},
            )
            self.assertEqual(response.status, 202)
            started = (await response.json())["run"]
            for _ in range(100):
                progress = await self.client.get(
                    f"/db/ask/{started['id']}/progress",
                    headers=headers,
                )
                run = (await progress.json())["run"]
                if run["status"] != "running":
                    return run
                await asyncio.sleep(0.01)
            self.fail("conversation clarification workflow timed out")

        def _llm_unsure(prompt, cfg, history=None):
            if "应映射到数据库哪张表" in prompt:
                return {"table": None}
            return {"result": None}

        with mock.patch.object(dc, "_llm_ask_json", side_effect=_llm_unsure):
            first = await ask_and_wait("删除记录")
            greeting = await ask_and_wait("你好", first["sessionId"])
            resumed = await ask_and_wait("items", first["sessionId"])

        self.assertEqual(first["result"]["kind"], "clarification")
        self.assertEqual(greeting["result"]["kind"], "conversation")
        self.assertIn("仍然保留", greeting["result"]["narrative"])
        self.assertEqual(resumed["result"]["kind"], "clarification")
        self.assertEqual(resumed["result"]["clarification"]["missing"], "filter_condition")
        self.assertIn("目标表：items", resumed["result"]["clarification"]["original_question"])

        events = db_audit_store.list_events(
            database_key=desktop_bridge._db_semantic_key(
                desktop_bridge._DB_AGENT_DBS[self.db_id],
            ),
        )
        conversation_event = next(
            event for event in events if event["action"] == "conversation"
        )
        self.assertEqual(conversation_event["summary"], "基础沟通")
        self.assertEqual(conversation_event["outcome"], "succeeded")

    async def test_read_time_clarifications_persist_across_http_rounds(self):
        db_path = Path(desktop_bridge._DB_AGENT_DBS[self.db_id]["path"])
        with closing(sqlite3.connect(db_path)) as conn:
            conn.execute("ALTER TABLE items ADD COLUMN created_at TEXT")
            conn.execute("ALTER TABLE items ADD COLUMN updated_at TEXT")
            conn.commit()
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}

        async def ask_and_wait(question: str, session_id: str = "") -> dict:
            response = await self.client.post(
                "/db/ask",
                headers=headers,
                json={"dbId": self.db_id, "question": question, "sessionId": session_id},
            )
            self.assertEqual(response.status, 202)
            started = (await response.json())["run"]
            for _ in range(100):
                progress = await self.client.get(
                    f"/db/ask/{started['id']}/progress",
                    headers=headers,
                )
                run = (await progress.json())["run"]
                if run["status"] != "running":
                    return run
                await asyncio.sleep(0.01)
            self.fail("read clarification workflow timed out")

        first = await ask_and_wait("统计 items 最近的趋势")
        self.assertEqual(first["result"]["clarification"]["missing"], "time_field")

        second = await ask_and_wait("items.created_at", first["sessionId"])
        self.assertEqual(second["result"]["kind"], "clarification")
        self.assertEqual(second["result"]["clarification"]["missing"], "time_range")
        self.assertIn("时间字段：items.created_at", second["result"]["clarification"]["original_question"])


class SemanticApiTests(AioHTTPTestCase):
    async def get_application(self):
        app = web.Application(middlewares=[desktop_bridge.cors_middleware])
        app.router.add_get("/db/semantics", desktop_bridge.db_semantics_handler)
        app.router.add_post("/db/semantics", desktop_bridge.db_semantics_handler)
        app.router.add_delete(
            "/db/semantics/{semantic_id}", desktop_bridge.db_semantics_delete_handler,
        )
        app.router.add_get(
            "/db/semantics/export", desktop_bridge.db_semantics_export_handler,
        )
        app.router.add_post(
            "/db/semantics/import/preflight",
            desktop_bridge.db_semantics_import_preflight_handler,
        )
        app.router.add_post(
            "/db/semantics/import/apply",
            desktop_bridge.db_semantics_import_apply_handler,
        )
        return app

    async def asyncSetUp(self):
        await super().asyncSetUp()
        self.audit_tmp = tempfile.TemporaryDirectory()
        self.audit_patch_data = mock.patch.object(
            db_audit_store, "_DATA_DIR", Path(self.audit_tmp.name),
        )
        self.audit_patch_path = mock.patch.object(
            db_audit_store, "_DB_PATH", Path(self.audit_tmp.name) / "audit.db",
        )
        self.audit_patch_data.start()
        self.audit_patch_path.start()
        db_audit_store.init_db()
        self.original_audit_store = desktop_bridge._audit_store
        desktop_bridge._audit_store = db_audit_store
        self.tmp = tempfile.TemporaryDirectory()
        path = Path(self.tmp.name) / "semantic-api.db"
        _make_db(path, rows=2)
        with closing(sqlite3.connect(path)) as conn:
            conn.execute("ALTER TABLE items ADD COLUMN created_at TEXT")
            conn.execute("ALTER TABLE items ADD COLUMN amount REAL DEFAULT 10")
            conn.execute(
                "CREATE TABLE holidays (holiday_date DATE, name TEXT, is_working INTEGER)"
            )
            conn.commit()
        self.db_id = "semantic-api-test"
        desktop_bridge._DB_AGENT_DBS[self.db_id] = {
            "id": self.db_id,
            "name": path.name,
            "path": str(path),
            "tables": ["items", "holidays"],
            "attachedAt": 0,
        }
        self.store_tmp = tempfile.TemporaryDirectory()
        self.store_patch_data = mock.patch.object(
            db_semantic_store, "_DATA_DIR", Path(self.store_tmp.name),
        )
        self.store_patch_path = mock.patch.object(
            db_semantic_store, "_DB_PATH", Path(self.store_tmp.name) / "semantics.db",
        )
        self.store_patch_data.start()
        self.store_patch_path.start()
        db_semantic_store.init_db()
        self.original_store = desktop_bridge._semantic_store
        desktop_bridge._semantic_store = db_semantic_store
        with desktop_bridge._SEMANTIC_IMPORTS_LOCK:
            desktop_bridge._SEMANTIC_IMPORTS.clear()

    async def asyncTearDown(self):
        desktop_bridge._DB_AGENT_CACHE.pop(self.db_id, None)
        desktop_bridge._DB_AGENT_DBS.pop(self.db_id, None)
        with desktop_bridge._SEMANTIC_IMPORTS_LOCK:
            desktop_bridge._SEMANTIC_IMPORTS.clear()
        desktop_bridge._semantic_store = self.original_store
        desktop_bridge._audit_store = self.original_audit_store
        self.audit_patch_path.stop()
        self.audit_patch_data.stop()
        self.audit_tmp.cleanup()
        self.store_patch_path.stop()
        self.store_patch_data.stop()
        self.store_tmp.cleanup()
        self.tmp.cleanup()
        await super().asyncTearDown()

    async def test_semantic_crud_validates_schema_and_invalidates_agent_cache(self):
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        desktop_bridge._DB_AGENT_CACHE[self.db_id] = object()
        response = await self.client.post(
            "/db/semantics",
            headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {"kind": "table_alias", "term": "商品", "table": "items"},
            },
        )
        self.assertEqual(response.status, 200)
        saved = (await response.json())["entry"]
        self.assertNotIn(self.db_id, desktop_bridge._DB_AGENT_CACHE)
        audit_events = db_audit_store.list_events(
            database_key=desktop_bridge._db_semantic_key(
                desktop_bridge._DB_AGENT_DBS[self.db_id]
            ),
            category="semantic_change",
        )
        self.assertEqual(audit_events[0]["outcome"], "succeeded")
        self.assertEqual(audit_events[1]["outcome"], "approved")
        self.assertEqual(audit_events[0]["correlation_id"], audit_events[1]["correlation_id"])
        self.assertEqual(len(audit_events[0]["correlation_id"]), 20)
        self.assertEqual(db_audit_store.reconciliation_status()["unresolved_count"], 0)
        self.assertNotIn("商品", json.dumps(audit_events, ensure_ascii=False))

        listed = await self.client.get(
            f"/db/semantics?dbId={self.db_id}", headers=headers,
        )
        listed_payload = await listed.json()
        self.assertEqual(listed_payload["entries"][0]["term"], "商品")
        self.assertTrue(listed_payload["timezone_runtime"]["available"])
        self.assertEqual(listed_payload["timezone_runtime"]["iana_version"], "2026c")

        time_response = await self.client.post(
            "/db/semantics",
            headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {
                    "kind": "time_field", "term": "入库时间", "table": "items",
                    "column": "created_at", "default_grain": "month",
                },
            },
        )
        self.assertEqual(time_response.status, 200)
        self.assertEqual((await time_response.json())["entry"]["default_grain"], "month")

        dimension_response = await self.client.post(
            "/db/semantics",
            headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {
                    "kind": "dimension", "term": "商品分类", "table": "items",
                    "column": "value", "hierarchy": {"name": "商品层级", "level": 1},
                    "filters": [{"column": "value", "operator": "neq", "value": "archived"}],
                },
            },
        )
        self.assertEqual(dimension_response.status, 200)
        self.assertEqual(
            (await dimension_response.json())["entry"]["hierarchy"],
            {"name": "商品层级", "level": 1},
        )
        self.assertEqual(
            (await dimension_response.json())["entry"]["filters"][0]["operator"], "neq",
        )

        ratio_response = await self.client.post(
            "/db/semantics",
            headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {
                    "kind": "ratio_metric", "term": "客单价", "table": "items",
                    "formula": {
                        "operator": "divide",
                        "numerator": {"aggregation": "sum", "column": "amount"},
                        "denominator": {"aggregation": "count"},
                        "scale": 1,
                        "zero_division": "null",
                    },
                },
            },
        )
        self.assertEqual(ratio_response.status, 200)
        self.assertEqual((await ratio_response.json())["entry"]["formula"]["operator"], "divide")

        filtered_metric = await self.client.post(
            "/db/semantics",
            headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {
                    "kind": "metric", "term": "大额商品数", "table": "items",
                    "aggregation": "count",
                    "filters": [{"column": "amount", "operator": "gte", "value": 10}],
                },
            },
        )
        self.assertEqual(filtered_metric.status, 200)
        self.assertEqual(
            (await filtered_metric.json())["entry"]["filters"],
            [{"column": "amount", "operator": "gte", "value": 10}],
        )

        calendar_response = await self.client.post(
            "/db/semantics",
            headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {
                    "kind": "business_calendar", "term": "公司日历", "table": "items",
                    "column": "created_at",
                    "calendar": {
                        "fiscal_year_start_month": 4,
                        "fiscal_year_start_day": 1,
                        "fiscal_year_label": "start_year",
                        "timezone": "Asia/Shanghai",
                        "week_start": 1,
                        "weekend_days": [6, 7],
                        "holiday_table": "holidays",
                        "holiday_date_column": "holiday_date",
                        "holiday_name_column": "name",
                        "working_override_column": "is_working",
                    },
                },
            },
        )
        self.assertEqual(calendar_response.status, 200)
        self.assertEqual(
            (await calendar_response.json())["entry"]["calendar"]["weekend_days"],
            [6, 7],
        )
        calendar_export_response = await self.client.get(
            f"/db/semantics/export?dbId={self.db_id}", headers=headers,
        )
        self.assertEqual(calendar_export_response.status, 200)
        calendar_catalog = (await calendar_export_response.json())["catalog"]
        exported_calendar = next(
            item["calendar"] for item in calendar_catalog["entries"]
            if item["kind"] == "business_calendar"
        )
        self.assertEqual(exported_calendar["fiscal_year_label"], "start_year")
        self.assertEqual(exported_calendar["fiscal_year_label_source"], "explicit")
        exported_time = next(
            item for item in calendar_catalog["entries"]
            if item["kind"] == "time_field"
        )
        exported_dimension = next(
            item for item in calendar_catalog["entries"]
            if item["kind"] == "dimension"
        )
        self.assertEqual(exported_time["default_grain"], "month")
        self.assertEqual(
            exported_dimension["hierarchy"], {"name": "商品层级", "level": 1},
        )
        self.assertEqual(exported_dimension["filters"][0]["value"], "archived")

        invalid = await self.client.post(
            "/db/semantics",
            headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {"kind": "column_alias", "term": "缺失金额", "table": "items", "column": "missing_amount"},
            },
        )
        self.assertEqual(invalid.status, 400)

        invalid_time = await self.client.post(
            "/db/semantics",
            headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {"kind": "time_field", "term": "错误时间", "table": "items", "column": "value"},
            },
        )
        self.assertEqual(invalid_time.status, 400)

        invalid_formula = await self.client.post(
            "/db/semantics",
            headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {
                    "kind": "ratio_metric", "term": "自由公式", "table": "items",
                    "formula": {
                        "operator": "sql",
                        "numerator": {"aggregation": "sum", "column": "amount"},
                        "denominator": {"aggregation": "count"},
                        "scale": 1,
                    },
                },
            },
        )
        self.assertEqual(invalid_formula.status, 400)

        invalid_calendar = await self.client.post(
            "/db/semantics",
            headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {
                    "kind": "business_calendar", "term": "危险日历", "table": "items",
                    "column": "created_at",
                    "calendar": {"timezone": "UTC; DROP TABLE", "weekend_days": [6, 7]},
                },
            },
        )
        self.assertEqual(invalid_calendar.status, 400)

        deleted = await self.client.delete(
            f"/db/semantics/{saved['id']}?dbId={self.db_id}", headers=headers,
        )
        self.assertEqual(deleted.status, 200)

    async def test_column_scope_hides_and_cannot_overwrite_hidden_semantics(self):
        admin_headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        hidden_response = await self.client.post(
            "/db/semantics", headers=admin_headers,
            json={
                "dbId": self.db_id,
                "entry": {
                    "kind": "column_alias", "term": "内部金额",
                    "table": "items", "column": "amount",
                },
            },
        )
        self.assertEqual(hidden_response.status, 200, await hidden_response.text())
        hidden = (await hidden_response.json())["entry"]
        entry = desktop_bridge._DB_AGENT_DBS[self.db_id]
        database_ref = desktop_bridge._database_scope_ref(entry)
        principal = {
            "kind": "credential", "role": "operator", "label": "字段操作员",
            "credentialRef": "f" * 16, "expiresAt": "2026-08-18T00:00:00+00:00",
            "databaseScope": {
                "mode": "restricted", "databaseRefs": [database_ref],
                "tableScopes": {database_ref: ["items"]},
                "columnScopes": {database_ref: {"items": ["id", "value"]}},
            },
        }
        scoped_headers = {"X-DBQuill-Token": "field-scoped-test-token"}
        with mock.patch.object(
            desktop_bridge, "_principal_for_token", return_value=principal,
        ):
            listed = await self.client.get(
                f"/db/semantics?dbId={self.db_id}", headers=scoped_headers,
            )
            self.assertEqual(listed.status, 200, await listed.text())
            payload = await listed.json()
            self.assertEqual(payload["entries"], [])
            self.assertEqual(payload["invalid"], [])

            overwrite = await self.client.post(
                "/db/semantics", headers=scoped_headers,
                json={
                    "dbId": self.db_id,
                    "entry": {
                        "id": hidden["id"], "kind": "column_alias",
                        "term": "公开值", "table": "items", "column": "value",
                    },
                },
            )
            self.assertEqual(overwrite.status, 400)

            preflight = await self.client.post(
                "/db/semantics/import/preflight", headers=scoped_headers,
                json={
                    "dbId": self.db_id,
                    "catalog": {
                        "format": "dbagent-semantic-catalog", "schema_version": 8,
                        "entries": [{
                            "kind": "column_alias", "term": "内部金额",
                            "table": "items", "column": "value",
                        }],
                    },
                },
            )
            self.assertEqual(preflight.status, 400)

        stored = db_semantic_store.list_entries(
            desktop_bridge._db_semantic_key(entry),
        )
        self.assertEqual(len(stored), 1)
        self.assertEqual(stored[0]["column"], "amount")

    async def test_versioned_export_and_two_phase_import_detect_catalog_drift(self):
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}
        created = await self.client.post(
            "/db/semantics", headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {
                    "kind": "table_alias", "term": "商品", "table": "items",
                    "description": "旧定义",
                },
            },
        )
        original = (await created.json())["entry"]

        exported_response = await self.client.get(
            f"/db/semantics/export?dbId={self.db_id}", headers=headers,
        )
        self.assertEqual(exported_response.status, 200)
        catalog = (await exported_response.json())["catalog"]
        self.assertEqual(catalog["format"], "dbquill-semantic-catalog")
        self.assertEqual(catalog["schema_version"], 8)
        self.assertEqual(catalog["semantic_version"], "2.8")
        self.assertNotIn("path", catalog["source"])
        self.assertNotIn("conn", catalog["source"])
        self.assertNotIn("id", catalog["entries"][0])
        self.assertNotIn("createdAt", catalog["entries"][0])

        catalog["entries"][0]["description"] = "导入覆盖"
        catalog["entries"].append({
            "kind": "dimension", "term": "商品分类", "table": "items",
            "column": "value", "description": "导入新增",
        })
        first_preflight_response = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog},
        )
        self.assertEqual(first_preflight_response.status, 200)
        first_preflight = (await first_preflight_response.json())["preflight"]
        self.assertEqual(first_preflight["counts"], {"add": 1, "update": 1, "skip": 0})

        drift = await self.client.post(
            "/db/semantics", headers=headers,
            json={
                "dbId": self.db_id,
                "entry": {
                    "kind": "column_alias", "term": "中途变更", "table": "items",
                    "column": "amount",
                },
            },
        )
        self.assertEqual(drift.status, 200)
        stale_apply = await self.client.post(
            "/db/semantics/import/apply", headers=headers,
            json={"dbId": self.db_id, "token": first_preflight["token"]},
        )
        self.assertEqual(stale_apply.status, 400)
        self.assertIn("预检后已发生变化", (await stale_apply.json())["error"])

        second_preflight_response = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog},
        )
        second_preflight = (await second_preflight_response.json())["preflight"]
        applied = await self.client.post(
            "/db/semantics/import/apply", headers=headers,
            json={"dbId": self.db_id, "token": second_preflight["token"]},
        )
        self.assertEqual(applied.status, 200)
        self.assertEqual((await applied.json())["imported"], 2)

        reused = await self.client.post(
            "/db/semantics/import/apply", headers=headers,
            json={"dbId": self.db_id, "token": second_preflight["token"]},
        )
        self.assertEqual(reused.status, 400)
        listed = await self.client.get(
            f"/db/semantics?dbId={self.db_id}", headers=headers,
        )
        entries = (await listed.json())["entries"]
        by_term = {item["term"]: item for item in entries}
        self.assertEqual(set(by_term), {"商品", "商品分类", "中途变更"})
        self.assertEqual(by_term["商品"]["id"], original["id"])
        self.assertEqual(by_term["商品"]["description"], "导入覆盖")

    async def test_import_preflight_rejects_duplicate_terms_schema_conflicts_and_size(self):
        headers = {"X-DBQuill-Token": desktop_bridge.BRIDGE_TOKEN}

        def catalog(entries, version=1):
            return {
                "format": "dbagent-semantic-catalog",
                "schema_version": version,
                "entries": entries,
            }

        duplicate = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog([
                {"kind": "table_alias", "term": "商品", "table": "items"},
                {"kind": "dimension", "term": "商品", "table": "items", "column": "value"},
            ])},
        )
        self.assertEqual(duplicate.status, 400)
        self.assertIn("重复", (await duplicate.json())["error"])

        schema_conflict = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog([
                {"kind": "column_alias", "term": "缺失字段", "table": "items", "column": "missing"},
            ])},
        )
        self.assertEqual(schema_conflict.status, 400)
        self.assertIn("不存在", (await schema_conflict.json())["error"])

        legacy = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog([
                {"kind": "metric", "term": "旧版商品数", "table": "items", "aggregation": "count"},
            ], version=1)},
        )
        self.assertEqual(legacy.status, 200)

        version_two = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog([
                {"kind": "metric", "term": "二版商品数", "table": "items", "aggregation": "count", "filters": []},
            ], version=2)},
        )
        self.assertEqual(version_two.status, 200)

        version_three = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog([
                {
                    "kind": "business_calendar", "term": "三版日历", "table": "items",
                    "column": "created_at",
                    "calendar": {"timezone": "UTC", "weekend_days": [6, 7]},
                },
            ], version=3)},
        )
        self.assertEqual(version_three.status, 200)

        version_four = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog([
                {"kind": "table_alias", "term": "四版商品", "table": "items"},
            ], version=4)},
        )
        self.assertEqual(version_four.status, 200)

        version_five = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog([
                {
                    "kind": "business_calendar", "term": "五版日期日历",
                    "table": "holidays", "column": "holiday_date",
                    "calendar": {
                        "timezone": "UTC", "weekend_days": [6, 7],
                        "storage_basis": "declared_date",
                    },
                },
            ], version=5)},
        )
        self.assertEqual(version_five.status, 200)

        version_six = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog([
                {
                    "kind": "time_field", "term": "六版入库时间", "table": "items",
                    "column": "created_at", "default_grain": "month",
                },
            ], version=6)},
        )
        self.assertEqual(version_six.status, 200)

        version_seven = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog([
                {"kind": "table_alias", "term": "七版商品", "table": "items"},
            ], version=7)},
        )
        self.assertEqual(version_seven.status, 200)

        version_eight = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog([
                {
                    "kind": "dimension", "term": "八版商品分类", "table": "items",
                    "column": "value",
                    "filters": [{"column": "value", "operator": "neq", "value": "archived"}],
                },
            ], version=8)},
        )
        self.assertEqual(version_eight.status, 200)

        unsupported = await self.client.post(
            "/db/semantics/import/preflight", headers=headers,
            json={"dbId": self.db_id, "catalog": catalog([], version=9)},
        )
        self.assertEqual(unsupported.status, 400)

        with mock.patch.object(desktop_bridge, "_SEMANTIC_IMPORT_MAX_BYTES", 256):
            oversized = await self.client.post(
                "/db/semantics/import/preflight", headers=headers,
                json={"dbId": self.db_id, "catalog": catalog([
                    {
                        "kind": "table_alias", "term": "商品", "table": "items",
                        "description": "x" * 512,
                    },
                ])},
            )
        self.assertIn(oversized.status, {400, 413})
        listed = await self.client.get(
            f"/db/semantics?dbId={self.db_id}", headers=headers,
        )
        self.assertEqual((await listed.json())["entries"], [])


class ClarificationStoreTests(unittest.TestCase):
    def test_pending_clarification_persists_and_is_deleted_with_session(self):
        store = desktop_bridge._sess_store
        self.assertIsNotNone(store)
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(store, "_DB_PATH", Path(tmp) / "sessions.db"):
            store.init_db()
            store.upsert_session("clarification-store", db_id="db", last_question="删除记录", count=1)
            payload = {
                "missing": "target_table",
                "original_question": "删除记录",
                "candidates": [{"label": "items", "prompt": "删除记录；目标表：items"}],
            }
            store.set_pending_clarification("clarification-store", payload)
            self.assertEqual(store.get_pending_clarification("clarification-store"), payload)
            self.assertTrue(store.delete_session("clarification-store"))
            self.assertIsNone(store.get_pending_clarification("clarification-store"))


class ModelSettingsUiTests(AioHTTPTestCase):
    """2026-08-20 紧凑型前端：模型测试连接端点与静态桌面载荷回归。"""

    async def get_application(self):
        app = web.Application(middlewares=[desktop_bridge.cors_middleware])
        app.router.add_post("/model-profiles/test", desktop_bridge.model_profile_test_handler)
        return app

    def _headers(self, token=None):
        return {"X-DBQuill-Token": token or desktop_bridge.BRIDGE_TOKEN}

    async def test_admin_test_connection_success_and_model_presence(self):
        fake = mock.MagicMock(status_code=200)
        fake.json.return_value = {"data": [{"id": "demo-model"}]}
        with mock.patch("requests.get", return_value=fake) as get:
            response = await self.client.post(
                "/model-profiles/test",
                headers=self._headers(),
                json={
                    "baseUrl": "https://api.example.com/v1",
                    "apiKey": "sk-test-secret",
                    "model": "demo-model",
                },
            )
        self.assertEqual(response.status, 200)
        data = await response.json()
        self.assertTrue(data["ok"])
        self.assertTrue(data["hasModel"])
        self.assertNotIn("sk-test-secret", response._body.decode("utf-8", "ignore"))
        called_url = get.call_args.args[0] if get.call_args.args else get.call_args[0][0]
        self.assertEqual(called_url, "https://api.example.com/v1/models")
        auth_header = get.call_args.kwargs.get("headers", {}).get("Authorization", "")
        self.assertEqual(auth_header, "Bearer sk-test-secret")

    async def test_admin_test_connection_http_error_reported_without_key(self):
        fake = mock.MagicMock(status_code=401)
        with mock.patch("requests.get", return_value=fake):
            response = await self.client.post(
                "/model-profiles/test",
                headers=self._headers(),
                json={"baseUrl": "https://api.example.com", "apiKey": "sk-x", "model": "m"},
            )
        self.assertEqual(response.status, 200)
        data = await response.json()
        self.assertFalse(data["ok"])
        self.assertEqual(data["httpStatus"], 401)
        self.assertNotIn("sk-x", response._body.decode("utf-8", "ignore"))

    async def test_admin_test_connection_missing_fields_rejected(self):
        response = await self.client.post(
            "/model-profiles/test",
            headers=self._headers(),
            json={"baseUrl": "https://api.example.com"},
        )
        self.assertEqual(response.status, 400)

    async def test_viewer_cannot_use_test_connection(self):
        fake = mock.MagicMock(status_code=200)
        fake.json.return_value = {"data": []}
        with mock.patch("requests.get", return_value=fake) as get:
            response = await self.client.post(
                "/model-profiles/test",
                headers=self._headers(desktop_bridge._ROLE_TOKENS["viewer"]),
                json={"baseUrl": "https://api.example.com", "apiKey": "sk-v"},
            )
        self.assertEqual(response.status, 403)
        get.assert_not_called()

    def test_model_profiles_list_reports_has_key(self):
        with mock.patch.object(
            desktop_bridge.manager, "_read_unlocked",
            return_value={"version": 1, "profiles": [
                {"key": "profile-a", "name": "A", "model": "m1", "api_key": "secret-a"},
                {"key": "profile-b", "name": "B", "model": "m2", "api_key": ""},
            ]},
        ):
            profiles = desktop_bridge.manager.list_model_profiles()
        by_key = {p["key"]: p for p in profiles}
        self.assertTrue(by_key["profile-a"]["hasKey"])
        self.assertFalse(by_key["profile-b"]["hasKey"])

    def test_desktop_markup_contains_compact_conversation_controls(self):
        html = (Path(__file__).resolve().parent / "desktop" / "static" / "db.html").read_text(
            encoding="utf-8",
        )
        for needle in (
            'id="sideNewChatBtn"', 'id="settingsBtn"',
            'id="newChatSheet"', 'id="settingsModal"', 'id="dbChipBtn"',
            'id="modelChip"', 'id="newChatDbList"', 'id="modelProfileList"',
            "/model-profiles/test",
            # 2026-08-20 会话右键菜单 + 最近/历史分区 + 回答卡默认收起区块
            'id="sessCtxMenu"', "最近对话", "历史对话", "side-section-label",
            'data-act="rename"', 'data-act="del"', ">重命名<", ">删除<",
            'class="sec-details"', "tw-caret", "检索证据", "执行步骤",
            'class="read-report"', "调查报告", "覆盖范围", "未确认项",
            # 2026-08-20 SQL 默认收起 + 操作计划条语义着色
            "（点击展开）", 'class="sql-block"', 'class="op-tag ok"', 'class="op-tag warn"',
            'class="op-tag danger"', ">只读</b>", ">写入</b>", ">自动选择</b>", "statusLabels.executed",
        ):
            self.assertIn(needle, html)
        self.assertNotIn('id="dbSelect"', html)
        self.assertNotIn('id="modelSelect"', html)
        self.assertNotIn('id="newChatBtn"', html)
        self.assertNotIn('id="topNewChatBtn"', html)
        self.assertNotIn('id="uploadBtn"', html)
        # 2026-08-20 精简：徽标/编号艺术字移除，品牌为美术字 wordmark
        self.assertNotIn('id="roleBadge"', html)
        self.assertNotIn('id="sessCnt"', html)
        self.assertNotIn('class="view-no"', html)
        self.assertNotIn('welcome-kicker', html)
        self.assertNotIn('brand-mark', html)
        self.assertNotIn('brand-sub', html)
        self.assertNotIn('<span>01</span>', html)
        self.assertIn('class="brand-wordmark">DBQUILL', html)
        # 2026-08-20 视图英文艺术标题块整体移除，全站纸墨风
        self.assertNotIn('class="view-heading"', html)
        self.assertNotIn("Visual<br>", html)
        self.assertNotIn("Scheduled<br>", html)
        self.assertNotIn("自动扫描可分析字段", html)
        self.assertNotIn("以明确节奏执行只读查询", html)
        # 2026-08-20 顶栏连接库移除 + 无密钥提醒条
        self.assertNotIn('id="connectBtn"', html)
        self.assertIn('id="keyNotice"', html)
        self.assertIn('id="keyNoticeSetup"', html)
        self.assertIn("尚未配置模型 API Key", html)
        self.assertIn("p.hasKey", html)
        # 2026-08-20 提醒条防误报：hasKey 缺失（旧后端）或接口失败不提醒
        self.assertIn("modelsLoaded", html)
        self.assertIn("p.hasKey === false", html)
        for label in (">新对话<", ">设置<", ">上传数据<", ">连接库<"):
            self.assertIn(label, html)

class NaturalFlowFixTests(unittest.TestCase):
    """2026-08-20 真实使用反馈修复：自由表述写入、人名双列、主题检索与历史数据预览。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "flow.db"
        with closing(sqlite3.connect(self.path)) as conn:
            conn.executescript(
                """
                CREATE TABLE customers (
                    id INTEGER PRIMARY KEY, name TEXT NOT NULL,
                    city TEXT, region TEXT, tier TEXT, signup_date DATE, is_active INTEGER
                );
                CREATE TABLE people (
                    id INTEGER PRIMARY KEY, name_zh TEXT, name_en TEXT, status TEXT
                );
                CREATE TABLE papers (
                    id INTEGER PRIMARY KEY, title TEXT, abstract TEXT, year INTEGER
                );
                """
            )
            conn.executemany(
                "INSERT INTO customers(id, name, city, region, tier, signup_date, is_active)"
                " VALUES (?, ?, ?, ?, ?, ?, ?)",
                [(1, "张三", "上海", "华东", "gold", "2026-08-01", 1)],
            )
            conn.commit()
        self.agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=5)

    def tearDown(self):
        self.tmp.cleanup()

    @staticmethod
    def _write_llm_mock(parse=None, sql=None):
        def _dispatch(prompt, cfg, history=None):
            if "解析用户的自然语言数据库写请求" in prompt:
                return parse if parse is not None else {"result": None}
            if "数据库操作助手" in prompt:
                return {"sql": sql or "", "summary_zh": "写操作"}
            raise AssertionError("unexpected prompt: " + prompt[:60])
        return _dispatch

    def test_llm_write_rewrite_unlocks_freeform_update(self):
        mock_json = self._write_llm_mock(
            parse={
                "operation": "update", "table": "customers",
                "set": [{"column": "city", "value": "北京"}],
                "where": [{"column": "name", "value": "张三"}],
            },
            sql="UPDATE customers SET city='北京' WHERE name='张三'",
        )
        with mock.patch.object(dc, "_llm_ask_json", side_effect=mock_json):
            answer = self.agent.ask("把张三的城市改成北京")
        self.assertEqual(answer.kind, "write_pending")
        self.assertIn("UPDATE customers", answer.sql or "")
        self.assertIn("name", answer.sql or "")

    def test_llm_write_rewrite_unlocks_delete_by_name(self):
        mock_json = self._write_llm_mock(
            parse={
                "operation": "delete", "table": "customers",
                "set": [],
                "where": [{"column": "name", "value": "张三"}],
            },
            sql="DELETE FROM customers WHERE name='张三'",
        )
        with mock.patch.object(dc, "_llm_ask_json", side_effect=mock_json):
            answer = self.agent.ask("删除客户张三")
        self.assertEqual(answer.kind, "write_pending")
        self.assertIn("DELETE FROM customers", answer.sql or "")

    def test_llm_write_rewrite_rejects_invented_table(self):
        mock_json = self._write_llm_mock(
            parse={
                "operation": "update", "table": "nope_table",
                "set": [{"column": "city", "value": "北京"}],
                "where": [{"column": "name", "value": "张三"}],
            },
        )
        with mock.patch.object(dc, "_llm_ask_json", side_effect=mock_json):
            answer = self.agent.ask("把张三的城市改成北京")
        self.assertEqual(answer.kind, "clarification")

    def test_llm_write_rewrite_without_progress_keeps_clarification(self):
        mock_json = self._write_llm_mock(parse=None)
        with mock.patch.object(dc, "_llm_ask_json", side_effect=mock_json):
            answer = self.agent.ask("修改一下数据")
        self.assertEqual(answer.kind, "clarification")

    def test_llm_write_rewrite_requires_where_for_update(self):
        mock_json = self._write_llm_mock(
            parse={
                "operation": "update", "table": "customers",
                "set": [{"column": "city", "value": "北京"}],
                "where": [],
            },
        )
        with mock.patch.object(dc, "_llm_ask_json", side_effect=mock_json):
            answer = self.agent.ask("把张三的城市改成北京")
        self.assertEqual(answer.kind, "clarification")

    def test_person_name_and_knowledge_topic_hints(self):
        hints = self.agent.nl2sql._schema_semantic_hints(None)
        self.assertTrue(any("人名匹配提示" in h for h in hints))
        self.assertTrue(any("主题检索提示" in h for h in hints))
        self.assertFalse(any(
            "人名匹配提示" in h or "主题检索提示" in h
            for h in self.agent.nl2sql._schema_semantic_hints(["customers"])
        ))

    def test_intent_router_prompt_routes_paper_lists_to_query(self):
        self.assertIn("论文", dc.IntentRouter._PROMPT)

    def test_history_content_includes_data_preview(self):
        ans = dc.DBAnswer(
            kind="query", narrative="查询完成",
            columns=["标题", "年份"], rows=[["A论文", 2026]],
        )
        content = desktop_bridge._db_history_content(ans)
        self.assertIn("结果列：标题,年份", content)
        self.assertIn("行：A论文,2026", content)

    def test_session_display_snapshot_is_separate_from_three_row_model_context(self):
        rows = [[f"table_{index}", index] for index in range(12)]
        answer = dc.DBAnswer(
            kind="schema", narrative="当前数据库共有 12 张表。",
            columns=["表名", "行数"], rows=rows,
            report={
                "findings": ["一共 12 张表"],
                "scope": "当前已接入的数据库",
                "limitations": "",
            },
        )

        model_content = desktop_bridge._db_history_content(answer)
        snapshot = desktop_bridge._db_session_display_payload(answer)

        self.assertEqual(model_content.count("行："), 3)
        self.assertIsNotNone(snapshot)
        self.assertEqual(snapshot["row_count"], 12)
        self.assertEqual(snapshot["rows"], rows)
        self.assertEqual(snapshot["report"]["findings"], ["一共 12 张表"])
        self.assertFalse(snapshot["snapshot_limited"])

        large_answer = dc.DBAnswer(
            kind="query", narrative="查询完成",
            columns=["id"], rows=[[index] for index in range(80)],
        )
        large_snapshot = desktop_bridge._db_session_display_payload(large_answer)
        self.assertEqual(large_snapshot["row_count"], 80)
        self.assertEqual(len(large_snapshot["rows"]), 80)
        self.assertFalse(large_snapshot["snapshot_limited"])

        over_query_limit = dc.DBAnswer(
            kind="query", narrative="查询完成",
            columns=["id"], rows=[[index] for index in range(600)],
        )
        bounded_snapshot = desktop_bridge._db_session_display_payload(over_query_limit)
        self.assertEqual(bounded_snapshot["row_count"], 600)
        self.assertEqual(len(bounded_snapshot["rows"]), 500)
        self.assertTrue(bounded_snapshot["snapshot_limited"])

        write_answer = dc.DBAnswer(
            kind="write_pending", narrative="等待确认", confirm_id="secret-confirm-id",
        )
        self.assertIsNone(desktop_bridge._db_session_display_payload(write_answer))

    def test_pruned_async_run_stops_before_session_and_audit_side_effects(self):
        run_id = "pruned-run"
        desktop_bridge._DB_RUNS.pop(run_id, None)
        fake_agent = mock.MagicMock()
        fake_agent.ask.return_value = dc.DBAnswer(kind="schema", narrative="done")
        with mock.patch.object(desktop_bridge, "_db_get_agent", return_value=fake_agent), \
             mock.patch.object(desktop_bridge, "_db_session_append") as append, \
             mock.patch.object(desktop_bridge, "_audit_nl_terminal") as audit:
            desktop_bridge._db_ask_workflow(run_id, "missing-db", "question")
        append.assert_not_called()
        audit.assert_not_called()


class GroupedMetricsAndCancellationRegressionTests(unittest.TestCase):
    """Regression coverage for typed grouped metrics and cancel boundaries."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "commerce.db"
        with closing(sqlite3.connect(self.path)) as conn:
            conn.executescript(
                """
                CREATE TABLE customers (
                    id INTEGER PRIMARY KEY,
                    name TEXT NOT NULL,
                    region TEXT NOT NULL
                );
                CREATE TABLE orders (
                    id INTEGER PRIMARY KEY,
                    customer_id INTEGER NOT NULL REFERENCES customers(id),
                    status TEXT NOT NULL,
                    total_amount REAL NOT NULL
                );
                INSERT INTO customers(id, name, region) VALUES
                    (1, 'A', 'east'), (2, 'B', 'west'), (3, 'C', 'east');
                INSERT INTO orders(id, customer_id, status, total_amount) VALUES
                    (1, 1, 'paid', 100.0),
                    (2, 1, 'cancelled', 25.0),
                    (3, 2, 'paid', 50.0),
                    (4, 3, 'paid', 75.0),
                    (5, 3, 'pending', 20.0);
                """
            )
            conn.commit()
        connector = dc.DBConnector(str(self.path))
        self.schema = dc.SchemaDiscovery(connector).discover()
        self.executor = dc.NL2SQLExecutor(
            dc.SQLSecurity(connector), self.schema,
        )

    def tearDown(self):
        self.tmp.cleanup()

    def test_same_table_grouped_count_and_sum_are_compiled_without_model(self):
        with mock.patch.object(
            dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
        ):
            answer = self.executor.answer(
                "按 status 统计 orders 的订单数和 total_amount 合计，按 status 排序。"
            )
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.relational_plan["kind"], "grouped_metrics")
        self.assertIn("COUNT(*)", answer.sql or "")
        self.assertIn("SUM(", answer.sql or "")
        self.assertIn("GROUP BY", answer.sql or "")
        self.assertEqual(
            answer.rows,
            [["cancelled", 1, 25.0], ["paid", 3, 225.0], ["pending", 1, 20.0]],
        )

    def test_explicit_join_grouped_metrics_use_schema_fk_and_filter(self):
        question = (
            "通过 customers.id = orders.customer_id，按客户 region 统计 "
            "status='paid' 的订单数和 total_amount，按金额降序。"
        )
        with mock.patch.object(
            dc, "_llm_ask_json", side_effect=AssertionError("model must not run"),
        ):
            answer = self.executor.answer(question)
        self.assertEqual(answer.kind, "query")
        self.assertEqual(answer.relational_plan["kind"], "grouped_metrics")
        self.assertEqual(answer.relational_plan["joins"][0]["source"], "foreign_key")
        self.assertIn("JOIN", answer.sql or "")
        self.assertIn("WHERE", answer.sql or "")
        self.assertIn("DESC", answer.sql or "")
        self.assertEqual(answer.rows, [["east", 2, 175.0], ["west", 1, 50.0]])

    def test_original_executable_multi_statement_is_rejected_before_model(self):
        agent = dc.DBQuillAgent(db_path=str(self.path))
        with mock.patch.object(
            agent.router,
            "classify",
            side_effect=AssertionError("intent model must not run"),
        ):
            answer = agent.ask("请执行：SELECT COUNT(*) FROM orders; DROP TABLE orders;")
        self.assertEqual(answer.kind, "error")
        self.assertEqual(answer.error, "original_request_contains_multiple_sql_statements")
        self.assertEqual(answer.steps[0]["model_calls"], 0)
        with closing(sqlite3.connect(self.path)) as conn:
            self.assertEqual(conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0], 5)

    def test_sql_terms_in_explanatory_prose_are_not_blocked(self):
        self.assertEqual(
            dc.OriginalSQLRequestGuard.reject_reason("解释 SELECT 和 DROP 的区别"),
            "",
        )

    def test_remote_numeric_and_date_values_remain_json_typed(self):
        self.assertEqual(
            dc.SQLSecurity._safe_row((Decimal("12.50"), date(2026, 8, 26))),
            [12.5, "2026-08-26"],
        )

    def test_remote_schema_discovers_row_counts_primary_and_foreign_keys(self):
        class FakeCursor:
            def __init__(self, dialect):
                self.dialect = dialect
                self.rows = []

            def execute(self, sql, params=None):
                table = str((params or [""])[0])
                if "information_schema.tables" in sql:
                    self.rows = [("customers",), ("orders",)]
                elif "information_schema.columns" in sql:
                    if table == "customers":
                        key = "PRI" if self.dialect == "mysql" else ""
                        self.rows = [
                            ("id", "integer", "NO", key),
                            ("region", "varchar", "NO", ""),
                        ]
                    else:
                        key = "PRI" if self.dialect == "mysql" else ""
                        self.rows = [
                            ("id", "integer", "NO", key),
                            ("customer_id", "integer", "NO", ""),
                            ("total_amount", "numeric", "NO", ""),
                        ]
                elif "pg_catalog.pg_index" in sql:
                    self.rows = [("id",)]
                elif "SELECT COUNT(*)" in sql:
                    self.rows = [(3 if "customers" in sql else 5,)]
                elif "information_schema.key_column_usage" in sql \
                        or "pg_catalog.pg_constraint" in sql:
                    self.rows = (
                        [("customer_id", "customers", "id")]
                        if table == "orders" else []
                    )
                else:
                    raise AssertionError("unexpected discovery SQL: " + sql[:80])
                # mysqlclient-compatible execute returns a row count, not a
                # cursor.  This keeps the row-count regression honest.
                return len(self.rows)

            def fetchall(self):
                return list(self.rows)

            def fetchone(self):
                return self.rows[0] if self.rows else None

            def close(self):
                return None

        class FakeConnection:
            def __init__(self, dialect):
                self._cursor = FakeCursor(dialect)

            def cursor(self):
                return self._cursor

        for dialect in ("mysql", "postgresql"):
            with self.subTest(dialect=dialect):
                connector = types.SimpleNamespace(
                    dialect=dialect,
                    db_path=f"{dialect}://fixture",
                )
                discovery = dc.SchemaDiscovery(connector, sample_rows=0)
                snapshot = discovery._discover_remote(FakeConnection(dialect))
                self.assertTrue(snapshot.tables["customers"].columns[0].pk)
                self.assertEqual(snapshot.tables["orders"].row_count, 5)
                fk = next(
                    column for column in snapshot.tables["orders"].columns
                    if column.name == "customer_id"
                )
                self.assertEqual((fk.fk_table, fk.fk_column), ("customers", "id"))

    def test_remote_write_requires_explicit_connection_opt_in(self):
        connector = dc.RemoteDBConnector({
            "dialect": "mysql", "host": "127.0.0.1", "database": "fixture",
        })
        self.assertFalse(connector.write_enabled)
        with self.assertRaisesRegex(dc.WriteSecurityError, "未启用受控写入"):
            connector.connect_rw()

        enabled = dc.RemoteDBConnector({
            "dialect": "postgresql", "host": "127.0.0.1", "database": "fixture",
            "write_enabled": True,
        })
        self.assertTrue(enabled.write_enabled)

    def test_remote_dml_preview_rolls_back_and_confirmation_commits(self):
        for dialect in ("mysql", "postgresql"):
            with self.subTest(dialect=dialect):
                path = Path(self.tmp.name) / f"remote-{dialect}.db"
                with closing(sqlite3.connect(path)) as conn:
                    conn.executescript(
                        "CREATE TABLE items(id INTEGER PRIMARY KEY, value TEXT NOT NULL);"
                        "INSERT INTO items(id, value) VALUES (1, 'alpha'), (2, 'beta');"
                    )
                    conn.commit()

                connector = dc.RemoteDBConnector({
                    "dialect": dialect,
                    "host": "fixture.invalid",
                    "database": "fixture",
                    "write_enabled": True,
                })
                connector.connect_rw = lambda: sqlite3.connect(path)
                connector.begin_rw = lambda conn: conn.execute("BEGIN")
                connector.execute_rw = lambda conn, sql, params=None: conn.execute(sql, params or ())
                connector.commit_rw = lambda conn: conn.commit()
                connector.rollback_rw = lambda conn: conn.rollback()
                connector.close = lambda conn: conn.close()
                security = dc.WriteSecurity()
                previewer = dc.WritePreviewer(connector)
                agent = object.__new__(dc.DBQuillAgent)
                agent.connector = connector
                agent.write_security = security
                agent.write_previewer = previewer

                operations = [
                    ("UPDATE items SET value = 'gamma' WHERE id = 1", 1),
                    ("INSERT INTO items(id, value) VALUES (3, 'delta')", 1),
                    ("DELETE FROM items WHERE id = 2", 1),
                ]
                for sql, expected_affected in operations:
                    pending = dc._prepare_write_proposal(
                        connector, security, previewer, sql, "远程受控变更",
                    )
                    self.assertEqual(pending.kind, "write_pending")
                    self.assertEqual(pending.write["preview"]["affected"], expected_affected)
                    with closing(sqlite3.connect(path)) as check:
                        if sql.startswith("UPDATE"):
                            self.assertEqual(
                                check.execute("SELECT value FROM items WHERE id = 1").fetchone()[0],
                                "alpha",
                            )
                        elif sql.startswith("INSERT"):
                            self.assertIsNone(check.execute("SELECT id FROM items WHERE id = 3").fetchone())
                        else:
                            self.assertIsNotNone(check.execute("SELECT id FROM items WHERE id = 2").fetchone())
                    result = agent.confirm_write(pending.confirm_id, approve=True)
                    self.assertEqual(result.kind, "write_result")
                    self.assertEqual(result.operation["status"], "executed")

                with closing(sqlite3.connect(path)) as check:
                    self.assertEqual(check.execute("SELECT value FROM items WHERE id = 1").fetchone()[0], "gamma")
                    self.assertEqual(check.execute("SELECT value FROM items WHERE id = 3").fetchone()[0], "delta")
                    self.assertIsNone(check.execute("SELECT id FROM items WHERE id = 2").fetchone())

    def test_remote_ddl_is_rejected_before_preview_connection(self):
        connector = dc.RemoteDBConnector({
            "dialect": "mysql", "host": "fixture.invalid", "database": "fixture",
            "write_enabled": True,
        })
        connector.connect_rw = mock.Mock(side_effect=AssertionError("preview connection must not open"))
        result = dc._prepare_write_proposal(
            connector,
            dc.WriteSecurity(),
            dc.WritePreviewer(connector),
            "DROP TABLE items",
            "删除表",
        )
        self.assertEqual(result.kind, "error")
        self.assertIn("DDL 不支持回滚预览", result.error)
        connector.connect_rw.assert_not_called()

    def test_model_gateway_pre_cancel_does_not_open_http_request(self):
        import model_gateway

        cancel_event = threading.Event()
        cancel_event.set()
        config = {
            "model": "test-model",
            "api_key": "placeholder",
            "base_url": "https://example.invalid/v1",
            "api_mode": "chat_completions",
        }
        with mock.patch.object(model_gateway, "get_profile", return_value=config), \
             mock.patch.object(model_gateway.requests, "post") as post:
            result = "".join(model_gateway.stream_text(
                "test", "default", cancel_event=cancel_event,
            ))
        self.assertEqual(result, "!!!Error: Cancelled")
        post.assert_not_called()

    def test_bridge_cancel_is_immediate_and_worker_discards_late_answer(self):
        run_id = "cancel-regression"
        started = threading.Event()
        cancel_event = threading.Event()
        fake_agent = mock.MagicMock()

        def _ask(*_args, **_kwargs):
            started.set()
            self.assertTrue(cancel_event.wait(2))
            return dc.DBAnswer(kind="query", narrative="late", rows=[[1]])

        fake_agent.ask.side_effect = _ask
        desktop_bridge._DB_RUNS[run_id] = {
            "id": run_id,
            "dbId": "fixture-db",
            "question": "slow question",
            "sessionId": "fixture-session",
            "llmCfg": "",
            "accessScopeRef": "",
            "status": "running",
            "percent": 5,
            "stage": "intent",
            "label": "running",
            "error": "",
            "result": None,
            "cancelRequested": False,
            "createdAt": 0,
            "clarification": None,
        }
        desktop_bridge._DB_RUN_CANCEL_EVENTS[run_id] = cancel_event
        request = types.SimpleNamespace(match_info={"run_id": run_id})
        try:
            with mock.patch.object(desktop_bridge, "_db_get_agent", return_value=fake_agent), \
                 mock.patch.object(desktop_bridge, "_db_session_append") as append, \
                 mock.patch.object(desktop_bridge, "_audit_append"), \
                 mock.patch.object(desktop_bridge, "_audit_nl_terminal") as terminal, \
                 mock.patch.object(desktop_bridge, "_stored_access_scope_allowed", return_value=True):
                worker = threading.Thread(
                    target=desktop_bridge._db_ask_workflow,
                    args=(run_id, "fixture-db", "slow question", "fixture-session", []),
                )
                worker.start()
                self.assertTrue(started.wait(2))
                asyncio.run(desktop_bridge.db_cancel_handler(request))
                self.assertEqual(desktop_bridge._DB_RUNS[run_id]["status"], "cancelled")
                self.assertEqual(desktop_bridge._DB_RUNS[run_id]["percent"], 100)
                worker.join(timeout=2)
                self.assertFalse(worker.is_alive())
                append.assert_not_called()
                terminal.assert_called_once()
                self.assertIsNone(desktop_bridge._DB_RUNS[run_id]["result"])
        finally:
            desktop_bridge._DB_RUNS.pop(run_id, None)
            desktop_bridge._DB_RUN_CANCEL_EVENTS.pop(run_id, None)


class EntityRetrievalAndMemoryRegressionTests(unittest.TestCase):
    """实体跨表召回和短追问执行记忆回归。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "advisor.db"
        with closing(sqlite3.connect(self.path)) as conn:
            conn.executescript(
                """
                CREATE TABLE advisor_profiles (
                    id INTEGER PRIMARY KEY,
                    name TEXT,
                    university TEXT,
                    title TEXT,
                    research_summary TEXT
                );
                CREATE TABLE demo_research (
                    id INTEGER PRIMARY KEY,
                    roster_id INTEGER,
                    name TEXT,
                    university TEXT,
                    department TEXT,
                    research_direction TEXT,
                    evidence_note TEXT,
                    status TEXT
                );
                """
            )
            conn.executemany(
                "INSERT INTO advisor_profiles(name, university, title, research_summary) "
                "VALUES (?, ?, ?, ?)",
                [
                    ("徐老师", "测试大学", "教授", "日常工作"),
                    ("曹之老师", "测试大学", "教授", "学生工作"),
                ],
            )
            conn.execute(
                "INSERT INTO demo_research("
                "id, roster_id, name, university, department, research_direction, evidence_note, status"
                ") VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (22298, 77600, "肖仰华", "复旦大学", "", None, None, "unresolved"),
            )
            conn.commit()
        self.agent = dc.DBQuillAgent(db_path=str(self.path), sample_rows=3)

    def tearDown(self):
        self.tmp.cleanup()

    def test_person_entity_is_first_keyword_and_reaches_later_table(self):
        keywords = self.agent.rag._keywords("肖仰华老师的工作是什么")
        self.assertEqual(keywords[0], "肖仰华")
        evidence = self.agent.rag._recall(keywords)
        exact = [item for item in evidence if item.get("matched") == "肖仰华"]
        self.assertTrue(exact)
        self.assertEqual(exact[0]["table"], "demo_research")
        self.assertIn("unresolved", exact[0]["row"])

    def test_descriptive_columns_are_not_lost_to_physical_column_order(self):
        profile = self.agent.schema.tables["advisor_profiles"]
        research = self.agent.schema.tables["demo_research"]
        self.assertIn("research_summary", self.agent.rag._text_scan_columns(profile))
        self.assertIn("research_direction", self.agent.rag._text_scan_columns(research))

    def test_short_followup_inherits_previous_subject(self):
        resolved = self.agent.operation_planner.resolve_followup(
            "是什么",
            [
                {"role": "user", "content": "肖仰华工作"},
                {"role": "assistant", "content": "未查到结果"},
            ],
        )
        self.assertIn("肖仰华工作", resolved)
        self.assertIn("追问：是什么", resolved)
        self.assertEqual(self.agent.rag._entity_terms(resolved)[0], "肖仰华")

    def test_chained_pronoun_followup_uses_last_complete_topic_anchor(self):
        history = [
            {"role": "user", "content": "肖仰华老师的工作是什么"},
            {"role": "assistant", "content": "只知道关联高校，具体岗位未说明"},
            {"role": "user", "content": "是什么"},
            {"role": "assistant", "content": "具体职务信息未说明"},
        ]
        resolved = self.agent.operation_planner.resolve_followup(
            "那他的研究方向呢",
            history,
        )
        self.assertTrue(resolved.startswith("肖仰华老师的工作是什么"))
        self.assertIn("追问：那他的研究方向呢", resolved)
        self.assertEqual(self.agent.rag._entity_terms(resolved)[0], "肖仰华")
        evidence_followup = self.agent.operation_planner.resolve_followup("证据呢", history)
        self.assertTrue(evidence_followup.startswith("肖仰华老师的工作是什么"))
        self.assertIn("追问：证据呢", evidence_followup)

    def test_referential_correction_skips_prior_dependent_turns(self):
        history = [
            {"role": "user", "content": "肖仰华老师的工作"},
            {"role": "assistant", "content": "查到了研究方向"},
            {"role": "user", "content": "我的意思是他担任的工作"},
            {"role": "assistant", "content": "未找到相关内容"},
        ]
        resolved = self.agent.operation_planner.resolve_followup(
            "我的意思是他担任的工作",
            history,
        )
        self.assertTrue(resolved.startswith("肖仰华老师的工作"))
        self.assertIn("追问：我的意思是他担任的工作", resolved)

    def test_bare_person_work_request_asks_which_meaning(self):
        routed = dc.IntentResult(
            intent="query", confidence=0.9, reasoning="询问特定人物的工作", source="model",
        )
        with mock.patch.object(self.agent.router, "classify", return_value=routed), \
             mock.patch.object(
                 self.agent.nl2sql,
                 "answer",
                 side_effect=AssertionError("工作含义澄清前不应执行 SQL"),
             ):
            answer = self.agent.ask("肖仰华老师的工作")
        self.assertEqual(answer.kind, "clarification")
        self.assertEqual(answer.clarification["missing"], "person_work_scope")
        self.assertEqual(
            [item["label"] for item in answer.clarification["candidates"]],
            ["任职/职务", "研究工作/成果"],
        )

        merged = self.agent.operation_planner.resolve_followup(
            "我的意思是他担任的工作",
            [
                {"role": "user", "content": "肖仰华老师的工作"},
                {"role": "assistant", "content": answer.narrative},
            ],
            answer.clarification,
        )
        self.assertIn("肖仰华老师的工作", merged)
        self.assertIn("工作含义：我的意思是他担任的工作", merged)

    def test_canonical_profile_table_precedes_free_text_mentions(self):
        path = Path(self.tmp.name) / "profile_priority.db"
        with closing(sqlite3.connect(path)) as conn:
            conn.executescript(
                """
                CREATE TABLE activity_audits (id INTEGER PRIMARY KEY, audit_summary TEXT);
                CREATE TABLE claims (id INTEGER PRIMARY KEY, claim_text TEXT);
                CREATE TABLE scholars (
                    id INTEGER PRIMARY KEY,
                    name_zh TEXT,
                    unit TEXT,
                    title TEXT,
                    advisor_status TEXT,
                    short_bio TEXT
                );
                """
            )
            conn.execute(
                "INSERT INTO activity_audits(audit_summary) VALUES (?)",
                ("肖仰华的活动审计摘要",),
            )
            conn.execute(
                "INSERT INTO claims(claim_text) VALUES (?)",
                ("声明文本中提到肖仰华",),
            )
            conn.execute(
                "INSERT INTO scholars(name_zh, unit, title, advisor_status, short_bio) "
                "VALUES (?, ?, ?, ?, ?)",
                (
                    "肖仰华", "计算与智能创新学院", "教授、博导", "博导",
                    "复旦大学教授、博士生导师，实验室主任。",
                ),
            )
            conn.commit()
        agent = dc.DBQuillAgent(db_path=str(path), sample_rows=0)
        evidence = agent.rag._recall(["肖仰华"])
        self.assertTrue(evidence)
        self.assertEqual(evidence[0]["table"], "scholars")
        self.assertIn("教授、博导", evidence[0]["row"])

    def test_empty_sql_result_falls_back_only_to_exact_entity_evidence(self):
        empty = dc.DBAnswer(
            kind="query",
            narrative="查询完成，返回 0 行。",
            sql="SELECT name FROM advisor_profiles WHERE name = '肖仰华'",
            columns=["name"],
            rows=[],
        )
        router_result = dc.IntentResult(
            intent="query", confidence=0.95, reasoning="要求查询具体人物", source="model",
        )
        with mock.patch.object(self.agent.router, "classify", return_value=router_result), \
             mock.patch.object(self.agent.multi_metric_query, "answer", return_value=None), \
             mock.patch.object(self.agent.trend_query, "answer", return_value=None), \
             mock.patch.object(self.agent.dimension_query, "answer", return_value=None), \
             mock.patch.object(self.agent.calendar_query, "answer", return_value=None), \
             mock.patch.object(self.agent.nl2sql, "answer", return_value=empty), \
             mock.patch.object(
                 dc,
                 "_llm_ask_json",
                 return_value={
                     "answer_zh": "数据库中有肖仰华的待研究记录，但工作信息尚未补全。"
                 },
             ):
            answer = self.agent.ask("肖仰华的工作是什么")

        self.assertEqual(answer.kind, "retrieve")
        self.assertEqual(answer.operation["action"], "retrieve")
        self.assertEqual(answer.operation["target_tables"], ["demo_research"])
        self.assertTrue(any(
            item.get("tool") == "empty_query_entity_fallback" for item in answer.steps
        ))
        self.assertTrue(any(
            item.get("table") == "demo_research" and item.get("matched") == "肖仰华"
            for item in answer.evidence
        ))
        self.assertTrue(all(item.get("matched") == "肖仰华" for item in answer.evidence))
        self.assertIn("尚未补全", answer.narrative)

    def test_entity_evidence_question_does_not_guess_a_cross_table_join(self):
        model_misroute = dc.IntentResult(
            intent="query", confidence=0.9, reasoning="需要查询并核验", source="model",
        )
        with mock.patch.object(self.agent.router, "classify", return_value=model_misroute), \
             mock.patch.object(
                 self.agent.nl2sql,
                 "answer",
                 side_effect=AssertionError("实体证据问题不应猜测 JOIN"),
             ), \
             mock.patch.object(
                 dc,
                 "_llm_ask_json",
                 return_value={
                     "answer_zh": (
                         "已知：记录中的高校是复旦大学，状态为 unresolved。"
                         "未知：工作岗位、研究方向和证据均为空，因此证据不足。"
                     )
                 },
             ):
            answer = self.agent.ask(
                "肖仰华的研究状态是什么，数据库里是否有足够证据"
                "说明他的工作？请区分已知和未知"
            )

        self.assertEqual(answer.kind, "retrieve")
        self.assertEqual(answer.operation["target_tables"], ["demo_research"])
        self.assertEqual(
            next(step for step in answer.steps if step.get("tool") == "intent")["source"],
            "deterministic",
        )
        self.assertIn("证据不足", answer.narrative)
        self.assertIn("<NULL>", answer.evidence[0]["row"])


class LauncherPortSelectionTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        launcher_path = Path(__file__).resolve().parents[3] / "dbquill_launcher.pyw"
        spec = importlib.util.spec_from_file_location("dbquill_launcher_tests", launcher_path)
        cls.launcher = importlib.util.module_from_spec(spec)
        assert spec.loader is not None
        spec.loader.exec_module(cls.launcher)

    def test_reuses_current_project_bridge(self):
        with mock.patch.object(self.launcher, "_status_ok", side_effect=lambda port: port == 14172), \
             mock.patch.object(self.launcher, "_port_occupied", return_value=True):
            self.assertEqual(self.launcher._select_bridge_port(), (14172, True))

    def test_status_reuse_requires_matching_transport_protocol(self):
        compatible = {
            "ok": True,
            "authRequired": True,
            "appRoot": self.launcher.APP_ROOT,
            "bridgeProtocol": self.launcher.EXPECTED_BRIDGE_PROTOCOL,
            "uploadProtocol": self.launcher.EXPECTED_UPLOAD_PROTOCOL,
        }
        with mock.patch.object(self.launcher, "_status_info", return_value=compatible):
            self.assertTrue(self.launcher._status_ok(14169))
        incompatible = dict(compatible)
        incompatible.pop("uploadProtocol")
        with mock.patch.object(self.launcher, "_status_info", return_value=incompatible):
            self.assertFalse(self.launcher._status_ok(14169))

    def test_launcher_and_bridge_protocol_constants_match(self):
        self.assertEqual(
            self.launcher.EXPECTED_BRIDGE_PROTOCOL,
            desktop_bridge.BRIDGE_PROTOCOL_VERSION,
        )
        self.assertEqual(
            self.launcher.EXPECTED_UPLOAD_PROTOCOL,
            desktop_bridge.UPLOAD_PROTOCOL,
        )

    def test_uses_free_fallback_when_default_belongs_to_another_project(self):
        occupied = {14169, 14170}
        with mock.patch.object(self.launcher, "_status_ok", return_value=False), \
             mock.patch.object(self.launcher, "_port_occupied", side_effect=lambda port: port in occupied):
            self.assertEqual(self.launcher._select_bridge_port(), (14171, False))

    def test_explicit_python_runtime_has_priority(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            configured = Path(temp_dir) / "python.exe"
            configured.write_bytes(b"test executable placeholder")
            with mock.patch.dict(os.environ, {"DBQUILL_PYTHON": str(configured)}):
                self.assertEqual(
                    self.launcher._select_python(),
                    str(configured.resolve()),
                )

    def test_legacy_python_runtime_override_remains_a_fallback(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            configured = Path(temp_dir) / "python.exe"
            configured.write_bytes(b"test executable placeholder")
            with mock.patch.dict(
                os.environ,
                {"DBAGENT_PYTHON": str(configured)},
                clear=True,
            ):
                self.assertEqual(
                    self.launcher._select_python(),
                    str(configured.resolve()),
                )


if __name__ == "__main__":
    unittest.main()
