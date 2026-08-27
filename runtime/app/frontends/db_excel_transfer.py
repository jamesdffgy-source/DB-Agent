"""Transactional Excel import/export for the DBQuill desktop bridge.

The workbook contract is intentionally narrow: one worksheet maps to one
existing table, row one contains physical column names, and subsequent rows are
inserted only after a fresh schema check and explicit confirmation.
"""
from __future__ import annotations

import base64
import hashlib
import json
import math
import re
import threading
import time
import uuid
from datetime import date, datetime, time as dt_time
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Optional


IMPORT_TTL_SECONDS = 5 * 60
IMPORT_MAX_FILE_BYTES = 25 * 1024 * 1024
IMPORT_MAX_SHEETS = 64
IMPORT_MAX_COLUMNS = 128
IMPORT_MAX_ROWS = 10_000
EXPORT_MAX_TABLES = 100
EXPORT_MAX_ROWS = 100_000
MANIFEST_SHEET = "__DBQuill__"
MANIFEST_FORMAT = "DBQUILL_EXPORT_V1"

_PENDING: Dict[str, dict] = {}
_LOCK = threading.RLock()


class ExcelTransferError(ValueError):
    pass


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _schema_fingerprint(schema: Any) -> str:
    payload = []
    for table_name, table in sorted(schema.tables.items(), key=lambda item: item[0].casefold()):
        payload.append({
            "table": table_name,
            "columns": [{
                "name": column.name,
                "type": column.type,
                "nullable": bool(column.nullable),
                "pk": bool(column.pk),
                "default": str(column.default_sql or ""),
                "automatic": bool(getattr(column, "automatic", False)),
            } for column in table.columns],
        })
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _preview_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, str)):
        return value if not isinstance(value, str) else value[:240]
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, (date, datetime, dt_time)):
        return value.isoformat()
    if isinstance(value, (bytes, bytearray, memoryview)):
        return f"<BLOB {len(value)} bytes>"
    return str(value)[:240]


def _manifest_mapping(workbook: Any) -> dict[str, str]:
    if MANIFEST_SHEET not in workbook.sheetnames:
        return {}
    sheet = workbook[MANIFEST_SHEET]
    rows = sheet.iter_rows(values_only=True)
    first = next(rows, ())
    if not first or str(first[0] or "") != MANIFEST_FORMAT:
        raise ExcelTransferError("Excel 内置映射清单格式无效")
    mapping: dict[str, str] = {}
    for row in rows:
        sheet_name = str(row[0] or "").strip() if len(row) > 0 else ""
        table_name = str(row[1] or "").strip() if len(row) > 1 else ""
        if sheet_name and table_name:
            mapping[sheet_name] = table_name
    return mapping


def _column_is_required(column: Any) -> bool:
    return bool(
        not column.nullable
        and not getattr(column, "automatic", False)
        and not str(column.default_sql or "").strip()
    )


def _coerce_value(value: Any, column: Any, *, row_number: int, table_name: str) -> Any:
    if value is None:
        if not column.nullable and not getattr(column, "automatic", False):
            raise ExcelTransferError(
                f"{table_name} 第 {row_number} 行的必填字段 {column.name} 为空"
            )
        return None
    type_name = str(column.type or "TEXT").upper()
    try:
        if "BLOB" in type_name or "BINARY" in type_name or "BYTEA" in type_name:
            if isinstance(value, (bytes, bytearray, memoryview)):
                return bytes(value)
            text = str(value)
            if text.startswith("dbquill:base64:"):
                return base64.b64decode(text.removeprefix("dbquill:base64:"), validate=True)
            raise ValueError("BLOB 必须来自 DBQuill 导出的 base64 单元格")
        if any(token in type_name for token in ("INT", "SERIAL")):
            if isinstance(value, bool):
                return int(value)
            number = float(value)
            if not math.isfinite(number) or not number.is_integer():
                raise ValueError("不是整数")
            return int(number)
        if any(token in type_name for token in (
            "REAL", "FLOA", "DOUB", "NUMERIC", "DECIMAL", "MONEY",
        )):
            number = float(value)
            if not math.isfinite(number):
                raise ValueError("数值不是有限数")
            return number
        if "BOOL" in type_name:
            if isinstance(value, bool):
                return value
            text = str(value).strip().casefold()
            if text in {"1", "true", "yes", "是"}:
                return True
            if text in {"0", "false", "no", "否"}:
                return False
            raise ValueError("不是布尔值")
        if isinstance(value, (date, datetime, dt_time)):
            return value.isoformat()
        return str(value) if not isinstance(value, str) else value
    except (TypeError, ValueError) as exc:
        raise ExcelTransferError(
            f"{table_name} 第 {row_number} 行字段 {column.name} 与 {column.type or 'TEXT'} 不兼容：{exc}"
        ) from exc


def _parse_workbook(
    path: Path,
    schema: Any,
    *,
    allowed_tables: Optional[frozenset[str]],
    allowed_columns: dict[str, frozenset[str]],
    include_rows: bool,
) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExcelTransferError("缺少 openpyxl，无法处理 Excel") from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise ExcelTransferError(f"无法读取 .xlsx：{exc}") from exc
    try:
        manifest = _manifest_mapping(workbook)
        worksheet_names = [name for name in workbook.sheetnames if name != MANIFEST_SHEET]
        if not worksheet_names:
            raise ExcelTransferError("Excel 中没有可导入的工作表")
        if len(worksheet_names) > IMPORT_MAX_SHEETS:
            raise ExcelTransferError(f"工作表不能超过 {IMPORT_MAX_SHEETS} 张")
        tables_by_folded = {name.casefold(): name for name in schema.tables}
        plans = []
        total_rows = 0
        seen_tables: set[str] = set()
        for worksheet_name in worksheet_names:
            sheet = workbook[worksheet_name]
            target_name = manifest.get(worksheet_name, worksheet_name)
            actual_table = tables_by_folded.get(target_name.casefold())
            if actual_table is None:
                raise ExcelTransferError(f"工作表 {worksheet_name} 找不到同名目标表 {target_name}")
            folded_table = actual_table.casefold()
            if folded_table in seen_tables:
                raise ExcelTransferError(f"目标表 {actual_table} 在 Excel 中重复出现")
            seen_tables.add(folded_table)
            if allowed_tables is not None and folded_table not in allowed_tables:
                raise ExcelTransferError(f"当前凭据无权写入表 {actual_table}")
            if int(sheet.max_column or 0) > IMPORT_MAX_COLUMNS:
                raise ExcelTransferError(
                    f"工作表 {worksheet_name} 超过 {IMPORT_MAX_COLUMNS} 列限制"
                )
            iterator = sheet.iter_rows()
            header_cells = next(iterator, None)
            if not header_cells:
                raise ExcelTransferError(f"工作表 {worksheet_name} 缺少表头")
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
            while headers and not headers[-1]:
                headers.pop()
            if not headers or any(not item for item in headers):
                raise ExcelTransferError(f"工作表 {worksheet_name} 的第一行必须是连续字段名")
            if len(headers) > IMPORT_MAX_COLUMNS:
                raise ExcelTransferError(f"单表字段不能超过 {IMPORT_MAX_COLUMNS} 个")
            if len({item.casefold() for item in headers}) != len(headers):
                raise ExcelTransferError(f"工作表 {worksheet_name} 存在重复字段名")
            table = schema.tables[actual_table]
            columns_by_folded = {column.name.casefold(): column for column in table.columns}
            columns = []
            for header in headers:
                column = columns_by_folded.get(header.casefold())
                if column is None:
                    raise ExcelTransferError(f"表 {actual_table} 不存在字段 {header}")
                visible = allowed_columns.get(folded_table)
                if visible is not None and column.name.casefold() not in visible:
                    raise ExcelTransferError(f"当前凭据无权写入字段 {actual_table}.{column.name}")
                columns.append(column)
            included = {column.name.casefold() for column in columns}
            missing_required = [
                column.name for column in table.columns
                if _column_is_required(column) and column.name.casefold() not in included
            ]
            if missing_required:
                raise ExcelTransferError(
                    f"表 {actual_table} 缺少必填字段：{', '.join(missing_required)}"
                )
            rows = []
            preview = []
            row_count = 0
            for row_number, cells in enumerate(iterator, start=2):
                if any(cell.value is not None for cell in cells[len(headers):]):
                    raise ExcelTransferError(
                        f"工作表 {worksheet_name} 第 {row_number} 行包含表头之外的数据"
                    )
                selected = cells[:len(headers)]
                if not any(cell.value is not None for cell in selected):
                    continue
                for cell in selected:
                    if cell.data_type == "f":
                        raise ExcelTransferError(
                            f"工作表 {worksheet_name} 第 {row_number} 行包含公式；请改为确定值"
                        )
                    if cell.data_type == "e":
                        raise ExcelTransferError(
                            f"工作表 {worksheet_name} 第 {row_number} 行包含 Excel 错误值"
                        )
                normalized = [
                    _coerce_value(cell.value, column, row_number=row_number, table_name=actual_table)
                    for cell, column in zip(selected, columns)
                ]
                row_count += 1
                total_rows += 1
                if total_rows > IMPORT_MAX_ROWS:
                    raise ExcelTransferError(f"单次导入不能超过 {IMPORT_MAX_ROWS} 行")
                if include_rows:
                    rows.append(normalized)
                if len(preview) < 3:
                    preview.append([_preview_value(value) for value in normalized])
            if row_count == 0:
                raise ExcelTransferError(f"工作表 {worksheet_name} 没有可导入的数据行")
            plans.append({
                "sheet": worksheet_name,
                "table": actual_table,
                "columns": [column.name for column in columns],
                "rows": rows,
                "rowCount": row_count,
                "preview": preview,
            })
        digest_payload = [{
            "table": item["table"], "columns": item["columns"], "rowCount": item["rowCount"],
        } for item in plans]
        plan_digest = hashlib.sha256(json.dumps(
            digest_payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"),
        ).encode("utf-8")).hexdigest()
        return {"tables": plans, "rowCount": total_rows, "planDigest": plan_digest}
    finally:
        workbook.close()


def _sweep_pending(now: Optional[float] = None) -> None:
    current = time.time() if now is None else now
    expired = []
    with _LOCK:
        for confirm_id, item in list(_PENDING.items()):
            if float(item.get("expiresAt") or 0) <= current:
                expired.append(_PENDING.pop(confirm_id))
    for item in expired:
        try:
            Path(item["path"]).unlink(missing_ok=True)
        except OSError:
            pass


def prepare_import(
    path: str,
    *,
    db_id: str,
    database_ref: str,
    access_scope_ref: str,
    agent: Any,
) -> dict:
    source = Path(path).resolve()
    if source.suffix.lower() != ".xlsx" or not source.is_file():
        raise ExcelTransferError("只支持有效的 .xlsx 文件")
    if source.stat().st_size > IMPORT_MAX_FILE_BYTES:
        raise ExcelTransferError("Excel 导入文件不能超过 25MB")
    if agent.security.row_filters:
        raise ExcelTransferError("行级授权凭据不能执行批量 Excel 写入")
    if getattr(agent.connector, "dialect", "sqlite") != "sqlite" \
            and not getattr(agent.connector, "write_enabled", False):
        raise ExcelTransferError("远程数据库当前为只读连接，不能并入 Excel")
    summary = _parse_workbook(
        source,
        agent.schema,
        allowed_tables=agent.security.allowed_tables,
        allowed_columns=agent.security.allowed_columns,
        include_rows=False,
    )
    _sweep_pending()
    confirm_id = uuid.uuid4().hex
    expires_at = time.time() + IMPORT_TTL_SECONDS
    record = {
        "confirmId": confirm_id,
        "dbId": db_id,
        "databaseRef": database_ref,
        "accessScopeRef": access_scope_ref,
        "path": str(source),
        "fileSha256": _sha256_file(source),
        "schemaFingerprint": _schema_fingerprint(agent.schema),
        "planDigest": summary["planDigest"],
        "rowCount": summary["rowCount"],
        "tableCount": len(summary["tables"]),
        "expiresAt": expires_at,
    }
    with _LOCK:
        _PENDING[confirm_id] = record
    return {
        "confirmId": confirm_id,
        "expiresAt": expires_at,
        "rowCount": summary["rowCount"],
        "tableCount": len(summary["tables"]),
        "tables": [{key: value for key, value in item.items() if key != "rows"}
                   for item in summary["tables"]],
        "requiresAdmin": summary["rowCount"] > 100,
    }


def pending_import(confirm_id: str) -> Optional[dict]:
    _sweep_pending()
    with _LOCK:
        item = _PENDING.get(str(confirm_id or ""))
        return dict(item) if item is not None else None


def cancel_import(confirm_id: str) -> bool:
    with _LOCK:
        item = _PENDING.pop(str(confirm_id or ""), None)
    if item is None:
        return False
    try:
        Path(item["path"]).unlink(missing_ok=True)
    except OSError:
        pass
    return True


def execute_import(
    confirm_id: str,
    *,
    db_id: str,
    database_ref: str,
    access_scope_ref: str,
    agent: Any,
) -> dict:
    _sweep_pending()
    with _LOCK:
        record = _PENDING.pop(str(confirm_id or ""), None)
    if record is None:
        raise ExcelTransferError("Excel 导入确认单不存在或已过期")
    source = Path(record["path"])
    try:
        if record["dbId"] != db_id or record["databaseRef"] != database_ref \
                or record["accessScopeRef"] != access_scope_ref:
            raise ExcelTransferError("Excel 导入确认单与当前数据库或权限范围不匹配")
        if not source.is_file() or _sha256_file(source) != record["fileSha256"]:
            raise ExcelTransferError("Excel 文件在确认前发生变化")
        if _schema_fingerprint(agent.schema) != record["schemaFingerprint"]:
            raise ExcelTransferError("数据库结构在确认前发生变化，请重新预检")
        if agent.security.row_filters:
            raise ExcelTransferError("行级授权凭据不能执行批量 Excel 写入")
        parsed = _parse_workbook(
            source,
            agent.schema,
            allowed_tables=agent.security.allowed_tables,
            allowed_columns=agent.security.allowed_columns,
            include_rows=True,
        )
        if parsed["planDigest"] != record["planDigest"]:
            raise ExcelTransferError("Excel 导入计划在确认前发生变化")
        connector = agent.connector
        connection = None
        try:
            connection = connector.connect_rw()
            if getattr(connector, "dialect", "sqlite") == "sqlite":
                connection.execute("PRAGMA foreign_keys = ON")
            connector.begin_rw(connection)
            placeholder = "?" if getattr(connector, "dialect", "sqlite") == "sqlite" else "%s"
            for table in parsed["tables"]:
                quote = connector.quote_identifier
                columns_sql = ", ".join(quote(column) for column in table["columns"])
                values_sql = ", ".join(placeholder for _ in table["columns"])
                sql = f"INSERT INTO {quote(table['table'])} ({columns_sql}) VALUES ({values_sql})"
                cursor = connection.cursor()
                try:
                    cursor.executemany(sql, table["rows"])
                finally:
                    cursor.close()
            connector.commit_rw(connection)
        except Exception as exc:
            if connection is not None:
                try:
                    connector.rollback_rw(connection)
                except Exception:
                    pass
            raise ExcelTransferError(f"Excel 导入失败，全部变更已回滚：{exc}") from exc
        finally:
            if connection is not None:
                connector.close(connection)
        return {
            "affected": parsed["rowCount"],
            "tableCount": len(parsed["tables"]),
            "tables": [{"table": item["table"], "rows": item["rowCount"]}
                       for item in parsed["tables"]],
        }
    finally:
        try:
            source.unlink(missing_ok=True)
        except OSError:
            pass


def _safe_sheet_title(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", str(name or "table")).strip(" '")[:31] or "table"
    candidate = base
    index = 2
    while candidate.casefold() in used or candidate == MANIFEST_SHEET:
        suffix = f"_{index}"
        candidate = base[:31 - len(suffix)] + suffix
        index += 1
    used.add(candidate.casefold())
    return candidate


def _excel_cell(worksheet: Any, value: Any) -> Any:
    from openpyxl.cell import WriteOnlyCell
    if isinstance(value, (bytes, bytearray, memoryview)):
        value = "dbquill:base64:" + base64.b64encode(bytes(value)).decode("ascii")
    elif isinstance(value, Decimal):
        value = str(value)
    elif isinstance(value, (dict, list, tuple, set)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    elif value is not None and not isinstance(value, (str, int, float, bool, date, datetime, dt_time)):
        value = str(value)
    cell = WriteOnlyCell(worksheet, value=value)
    if isinstance(value, str):
        # Database text beginning with '=' must stay text, not become a formula.
        cell.data_type = "s"
    return cell


def _sweep_exports(directory: Path) -> None:
    cutoff = time.time() - 24 * 60 * 60
    for path in directory.glob("dbquill-export-*.xlsx"):
        try:
            if path.stat().st_mtime < cutoff:
                path.unlink()
        except OSError:
            pass


def export_workbook(output_dir: str, *, agent: Any) -> dict:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ExcelTransferError("缺少 openpyxl，无法导出 Excel") from exc
    tables = list(agent.schema.tables.items())
    if not tables:
        raise ExcelTransferError("当前权限范围内没有可导出的表")
    if len(tables) > EXPORT_MAX_TABLES:
        raise ExcelTransferError(f"一次最多导出 {EXPORT_MAX_TABLES} 张表")
    destination = Path(output_dir).resolve()
    destination.mkdir(parents=True, exist_ok=True)
    _sweep_exports(destination)
    output = destination / f"dbquill-export-{uuid.uuid4().hex[:12]}.xlsx"
    workbook = Workbook(write_only=True)
    connector = agent.connector
    connection = None
    total_rows = 0
    manifest_rows = []
    used_titles: set[str] = set()
    try:
        connection = connector.connect()
        if getattr(connector, "dialect", "sqlite") == "sqlite":
            import dbquill_core as core
            row_internal = core._prepare_sqlite_row_views(
                connection, agent.security.row_filters, agent.security.allowed_columns,
            ) if agent.security.row_filters else {}
            core._install_sqlite_scope_authorizer(
                connection,
                allowed_tables=agent.security.allowed_tables,
                allowed_columns=agent.security.allowed_columns,
                row_internal_columns=row_internal,
                allow_writes=False,
                unavailable_error="当前 SQLite 运行时无法安全导出授权数据",
            )
        elif agent.security.allowed_tables is not None or agent.security.allowed_columns \
                or agent.security.row_filters:
            raise ExcelTransferError("远程数据库暂不支持受限凭据的 Excel 导出")
        for table_name, table in tables:
            title = _safe_sheet_title(table_name, used_titles)
            manifest_rows.append((title, table_name))
            worksheet = workbook.create_sheet(title=title)
            columns = [column.name for column in table.columns]
            worksheet.append([_excel_cell(worksheet, column) for column in columns])
            quote = connector.quote_identifier
            sql = (
                "SELECT " + ", ".join(quote(column) for column in columns)
                + " FROM " + quote(table_name)
            )
            cursor = connection.cursor()
            try:
                cursor.execute(sql)
                while True:
                    rows = cursor.fetchmany(500)
                    if not rows:
                        break
                    total_rows += len(rows)
                    if total_rows > EXPORT_MAX_ROWS:
                        raise ExcelTransferError(
                            f"导出超过 {EXPORT_MAX_ROWS} 行上限；未生成不完整文件"
                        )
                    for row in rows:
                        worksheet.append([_excel_cell(worksheet, value) for value in row])
            finally:
                cursor.close()
        manifest = workbook.create_sheet(title=MANIFEST_SHEET)
        manifest.sheet_state = "hidden"
        manifest.append([_excel_cell(manifest, MANIFEST_FORMAT), _excel_cell(manifest, "table")])
        for sheet_name, table_name in manifest_rows:
            manifest.append([_excel_cell(manifest, sheet_name), _excel_cell(manifest, table_name)])
        workbook.save(output)
        return {"path": str(output), "rowCount": total_rows, "tableCount": len(tables)}
    except Exception:
        try:
            output.unlink(missing_ok=True)
        except OSError:
            pass
        raise
    finally:
        if connection is not None:
            connector.close(connection)
        try:
            workbook.close()
        except Exception:
            pass
